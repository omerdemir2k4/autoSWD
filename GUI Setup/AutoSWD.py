#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Enhanced Predictor for EEG SWD using Hybrid CNN + 1DConv Token-Level Classifier
==============================================================================

This script:
  - Loads a 2-channel model from an H5 file.
  - Loads a 2-channel EDF file, prompts for start time.
  - Divides both channels into overlapping intervals.
  - Computes CWT spectrograms for each channel, stacks into (N,32,32,2).
  - Applies median-IQR normalization per interval and channel.
  - Groups intervals into sequences for temporal context.
  - Runs the model to produce token-level predictions.
  - Merges, refines, and displays SWD events in an interactive PyQt5 GUI.
  
All original GUI options, editing tools, navigation, and export features are retained.
"""

import sys
import os
import re
import gc
import hashlib
import time
import numpy as np
import pandas as pd
import matplotlib

# Mac OS Optimization: Use Qt5Agg backend specifically to avoid Bus Error 10
if sys.platform == 'darwin':
    matplotlib.use('Qt5Agg')
    # Fix for missing font family warning on Mac
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'DejaVu Sans']

import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
from datetime import datetime, timedelta
from scipy.ndimage import zoom
import mne
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QSlider, QPushButton, QMessageBox,
    QFileDialog, QInputDialog, QDoubleSpinBox,
    QSpinBox, QCheckBox, QGroupBox, QProgressBar,
    QComboBox, QDialog, QListWidget, QListWidgetItem,
    QSplitter, QFrame, QShortcut, QTabWidget, QToolBar,
    QStatusBar, QTableWidget, QTableWidgetItem, QHeaderView,
    QLineEdit, QScrollArea, QGridLayout, QAction, QMenu, QSizePolicy,
    QTimeEdit, QAbstractItemView
)
from PyQt5.QtCore import Qt, pyqtSignal, QTimer, QEvent, QObject, QThread, QPropertyAnimation, QEasingCurve, QRect, QPoint, QSize, QTime
from PyQt5.QtGui import QKeySequence, QPalette, QColor, QPainter, QFont, QIcon, QPen, QCursor
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from tensorflow.keras.models import load_model
import concurrent.futures
from multiprocessing import cpu_count
from scipy.signal import welch
from scipy.signal import cwt
from scipy.signal.windows import dpss
from scipy.signal import periodogram
from scipy.signal import resample, resample_poly
import tensorflow as tf
import json
import importlib
import traceback
import h5py
import zipfile
import tempfile
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import ctypes
from ctypes import wintypes
from joblib import Parallel, delayed
import multiprocessing
try:
    import keras as _keras
    _HAS_KERAS3 = hasattr(_keras, 'config') and hasattr(_keras.config, 'enable_unsafe_deserialization')
except Exception:
    _keras = None
    _HAS_KERAS3 = False
try:
    from keras.saving import register_keras_serializable
except Exception:
    try:
        from tensorflow.keras.utils import register_keras_serializable
    except Exception:
        def register_keras_serializable(*_args, **_kwargs):
            def _decorator(cls):
                return cls
            return _decorator
try:
    import torch
except Exception:
    torch = None
# -----------------------------------------------------------------------------
# NumPy compatibility shim (for NumPy 2.x deprecated aliases)
# -----------------------------------------------------------------------------
try:
    # Accessing will raise AttributeError on NumPy 2.x
    _ = np.bool
except AttributeError:  # NumPy 2.x: restore deprecated aliases used by legacy code/models
    np.bool = bool
    np.int = int
    np.float = float
    np.object = object
    np.complex = complex

# -----------------------------------------------------------------------------
# PyTorch <> NumPy compatibility guard
#   Older PyTorch builds can misbehave with NumPy 2.x (warning: _ARRAY_API not found)
#   If detected, disable UNET preloading and notify the user at runtime.
# -----------------------------------------------------------------------------
def _parse_version_tuple(ver_str):
    parts = re.split(r"[^0-9]+", ver_str)
    nums = [int(p) for p in parts if p.isdigit()]
    while len(nums) < 3:
        nums.append(0)
    return tuple(nums[:3])

_NUMPY_VER = _parse_version_tuple(np.__version__)
_TORCH_VER = _parse_version_tuple(getattr(torch, "__version__", "0.0.0")) if torch is not None else (0, 0, 0)

_NUMPY_MAJOR = _NUMPY_VER[0]
_TORCH_MAJOR_MINOR = _TORCH_VER[:2]

_UNET_TORCH_DISABLED = False
if (_NUMPY_MAJOR >= 2) and (torch is not None) and (_TORCH_MAJOR_MINOR < (2, 3)):
    # Known problematic combo: NumPy >=2 with torch < 2.3
    _UNET_TORCH_DISABLED = True
    print("Warning: Disabling UNET (PyTorch) due to incompatible versions: NumPy", np.__version__, "with PyTorch", torch.__version__)
    torch = None
# -----------------------------------------------------------------------------
# Keras compatibility shim
# -----------------------------------------------------------------------------

class SafeReshape(tf.keras.layers.Reshape):
    """Reshape that coerces any string dims (e.g., '100') to integers.
    Some legacy H5 models may serialize target_shape with string numbers,
    which can error in newer Keras when deserializing or calling layers.
    """
    def __init__(self, target_shape, **kwargs):
        def _coerce_dim(d):
            if isinstance(d, str):
                try:
                    return int(d)
                except Exception:
                    return d
            return d

        if isinstance(target_shape, (list, tuple)):
            target_shape = tuple(_coerce_dim(d) for d in target_shape)
        elif isinstance(target_shape, str):
            try:
                target_shape = (int(target_shape),)
            except Exception:
                target_shape = (target_shape,)
        super().__init__(target_shape=target_shape, **kwargs)

    def get_config(self):
        cfg = super().get_config()
        def _coerce_dim(d):
            if isinstance(d, str):
                try:
                    return int(d)
                except Exception:
                    return d
            return d
        cfg['target_shape'] = tuple(_coerce_dim(d) for d in cfg.get('target_shape', ()))
        return cfg

@register_keras_serializable(package="custom")
class ChannelSlice(tf.keras.layers.Layer):
    """Serializable channel slicer compatible with Keras 3 deserialization."""
    def __init__(self, index, **kwargs):
        super().__init__(**kwargs)
        self.index = int(index)

    def call(self, x):
        return x[..., self.index:self.index+1]

    def compute_output_shape(self, input_shape):
        try:
            return input_shape[:-1] + (1,)
        except Exception:
            return input_shape

    def get_config(self):
        cfg = super().get_config()
        cfg.update({"index": self.index})
        return cfg

class SavedModelPredictor:
    """Lightweight wrapper to use a TF SavedModel with a Keras-like predict().

    Expects a directory produced by Keras 3 model.export(). The signature key
    is typically 'serve' but we also try 'serving_default'."""
    def __init__(self, saved_model_dir: str):
        self.saved_model_dir = saved_model_dir
        obj = tf.saved_model.load(saved_model_dir)
        serve = None
        try:
            serve = obj.signatures.get('serve', None)
        except Exception:
            serve = None
        if serve is None:
            try:
                serve = obj.signatures.get('serving_default', None)
            except Exception:
                serve = None
        if serve is None:
            raise RuntimeError("SavedModel does not contain a 'serve' signature")
        self._fn = serve

    def predict(self, x, verbose=0):
        xt = tf.convert_to_tensor(x, dtype=tf.float32)
        outputs = self._fn(xt)
        # Take the first tensor from the outputs dict
        first = next(iter(outputs.values()))
        return first.numpy()

def _enable_unsafe_deserialization():
    if _keras is not None and _HAS_KERAS3:
        try:
            _keras.config.enable_unsafe_deserialization()
        except Exception:
            pass

def _disable_unsafe_deserialization():
    if _keras is not None and _HAS_KERAS3:
        try:
            _keras.config.disable_unsafe_deserialization()
        except Exception:
            pass

def _coerce_numeric_strings(obj):
    """Recursively convert strings like '100' or '32.0' to int/float in JSON-like structures."""
    if isinstance(obj, dict):
        return {k: _coerce_numeric_strings(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_coerce_numeric_strings(v) for v in obj]
    if isinstance(obj, tuple):
        return tuple(_coerce_numeric_strings(v) for v in obj)
    if isinstance(obj, str):
        s = obj.strip()
        # Try int
        if s.isdigit() or (s.startswith('-') and s[1:].isdigit()):
            try:
                return int(s)
            except Exception:
                return obj
        # Try float
        try:
            # Avoid converting non-numeric like 'None'
            if s.lower() in ('none', 'nan', 'inf', '-inf'):
                return obj
            f = float(s)
            # If it is integral like '32.0', cast to int
            if abs(f - int(f)) < 1e-9:
                return int(f)
            return f
        except Exception:
            return obj
    return obj

def _sanitize_timedistributed_configs(obj):
    """Specifically fix TimeDistributed layer configs that have string shapes."""
    if isinstance(obj, dict):
        # Fix TimeDistributed layer configs
        if obj.get('class_name') == 'TimeDistributed':
            config = obj.get('config', {})
            if 'layer' in config and isinstance(config['layer'], dict):
                layer_config = config['layer'].get('config', {})
                # Fix input_shape in nested layer configs
                if 'input_shape' in layer_config:
                    layer_config['input_shape'] = _coerce_numeric_strings(layer_config['input_shape'])
                # Fix batch_shape in nested layer configs
                if 'batch_shape' in layer_config:
                    layer_config['batch_shape'] = _coerce_numeric_strings(layer_config['batch_shape'])
                # Fix any other shape-related fields
                for key in ['target_shape', 'output_shape', 'shape']:
                    if key in layer_config:
                        layer_config[key] = _coerce_numeric_strings(layer_config[key])
            # Also fix the TimeDistributed layer's own config
            for key in ['input_shape', 'batch_shape', 'target_shape', 'output_shape', 'shape']:
                if key in config:
                    config[key] = _coerce_numeric_strings(config[key])
        # Fix any layer with shape-related fields
        elif 'config' in obj and isinstance(obj['config'], dict):
            config = obj['config']
            for key in ['input_shape', 'batch_shape', 'target_shape', 'output_shape', 'shape']:
                if key in config:
                    config[key] = _coerce_numeric_strings(config[key])
        # Recursively process all nested structures
        return {k: _sanitize_timedistributed_configs(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_sanitize_timedistributed_configs(v) for v in obj]
    elif isinstance(obj, tuple):
        return tuple(_sanitize_timedistributed_configs(v) for v in obj)
    else:
        return obj

def load_model_with_sanitization(model_path):
    """Fallback loader that sanitizes JSON config numeric strings and reloads weights.

    This addresses errors like "Cannot convert '100' to a shape" from legacy H5 models.
    """
    with h5py.File(model_path, 'r') as f:
        # Obtain model config as JSON string
        cfg_bytes = None
        if 'model_config' in f.attrs:
            cfg_bytes = f.attrs['model_config']
        elif 'model_config' in f:
            cfg_bytes = f['model_config'][()]
        if cfg_bytes is None:
            raise ValueError('No model_config found in H5 file')
        if isinstance(cfg_bytes, bytes):
            cfg_str = cfg_bytes.decode('utf-8')
        elif isinstance(cfg_bytes, str):
            cfg_str = cfg_bytes
        else:
            cfg_str = cfg_bytes.tobytes().decode('utf-8')

    cfg_json = json.loads(cfg_str)
    cfg_json = _coerce_numeric_strings(cfg_json)
    cfg_json = _sanitize_timedistributed_configs(cfg_json)
    sanitized_cfg_str = json.dumps(cfg_json)

    # Reconstruct model from sanitized config and load weights
    custom_objects = {
        'SafeReshape': SafeReshape,
        'Reshape': SafeReshape,  # override builtin Reshape if target_shape has strings
        'ChannelSlice': ChannelSlice,
        # Map legacy 'Functional' identifier to Keras Model for deserialization
        'Functional': tf.keras.models.Model,
    }
    _enable_unsafe_deserialization()
    try:
        model = tf.keras.models.model_from_json(sanitized_cfg_str, custom_objects=custom_objects)
        model.load_weights(model_path)
    finally:
        _disable_unsafe_deserialization()
    return model

def load_keras_with_sanitization(keras_path):
    """Sanitize and load a Keras v3 .keras archive by fixing numeric strings in config.

    This extracts the JSON config from the zip, coerces numeric strings to numbers,
    recreates the model, then loads the internal H5 weights file.
    """
    with zipfile.ZipFile(keras_path, 'r') as zf:
        # Find config JSON
        cfg_name = None
        for name in zf.namelist():
            if name.endswith('config.json'):
                cfg_name = name
                break
        if cfg_name is None:
            raise ValueError('config.json not found inside .keras archive')
        cfg_bytes = zf.read(cfg_name)
        cfg_str = cfg_bytes.decode('utf-8')

        # Find weights file
        weights_name = None
        for name in zf.namelist():
            if name.endswith('.h5') or name.endswith('.hdf5'):
                weights_name = name
                break
        if weights_name is None:
            raise ValueError('weights file (.h5) not found inside .keras archive')

        # Sanitize config
        cfg_json = json.loads(cfg_str)
        cfg_json = _coerce_numeric_strings(cfg_json)
        cfg_json = _sanitize_timedistributed_configs(cfg_json)
        sanitized_cfg_str = json.dumps(cfg_json)

        # Build model and load weights
        custom_objects = {
            'SafeReshape': SafeReshape,
            'Reshape': SafeReshape,
            'ChannelSlice': ChannelSlice,
            'Functional': tf.keras.models.Model,
        }
        _enable_unsafe_deserialization()
        try:
            model = tf.keras.models.model_from_json(sanitized_cfg_str, custom_objects=custom_objects)
            # Extract weights to a temp file
            with tempfile.NamedTemporaryFile(suffix='.h5', delete=True) as tmp:
                tmp.write(zf.read(weights_name))
                tmp.flush()
                model.load_weights(tmp.name)
        finally:
            _disable_unsafe_deserialization()
    return model
# -----------------------------------------------------------------------------
# Helper Functions
# -----------------------------------------------------------------------------

def robust_parse_hms_to_seconds(time_str: str) -> float:
    """Robustly parse time string in HH:MM:SS format to seconds."""
    try:
        # Try standard format first
        h, m, s = map(int, time_str.split(':'))
        return h * 3600 + m * 60 + s
    except ValueError:
        try:
            # Try with decimal seconds
            h, m, s = time_str.split(':')
            s, ms = s.split('.')
            return int(h) * 3600 + int(m) * 60 + int(s) + float('0.' + ms)
        except ValueError:
            # If all parsing fails, return 0
            return 0

def divide_into_intervals(data, sfreq, interval_length, overlap_length):
    """
    Divide a 1D signal array into overlapping intervals - optimized version using strides.
    Returns array of shape (n_intervals, window_size).
    """
    step = int((interval_length - overlap_length) * sfreq)
    window_size = int(interval_length * sfreq)
    
    # Use numpy's stride tricks for efficient windowing
    n_intervals = (len(data) - window_size) // step + 1
    if n_intervals <= 0:
        return np.empty((0, window_size))
        
    shape = (n_intervals, window_size)
    strides = (data.strides[0] * step, data.strides[0])
    
    # Create strided view (no data copying)
    intervals = np.lib.stride_tricks.as_strided(data, shape=shape, strides=strides)
    
    # Return a copy to avoid memory issues during serialization
    return intervals.copy()

def compute_cwt_features(interval, sfreq,
                         target_freq_bins=32, target_time_bins=32,
                         fmin=6.0, fmax=30.0, n_cycles=2.0):
    """
    Compute Morlet CWT power spectrogram, convert to dB, resize to (freq_bins, time_bins).
    """
    freqs = np.linspace(fmin, fmax, target_freq_bins)
    n_cycles_arr = np.full_like(freqs, n_cycles)
    arr = interval[np.newaxis, np.newaxis, :]
    power = mne.time_frequency.tfr_array_morlet(
        arr, sfreq=sfreq, freqs=freqs, n_cycles=n_cycles_arr,
        output='power', decim=1, n_jobs=1, verbose=False
    )[0,0]
    power_db = 10 * np.log10(power + np.finfo(float).eps)
    if power_db.shape[1] != target_time_bins:
        power_db = zoom(power_db, (1, target_time_bins / power_db.shape[1]), order=1)
    return power_db

def compute_multitaper_psd(signal, sfreq, fmin=1.0, fmax=50.0, bandwidth=2, normalize=False):
    """
    Compute multitaper PSD using MNE's psd_array_multitaper implementation.
    """
    from mne.time_frequency import psd_array_multitaper
    
    if signal.ndim == 1:
        signal = signal[np.newaxis, :]
    
    psds, freqs = psd_array_multitaper(
        signal, 
        sfreq=sfreq, 
        fmin=fmin, 
        fmax=fmax, 
        bandwidth=bandwidth,
        adaptive=True,
        normalization='full',
        verbose=False
    )
    
    psd = psds[0] if psds.ndim > 1 else psds
    
    if normalize:
        total = np.trapz(psd, freqs)
        if total > 0:
            psd = psd / total
    
    return freqs, psd

_HW_CONFIG_CACHE = None


def assess_hardware(force=False):
    """Assess the host machine and derive an optimized CWT-preprocessing config.

    The CWT feature extraction (Morlet transform over thousands of short
    intervals) is the heaviest part of the pipeline. Its runtime depends almost
    entirely on how much parallelism the machine can sustain and how much memory
    is available to hold vectorized batches. Rather than hard-coding a single
    worker/batch count that is either wasteful on strong machines or crippling
    on weak ones, we probe the hardware once and pick a tier-appropriate config.

    Returns a dict with keys:
      - 'tier':        'low' | 'medium' | 'high'
      - 'n_workers':   joblib parallel workers (1 -> run serially, no spawn cost)
      - 'batch_size':  intervals per vectorized MNE chunk
      - 'backend':     joblib backend ('loky' processes or 'threading')
      - 'cpu_count', 'physical_cores', 'ram_gb', 'has_gpu'
    The result is cached after the first call (pass force=True to recompute).
    """
    global _HW_CONFIG_CACHE
    if _HW_CONFIG_CACHE is not None and not force:
        return _HW_CONFIG_CACHE

    # --- Logical CPU cores ---
    try:
        logical = int(multiprocessing.cpu_count())
    except Exception:
        logical = 2

    # --- Physical cores (best effort) ---
    physical = None
    try:
        import psutil
        physical = psutil.cpu_count(logical=False)
    except Exception:
        physical = None
    if not physical:
        physical = max(1, logical // 2)

    # --- Total RAM in GB (best effort) ---
    ram_gb = None
    try:
        import psutil
        ram_gb = psutil.virtual_memory().total / (1024 ** 3)
    except Exception:
        try:
            if hasattr(os, 'sysconf') and 'SC_PHYS_PAGES' in os.sysconf_names:
                ram_gb = (os.sysconf('SC_PHYS_PAGES') * os.sysconf('SC_PAGE_SIZE')) / (1024 ** 3)
        except Exception:
            ram_gb = None
    if ram_gb is None and sys.platform == 'win32':
        # Windows fallback (no psutil): GlobalMemoryStatusEx.
        try:
            class _MEMORYSTATUSEX(ctypes.Structure):
                _fields_ = [
                    ('dwLength', ctypes.c_ulong),
                    ('dwMemoryLoad', ctypes.c_ulong),
                    ('ullTotalPhys', ctypes.c_ulonglong),
                    ('ullAvailPhys', ctypes.c_ulonglong),
                    ('ullTotalPageFile', ctypes.c_ulonglong),
                    ('ullAvailPageFile', ctypes.c_ulonglong),
                    ('ullTotalVirtual', ctypes.c_ulonglong),
                    ('ullAvailVirtual', ctypes.c_ulonglong),
                    ('ullAvailExtendedVirtual', ctypes.c_ulonglong),
                ]
            stat = _MEMORYSTATUSEX()
            stat.dwLength = ctypes.sizeof(_MEMORYSTATUSEX)
            if ctypes.windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(stat)):
                ram_gb = stat.ullTotalPhys / (1024 ** 3)
        except Exception:
            ram_gb = None
    if ram_gb is None:
        ram_gb = 8.0  # conservative default when detection fails

    # --- GPU presence (TensorFlow or Torch) ---
    has_gpu = False
    try:
        has_gpu = bool(tf.config.list_physical_devices('GPU'))
    except Exception:
        has_gpu = False
    if not has_gpu and torch is not None:
        try:
            has_gpu = bool(torch.cuda.is_available())
        except Exception:
            pass

    # --- Tier classification ---
    if logical >= 8 and ram_gb >= 12:
        tier = 'high'
    elif logical >= 4 and ram_gb >= 6:
        tier = 'medium'
    else:
        tier = 'low'

    if tier == 'high':
        # Strong machine: use most cores and large vectorized batches.
        n_workers = max(1, min(logical - 1, 16))
        batch_size = 96
        backend = 'loky'
    elif tier == 'medium':
        # Typical laptop/desktop: balanced parallelism.
        n_workers = max(1, min(logical - 1, 6))
        batch_size = 64
        backend = 'loky'
    else:
        # Weak machine: process-spawn overhead and memory copies dominate, so
        # keep parallelism minimal and batches small. On the very weakest
        # configs run serially with threads to avoid loky's process cost.
        if logical <= 2 or ram_gb < 4:
            n_workers = 1
            backend = 'threading'
        else:
            n_workers = 2
            backend = 'loky'
        batch_size = 32

    _HW_CONFIG_CACHE = {
        'tier': tier,
        'n_workers': int(n_workers),
        'batch_size': int(batch_size),
        'backend': backend,
        'cpu_count': int(logical),
        'physical_cores': int(physical),
        'ram_gb': float(round(ram_gb, 1)),
        'has_gpu': bool(has_gpu),
    }
    try:
        print(f"[Hardware] tier={tier} cores={logical} ram={ram_gb:.1f}GB gpu={has_gpu} "
              f"-> workers={n_workers} batch={batch_size} backend={backend}")
    except Exception:
        pass
    return _HW_CONFIG_CACHE


def compute_cwt_features_batch(intervals_pairs, sfreq, n_workers=None,
                               batch_size=None, backend=None):
    """
    Compute CWT features for multiple interval pairs in parallel using joblib.
    Optimized version: processes data in chunks using vectorized MNE operations for significant speedup.

    Worker count, chunk size and joblib backend default to a hardware-aware
    configuration (see assess_hardware) so weak machines avoid parallel overhead
    while strong machines exploit all available cores.
    """
    hw = assess_hardware()
    if n_workers is None:
        n_workers = hw['n_workers']
    if batch_size is None:
        batch_size = hw['batch_size']
    if backend is None:
        backend = hw['backend']

    # Split into chunks
    chunks = [intervals_pairs[i:i + batch_size] for i in range(0, len(intervals_pairs), batch_size)]
    
    def process_chunk(chunk):
        if not chunk:
            return None
            
        # Stack chunk into (Batch, 2, Time)
        # iv0 and iv1 are 1D arrays of shape (T,)
        # chunk is list of tuples (iv0, iv1)
        data = np.stack([np.stack(pair) for pair in chunk]) # Shape: (B, 2, T)
        
        # MNE CWT parameters
        freqs = np.linspace(6.0, 30.0, 32)
        n_cycles = np.full_like(freqs, 2.0)
        
        # Compute CWT: returns (Batch, 2, Freqs, Times)
        # tfr_array_morlet is vectorized over the first dim (epochs)
        power = mne.time_frequency.tfr_array_morlet(
            data, sfreq=sfreq, freqs=freqs, n_cycles=n_cycles,
            output='power', decim=1, n_jobs=1, verbose=False
        )
        
        # Log transform (avoid log(0))
        power = 10 * np.log10(power + np.finfo(float).eps)
        
        # Resize to target shape (32, 32)
        # Current shape: (B, 2, 32, T_in)
        # Target shape: (B, 2, 32, 32)
        # We only need to resize the last dimension (Time)
        target_time = 32
        current_time = power.shape[-1]
        zoom_factor = target_time / current_time
        
        # Apply zoom to the last axis
        # zoom(input, zoom_factors) -> factors match input rank
        # (B, 2, F, T) -> zoom=(1, 1, 1, factor)
        power_resized = zoom(power, (1, 1, 1, zoom_factor), order=1)
        
        # Normalization per spectrogram (Median-IQR)
        # Normalize across (Freq, Time) dimensions (last two axes)
        # shape is (B, 2, 32, 32)
        m = np.median(power_resized, axis=(2,3), keepdims=True)
        q75 = np.percentile(power_resized, 75, axis=(2,3), keepdims=True)
        q25 = np.percentile(power_resized, 25, axis=(2,3), keepdims=True)
        iqr = np.maximum(q75 - q25, 1e-6)
        
        norm_power = (power_resized - m) / iqr
        
        # Transpose to model input format: (Batch, 32, 32, 2)
        # Current: (Batch, 2, 32, 32) -> (0, 1, 2, 3)
        # Target: (0, 2, 3, 1)
        return np.transpose(norm_power, (0, 2, 3, 1))

    # Process chunks. On weak hardware n_workers collapses to 1, in which case
    # we run serially to avoid joblib's per-task dispatch and process overhead.
    if n_workers <= 1:
        results = [process_chunk(chunk) for chunk in chunks]
    else:
        results = Parallel(n_jobs=n_workers, backend=backend)(
            delayed(process_chunk)(chunk) for chunk in chunks
        )
    
    # Concatenate results
    valid_results = [r for r in results if r is not None]
    if not valid_results:
        return np.empty((0, 32, 32, 2), dtype=np.float32)
        
    return np.concatenate(valid_results, axis=0).astype(np.float32)

def process_interval_batch(batch, sfreq):
    """
    Legacy batch processor - deprecated in favor of joblib pipeline.
    """
    n = len(batch)
    out = np.zeros((n, 32, 32, 2), dtype=np.float32)
    for i, (iv0, iv1) in enumerate(batch):
        c0 = compute_cwt_features(iv0, sfreq)
        c1 = compute_cwt_features(iv1, sfreq)
        # median-IQR normalization per channel
        def norm(x):
            m = np.median(x)
            q75, q25 = np.percentile(x, [75,25])
            iqr = max(q75-q25, 1e-6)
            return (x - m) / iqr
        n0 = norm(c0)
        n1 = norm(c1)
        out[i,...,0] = n0
        out[i,...,1] = n1
    return out

def process_intervals_parallel(intervals_pairs, sfreq, progress_callback=None):
    """
    Process all interval pairs in parallel using joblib (replaced ProcessPoolExecutor).
    Ignores progress_callback for the computation phase as joblib handles it internally/atomically.
    """
    # Simply delegate to the robust joblib function
    return compute_cwt_features_batch(intervals_pairs, sfreq)


def group_intervals_with_overlap(intervals, sequence_length, overlap_ratio=0.0):
    """
    Slide a window of size sequence_length over intervals (no overlap if overlap_ratio=0).
    Returns np.array of shape (n_sequences, seq_len, 32,32,2).
    """
    stride = max(1, int(sequence_length * (1-overlap_ratio)))
    sequences = []
    for i in range(0, len(intervals)-sequence_length+1, stride):
        sequences.append(intervals[i:i+sequence_length])
    return np.array(sequences)

def merge_adjacent_swds(token_preds, threshold=0.5, max_gap_intervals=3, require_overlap_agreement=True):
    """
    Merge token-level predictions into SWD regions based on threshold and gap allowance.
    
    If require_overlap_agreement=True, count overlapping regions (50% overlap) as SWD 
    only when both overlapping tokens predict SWD. This ensures the overlapping portion 
    between consecutive tokens is only counted when both predictions agree.
    """
    binary = (token_preds >= threshold).astype(int)
    
    if require_overlap_agreement:
        # For overlap-aware merging: mark regions where overlapping tokens both predict SWD
        # With 50% overlap, token i and token i+1 overlap for 50% of their duration.
        # Count a token as SWD only if both it and the next overlapping token predict SWD.
        # This ensures the overlapping region (0.3s out of 0.6s) is only counted when both agree.
        overlap_binary = np.zeros_like(binary)
        for i in range(len(binary)):
            if i == len(binary) - 1:
                # Last token: check if it and previous token both predict SWD
                if binary[i] == 1 and binary[i-1] == 1:
                    overlap_binary[i] = 1
            else:
                # For token i: count as SWD if both token i and token i+1 predict SWD
                # This marks the overlapping region between consecutive tokens (50% overlap)
                if binary[i] == 1 and binary[i+1] == 1:
                    overlap_binary[i] = 1
        binary = overlap_binary
    
    regions = []
    in_swd = False
    for i, b in enumerate(binary):
        if b and not in_swd:
            in_swd = True
            start = i
        elif not b and in_swd:
            in_swd = False
            regions.append({'start_idx': start, 'end_idx': i-1})
    if in_swd:
        regions.append({'start_idx': start, 'end_idx': len(binary)-1})
    # merge small gaps
    merged = [regions[0]] if regions else []
    for r in regions[1:]:
        prev = merged[-1]
        if r['start_idx'] - prev['end_idx'] <= max_gap_intervals:
            prev['end_idx'] = r['end_idx']
        else:
            merged.append(r)
    return merged

import pandas as pd
from datetime import timedelta

def export_to_excel(events, sfreq, recording_start, filename):
    """
    Export list of {'start_sample','end_sample'} to Excel with EXACT textual format:
      - Start Time, End Time: HH:MM:SS,mmm  (comma before milliseconds, 3 digits)
      - Duration:             xx,yy         (comma decimal, 2 decimals)
    The cells are forced to Text format so Excel won't auto-reformat.
    """
    rows = []
    for ev in events:
        st = recording_start + timedelta(seconds=ev['start_sample'] / sfreq)
        et = recording_start + timedelta(seconds=ev['end_sample'] / sfreq)
        dur = (ev['end_sample'] - ev['start_sample']) / sfreq

        # format milliseconds correctly as 3 digits
        st_ms = f"{int(st.microsecond // 1000):03d}"
        et_ms = f"{int(et.microsecond // 1000):03d}"

        start_time = f"{st:%H:%M:%S},{st_ms}"
        end_time   = f"{et:%H:%M:%S},{et_ms}"

        # duration with comma as decimal separator, 2 decimals
        duration = f"{dur:.2f}".replace('.', ',')

        rows.append({
            'Start Time': start_time,
            'End Time': end_time,
            'Duration': duration
        })

    df = pd.DataFrame(rows, columns=['Start Time', 'End Time', 'Duration'])

    # Write as text, set column widths, and force Text format (@) to prevent Excel auto-parsing
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='SWD_Analysis')

        ws = writer.sheets['SWD_Analysis']
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12

        # Force Text format on all data cells
        for col in ('A', 'B', 'C'):
            for row in range(2, len(df) + 2):  # header is row 1
                cell = ws[f'{col}{row}']
                cell.number_format = '@'  # Text

    return df


from scipy.signal import butter, filtfilt
import numpy as np
from PyQt5.QtCore import pyqtSignal

class PreprocessWorker(QObject):
    progress = pyqtSignal(int, int, float, str)  # done, total, eta, phase
    finished = pyqtSignal(object)
    error = pyqtSignal(str)

    def __init__(self, eeg0, eeg1, sfreq, interval_length, overlap_length, sequence_length, model, power_percentile,
                 require_overlap_agreement=True):
        super().__init__()
        self.eeg0 = eeg0
        self.eeg1 = eeg1
        self.sfreq = sfreq
        self.interval_length = interval_length
        self.overlap_length = overlap_length
        self.sequence_length = sequence_length
        self.model = model
        self.power_percentile = power_percentile
        self.require_overlap_agreement = bool(require_overlap_agreement)

    def run(self):
        try:
            # Features phase
            intervals0 = divide_into_intervals(self.eeg0, self.sfreq, self.interval_length, self.overlap_length)
            intervals1 = divide_into_intervals(self.eeg1, self.sfreq, self.interval_length, self.overlap_length)
            intervals_pairs = list(zip(intervals0, intervals1))

            def progress_callback(done, total, eta):
                self.progress.emit(done, total, eta if eta is not None else 0.0, 'features')

            features = process_intervals_parallel(intervals_pairs, self.sfreq, progress_callback)
            sequences = group_intervals_with_overlap(features, self.sequence_length)

            # Predictions phase (map to last 20%)
            total_batches = max(1, len(sequences))
            # We simulate progress by batches, but Keras predict returns once; emit start and finish
            self.progress.emit(0, total_batches, 0.0, 'predictions')
            token_preds = self.model.predict(sequences, verbose=0).reshape(-1)
            self.progress.emit(total_batches, total_batches, 0.0, 'predictions')

            # Merge and refine
            # Overlap agreement (both overlapping tokens must predict SWD).
            # Honour the user-configurable flag passed in by the caller.
            merged_swd = merge_adjacent_swds(token_preds, threshold=0.5, max_gap_intervals=3,
                                             require_overlap_agreement=self.require_overlap_agreement)
            step = int((self.interval_length - self.overlap_length) * self.sfreq)
            for region in merged_swd:
                start_idx = region['start_idx']
                end_idx = region['end_idx']
                region['start_sample'] = start_idx * step
                region['end_sample'] = (end_idx + 1) * step + int(self.interval_length * self.sfreq) - step
                region['duration_sec'] = (region['end_sample'] - region['start_sample']) / self.sfreq
                region['is_auto'] = True

            refined_swd = refine_swd_borders(self.eeg0, self.sfreq, merged_swd, self.power_percentile)
            refined_swd = [r for r in refined_swd if r['duration_sec'] >= 1.0]
            try:
                refined_swd = merge_overlapping_events(refined_swd, self.sfreq, min_overlap_sec=-1.0)
            except Exception:
                pass

            self.finished.emit((token_preds, merged_swd, refined_swd))
        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")

def bandpass_filter(data, sfreq, low, high):
    nyq = 0.5 * sfreq
    b, a = butter(4, [low / nyq, high / nyq], btype='band')
    return filtfilt(b, a, data)

def compute_rms_power(signal, win_size):
    return np.sqrt(np.convolve(signal**2, np.ones(win_size)/win_size, mode='same'))

def refine_swd_borders(eeg0, sfreq, events, power_percentile=25):
    refined_events = []
    
    # Bandpass filter in 10–20 Hz and compute smoothed power
    filtered = bandpass_filter(eeg0, sfreq, 10, 20)
    rms_win = int(0.2 * sfreq)  # 200 ms smoothing
    power = compute_rms_power(filtered, rms_win)
    
    # Process each event
    for event in events:
        start = event['start_sample']
        end = event['end_sample']
        
        # Calculate power threshold using only this event's region
        event_power = power[start:end]
        power_threshold = np.percentile(event_power, power_percentile)
        
        # Define border windows (1.5 seconds on each end)
        border_win = int(1.5 * sfreq)
        start_window = event_power[:border_win]
        end_window = event_power[-border_win:]
        
        # Find regions above threshold in start window
        start_above = start_window > power_threshold
        start_segments = []
        seg_start = None
        
        for i, is_above in enumerate(start_above):
            if is_above and seg_start is None:
                seg_start = i
            elif not is_above and seg_start is not None:
                start_segments.append((seg_start, i))
                seg_start = None
        
        if seg_start is not None:
            start_segments.append((seg_start, len(start_window)))
            
        # Find regions above threshold in end window
        end_above = end_window > power_threshold
        end_segments = []
        seg_start = None
        
        for i, is_above in enumerate(end_above):
            if is_above and seg_start is None:
                seg_start = i
            elif not is_above and seg_start is not None:
                end_segments.append((seg_start, i))
                seg_start = None
                
        if seg_start is not None:
            end_segments.append((seg_start, len(end_window)))
        
        # Get new start and end points
        new_start = start
        new_end = end
        
        # If we found segments in start window, use the first one
        if start_segments:
            new_start = start + start_segments[0][0]
            
        # If we found segments in end window, use the last one
        if end_segments:
            new_end = start + (len(event_power) - border_win) + end_segments[-1][1]
        
        # Create refined event
        refined_event = event.copy()
        refined_event['start_sample'] = new_start
        refined_event['end_sample'] = new_end
        refined_event['duration_sec'] = (new_end - new_start) / sfreq
        
        # Only keep events that are at least 1 second long
        if refined_event['duration_sec'] >= 1.0:
            refined_events.append(refined_event)
    
    return refined_events


def merge_overlapping_events(events, sfreq, min_overlap_sec=0.1):
    """
    Merge events that overlap in time.
    Returns a new list of merged events.
    """
    if not events:
        return []
        
    # Sort events by start time
    sorted_events = sorted(events, key=lambda x: x['start_sample'])
    merged = []
    current = sorted_events[0].copy()
    
    for next_event in sorted_events[1:]:
        # Calculate overlap between current and next event
        current_end = current['end_sample']
        next_start = next_event['start_sample']
        
        # Check if events overlap (positive value means they overlap)
        overlap = (current_end - next_start) / sfreq
        
        if overlap >= min_overlap_sec:
            # Merge events - take the earliest start and latest end
            current['start_sample'] = min(current['start_sample'], next_event['start_sample'])
            current['end_sample'] = max(current['end_sample'], next_event['end_sample'])
            
            # Recalculate duration
            current['duration_sec'] = (current['end_sample'] - current['start_sample']) / sfreq
            
            # If merging manual and auto event, prefer to keep it as manual
            if 'is_auto' in current and 'is_auto' in next_event:
                current['is_auto'] = current['is_auto'] and next_event['is_auto']
        else:
            # No overlap, add current to results and move to next
            merged.append(current)
            current = next_event.copy()
    
    # Add the last event
    merged.append(current)
    return merged

# -----------------------------------------------------------------------------
# UNET helpers (pre/post-processing)
# -----------------------------------------------------------------------------

def unet_find_continuous_intervals(binary_array):
    intervals = []
    start_idx = None
    arr = np.array(binary_array).astype(int)
    for i, value in enumerate(arr):
        if value == 1 and start_idx is None:
            start_idx = i
        elif value == 0 and start_idx is not None:
            intervals.append((start_idx, i - 1))
            start_idx = None
    if start_idx is not None:
        intervals.append((start_idx, len(arr) - 1))
    return intervals

def unet_merge_intervals(intervals, gap_threshold):
    if not intervals:
        return []
    merged = [intervals[0]]
    for current in intervals[1:]:
        prev = merged[-1]
        if current[0] - prev[1] - 1 < gap_threshold:
            merged[-1] = (prev[0], current[1])
        else:
            merged.append(current)
    return merged

def unet_postprocess_predictions(predictions, fs_hz=100.0, min_duration=1.0, gap_threshold=0.75):
    """
    Post-processing adapted from app_attention_res_UNet.py:
      - threshold at 0.5 (configurable)
      - keep intervals >= min_duration seconds
      - merge gaps < gap_threshold seconds
      - iteratively expand with nearby short intervals and merge gaps
    Returns intervals as list of (start_idx, end_idx) in samples at fs_hz.
    """
    binary_pred = (np.array(predictions) > 0.5).astype(np.int32)
    all_intervals = unet_find_continuous_intervals(binary_pred)
    min_len = int(round(min_duration * fs_hz))
    valid_intervals = [(s, e) for s, e in all_intervals if (e - s + 1) >= min_len]
    removed_intervals = [(s, e) for s, e in all_intervals if (e - s + 1) < min_len]
    merged_intervals = unet_merge_intervals(valid_intervals, gap_threshold=min_len)
    prev_intervals = None
    gap_thresh_samples = int(round(gap_threshold * fs_hz))
    while merged_intervals != prev_intervals:
        prev_intervals = merged_intervals.copy()
        expanded = []
        for (s, e) in merged_intervals:
            new_s = s
            new_e = e
            for (rs, re) in removed_intervals:
                if re < s and (s - re) < gap_thresh_samples:
                    new_s = min(new_s, rs)
                if rs > e and (rs - e) < gap_thresh_samples:
                    new_e = max(new_e, re)
            expanded.append((new_s, new_e))
        merged_intervals = unet_merge_intervals(expanded, gap_threshold=gap_thresh_samples)
    return merged_intervals

class PSDWorker(QObject):
    """Worker to compute multitaper PSDs in a background thread.

    Computes a PSD for each provided segment (e.g. full / first / middle / last)
    so the same worker pipeline serves both normal and advanced analysis modes.
    """
    finished = pyqtSignal(object)  # {'freqs': np.ndarray, 'segments': {name: {'psd_ch1','psd_ch2','start_sample','end_sample'}}}
    error = pyqtSignal(str)

    def __init__(self, eeg0, eeg1, sfreq, segments, fmin, fmax, bandwidth):
        super().__init__()
        self.eeg0 = eeg0
        self.eeg1 = eeg1
        self.sfreq = sfreq
        # segments: list of dicts with keys: 'name', 'start_sample', 'end_sample'
        self.segments = list(segments)
        self.fmin = fmin
        self.fmax = fmax
        self.bandwidth = bandwidth

    def run(self):
        try:
            n_total = len(self.eeg0)
            ref_freqs = None
            seg_results = {}
            for seg in self.segments:
                start = max(0, int(seg['start_sample']))
                end = min(n_total, int(seg['end_sample']))
                if end <= start:
                    raise ValueError(f"Invalid segment range for '{seg.get('name', '?')}': start={start}, end={end}")
                signal_ch1 = self.eeg0[start:end].copy()
                signal_ch2 = self.eeg1[start:end].copy()

                freqs1, psd1 = compute_multitaper_psd(
                    signal_ch1, self.sfreq, self.fmin, self.fmax, self.bandwidth, normalize=False)
                freqs2, psd2 = compute_multitaper_psd(
                    signal_ch2, self.sfreq, self.fmin, self.fmax, self.bandwidth, normalize=False)

                # Reference frequency axis: keep the longest one as reference and
                # interpolate all others onto it so plotting/exporting stays simple.
                if ref_freqs is None or len(freqs1) > len(ref_freqs):
                    ref_freqs = freqs1

                seg_results[seg['name']] = {
                    'freqs': freqs1,
                    'psd_ch1': psd1,
                    'psd_ch2': psd2,
                    'start_sample': start,
                    'end_sample': end,
                }

            # Resample PSDs to the common reference frequency axis.
            for name, data in seg_results.items():
                f_local = data['freqs']
                if len(f_local) != len(ref_freqs) or not np.allclose(f_local, ref_freqs):
                    data['psd_ch1'] = np.interp(ref_freqs, f_local, data['psd_ch1'])
                    data['psd_ch2'] = np.interp(ref_freqs, f_local, data['psd_ch2'])
                data.pop('freqs', None)

            self.finished.emit({
                'freqs': ref_freqs,
                'segments': seg_results,
            })
        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")

class UNETWorker(QObject):
    """Worker to run UNET preprocessing, inference and postprocessing in background"""
    finished = pyqtSignal(object)  # {'predictions': np.ndarray, 'intervals': list, 'confidence': np.ndarray}
    error = pyqtSignal(str)

    def __init__(self, signal, sfreq, model, unet_fs=100.0, segment_len=1024, min_duration=1.0, gap_threshold=0.75):
        super().__init__()
        self.signal = signal
        self.sfreq = sfreq
        self.model = model
        self.unet_fs = float(unet_fs)
        self.segment_len = int(segment_len)
        self.min_duration = float(min_duration)
        self.gap_threshold = float(gap_threshold)

    def run(self):
        try:
            if torch is None:
                raise RuntimeError('PyTorch is not installed')
            if self.model is None:
                raise RuntimeError('UNET model is not loaded')

            target_fs = self.unet_fs
            # Use resample_poly for faster and more memory-efficient resampling
            # Calculate up/down factors
            import math
            gcd = math.gcd(int(target_fs), int(self.sfreq))
            up = int(target_fs) // gcd
            down = int(self.sfreq) // gcd
            
            # Estimate new length
            resampled_len = int(len(self.signal) * up / down)
            
            if resampled_len < self.segment_len:
                raise RuntimeError(f'Signal too short for UNET (need >= {self.segment_len} samples at {int(target_fs)} Hz).')

            # Use resample_poly (FIR filter) instead of resample (FFT)
            # This is significantly faster for long signals and avoids large memory allocation
            sig_100 = resample_poly(self.signal, up, down)
            
            # Ensure exact length match if needed, though resample_poly is usually precise
            # Truncate or pad is handled by the logic below
            
            n_segments = len(sig_100) // self.segment_len
            if n_segments <= 0:
                raise RuntimeError('Not enough data for UNET prediction.')
                
            used_len = n_segments * self.segment_len
            data_used = sig_100[:used_len]
            segments = data_used.reshape(n_segments, self.segment_len)

            # Vectorized standardization (much faster than loop)
            # Compute mean and std along the last axis (time)
            means = segments.mean(axis=1, keepdims=True)
            stds = segments.std(axis=1, keepdims=True)
            # Avoid division by zero
            stds[stds == 0] = 1.0
            segments = (segments - means) / stds

            # To tensor (batch, 1, 1024)
            # Use a large batch size for inference to speed up but avoid OOM
            # Modern GPUs/CPUs can handle large batches of 1D signals
            batch_size = 1024 
            predictions_list = []
            
            # Process in batches
            total_batches = (n_segments + batch_size - 1) // batch_size
            
            with torch.no_grad():
                for i in range(total_batches):
                    start_idx = i * batch_size
                    end_idx = min((i + 1) * batch_size, n_segments)
                    batch_segments = segments[start_idx:end_idx]
                    
                    # Create tensor directly from numpy array
                    inp = torch.from_numpy(batch_segments).float().unsqueeze(1)
                    
                    # Inference
                    out = self.model(inp)
                    out_np = out.detach().cpu().numpy()
                    predictions_list.append(out_np)
            
            # Concatenate results
            if predictions_list:
                out_np = np.concatenate(predictions_list, axis=0)
            else:
                out_np = np.array([])

            out_np = np.squeeze(out_np)
            if out_np.ndim == 1:
                out_np = out_np.reshape(1, -1)
            
            # Ensure shape consistency
            if out_np.size != n_segments * self.segment_len:
                 out_np = np.reshape(out_np, (n_segments, self.segment_len))

            predictions_flat = out_np.reshape(-1)
            confidence = (predictions_flat - predictions_flat.min()) / (predictions_flat.max() - predictions_flat.min() + 1e-8)

            intervals = unet_postprocess_predictions(
                predictions_flat,
                fs_hz=self.unet_fs,
                min_duration=self.min_duration,
                gap_threshold=self.gap_threshold
            )

            self.finished.emit({
                'predictions': predictions_flat,
                'intervals': intervals,
                'confidence': confidence
            })
        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")


class ResampleWorker(QObject):
    """Worker to downsample an MNE Raw object in the background so the UI stays
    responsive. Resampling applies an anti-aliasing low-pass filter (cutoff at
    the new Nyquist frequency = target_fs / 2) before decimation."""
    finished = pyqtSignal(object)  # emits the resampled Raw object
    error = pyqtSignal(str)

    def __init__(self, raw, target_fs):
        super().__init__()
        self.raw = raw
        self.target_fs = float(target_fs)

    def run(self):
        try:
            try:
                # 'polyphase' uses scipy.signal.resample_poly with an explicit
                # anti-aliasing FIR filter and is far more memory-efficient than
                # the default FFT method for long recordings.
                self.raw.resample(self.target_fs, method='polyphase', verbose=False)
            except (TypeError, ValueError):
                # Older MNE versions lack the 'method' argument; the default
                # FFT-based resample is likewise band-limited (anti-aliased) at
                # the new Nyquist frequency.
                self.raw.resample(self.target_fs, npad='auto', verbose=False)
            self.finished.emit(self.raw)
        except Exception as e:
            import traceback
            self.error.emit(f"{e}\n{traceback.format_exc()}")

# ==============================================================================
# UI Helper Classes
# ==============================================================================

class ToastNotification(QWidget):
    """Modern toast notification that appears at the top-right of the parent window"""
    
    def __init__(self, parent, message, notification_type="info", duration=3000):
        super().__init__(parent)
        self.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAttribute(Qt.WA_ShowWithoutActivating)
        
        self.notification_type = notification_type
        self.duration = duration
        
        # Icon mapping
        icons = {
            "success": "✓",
            "warning": "⚠",
            "error": "✗",
            "info": "ℹ"
        }
        
        # Color mapping
        colors = {
            "success": ("#10B981", "#ECFDF5", "#047857"),
            "warning": ("#F59E0B", "#FFFBEB", "#D97706"),
            "error": ("#EF4444", "#FEF2F2", "#DC2626"),
            "info": ("#3B82F6", "#EFF6FF", "#2563EB")
        }
        
        bg_color, text_bg, border_color = colors.get(notification_type, colors["info"])
        icon = icons.get(notification_type, "ℹ")
        
        # Layout
        layout = QHBoxLayout(self)
        # Single, compact internal margin to avoid double spacing look
        layout.setContentsMargins(12, 10, 12, 10)
        
        # Icon label
        icon_label = QLabel(icon)
        icon_label.setStyleSheet(f"""
            QLabel {{
                color: {border_color};
                font-size: 18px;
                font-weight: bold;
                padding-right: 8px;
            }}
        """)
        icon_label.setAutoFillBackground(False)
        layout.addWidget(icon_label)
        
        # Message label
        msg_label = QLabel(message)
        msg_label.setStyleSheet(f"""
            QLabel {{
                color: {border_color};
                font-size: 13px;
                padding: 0px;
            }}
        """)
        msg_label.setAutoFillBackground(False)
        msg_label.setWordWrap(True)
        layout.addWidget(msg_label, 1)
        
        # Apply main style
        self.setStyleSheet(f"""
            ToastNotification {{
                background-color: {text_bg};
                border: 2px solid {border_color};
                border-radius: 8px;
            }}
        """)
        
        # Set fixed width and adjust height to content
        self.setFixedWidth(350)
        self.adjustSize()
        
        # Animation
        self.opacity_effect = None
        self.fade_animation = None
        
    def show_toast(self):
        """Show the toast with fade-in animation"""
        # Position at top-right
        parent = self.parent()
        if parent:
            parent_rect = parent.rect()
            # Reduce outer offset so toast appears with a single outer margin
            x = parent_rect.width() - self.width() - 12
            y = 12
            self.move(parent.mapToGlobal(QPoint(x, y)))
        
        self.show()
        self.raise_()
        
        # Auto-hide after duration
        QTimer.singleShot(self.duration, self.fade_out)
        
    def fade_out(self):
        """Fade out and close"""
        QTimer.singleShot(300, self.close)


class StatCard(QFrame):
    """Statistical information card widget"""
    
    def __init__(self, title, value, icon="", color="#3B82F6", parent=None, dark_mode=False):
        super().__init__(parent)
        self.setFrameStyle(QFrame.StyledPanel)
        self.color = color
        self.dark_mode = dark_mode
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(8)
        
        # Title with icon
        title_layout = QHBoxLayout()
        if icon:
            self.icon_label = QLabel(icon)
            self.icon_label.setStyleSheet(f"font-size: 24px; color: {color};")
            title_layout.addWidget(self.icon_label)
        
        self.title_label = QLabel(title)
        title_color = "#8B949E" if dark_mode else "#6B7280"
        self.title_label.setStyleSheet(f"font-size: 12px; color: {title_color}; font-weight: 600;")
        title_layout.addWidget(self.title_label)
        title_layout.addStretch()
        layout.addLayout(title_layout)
        
        # Value
        self.value_label = QLabel(str(value))
        self.value_label.setStyleSheet(f"font-size: 28px; font-weight: bold; color: {color};")
        layout.addWidget(self.value_label)
        
        # Style the card based on mode
        if dark_mode:
            self.setStyleSheet(f"""
                StatCard {{
                    background-color: #161B22;
                    border: 1px solid #30363D;
                    border-radius: 8px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                StatCard {{
                    background-color: white;
                    border: 1px solid #E5E7EB;
                    border-radius: 8px;
                }}
            """)
        
    def update_value(self, value):
        """Update the displayed value"""
        self.value_label.setText(str(value))


class SpinnerWidget(QWidget):
    """Animated spinner/loading indicator widget"""
    
    def __init__(self, size=32, color="#3B82F6", parent=None):
        super().__init__(parent)
        self.size = size
        self.color = QColor(color)
        self.angle = 0
        self.setFixedSize(size, size)
        
        # Animation timer
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.rotate)
        
    def start(self):
        """Start the spinner animation"""
        self.timer.start(50)  # Update every 50ms
        self.show()
        
    def stop(self):
        """Stop the spinner animation"""
        self.timer.stop()
        self.hide()
        
    def rotate(self):
        """Rotate the spinner"""
        self.angle = (self.angle + 30) % 360
        self.update()
        
    def paintEvent(self, event):
        """Custom paint for the spinner"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        # Draw spinner arc
        center = self.size / 2
        radius = (self.size - 4) / 2
        
        # Create gradient for smooth fade effect
        for i in range(12):
            angle = self.angle + (i * 30)
            opacity = 255 - (i * 20)
            if opacity < 50:
                opacity = 50
            
            color = QColor(self.color)
            color.setAlpha(opacity)
            
            painter.setPen(QPen(color, 3, Qt.SolidLine, Qt.RoundCap))
            
            # Calculate start and end points
            import math
            rad = math.radians(angle)
            x1 = center + (radius * 0.6) * math.cos(rad)
            y1 = center + (radius * 0.6) * math.sin(rad)
            x2 = center + radius * math.cos(rad)
            y2 = center + radius * math.sin(rad)
            
            painter.drawLine(int(x1), int(y1), int(x2), int(y2))


class LoadingOverlay(QWidget):
    """Full-screen loading overlay with spinner and message"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setAttribute(Qt.WA_ShowWithoutActivating)
        
        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)
        
        # Container
        container = QFrame()
        container.setStyleSheet("""
            QFrame {
                background-color: rgba(255, 255, 255, 0.95);
                border-radius: 12px;
                padding: 20px;
            }
        """)
        container_layout = QVBoxLayout(container)
        container_layout.setAlignment(Qt.AlignCenter)
        container_layout.setSpacing(15)
        
        # Spinner
        self.spinner = SpinnerWidget(48, "#3B82F6")
        container_layout.addWidget(self.spinner, alignment=Qt.AlignCenter)
        
        # Message
        self.message_label = QLabel("Loading...")
        self.message_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #1F2937;")
        self.message_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(self.message_label)
        
        # Progress info
        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet("font-size: 12px; color: #6B7280;")
        self.progress_label.setAlignment(Qt.AlignCenter)
        container_layout.addWidget(self.progress_label)
        
        layout.addWidget(container)
        
    def show_loading(self, message="Loading...", progress=""):
        """Show the loading overlay with message"""
        self.message_label.setText(message)
        self.progress_label.setText(progress)
        
        # Position in center of parent
        if self.parent():
            parent_rect = self.parent().rect()
            self.setGeometry(parent_rect)
        
        self.show()
        self.raise_()
        self.spinner.start()
        
    def hide_loading(self):
        """Hide the loading overlay"""
        self.spinner.stop()
        self.hide()


class KeyboardShortcutsDialog(QDialog):
    """Dialog showing keyboard shortcuts"""
    
    def __init__(self, key_assignments, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🎮 Controls")
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)
        
        layout = QVBoxLayout(self)
        
        # Header
        header = QLabel("Controls")
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding: 10px;")
        layout.addWidget(header)
        
        # Shortcuts table
        table = QTableWidget()
        table.setColumnCount(2)
        table.setHorizontalHeaderLabels(["Action", "Shortcut"])
        table.horizontalHeader().setStretchLastSection(True)
        table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        table.setEditTriggers(QTableWidget.NoEditTriggers)
        table.setSelectionMode(QTableWidget.NoSelection)
        
        # Add rows
        table.setRowCount(len(key_assignments) + 4)  # +4 for additional shortcuts
        row = 0
        
        for action, key in key_assignments.items():
            table.setItem(row, 0, QTableWidgetItem(action))
            key_name = QKeySequence(key).toString() if isinstance(key, int) else str(key)
            table.setItem(row, 1, QTableWidgetItem(key_name))
            row += 1
        
        # Add other common shortcuts
        other_shortcuts = [
            ("Export to Excel", "Ctrl+S"),
            ("Show Shortcuts", "F1"),
            ("Scroll View", "Mouse Wheel"),
            ("Zoom Y-Axis", "+/- Buttons")
        ]
        
        for action, shortcut in other_shortcuts:
            table.setItem(row, 0, QTableWidgetItem(action))
            table.setItem(row, 1, QTableWidgetItem(shortcut))
            row += 1
        
        layout.addWidget(table)
        
        # Close button
        close_btn = QPushButton("Close")
        close_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn)


class TimelineMinimap(QWidget):
    """Mini timeline overview with event markers"""
    
    clicked = pyqtSignal(float)  # Emits time in seconds when clicked
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(50)
        self.total_duration = 1.0
        self.current_pos = 0.0
        self.window_size = 1.0
        self.events = []  # List of (start, end, type) tuples
        self.dark_mode = False
        self.auto_color_hex = "#FF7F50"  # Default coral; will be overridden by model theme
        self.setMouseTracking(True)
        self.setToolTip("Click to jump to position")
        
    def set_data(self, total_duration, current_pos, window_size, events):
        """Update timeline data"""
        self.total_duration = max(total_duration, 1.0)
        self.current_pos = current_pos
        self.window_size = window_size
        self.events = events
        self.update()
    
    def set_dark_mode(self, dark):
        """Set dark mode"""
        self.dark_mode = dark
        self.update()
    
    def set_auto_color(self, hex_color):
        """Set color used for automatic events (model-themed)."""
        try:
            if isinstance(hex_color, str) and hex_color.startswith('#'):
                self.auto_color_hex = hex_color
        except Exception:
            pass
        self.update()
        
    def paintEvent(self, event):
        """Custom paint for the timeline"""
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        width = self.width()
        height = self.height()
        
        # Background
        bg_color = QColor("#161B22") if self.dark_mode else QColor("#F3F4F6")
        painter.fillRect(0, 0, width, height, bg_color)
        
        # Draw border
        border_color = QColor("#30363D") if self.dark_mode else QColor("#E5E7EB")
        painter.setPen(QPen(border_color, 1))
        painter.drawRect(0, 0, width - 1, height - 1)
        
        # Draw events
        # Comment markers are drawn *after* events so they sit on top as
        # thin vertical ticks (width 1-2 px) spanning the full mini-bar.
        comment_entries = []
        for ev in self.events:
            start, end, event_type = ev[0], ev[1], ev[2]
            if event_type == 'comment':
                comment_entries.append(ev)
                continue
            x1 = int((start / self.total_duration) * width)
            x2 = int((end / self.total_duration) * width)
            w = max(2, x2 - x1)

            if event_type == 'auto':
                color = QColor(self.auto_color_hex)
            elif event_type == 'manual':
                color = QColor("#5B7FC7") if self.dark_mode else QColor("#4169E1")
            else:
                continue

            painter.fillRect(x1, 10, w, height - 20, color)

        # Comment markers: subtle amber vertical ticks, always on top.
        if comment_entries:
            cmt_color = QColor("#F5B041") if self.dark_mode else QColor("#D97706")
            pen = QPen(cmt_color, 2)
            painter.setPen(pen)
            for ev in comment_entries:
                start = ev[0]
                x = int((start / self.total_duration) * width)
                painter.drawLine(x, 2, x, height - 4)
        
        # Draw current view window
        x1 = int((self.current_pos / self.total_duration) * width)
        x2 = int(((self.current_pos + self.window_size) / self.total_duration) * width)
        
        # Window outline and fill
        window_color = QColor("#238636") if self.dark_mode else QColor("#3B82F6")
        window_fill = QColor(35, 134, 54, 60) if self.dark_mode else QColor(59, 130, 246, 50)
        painter.setPen(QPen(window_color, 2))
        painter.setBrush(window_fill)
        painter.drawRect(x1, 0, x2 - x1, height - 1)
        
    def mousePressEvent(self, event):
        """Handle clicks to jump to position"""
        if event.button() == Qt.LeftButton:
            x = event.pos().x()
            time_sec = (x / self.width()) * self.total_duration
            self.clicked.emit(time_sec)


class PSDAnalysisDialog(QDialog):
    """Dialog for PSD analysis of SWD events with interactive graphs"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📊 Multitaper PSD Analysis")
        self.setGeometry(40, 40, 1300, 950)
        self.saved_psds = {}  # Dictionary to store saved PSDs
        self.current_freqs = None
        self.current_psd = None
        self.psd_thread = None
        self.psd_worker = None
        
        self.init_ui()
        
    def init_ui(self):
        # Layout: the plot fills the top (full width, maximum size) and every
        # control lives in a single compact row underneath. Values are read via
        # an aesthetic on-plot hover annotation (no matplotlib toolbar).
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(8, 8, 8, 8)
        main_layout.setSpacing(8)

        _spin_h = 28  # shared control height for a consistent, compact look

        # Filter syncing used to be a checkbox ("Match main viewer filter").
        # It has been removed from the UI; syncing stays enabled by default.
        self._sync_bandpass_enabled = True

        _group_css = (
            "QGroupBox { font-weight: 600; margin-top: 8px; }"
            "QGroupBox::title { subcontrol-origin: margin; left: 8px; padding: 0 4px; }"
        )

        # =============================================================
        #  TOP: plot area (canvas only), maximised
        # =============================================================
        self.fig, (self.ax1, self.ax2) = plt.subplots(2, 1, figsize=(12, 9))
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.canvas.setMinimumHeight(460)
        main_layout.addWidget(self.canvas, 1)

        # Aesthetic hover: annotation + crosshair created on the fly. These are
        # cleared and rebuilt on every redraw (axes are cleared there).
        self._hover_artists = []
        self.canvas.mpl_connect('motion_notify_event', self._on_plot_hover)
        self.canvas.mpl_connect('axes_leave_event', lambda _e: self._clear_hover())
        self.canvas.mpl_connect('figure_leave_event', lambda _e: self._clear_hover())

        # =============================================================
        #  BOTTOM: one compact row of control groups
        # =============================================================
        # ---- Frequency & channels ----
        freq_box = QGroupBox("Frequency && Channels")
        freq_box.setStyleSheet(_group_css)
        freq_grid = QGridLayout(freq_box)
        freq_grid.setContentsMargins(10, 8, 10, 8)
        freq_grid.setHorizontalSpacing(8)
        freq_grid.setVerticalSpacing(6)

        freq_grid.addWidget(QLabel("Range:"), 0, 0)
        range_row = QHBoxLayout()
        range_row.setSpacing(4)
        self.fmin_spin = QDoubleSpinBox()
        self.fmin_spin.setRange(0.1, 100)
        self.fmin_spin.setValue(1.0)
        self.fmin_spin.setFixedWidth(66)
        self.fmin_spin.setFixedHeight(_spin_h)
        range_row.addWidget(self.fmin_spin)
        range_row.addWidget(QLabel("–"))
        self.fmax_spin = QDoubleSpinBox()
        self.fmax_spin.setRange(1, 200)
        self.fmax_spin.setValue(50.0)
        self.fmax_spin.setFixedWidth(66)
        self.fmax_spin.setFixedHeight(_spin_h)
        range_row.addWidget(self.fmax_spin)
        range_row.addWidget(QLabel("Hz"))
        range_row.addStretch()
        freq_grid.addLayout(range_row, 0, 1)

        freq_grid.addWidget(QLabel("Bandwidth:"), 1, 0)
        self.bandwidth_spin = QDoubleSpinBox()
        self.bandwidth_spin.setRange(1.0, 10.0)
        self.bandwidth_spin.setValue(2.0)
        self.bandwidth_spin.setFixedWidth(66)
        self.bandwidth_spin.setFixedHeight(_spin_h)
        self.bandwidth_spin.setToolTip("Multitaper time-bandwidth product (frequency smoothing).")
        freq_grid.addWidget(self.bandwidth_spin, 1, 1)

        freq_grid.addWidget(QLabel("Channels:"), 2, 0)
        ch_row = QHBoxLayout()
        ch_row.setSpacing(8)
        self.show_ch1_cb = QCheckBox("Ch1")
        self.show_ch1_cb.setChecked(True)
        self.show_ch2_cb = QCheckBox("Ch2")
        self.show_ch2_cb.setChecked(True)
        self.normalize_cb = QCheckBox("Norm")
        self.normalize_cb.setToolTip("Scale each PSD so the area under the curve equals 1.")
        ch_row.addWidget(self.show_ch1_cb)
        ch_row.addWidget(self.show_ch2_cb)
        ch_row.addWidget(self.normalize_cb)
        ch_row.addStretch()
        freq_grid.addLayout(ch_row, 2, 1)
        freq_grid.setColumnStretch(1, 1)

        # ---- Bandpass filter (mirrors the main viewer's plot filter) ----
        _parent = self.parent()
        parent_low = float(getattr(_parent, 'filter_low_cut_hz', 1.0) or 1.0) if _parent else 1.0
        parent_high = float(getattr(_parent, 'filter_high_cut_hz', 40.0) or 40.0) if _parent else 40.0
        parent_filter_on = bool(getattr(_parent, 'filter_signal_enabled', False)) if _parent else False

        bp_box = QGroupBox("Bandpass (pre-PSD)")
        bp_box.setStyleSheet(_group_css)
        bp_grid = QGridLayout(bp_box)
        bp_grid.setContentsMargins(10, 8, 10, 8)
        bp_grid.setHorizontalSpacing(8)
        bp_grid.setVerticalSpacing(6)

        self.bandpass_cb = QCheckBox("Apply filter")
        self.bandpass_cb.setChecked(parent_filter_on)
        self.bandpass_cb.setToolTip(
            "Apply a zero-phase bandpass filter to the EEG signal before computing PSD.\n"
            "Mirrors the main viewer's plot bandpass filter (Settings ▸ Analysis).")
        bp_grid.addWidget(self.bandpass_cb, 0, 0, 1, 2)

        bp_grid.addWidget(QLabel("Low:"), 1, 0)
        self.bp_low_spin = QDoubleSpinBox()
        self.bp_low_spin.setRange(0.0, 500.0)
        self.bp_low_spin.setSingleStep(0.5)
        self.bp_low_spin.setDecimals(2)
        self.bp_low_spin.setValue(parent_low)
        self.bp_low_spin.setSuffix(" Hz")
        self.bp_low_spin.setFixedWidth(88)
        self.bp_low_spin.setFixedHeight(_spin_h)
        self.bp_low_spin.setToolTip("High-pass cutoff (set to 0 to disable the high-pass).")
        bp_grid.addWidget(self.bp_low_spin, 1, 1)

        bp_grid.addWidget(QLabel("High:"), 2, 0)
        self.bp_high_spin = QDoubleSpinBox()
        self.bp_high_spin.setRange(0.0, 1000.0)
        self.bp_high_spin.setSingleStep(0.5)
        self.bp_high_spin.setDecimals(2)
        self.bp_high_spin.setValue(parent_high)
        self.bp_high_spin.setSuffix(" Hz")
        self.bp_high_spin.setFixedWidth(88)
        self.bp_high_spin.setFixedHeight(_spin_h)
        self.bp_high_spin.setToolTip("Low-pass cutoff (set to 0 to disable the low-pass).")
        bp_grid.addWidget(self.bp_high_spin, 2, 1)
        bp_grid.setColumnStretch(1, 1)

        # ---- Segmentation (advanced) ----
        seg_box = QGroupBox("Segmentation")
        seg_box.setStyleSheet(_group_css)
        seg_v = QVBoxLayout(seg_box)
        seg_v.setContentsMargins(10, 8, 10, 8)
        seg_v.setSpacing(6)

        self.advanced_cb = QCheckBox("Advanced (pre/first/middle/last/post)")
        self.advanced_cb.setToolTip(
            "When enabled, the event is split into segments of n seconds each "
            "(beginning, middle and end) and a separate PSD is computed for every segment.")
        seg_v.addWidget(self.advanced_cb)

        seg_grid = QGridLayout()
        seg_grid.setHorizontalSpacing(8)
        seg_grid.setVerticalSpacing(6)
        seg_grid.addWidget(QLabel("Segment:"), 0, 0)
        self.segment_seconds_spin = QDoubleSpinBox()
        self.segment_seconds_spin.setRange(0.1, 600.0)
        self.segment_seconds_spin.setSingleStep(0.5)
        self.segment_seconds_spin.setDecimals(2)
        self.segment_seconds_spin.setValue(2.0)
        self.segment_seconds_spin.setSuffix(" s")
        self.segment_seconds_spin.setFixedWidth(84)
        self.segment_seconds_spin.setFixedHeight(_spin_h)
        self.segment_seconds_spin.setEnabled(False)
        seg_grid.addWidget(self.segment_seconds_spin, 0, 1)

        seg_grid.addWidget(QLabel("Pre/Post:"), 0, 2)
        self.prepost_seconds_spin = QDoubleSpinBox()
        self.prepost_seconds_spin.setRange(0.0, 600.0)
        self.prepost_seconds_spin.setSingleStep(0.5)
        self.prepost_seconds_spin.setDecimals(2)
        self.prepost_seconds_spin.setValue(2.0)
        self.prepost_seconds_spin.setSuffix(" s")
        self.prepost_seconds_spin.setFixedWidth(84)
        self.prepost_seconds_spin.setFixedHeight(_spin_h)
        self.prepost_seconds_spin.setEnabled(False)
        self.prepost_seconds_spin.setToolTip(
            "Length of the pre-SWD and post-SWD windows. Set to 0 to disable.")
        seg_grid.addWidget(self.prepost_seconds_spin, 0, 3)
        seg_v.addLayout(seg_grid)

        # Per-segment visibility toggles (used for plot only).
        vis_row = QHBoxLayout()
        vis_row.setSpacing(8)
        vis_row.addWidget(QLabel("Show:"))
        self.show_pre_cb = QCheckBox("Pre")
        self.show_first_cb = QCheckBox("First")
        self.show_middle_cb = QCheckBox("Middle")
        self.show_last_cb = QCheckBox("Last")
        self.show_post_cb = QCheckBox("Post")
        for _cb in (self.show_pre_cb, self.show_first_cb, self.show_middle_cb,
                    self.show_last_cb, self.show_post_cb):
            _cb.setChecked(True)
            _cb.setEnabled(False)
            vis_row.addWidget(_cb)
        vis_row.addStretch()
        seg_v.addLayout(vis_row)

        # ---- Actions ----
        actions_box = QGroupBox("Actions")
        actions_box.setStyleSheet(_group_css)
        actions_v = QVBoxLayout(actions_box)
        actions_v.setContentsMargins(10, 8, 10, 8)
        actions_v.setSpacing(6)

        self.compute_btn = QPushButton("🔄 Compute PSD")
        self.compute_btn.setMinimumHeight(34)
        self.compute_btn.setStyleSheet(
            "QPushButton { background-color: #2563EB; color: white; font-weight: 600;"
            " border: none; border-radius: 6px; padding: 6px 12px; }"
            "QPushButton:hover { background-color: #1D4ED8; }"
            "QPushButton:disabled { background-color: #9CA3AF; }"
        )
        actions_v.addWidget(self.compute_btn)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.save_btn = QPushButton("💾 Save")
        self.save_btn.setMinimumHeight(30)
        btn_row.addWidget(self.save_btn)
        self.export_btn = QPushButton("📤 Export")
        self.export_btn.setMinimumHeight(30)
        btn_row.addWidget(self.export_btn)
        self.validate_btn = QPushButton("🧪 Validate")
        self.validate_btn.setMinimumHeight(30)
        self.validate_btn.setToolTip(
            "Sanity check: compute the same window with scipy.signal.welch and "
            "compare against our MNE multitaper PSD. QA only — does not affect saved/exported PSDs.")
        btn_row.addWidget(self.validate_btn)
        actions_v.addLayout(btn_row)

        bins_row = QHBoxLayout()
        bins_row.setSpacing(6)
        bins_row.addWidget(QLabel("Export Bins:"))
        self.export_bins_spin = QSpinBox()
        self.export_bins_spin.setRange(10, 500)
        self.export_bins_spin.setValue(50)
        self.export_bins_spin.setFixedWidth(72)
        self.export_bins_spin.setFixedHeight(_spin_h)
        bins_row.addWidget(self.export_bins_spin)
        bins_row.addStretch()
        actions_v.addLayout(bins_row)

        # ---- Saved PSDs ----
        saved_group = QGroupBox("💾 Saved PSDs")
        saved_group.setStyleSheet(_group_css)
        saved_layout = QVBoxLayout(saved_group)
        saved_layout.setContentsMargins(10, 8, 10, 8)
        saved_layout.setSpacing(6)

        self.overlay_saved_cb = QCheckBox("Overlay on Plot")
        saved_layout.addWidget(self.overlay_saved_cb)

        self.saved_list = QListWidget()
        self.saved_list.setSelectionMode(QAbstractItemView.MultiSelection)
        self.saved_list.setMinimumHeight(64)
        saved_layout.addWidget(self.saved_list)

        self.clear_btn = QPushButton("🗑️ Clear All Saved")
        self.clear_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        saved_layout.addWidget(self.clear_btn)

        # ---- Assemble the compact bottom row ----
        # Boxes align to the top so differing heights don't stretch/overlap.
        for _b in (freq_box, bp_box, seg_box, actions_box, saved_group):
            _b.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Maximum)

        controls_row = QHBoxLayout()
        controls_row.setSpacing(10)
        controls_row.addWidget(freq_box)
        controls_row.addWidget(bp_box)
        controls_row.addWidget(seg_box)
        controls_row.addWidget(actions_box)
        controls_row.addWidget(saved_group, 1)  # saved list soaks up extra width

        controls_container = QWidget()
        controls_container.setLayout(controls_row)
        # Keep the controls a compact strip so the plot keeps the height.
        controls_container.setMaximumHeight(210)
        main_layout.addWidget(controls_container, 0)

        # Connect signals
        self.compute_btn.clicked.connect(self.compute_psd)
        self.save_btn.clicked.connect(self.save_current_psd)
        self.export_btn.clicked.connect(self.export_psds)
        self.validate_btn.clicked.connect(self.run_self_check)
        self.clear_btn.clicked.connect(self.clear_saved_psds)
        self.show_ch1_cb.stateChanged.connect(self.update_plot_visibility)
        self.show_ch2_cb.stateChanged.connect(self.update_plot_visibility)
        self.normalize_cb.stateChanged.connect(self.update_plot_visibility)
        self.overlay_saved_cb.stateChanged.connect(self.update_plot_visibility)
        self.saved_list.itemSelectionChanged.connect(self.update_plot_visibility)
        self.advanced_cb.stateChanged.connect(self._on_advanced_toggled)
        self.show_pre_cb.stateChanged.connect(self.update_plot_visibility)
        self.show_first_cb.stateChanged.connect(self.update_plot_visibility)
        self.show_middle_cb.stateChanged.connect(self.update_plot_visibility)
        self.show_last_cb.stateChanged.connect(self.update_plot_visibility)
        self.show_post_cb.stateChanged.connect(self.update_plot_visibility)

        # Bandpass controls -> keep the main viewer's plot filter in sync.
        self.bandpass_cb.stateChanged.connect(self._on_bandpass_changed)
        self.bp_low_spin.valueChanged.connect(self._on_bandpass_changed)
        self.bp_high_spin.valueChanged.connect(self._on_bandpass_changed)

        # Initially disable save button
        self.save_btn.setEnabled(False)

        # Guard against feedback loops while programmatically updating the
        # bandpass widgets from the parent (and vice-versa).
        self._syncing_bandpass = False

        # Cache key for the bandpass-filtered EEG copies maintained per dialog.
        self._filter_cache = None  # (low, high, n_total) -> just a marker
        self._filtered_eeg0 = None
        self._filtered_eeg1 = None

    def _on_advanced_toggled(self, _state):
        """Enable/disable advanced controls and clear stale results when toggled."""
        enabled = self.advanced_cb.isChecked()
        self.segment_seconds_spin.setEnabled(enabled)
        self.prepost_seconds_spin.setEnabled(enabled)
        self.show_pre_cb.setEnabled(enabled)
        self.show_first_cb.setEnabled(enabled)
        self.show_middle_cb.setEnabled(enabled)
        self.show_last_cb.setEnabled(enabled)
        self.show_post_cb.setEnabled(enabled)
        # The current PSD belongs to a different mode; force the user to recompute.
        self.current_psd = None
        self.current_freqs = None
        self.ax1.clear()
        self.ax2.clear()
        self.canvas.draw()
        self.save_btn.setEnabled(False)

    def _on_bandpass_changed(self, *_args):
        """Propagate PSD bandpass settings to the main viewer's plot filter.

        The main viewer exposes a single "plot bandpass filter" (Settings ▸
        Analysis). To avoid two filters that silently disagree, the PSD dialog
        mirrors that filter: when 'Match main viewer filter' is checked, editing
        the passband here updates the viewer's plot filter (and its plot) too.
        """
        if getattr(self, '_syncing_bandpass', False):
            return
        if not getattr(self, '_sync_bandpass_enabled', True):
            return
        parent = self.parent()
        if parent is None:
            return
        try:
            parent.filter_signal_enabled = self.bandpass_cb.isChecked()
            parent.filter_low_cut_hz = float(self.bp_low_spin.value())
            parent.filter_high_cut_hz = float(self.bp_high_spin.value())
            # Keep the Settings dialog widgets consistent if it is open.
            settings = getattr(parent, 'settings_dialog', None)
            if settings is not None:
                if hasattr(settings, 'cb_filter_plot_tab'):
                    settings.cb_filter_plot_tab.setChecked(parent.filter_signal_enabled)
                if hasattr(settings, 'filter_low_tab'):
                    settings.filter_low_tab.setValue(parent.filter_low_cut_hz)
                if hasattr(settings, 'filter_high_tab'):
                    settings.filter_high_tab.setValue(parent.filter_high_cut_hz)
            if parent.filter_signal_enabled and hasattr(parent, '_apply_plot_bandpass_filter'):
                parent._apply_plot_bandpass_filter()
            if hasattr(parent, 'update_plot'):
                parent.update_plot()
        except Exception as e:
            print(f"PSD bandpass sync error: {e}")

    def _sync_bandpass_from_parent(self):
        """Refresh the PSD bandpass controls from the main viewer's plot filter.

        Called whenever the dialog is (re)opened so the PSD passband always
        starts out matching the external plot bandpass filter, resolving any
        discrepancy between the two controls.
        """
        if not getattr(self, '_sync_bandpass_enabled', True):
            return
        parent = self.parent()
        if parent is None:
            return
        self._syncing_bandpass = True
        try:
            self.bandpass_cb.setChecked(bool(getattr(parent, 'filter_signal_enabled', False)))
            low = getattr(parent, 'filter_low_cut_hz', None)
            high = getattr(parent, 'filter_high_cut_hz', None)
            if low is not None:
                self.bp_low_spin.setValue(float(low))
            if high is not None:
                self.bp_high_spin.setValue(float(high))
        except Exception as e:
            print(f"PSD bandpass sync error: {e}")
        finally:
            self._syncing_bandpass = False

    def set_event_data(self, eeg0, eeg1, sfreq, event):
        """Set the current event data for PSD analysis"""
        prev_n = len(self.eeg0) if getattr(self, 'eeg0', None) is not None else None
        prev_sfreq = getattr(self, 'sfreq', None)

        self.eeg0 = eeg0
        self.eeg1 = eeg1
        self.sfreq = sfreq
        self.event = event

        # Invalidate the bandpass-filter cache when the underlying recording
        # changes (different file or sample rate), but keep it across events
        # within the same recording so re-computing is fast.
        if (prev_n is None or prev_n != len(eeg0) or prev_sfreq != sfreq):
            self._filtered_eeg0 = None
            self._filtered_eeg1 = None
            self._filter_cache = None

        # Mirror the main viewer's plot bandpass filter so the PSD passband
        # matches what the user sees on the main plot (removes the discrepancy
        # between the two separate bandpass controls).
        self._sync_bandpass_from_parent()

        # Clear current PSD data to show empty plot for new event
        self.current_psd = None
        self.current_freqs = None

        # Clear the plot axes
        self.ax1.clear()
        self.ax2.clear()

        # Disable save button since there's no current PSD
        self.save_btn.setEnabled(False)

        # Theme the (empty) axes to match the current light/dark mode, then draw.
        self._style_empty_axes()
        self.canvas.draw()

        # Update window title with event info
        start_sample = event['start_sample']
        duration = event['duration_sec']
        self.setWindowTitle(f"PSD Analysis - Event at {start_sample/sfreq:.1f}s ({duration:.1f}s duration)")

    def _theme_colors(self):
        """Return the current (bg, plot_bg, grid, title, axis) colors for the plot."""
        dark_mode = getattr(self.parent(), 'dark_mode', False) if self.parent() else False
        if dark_mode:
            return dark_mode, '#0D1117', '#161B22', '#30363D', '#C0C0C0', '#B0B0B0'
        return dark_mode, 'white', 'white', '#CCCCCC', 'black', 'black'

    def _style_empty_axes(self):
        """Apply light/dark theme to the (data-less) axes so the empty PSD plot
        matches the rest of the dialog before anything is computed."""
        _dark, bg_color, plot_bg, grid_color, _title, axis_color = self._theme_colors()
        self.fig.patch.set_facecolor(bg_color)
        for ax in (self.ax1, self.ax2):
            ax.set_facecolor(plot_bg)
            ax.tick_params(colors=axis_color)
            for spine in ax.spines.values():
                spine.set_color(grid_color)
            ax.xaxis.label.set_color(axis_color)
            ax.yaxis.label.set_color(axis_color)

    def refresh_theme(self):
        """Re-theme the PSD plot to match the current light/dark mode.

        Called when the main window toggles dark mode while this dialog is open;
        the matplotlib canvas does not follow Qt stylesheets, so it must be
        redrawn explicitly."""
        try:
            if self.current_psd is not None and self.current_freqs is not None:
                self.redraw_current_psd_plot()
            else:
                self._style_empty_axes()
                self.canvas.draw()
        except Exception:
            pass

    def _build_segments(self):
        """Build the list of (name, start_sample, end_sample) segments for the current event.

        - Normal mode: a single 'full' segment spanning the whole event.
        - Advanced mode:
            * 'pre'    -> n_pp seconds before the SWD (clamped to recording start)
            * 'first'  -> first n seconds of the SWD
            * 'middle' -> middle n seconds of the SWD
            * 'last'   -> last n seconds of the SWD
            * 'post'   -> n_pp seconds after the SWD (clamped to recording end)

          If the event is shorter than n seconds, the first/middle/last segments are
          clamped to the event duration (so they may overlap). If pre/post would be
          empty (event at the very start/end of the recording, or n_pp == 0), the
          corresponding segment is skipped.
        """
        start = int(self.event['start_sample'])
        end = int(self.event['end_sample'])
        if not self.advanced_cb.isChecked():
            return [{'name': 'full', 'start_sample': start, 'end_sample': end}], None, None

        n_seconds = float(self.segment_seconds_spin.value())
        n_pp_seconds = float(self.prepost_seconds_spin.value())
        n_samples = max(1, int(round(n_seconds * self.sfreq)))
        n_pp_samples = max(0, int(round(n_pp_seconds * self.sfreq)))
        event_len = max(1, end - start)
        # Allow segments to overlap when the event is too short to fit three of them
        # without overlap; clamp the segment length to the event length.
        seg_len = min(n_samples, event_len)

        first_start = start
        first_end = start + seg_len

        last_end = end
        last_start = max(start, end - seg_len)

        mid = (start + end) // 2
        middle_start = max(start, mid - seg_len // 2)
        middle_end = min(end, middle_start + seg_len)
        # Re-center if we had to clamp the right edge.
        if middle_end - middle_start < seg_len:
            middle_start = max(start, middle_end - seg_len)

        # Pre-SWD window (clamped to recording bounds).
        n_total = len(self.eeg0) if self.eeg0 is not None else end + n_pp_samples
        pre_start = max(0, start - n_pp_samples)
        pre_end = start
        post_start = end
        post_end = min(n_total, end + n_pp_samples)

        segments = []
        if n_pp_samples > 0 and pre_end > pre_start:
            segments.append({'name': 'pre', 'start_sample': pre_start, 'end_sample': pre_end})
        segments.extend([
            {'name': 'first', 'start_sample': first_start, 'end_sample': first_end},
            {'name': 'middle', 'start_sample': middle_start, 'end_sample': middle_end},
            {'name': 'last', 'start_sample': last_start, 'end_sample': last_end},
        ])
        if n_pp_samples > 0 and post_end > post_start:
            segments.append({'name': 'post', 'start_sample': post_start, 'end_sample': post_end})

        return segments, n_seconds, n_pp_seconds

    def _resolve_bandpass(self):
        """Return the active (low_hz, high_hz) cutoffs or (None, None) if filtering is off.

        Cutoffs <= 0 disable the corresponding pass (e.g. low=0 -> low-pass only).
        Returns (None, None) when the bandpass checkbox is unchecked.
        """
        if not self.bandpass_cb.isChecked():
            return None, None
        low = float(self.bp_low_spin.value())
        high = float(self.bp_high_spin.value())
        nyq = float(self.sfreq) / 2.0 - 1e-3
        l = low if low > 0 else None
        h = high if high > 0 and high < nyq else (nyq if high >= nyq else None)
        if l is not None and h is not None and l >= h:
            # Make sure low cutoff is strictly below the high cutoff.
            h = min(max(l + 0.1, 0.2), nyq)
        return l, h

    def _ensure_filtered_signals(self):
        """Filter the full eeg0/eeg1 once with the active cutoffs and cache the result.

        Returns the (eeg0_to_use, eeg1_to_use) signals to compute PSDs on. If the
        bandpass checkbox is off the original signals are returned. The cache is
        invalidated whenever the cutoffs change.
        """
        l, h = self._resolve_bandpass()
        if l is None and h is None:
            return self.eeg0, self.eeg1
        cache_key = (l, h, len(self.eeg0) if self.eeg0 is not None else 0)
        if (self._filter_cache != cache_key
                or self._filtered_eeg0 is None
                or self._filtered_eeg1 is None):
            try:
                self._filtered_eeg0 = mne.filter.filter_data(
                    self.eeg0.astype(np.float64),
                    sfreq=self.sfreq, l_freq=l, h_freq=h, verbose=False)
                self._filtered_eeg1 = mne.filter.filter_data(
                    self.eeg1.astype(np.float64),
                    sfreq=self.sfreq, l_freq=l, h_freq=h, verbose=False)
                self._filter_cache = cache_key
            except Exception as e:
                QMessageBox.warning(
                    self, "Filter Error",
                    f"Failed to apply bandpass filter; using raw signal.\n{e}")
                self._filtered_eeg0 = None
                self._filtered_eeg1 = None
                self._filter_cache = None
                return self.eeg0, self.eeg1
        return self._filtered_eeg0, self._filtered_eeg1

    def compute_psd(self):
        """Compute multitaper PSD for the current event in a background thread"""
        if not hasattr(self, 'event'):
            QMessageBox.warning(self, "Error", "No event data available")
            return

        if self.psd_thread is not None:
            return

        fmin = self.fmin_spin.value()
        fmax = self.fmax_spin.value()
        bandwidth = self.bandwidth_spin.value()

        if fmin >= fmax:
            QMessageBox.warning(self, "Invalid Range", "Minimum frequency must be less than maximum frequency.")
            return

        segments, seg_seconds, prepost_seconds = self._build_segments()
        # Stash the active mode/segment-length so we can attach it to results.
        self._pending_mode = 'advanced' if self.advanced_cb.isChecked() else 'normal'
        self._pending_segment_seconds = seg_seconds
        self._pending_prepost_seconds = prepost_seconds
        bp_low, bp_high = self._resolve_bandpass()
        self._pending_bandpass = (bp_low, bp_high)

        # Warn (but proceed) if the event is shorter than the chosen segment length.
        if self._pending_mode == 'advanced':
            event_dur = self.event['duration_sec']
            if seg_seconds is not None and seg_seconds > event_dur:
                QMessageBox.information(
                    self, "Segment length adjusted",
                    f"Event duration is {event_dur:.2f}s but segment length is {seg_seconds:.2f}s.\n"
                    f"Segments will be clamped to the event duration.")

        # Resolve the source signals (filtered or raw) up-front so the worker is
        # cheap and keeps the heavy filter step off the GUI thread only on the
        # first call (subsequent calls reuse the cache).
        eeg0_src, eeg1_src = self._ensure_filtered_signals()

        self.compute_btn.setEnabled(False)
        self.save_btn.setEnabled(False)

        self.psd_thread = QThread(self)
        self.psd_worker = PSDWorker(
            eeg0_src, eeg1_src, self.sfreq, segments, fmin, fmax, bandwidth)
        self.psd_worker.moveToThread(self.psd_thread)
        self.psd_thread.started.connect(self.psd_worker.run)
        self.psd_worker.finished.connect(self._on_psd_finished)
        self.psd_worker.error.connect(self._on_psd_error)
        self.psd_worker.finished.connect(self._cleanup_psd_thread)
        self.psd_worker.error.connect(self._cleanup_psd_thread)
        self.psd_thread.start()

    def _cleanup_psd_thread(self, *_args):
        try:
            if self.psd_thread is not None:
                self.psd_thread.quit()
                self.psd_thread.wait()
        except Exception:
            pass
        finally:
            self.psd_thread = None
            self.psd_worker = None

        # Re-enable compute button
        self.compute_btn.setEnabled(True)

    def _on_psd_error(self, msg):
        QMessageBox.critical(self, "Error", f"Failed to compute PSD: {msg}")

    def _on_psd_finished(self, result):
        self.current_freqs = result['freqs']
        self.current_psd = {
            'mode': getattr(self, '_pending_mode', 'normal'),
            'segment_seconds': getattr(self, '_pending_segment_seconds', None),
            'prepost_seconds': getattr(self, '_pending_prepost_seconds', None),
            'bandpass': getattr(self, '_pending_bandpass', (None, None)),
            'segments': result['segments'],  # {name: {'psd_ch1','psd_ch2','start_sample','end_sample'}}
        }
        self.redraw_current_psd_plot()
        self.save_btn.setEnabled(True)
    
    # Map of segment name -> (display label, line style, color in light mode, color in dark mode)
    _SEGMENT_STYLES = {
        'full':   ('Full event',   '-',  '#3366CC', '#7BA3D0'),
        'pre':    ('Pre-SWD',      ':',  '#7F7F7F', '#B0B0B0'),
        'first':  ('First',        '-',  '#1F77B4', '#7BA3D0'),
        'middle': ('Middle',       '--', '#2CA02C', '#7FCB7F'),
        'last':   ('Last',         '-.', '#D62728', '#D67B7B'),
        'post':   ('Post-SWD',     ':',  '#9467BD', '#C5A3E0'),
    }

    # Stable ordering for segments wherever they appear (plot, export, summary).
    _SEGMENT_ORDER = ('full', 'pre', 'first', 'middle', 'last', 'post')

    def _segment_visible(self, name):
        """Whether a given segment should be drawn given the per-segment toggles."""
        if name == 'full':
            return True
        if name == 'pre':
            return self.show_pre_cb.isChecked()
        if name == 'first':
            return self.show_first_cb.isChecked()
        if name == 'middle':
            return self.show_middle_cb.isChecked()
        if name == 'last':
            return self.show_last_cb.isChecked()
        if name == 'post':
            return self.show_post_cb.isChecked()
        return True

    def _normalize_psd(self, psd, freqs):
        if not self.normalize_cb.isChecked():
            return psd
        area = np.trapz(psd, freqs)
        if area > 0:
            return psd / area
        return psd

    def _build_title_suffix(self, mode):
        """Compose a short summary string with the active advanced-mode parameters."""
        if mode != 'advanced' or self.current_psd is None:
            bp = (self.current_psd or {}).get('bandpass', (None, None)) if self.current_psd else (None, None)
            if bp and (bp[0] is not None or bp[1] is not None):
                lo = f"{bp[0]:.1f}" if bp[0] is not None else "DC"
                hi = f"{bp[1]:.1f}" if bp[1] is not None else "Nyq"
                return f" [BP {lo}-{hi}Hz]"
            return ''
        parts = []
        seg_s = self.current_psd.get('segment_seconds')
        if seg_s:
            parts.append(f"first/middle/last {seg_s:.2f}s")
        pp_s = self.current_psd.get('prepost_seconds')
        if pp_s:
            parts.append(f"±{pp_s:.2f}s pre/post")
        bp = self.current_psd.get('bandpass', (None, None)) or (None, None)
        if bp[0] is not None or bp[1] is not None:
            lo = f"{bp[0]:.1f}" if bp[0] is not None else "DC"
            hi = f"{bp[1]:.1f}" if bp[1] is not None else "Nyq"
            parts.append(f"BP {lo}-{hi}Hz")
        return f" - {' | '.join(parts)}" if parts else ''

    def redraw_current_psd_plot(self):
        if self.current_psd is None or self.current_freqs is None:
            return

        freqs = self.current_freqs
        normalized = self.normalize_cb.isChecked()
        mode = self.current_psd.get('mode', 'normal')
        segments = self.current_psd.get('segments', {})

        start_sample = self.event['start_sample'] if hasattr(self, 'event') else 0
        self.ax1.clear()
        self.ax2.clear()
        # Hover overlay artists lived on the cleared axes; drop stale refs.
        self._hover_artists = []

        # Get dark mode state from parent if available
        dark_mode = getattr(self.parent(), 'dark_mode', False) if self.parent() else False

        if dark_mode:
            bg_color = '#0D1117'
            plot_bg = '#161B22'
            grid_color = '#30363D'
        else:
            bg_color = 'white'
            plot_bg = 'white'
            grid_color = '#CCCCCC'

        self.fig.patch.set_facecolor(bg_color)
        self.ax1.set_facecolor(plot_bg)
        self.ax2.set_facecolor(plot_bg)

        title_color = '#C0C0C0' if dark_mode else 'black'
        axis_color = '#B0B0B0' if dark_mode else 'black'

        ch1_default_color = '#7BA3D0' if dark_mode else '#3366CC'
        ch2_default_color = '#D67B7B' if dark_mode else '#CC3333'

        def _segment_color(name, ch):
            style = self._SEGMENT_STYLES.get(name)
            if style is None:
                return ch1_default_color if ch == 1 else ch2_default_color
            return style[3] if dark_mode else style[2]

        def _segment_linestyle(name):
            style = self._SEGMENT_STYLES.get(name)
            return style[1] if style else '-'

        def _segment_label(name):
            if name == 'full':
                return 'Current'
            style = self._SEGMENT_STYLES.get(name)
            return style[0] if style else name.title()

        # Order segments deterministically for stable legends.
        ordered_names = [n for n in self._SEGMENT_ORDER if n in segments]
        # Append any unexpected names at the end to be safe.
        for n in segments.keys():
            if n not in ordered_names:
                ordered_names.append(n)

        # Plot a segment on a single axis.
        def _plot_segment(ax, name, ch_key, color_for_ch_default):
            seg = segments[name]
            psd = self._normalize_psd(seg[ch_key].copy(), freqs)
            color = color_for_ch_default if mode == 'normal' else _segment_color(name, 1 if ch_key == 'psd_ch1' else 2)
            linestyle = '-' if mode == 'normal' else _segment_linestyle(name)
            label = 'Current' if mode == 'normal' else _segment_label(name)
            ax.plot(freqs, psd, color=color, linewidth=2, linestyle=linestyle, label=label)

        # ---------------- Channel 1 ----------------
        if self.show_ch1_cb.isChecked():
            for name in ordered_names:
                if not self._segment_visible(name):
                    continue
                _plot_segment(self.ax1, name, 'psd_ch1', ch1_default_color)
            self._overlay_saved_on_axis(self.ax1, 'psd_ch1', freqs, normalized)

            self.ax1.set_ylabel('Power' + (' (normalized)' if normalized else ''),
                                 fontsize=11, color=axis_color)
            title_suffix = self._build_title_suffix(mode)
            self.ax1.set_title(
                f'Channel 1 - Multitaper PSD - Event at {start_sample/self.sfreq:.1f}s{title_suffix}',
                fontsize=12, fontweight='bold', color=title_color)
            self.ax1.grid(True, alpha=0.4, linestyle='--', linewidth=0.7, color=grid_color)
            self.ax1.tick_params(colors=axis_color)
            for spine in self.ax1.spines.values():
                spine.set_color(grid_color)
            self.ax1.xaxis.label.set_color(axis_color)
            self.ax1.yaxis.label.set_color(axis_color)
            self.ax1.legend(fontsize=8, loc='best')
            self.ax1.set_visible(True)
        else:
            self.ax1.set_visible(False)

        # ---------------- Channel 2 ----------------
        if self.show_ch2_cb.isChecked():
            for name in ordered_names:
                if not self._segment_visible(name):
                    continue
                _plot_segment(self.ax2, name, 'psd_ch2', ch2_default_color)
            self._overlay_saved_on_axis(self.ax2, 'psd_ch2', freqs, normalized)

            self.ax2.set_xlabel('Frequency (Hz)', fontsize=11, color=axis_color)
            self.ax2.set_ylabel('Power' + (' (normalized)' if normalized else ''),
                                 fontsize=11, color=axis_color)
            title_suffix = self._build_title_suffix(mode)
            self.ax2.set_title(
                f'Channel 2 - Multitaper PSD - Event at {start_sample/self.sfreq:.1f}s{title_suffix}',
                fontsize=12, fontweight='bold', color=title_color)
            self.ax2.grid(True, alpha=0.4, linestyle='--', linewidth=0.7, color=grid_color)
            self.ax2.tick_params(colors=axis_color)
            for spine in self.ax2.spines.values():
                spine.set_color(grid_color)
            self.ax2.xaxis.label.set_color(axis_color)
            self.ax2.yaxis.label.set_color(axis_color)
            self.ax2.legend(fontsize=8, loc='best')
            self.ax2.set_visible(True)
        else:
            self.ax2.set_visible(False)

        # If only one channel is shown, put the x-axis label on it
        if self.show_ch1_cb.isChecked() and not self.show_ch2_cb.isChecked():
            self.ax1.set_xlabel('Frequency (Hz)', fontsize=11, color=axis_color)

        # Synchronize y-limits between Channel 1 and Channel 2 axes so visual
        # comparison between the two channels is meaningful.
        if self.show_ch1_cb.isChecked() and self.show_ch2_cb.isChecked():
            try:
                y1_low, y1_high = self.ax1.get_ylim()
                y2_low, y2_high = self.ax2.get_ylim()
                shared = (min(y1_low, y2_low), max(y1_high, y2_high))
                self.ax1.set_ylim(shared)
                self.ax2.set_ylim(shared)
            except Exception:
                pass

        # Tidy the power (y) axis: PSD values are tiny (e.g. 1.2e-5), so plain
        # decimal ticks show a long run of zeros. Use scientific notation with a
        # shared exponent to drop the zeros, and coarsen the ticks so the step
        # between them is much larger (fewer, cleaner gridlines).
        if self.show_ch1_cb.isChecked():
            self._format_power_yaxis(self.ax1, axis_color)
        if self.show_ch2_cb.isChecked():
            self._format_power_yaxis(self.ax2, axis_color)

        self.fig.tight_layout()
        self.canvas.draw()

    def _format_power_yaxis(self, ax, axis_color):
        """Format a PSD power axis: scientific notation (no long zero runs) plus
        a much coarser tick step so the axis is easy to read."""
        try:
            from matplotlib.ticker import ScalarFormatter, MaxNLocator
            # Scientific notation with a common offset factor (e.g. x10^-5),
            # which removes the leading "0.0000..." zeros from every label.
            fmt = ScalarFormatter(useMathText=True)
            fmt.set_scientific(True)
            # Factor out the exponent for tiny (or huge) magnitudes so raw PSD
            # values lose their long zero runs, but leave normalized ~O(1)
            # values as plain numbers.
            fmt.set_powerlimits((-2, 5))
            ax.yaxis.set_major_formatter(fmt)
            # Far fewer ticks -> each step is roughly an order of magnitude
            # larger than matplotlib's dense default.
            ax.yaxis.set_major_locator(MaxNLocator(nbins=5, prune='both'))
            # Colour the offset (x10^-n) text to match the theme.
            try:
                ax.yaxis.get_offset_text().set_color(axis_color)
                ax.yaxis.get_offset_text().set_fontsize(9)
            except Exception:
                pass
        except Exception:
            pass

    def _clear_hover(self):
        """Remove any hover overlay artists and refresh the canvas."""
        if not getattr(self, '_hover_artists', None):
            return
        for art in self._hover_artists:
            try:
                art.remove()
            except Exception:
                pass
        self._hover_artists = []
        try:
            self.canvas.draw_idle()
        except Exception:
            pass

    def _on_plot_hover(self, event):
        """Show an aesthetic tooltip for the nearest curve point under the cursor.

        Draws a subtle vertical guide line, a highlighted marker on the closest
        curve, and a rounded floating label with the segment name, frequency and
        power value. Everything is redrawn on each move and cleared when the
        cursor leaves the axes.
        """
        ax = event.inaxes
        if ax not in (self.ax1, self.ax2) or event.xdata is None or event.ydata is None:
            self._clear_hover()
            return

        # Find the closest plotted point (by pixel distance) across all curves.
        best = None  # (pixel_dist, x, y, label, color)
        cursor_px = (event.x, event.y)
        for line in ax.get_lines():
            xdata = line.get_xdata()
            ydata = line.get_ydata()
            if xdata is None or len(xdata) == 0:
                continue
            xarr = np.asarray(xdata, dtype=float)
            idx = int(np.argmin(np.abs(xarr - event.xdata)))
            xv, yv = float(xarr[idx]), float(np.asarray(ydata)[idx])
            try:
                px, py = ax.transData.transform((xv, yv))
            except Exception:
                continue
            dist = float(np.hypot(px - cursor_px[0], py - cursor_px[1]))
            if best is None or dist < best[0]:
                best = (dist, xv, yv, line.get_label(), line.get_color())

        # Only show the tooltip when reasonably close to a curve.
        if best is None or best[0] > 45:
            self._clear_hover()
            return

        _dist, xv, yv, label, color = best
        self._clear_hover()

        dark_mode = getattr(self.parent(), 'dark_mode', False) if self.parent() else False
        guide_color = '#8899AA' if dark_mode else '#94A3B8'
        box_face = '#1B2430' if dark_mode else '#FFFFFF'
        text_color = '#E6EDF3' if dark_mode else '#0F172A'

        # Vertical guide line at the hovered frequency.
        vline = ax.axvline(xv, color=guide_color, linewidth=0.8, linestyle='--',
                           alpha=0.7, zorder=8)

        # Halo + solid marker on the nearest curve point.
        halo = ax.plot([xv], [yv], marker='o', markersize=12, color=color,
                       alpha=0.22, zorder=9, linestyle='None')[0]
        marker = ax.plot([xv], [yv], marker='o', markersize=6.5, color=color,
                         markeredgecolor='white', markeredgewidth=1.3,
                         zorder=10, linestyle='None')[0]

        clean_label = (label or '').lstrip('_').strip() or 'PSD'
        annotation = ax.annotate(
            f"{clean_label}\n{xv:.2f} Hz\n{yv:.4g}",
            xy=(xv, yv), xytext=(14, 14), textcoords='offset points',
            fontsize=9, fontweight='medium', color=text_color, zorder=11,
            ha='left', va='bottom',
            bbox=dict(boxstyle='round,pad=0.5', fc=box_face, ec=color,
                      lw=1.4, alpha=0.96),
            arrowprops=dict(arrowstyle='-', color=color, lw=1.0, alpha=0.7))

        self._hover_artists = [vline, halo, marker, annotation]
        try:
            self.canvas.draw_idle()
        except Exception:
            pass

    def _overlay_saved_on_axis(self, ax, ch_key, freqs, normalized):
        """Draw selected saved PSDs (and their segments) on the given axis."""
        if not (self.overlay_saved_cb.isChecked() and self.saved_list.selectedItems()):
            return

        overlay_colors = ['orange', 'green', 'purple', 'brown', 'pink', 'gray', 'olive', 'cyan']
        color_idx = 0
        for item in self.saved_list.selectedItems():
            psd_name = item.text()
            if psd_name not in self.saved_psds:
                continue
            saved_data = self.saved_psds[psd_name]
            saved_freqs = saved_data['freqs']
            saved_segments = saved_data.get('segments', {})

            # Order: full / pre / first / middle / last / post.
            ordered_saved = [n for n in self._SEGMENT_ORDER if n in saved_segments]
            for n in saved_segments.keys():
                if n not in ordered_saved:
                    ordered_saved.append(n)

            for seg_name in ordered_saved:
                if not self._segment_visible(seg_name):
                    continue
                seg_data = saved_segments[seg_name]
                psd_arr = seg_data.get(ch_key)
                if psd_arr is None:
                    continue
                if len(saved_freqs) != len(freqs) or not np.allclose(saved_freqs, freqs):
                    interp_psd = np.interp(freqs, saved_freqs, psd_arr)
                else:
                    interp_psd = psd_arr.copy()
                if normalized:
                    area = np.trapz(interp_psd, freqs)
                    if area > 0:
                        interp_psd = interp_psd / area
                overlay_color = overlay_colors[color_idx % len(overlay_colors)]
                color_idx += 1
                seg_label = '' if seg_name == 'full' else f' [{seg_name}]'
                ax.plot(freqs, interp_psd, color=overlay_color, linewidth=1.5,
                        linestyle='--', alpha=0.7,
                        label=f'Saved: {psd_name[:20]}{seg_label}')
    
    def update_plot_visibility(self):
        """Update plot visibility based on channel checkboxes without recomputing"""
        if hasattr(self, 'current_psd') and self.current_psd is not None and self.current_freqs is not None:
            self.redraw_current_psd_plot()
            
    def save_current_psd(self):
        """Save the current PSD to the saved list"""
        if self.current_psd is None:
            return

        mode = self.current_psd.get('mode', 'normal')
        # Sequential labels in the saved list: SWD 1, SWD 2, ...
        n = 1
        while f"SWD {n}" in self.saved_psds:
            n += 1
        name = f"SWD {n}"

        # Deep-copy the segment PSDs so subsequent computations don't mutate them.
        saved_segments = {}
        for seg_name, seg_data in self.current_psd.get('segments', {}).items():
            saved_segments[seg_name] = {
                'psd_ch1': seg_data['psd_ch1'].copy(),
                'psd_ch2': seg_data['psd_ch2'].copy(),
                'start_sample': int(seg_data['start_sample']),
                'end_sample': int(seg_data['end_sample']),
            }

        self.saved_psds[name] = {
            'freqs': self.current_freqs.copy(),
            'mode': mode,
            'segment_seconds': self.current_psd.get('segment_seconds'),
            'prepost_seconds': self.current_psd.get('prepost_seconds'),
            'bandpass': self.current_psd.get('bandpass', (None, None)),
            'segments': saved_segments,
            'event': self.event.copy(),
            'fmin': self.fmin_spin.value(),
            'fmax': self.fmax_spin.value(),
            'bandwidth': self.bandwidth_spin.value(),
            'normalized': self.normalize_cb.isChecked(),
        }

        self.saved_list.addItem(name)
        
    def export_psds(self):
        """Export saved PSDs to Excel with specified number of frequency bins.

        - Normal-mode PSDs are written to ``Channel_1_PSDs`` / ``Channel_2_PSDs``.
        - Advanced-mode PSDs are written to per-segment sheets:
          ``Pre_Channel_1_PSDs``, ``Pre_Channel_2_PSDs``,
          ``First_Channel_1_PSDs``, ``First_Channel_2_PSDs``,
          ``Middle_Channel_1_PSDs``, ``Middle_Channel_2_PSDs``,
          ``Last_Channel_1_PSDs``, ``Last_Channel_2_PSDs``,
          ``Post_Channel_1_PSDs``, ``Post_Channel_2_PSDs``.
        - When the "Normalize" checkbox is on, every exported PSD is divided by
          its own area (trapezoidal integral over the frequency axis).
        - A ``Summary`` sheet describes every saved PSD.
        """
        if not self.saved_psds:
            QMessageBox.warning(self, "Error", "No PSDs saved to export")
            return

        filename, _ = QFileDialog.getSaveFileName(
            self, "Export PSDs", "", "Excel Files (*.xlsx);;All Files (*)")

        if not filename:
            return

        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'

        try:
            target_bins = self.export_bins_spin.value()
            normalize_export = self.normalize_cb.isChecked()

            # Build target frequency axis from the union of all saved PSD ranges.
            all_fmin = min(d['freqs'][0] for d in self.saved_psds.values())
            all_fmax = max(d['freqs'][-1] for d in self.saved_psds.values())
            target_freqs = np.linspace(all_fmin, all_fmax, target_bins)

            def _maybe_normalize(psd_arr):
                if not normalize_export:
                    return psd_arr
                area = np.trapz(psd_arr, target_freqs)
                if area > 0:
                    return psd_arr / area
                return psd_arr

            # Per-segment dataframes (one per channel + segment combination).
            segment_dfs = {}  # key: (segment_name, ch_idx) -> DataFrame

            def _ensure_df(seg_name, ch_idx):
                key = (seg_name, ch_idx)
                if key not in segment_dfs:
                    segment_dfs[key] = pd.DataFrame(index=target_freqs)
                return segment_dfs[key]

            for name, psd_data in self.saved_psds.items():
                original_freqs = psd_data['freqs']
                for seg_name, seg_data in psd_data.get('segments', {}).items():
                    interp_psd1 = np.interp(target_freqs, original_freqs, seg_data['psd_ch1'])
                    interp_psd2 = np.interp(target_freqs, original_freqs, seg_data['psd_ch2'])
                    interp_psd1 = _maybe_normalize(interp_psd1)
                    interp_psd2 = _maybe_normalize(interp_psd2)
                    _ensure_df(seg_name, 1)[name] = interp_psd1
                    _ensure_df(seg_name, 2)[name] = interp_psd2

            # Friendly sheet names per segment.
            segment_label = {
                'full': '',
                'pre': 'Pre_',
                'first': 'First_',
                'middle': 'Middle_',
                'last': 'Last_',
                'post': 'Post_',
            }

            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Stable sheet ordering: full / pre / first / middle / last / post,
                # then anything else.
                order = list(self._SEGMENT_ORDER)
                segment_keys = sorted(
                    segment_dfs.keys(),
                    key=lambda k: (order.index(k[0]) if k[0] in order else 99, k[1]))
                for (seg_name, ch_idx) in segment_keys:
                    prefix = segment_label.get(seg_name, f'{seg_name.title()}_')
                    sheet = f"{prefix}Channel_{ch_idx}_PSDs"
                    segment_dfs[(seg_name, ch_idx)].to_excel(writer, sheet_name=sheet)

                # Summary sheet describing each saved PSD.
                summary_data = []
                for name, psd_data in self.saved_psds.items():
                    event_info = psd_data['event']
                    seg_durations = []
                    for seg_name in self._SEGMENT_ORDER:
                        if seg_name not in psd_data.get('segments', {}):
                            continue
                        seg_data = psd_data['segments'][seg_name]
                        dur = (seg_data['end_sample'] - seg_data['start_sample']) / self.sfreq
                        seg_durations.append(f"{seg_name}={dur:.2f}s")
                    bp = psd_data.get('bandpass', (None, None)) or (None, None)
                    bp_str = ''
                    if bp[0] is not None or bp[1] is not None:
                        lo = f"{bp[0]:.2f}" if bp[0] is not None else "DC"
                        hi = f"{bp[1]:.2f}" if bp[1] is not None else "Nyq"
                        bp_str = f"{lo}-{hi} Hz"
                    summary_data.append({
                        'PSD_Name': name,
                        'Mode': psd_data.get('mode', 'normal'),
                        'Segment_Length_s': psd_data.get('segment_seconds') or '',
                        'PrePost_Length_s': psd_data.get('prepost_seconds') or '',
                        'Bandpass_Filter': bp_str,
                        'Segments': ', '.join(seg_durations),
                        'Start_Sample': event_info['start_sample'],
                        'End_Sample': event_info['end_sample'],
                        'Duration_s': event_info['duration_sec'],
                        'Freq_Min_Hz': psd_data['fmin'],
                        'Freq_Max_Hz': psd_data['fmax'],
                        'Bandwidth': psd_data['bandwidth'],
                        'Normalized_On_Save': psd_data['normalized'],
                        'Normalized_On_Export': normalize_export,
                        'Exported_Bins': target_bins,
                    })

                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

            norm_note = " (normalized)" if normalize_export else ""
            QMessageBox.information(
                self, "Export Successful",
                f"Exported {len(self.saved_psds)} PSDs across {len(segment_dfs)} sheet(s) "
                f"with {target_bins} frequency bins{norm_note} to {filename}")

        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export: {str(e)}")
            
    def clear_saved_psds(self):
        """Clear all saved PSDs"""
        reply = QMessageBox.question(
            self, "Confirm Clear", 
            "Are you sure you want to clear all saved PSDs?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.saved_psds.clear()
            self.saved_list.clear()

    # ------------------------------------------------------------------
    # Sanity self-check (multitaper vs scipy.signal.welch).
    # ------------------------------------------------------------------
    def run_self_check(self):
        """Compare our MNE multitaper PSD against scipy.signal.welch on the
        current event's "full" segment, plus a raw-vs-filtered overlay if a
        bandpass is active. This is a QA tool only; results are not saved or
        exported.
        """
        if not hasattr(self, 'event') or self.eeg0 is None:
            QMessageBox.warning(self, "No Event", "No event/data is currently loaded for validation.")
            return

        try:
            start = max(0, int(self.event['start_sample']))
            end = min(len(self.eeg0), int(self.event['end_sample']))
            if end - start < 8:
                QMessageBox.warning(
                    self, "Segment Too Short",
                    f"Current event window has only {end - start} samples; cannot validate.")
                return

            fmin = float(self.fmin_spin.value())
            fmax = float(self.fmax_spin.value())
            bandwidth = float(self.bandwidth_spin.value())
            sfreq = float(self.sfreq)

            # Same processing path the production PSD uses.
            eeg0_proc, eeg1_proc = self._ensure_filtered_signals()
            single_channel = bool(getattr(self.parent(), 'single_channel_view', False))

            sig1 = np.asarray(eeg0_proc[start:end], dtype=np.float64).copy()
            sig2 = (None if single_channel
                    else np.asarray(eeg1_proc[start:end], dtype=np.float64).copy())

            # Multitaper PSD (production function, normalized=False here so we
            # compare absolute power densities; both are normalized later for
            # the visual overlay).
            f_mt, p_mt1 = compute_multitaper_psd(sig1, sfreq, fmin, fmax, bandwidth, normalize=False)
            p_mt2 = None
            if sig2 is not None:
                _, p_mt2 = compute_multitaper_psd(sig2, sfreq, fmin, fmax, bandwidth, normalize=False)

            # Welch reference. Pick a window that resolves the requested band
            # without exceeding the segment length.
            n_seg = end - start
            target_nperseg = max(64, int(round(sfreq / max(0.5, bandwidth / 2.0))))
            nperseg = int(min(n_seg, target_nperseg))
            noverlap = nperseg // 2
            f_w, p_w1 = welch(sig1, fs=sfreq, nperseg=nperseg, noverlap=noverlap, scaling='density')
            p_w2 = None
            if sig2 is not None:
                _, p_w2 = welch(sig2, fs=sfreq, nperseg=nperseg, noverlap=noverlap, scaling='density')

            # Restrict Welch to the multitaper band for fair comparison.
            band_mask = (f_w >= f_mt[0]) & (f_w <= f_mt[-1])
            f_w_band = f_w[band_mask]
            p_w1_band = p_w1[band_mask]
            p_w2_band = p_w2[band_mask] if p_w2 is not None else None

            metrics_ch1 = self._psd_agreement(f_mt, p_mt1, f_w_band, p_w1_band)
            metrics_ch2 = (self._psd_agreement(f_mt, p_mt2, f_w_band, p_w2_band)
                           if p_mt2 is not None else None)

            # Filter sanity: PSD of the raw (unfiltered) window so the user can
            # visually confirm passband attenuation when the bandpass is on.
            bp_low, bp_high = self._resolve_bandpass()
            filter_active = (bp_low is not None) or (bp_high is not None)
            f_raw, p_raw1 = (None, None)
            p_raw2 = None
            if filter_active:
                raw1 = np.asarray(self.eeg0[start:end], dtype=np.float64)
                f_raw, p_raw1 = welch(raw1, fs=sfreq, nperseg=nperseg, noverlap=noverlap, scaling='density')
                if sig2 is not None:
                    raw2 = np.asarray(self.eeg1[start:end], dtype=np.float64)
                    _, p_raw2 = welch(raw2, fs=sfreq, nperseg=nperseg, noverlap=noverlap, scaling='density')

            self._show_self_check_dialog(
                f_mt=f_mt, p_mt=(p_mt1, p_mt2),
                f_w=f_w_band, p_w=(p_w1_band, p_w2_band),
                f_raw=f_raw, p_raw=(p_raw1, p_raw2),
                metrics=(metrics_ch1, metrics_ch2),
                meta=dict(
                    n_seg=n_seg, sfreq=sfreq, fmin=fmin, fmax=fmax,
                    bandwidth=bandwidth, nperseg=nperseg, noverlap=noverlap,
                    bp=(bp_low, bp_high), single_channel=single_channel,
                    start_sample=start, end_sample=end,
                ),
            )
        except Exception as e:
            import traceback
            QMessageBox.critical(
                self, "Validate failed",
                f"Self-check could not run.\n\n{e}\n\n{traceback.format_exc()}")

    @staticmethod
    def _psd_agreement(f_mt, p_mt, f_w, p_w):
        """Return a dict of metrics comparing two PSDs over the same band.

        Both PSDs are interpolated onto the Welch grid (it tends to be coarser
        than multitaper for short segments) before metrics are computed.
        """
        if p_mt is None or p_w is None or len(f_w) < 2:
            return None
        p_mt_i = np.interp(f_w, f_mt, p_mt)
        a_mt = float(np.trapz(p_mt_i, f_w))
        a_w = float(np.trapz(p_w, f_w))
        area_ratio = (a_mt / a_w) if a_w > 0 else float('nan')
        peak_mt = float(f_w[int(np.argmax(p_mt_i))])
        peak_w = float(f_w[int(np.argmax(p_w))])
        # Normalized RMSE: divide each PSD by its area then compare shape.
        n_mt = p_mt_i / a_mt if a_mt > 0 else p_mt_i
        n_w = p_w / a_w if a_w > 0 else p_w
        rmse = float(np.sqrt(np.mean((n_mt - n_w) ** 2)))
        # Pearson correlation of normalized shapes.
        if np.std(n_mt) > 0 and np.std(n_w) > 0:
            corr = float(np.corrcoef(n_mt, n_w)[0, 1])
        else:
            corr = float('nan')

        # Verdict thresholds calibrated against synthetic signals across
        # 0.5-30s windows at 512/1000 Hz: area_ratio in [0.7, 1.3], peak
        # within 1 Hz, normalized RMSE < 0.30, correlation > 0.90.
        ok_area = 0.7 <= area_ratio <= 1.3
        ok_peak = abs(peak_mt - peak_w) <= 1.0
        ok_rmse = rmse < 0.30
        ok_corr = (not np.isnan(corr)) and corr > 0.90
        verdict = 'PASS' if (ok_area and ok_peak and ok_rmse and ok_corr) else 'WARN'
        return dict(
            area_mt=a_mt, area_w=a_w, area_ratio=area_ratio,
            peak_mt=peak_mt, peak_w=peak_w,
            rmse_normalized=rmse, corr=corr, verdict=verdict,
            ok_area=ok_area, ok_peak=ok_peak, ok_rmse=ok_rmse, ok_corr=ok_corr,
        )

    def _show_self_check_dialog(self, *, f_mt, p_mt, f_w, p_w, f_raw, p_raw, metrics, meta):
        """Build and show the Validate popup with overlay plots + metrics."""
        dlg = QDialog(self)
        dlg.setWindowTitle("PSD Sanity Check (multitaper vs Welch)")
        dlg.resize(1100, 800)
        v = QVBoxLayout(dlg)

        single = meta['single_channel']
        nrows = 1 if single else 2
        fig, axes = plt.subplots(nrows, 1, figsize=(11, 4 * nrows), squeeze=False)
        axes = axes[:, 0]

        dark_mode = bool(getattr(self.parent(), 'dark_mode', False))
        if dark_mode:
            fig.patch.set_facecolor('#0D1117')
            for a in axes:
                a.set_facecolor('#161B22')
                a.tick_params(colors='#B0B0B0')
                for s in a.spines.values():
                    s.set_color('#30363D')
            text_c = '#C0C0C0'; grid_c = '#30363D'
        else:
            text_c = 'black'; grid_c = '#CCCCCC'

        ch_arrays = [(p_mt[0], p_w[0], p_raw[0] if p_raw is not None else None, 'Channel 1')]
        if not single and p_mt[1] is not None:
            ch_arrays.append((p_mt[1], p_w[1], p_raw[1] if p_raw is not None else None, 'Channel 2'))

        for ax, (pm, pw, pr, name) in zip(axes, ch_arrays):
            ax.plot(f_mt, pm, color='#3366CC', linewidth=2.0, label='Our multitaper')
            ax.plot(f_w, pw, color='#D62728', linewidth=1.4, linestyle='--', label='scipy.welch')
            if pr is not None:
                ax.plot(f_raw, pr, color='#999999', linewidth=1.0, linestyle=':', alpha=0.8,
                        label='Welch (raw, pre-filter)')
            ax.set_yscale('log')
            ax.grid(True, alpha=0.4, linestyle='--', linewidth=0.7, color=grid_c)
            ax.set_ylabel('Power (V²/Hz)', color=text_c)
            ax.set_title(f'{name} — sanity check', color=text_c, fontweight='bold')
            ax.tick_params(colors=text_c)
            ax.xaxis.label.set_color(text_c)
            ax.yaxis.label.set_color(text_c)
            ax.legend(fontsize=8, loc='best')
        axes[-1].set_xlabel('Frequency (Hz)', color=text_c)

        fig.tight_layout()
        canvas = FigureCanvas(fig)
        v.addWidget(canvas, stretch=1)

        report = self._format_self_check_report(metrics, meta)
        report_lbl = QLabel(report)
        report_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        report_lbl.setStyleSheet(
            "QLabel { font-family: Consolas, 'Courier New', monospace; "
            "font-size: 11px; padding: 8px; "
            f"background-color: {'#0D1117' if dark_mode else '#F6F8FA'}; "
            f"color: {'#C0C0C0' if dark_mode else 'black'}; "
            "border: 1px solid #888; }"
        )
        v.addWidget(report_lbl)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        btn_row.addWidget(close_btn)
        v.addLayout(btn_row)

        dlg.exec_()

    @staticmethod
    def _format_self_check_report(metrics, meta):
        """Plain-text report for the self-check dialog."""
        m1, m2 = metrics
        bp_low, bp_high = meta['bp']
        bp_str = "off"
        if bp_low is not None or bp_high is not None:
            lo = f"{bp_low:.2f}" if bp_low is not None else "DC"
            hi = f"{bp_high:.2f}" if bp_high is not None else "Nyq"
            bp_str = f"{lo}–{hi} Hz"

        lines = []
        lines.append("=== PSD self-check ===")
        lines.append(f"Window: samples [{meta['start_sample']}, {meta['end_sample']})  "
                     f"= {meta['n_seg']} samples = {meta['n_seg']/meta['sfreq']:.3f} s "
                     f"@ {meta['sfreq']:.0f} Hz")
        lines.append(f"Multitaper: fmin={meta['fmin']:.2f}  fmax={meta['fmax']:.2f}  "
                     f"bandwidth={meta['bandwidth']:.2f} Hz  (NW={meta['n_seg']*meta['bandwidth']/(2*meta['sfreq']):.2f})")
        lines.append(f"Welch:      nperseg={meta['nperseg']}  noverlap={meta['noverlap']}  "
                     f"scaling=density")
        lines.append(f"Bandpass before PSD: {bp_str}")
        lines.append("")
        lines.append("                            Channel 1                      Channel 2")
        lines.append("                            ----------                     ----------")

        def col(m):
            if m is None:
                return f"{'(single channel)':<30}"
            ok = lambda b: "✓" if b else "✗"
            return (f"area_ratio={m['area_ratio']:.3f}{ok(m['ok_area'])}  "
                    f"peak Δ={abs(m['peak_mt']-m['peak_w']):.2f}Hz{ok(m['ok_peak'])}")

        def fmt(metric, key, spec):
            if metric is None:
                return 'n/a'
            return format(metric[key], spec)

        rmse1 = fmt(m1, 'rmse_normalized', '.4f')
        rmse2 = fmt(m2, 'rmse_normalized', '.4f')
        corr1 = fmt(m1, 'corr', '.4f')
        corr2 = fmt(m2, 'corr', '.4f')
        amt1 = (fmt(m1, 'area_mt', '.3e') + ' V²') if m1 else 'n/a'
        amt2 = (fmt(m2, 'area_mt', '.3e') + ' V²') if m2 else 'n/a'
        aw1 = (fmt(m1, 'area_w', '.3e') + ' V²') if m1 else 'n/a'
        aw2 = (fmt(m2, 'area_w', '.3e') + ' V²') if m2 else 'n/a'
        v1 = m1['verdict'] if m1 else 'n/a'
        v2 = m2['verdict'] if m2 else 'n/a'

        lines.append(f"{'Agreement vs Welch:':<26}  {col(m1):<30}  {col(m2)}")
        lines.append(f"{'Normalized-shape RMSE:':<26}  {rmse1:<30}  {rmse2}")
        lines.append(f"{'Pearson correlation:':<26}  {corr1:<30}  {corr2}")
        lines.append(f"{'Total power (multitaper):':<26}  {amt1:<30}  {amt2}")
        lines.append(f"{'Total power (Welch):':<26}  {aw1:<30}  {aw2}")
        lines.append(f"{'Verdict:':<26}  {v1:<30}  {v2}")
        lines.append("")
        lines.append("Thresholds: area_ratio∈[0.7,1.3], |Δpeak|≤1 Hz, "
                     "normalized RMSE<0.30, corr>0.90.")
        lines.append("Note: Welch and multitaper inherently differ in spectral smoothing; "
                     "small disagreements are expected, especially for short windows.")
        return "\n".join(lines)


# -----------------------------------------------------------------------------
# Channel Selection Dialog - For selecting channels and animal name
# -----------------------------------------------------------------------------
class ChannelSelectionDialog(QDialog):
    """Dialog for selecting two channels and setting the animal name"""
    
    def __init__(self, channel_names, parent=None, default_animal_name="Animal_1", 
                 allow_single_channel=True, title="Select Channels"):
        super().__init__(parent)
        self.channel_names = channel_names
        self.default_animal_name = default_animal_name
        self.allow_single_channel = allow_single_channel
        self.selected_channels = None
        self.animal_name = None
        self.unet_channel_index = 0
        self._name_manually_edited = False
        
        self.setWindowTitle(title)
        self.setModal(True)
        self.setMinimumWidth(400)
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Animal name
        name_group = QGroupBox("Animal Identification")
        name_layout = QHBoxLayout(name_group)
        name_layout.addWidget(QLabel("Animal Name:"))
        self.name_edit = QLineEdit(self.default_animal_name)
        self.name_edit.setPlaceholderText("Enter animal name...")
        self.name_edit.textEdited.connect(self._on_name_manually_edited)
        name_layout.addWidget(self.name_edit)
        layout.addWidget(name_group)
        
        # Channel selection
        channel_group = QGroupBox("Channel Selection")
        channel_layout = QGridLayout(channel_group)
        
        channel_layout.addWidget(QLabel("Channel 1 (Primary):"), 0, 0)
        self.ch1_combo = QComboBox()
        self.ch1_combo.addItems(self.channel_names)
        # Pre-select first channel (index 0)
        self.ch1_combo.setCurrentIndex(0)
        self.ch1_combo.currentIndexChanged.connect(self._update_ch2_options)
        self.ch1_combo.currentTextChanged.connect(self._update_animal_name_from_channel)
        channel_layout.addWidget(self.ch1_combo, 0, 1)
        
        channel_layout.addWidget(QLabel("Channel 2 (Secondary):"), 1, 0)
        self.ch2_combo = QComboBox()
        # Populate ch2 excluding the first channel (which is selected in ch1)
        if self.allow_single_channel:
            self.ch2_combo.addItem("Single Channel Mode")
        first_channel = self.channel_names[0] if self.channel_names else None
        for ch in self.channel_names:
            if ch != first_channel:
                self.ch2_combo.addItem(ch)
        # Pre-select second channel (first available after excluding ch1)
        if len(self.channel_names) > 1:
            # Index 1 if single channel mode exists, index 0 otherwise (both point to second channel)
            self.ch2_combo.setCurrentIndex(1 if self.allow_single_channel else 0)
        channel_layout.addWidget(self.ch2_combo, 1, 1)
        
        layout.addWidget(channel_group)
        
        # UNET channel selection
        unet_group = QGroupBox("UNET Processing Channel")
        unet_layout = QHBoxLayout(unet_group)
        unet_layout.addWidget(QLabel("Use for UNET:"))
        self.unet_combo = QComboBox()
        self.unet_combo.addItems(["Channel 1", "Channel 2"])
        unet_layout.addWidget(self.unet_combo)
        layout.addWidget(unet_group)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        ok_btn = QPushButton("OK")
        ok_btn.setDefault(True)
        ok_btn.clicked.connect(self._accept)
        btn_layout.addWidget(ok_btn)
        
        layout.addLayout(btn_layout)
        
    def _on_name_manually_edited(self):
        self._name_manually_edited = True

    def _update_animal_name_from_channel(self, channel_text):
        """Auto-update animal name to first word of selected channel 1"""
        if not self._name_manually_edited and channel_text.strip():
            first_word = channel_text.split()[0].strip()
            if first_word:
                self.name_edit.setText(first_word)

    def _update_ch2_options(self):
        """Update channel 2 options when channel 1 changes"""
        current_ch1 = self.ch1_combo.currentText()
        current_ch2 = self.ch2_combo.currentText()
        
        self.ch2_combo.blockSignals(True)
        self.ch2_combo.clear()
        
        if self.allow_single_channel:
            self.ch2_combo.addItem("Single Channel Mode")
            
        for ch in self.channel_names:
            if ch != current_ch1:
                self.ch2_combo.addItem(ch)
                
        # Try to restore previous selection
        idx = self.ch2_combo.findText(current_ch2)
        if idx >= 0:
            self.ch2_combo.setCurrentIndex(idx)
            
        self.ch2_combo.blockSignals(False)
        
    def _accept(self):
        """Validate and accept the dialog"""
        self.animal_name = self.name_edit.text().strip()
        if not self.animal_name:
            QMessageBox.warning(self, "Invalid Name", "Please enter an animal name.")
            return
            
        ch1 = self.ch1_combo.currentText()
        ch2 = self.ch2_combo.currentText()
        
        ch1_idx = self.channel_names.index(ch1)
        
        if ch2 == "Single Channel Mode":
            ch2_idx = ch1_idx  # Same channel
            self.selected_channels = (ch1, None)
            self.selected_channel_indices = (ch1_idx, None)
        else:
            ch2_idx = self.channel_names.index(ch2)
            self.selected_channels = (ch1, ch2)
            self.selected_channel_indices = (ch1_idx, ch2_idx)
            
        self.unet_channel_index = self.unet_combo.currentIndex()
        self.accept()
        
    def get_selection(self):
        """Return the selection as a dictionary"""
        return {
            'animal_name': self.animal_name,
            'channels': self.selected_channels,
            'channel_indices': self.selected_channel_indices,
            'unet_channel': self.unet_channel_index
        }


# -----------------------------------------------------------------------------
# Comment import helpers (for LabChart .adicht files and plain text/CSV files)
# -----------------------------------------------------------------------------
# Comments are stored as a list of dicts:
#   {'time_sec': float (offset from recording_start, seconds),
#    'hhmm': 'HH:MM',
#    'hhmmss': 'HH:MM:SS',
#    'text': str}
#
# The Excel analysis code needs HH:MM times. Because the EDF does not carry the
# adicht comments, this loader lets the user import them from the original
# .adicht (preferred) or from a plain-text/CSV dump exported from LabChart.
# Parsing is intentionally defensive: any failure raises with a clear message
# so the UI can show it without crashing.

def _seconds_to_hhmm_str(total_seconds, include_seconds=False):
    """Convert an absolute second count (since midnight) to HH:MM[:SS]."""
    total_seconds = int(round(total_seconds)) % (24 * 3600)
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    if include_seconds:
        return f"{h:02d}:{m:02d}:{s:02d}"
    return f"{h:02d}:{m:02d}"


def _comment_entry(offset_sec, text, recording_start_dt):
    """Build a comment entry dict from an offset (seconds from recording start)."""
    try:
        offset_sec = float(offset_sec)
    except Exception:
        offset_sec = 0.0
    abs_dt = None
    try:
        if recording_start_dt is not None:
            abs_dt = recording_start_dt + timedelta(seconds=offset_sec)
    except Exception:
        abs_dt = None
    if abs_dt is not None:
        hhmm = abs_dt.strftime('%H:%M')
        hhmmss = abs_dt.strftime('%H:%M:%S')
    else:
        # Fall back to offset-as-clock (treat as seconds-since-midnight)
        hhmm = _seconds_to_hhmm_str(offset_sec, include_seconds=False)
        hhmmss = _seconds_to_hhmm_str(offset_sec, include_seconds=True)
    return {
        'time_sec': offset_sec,
        'hhmm': hhmm,
        'hhmmss': hhmmss,
        'text': str(text) if text is not None else '',
    }


def parse_adicht_comments(filepath, recording_start_dt=None):
    """Parse comments from a LabChart .adicht file.

    Requires the optional `adi` Python package (ADInstruments SDK wrapper).
    Raises ImportError with install instructions if the package is missing.
    Raises RuntimeError on any parsing failure.
    """
    try:
        import adi  # type: ignore
    except Exception as e:
        raise ImportError(
            "Reading .adicht files requires the 'adi' Python package.\n"
            "Install it with:  pip install adi-reader\n"
            f"(Import error: {e})"
        )

    try:
        f = adi.read_file(filepath)
    except Exception as e:
        raise RuntimeError(f"Failed to open .adicht file: {e}")

    comments = []
    try:
        n_records = getattr(f, 'n_records', None)
        records = getattr(f, 'records', None)
        if records is None and n_records:
            records = [f.get_record(i) for i in range(n_records)]
        if records is None:
            raise RuntimeError("adi library returned no records for this file")

        # Offset from the start of record 0 accumulates across earlier records
        # so multi-record files still produce sensible absolute times.
        prev_records_duration = 0.0
        for rec_idx, rec in enumerate(records):
            tick_dt = float(getattr(rec, 'tick_dt', 0.0) or 0.0)
            n_ticks = int(getattr(rec, 'n_ticks', 0) or 0)
            rec_comments = getattr(rec, 'comments', None) or []
            for c in rec_comments:
                tick_pos = int(getattr(c, 'tick_position', 0) or 0)
                text = getattr(c, 'text', '') or getattr(c, 'comment', '')
                offset_in_rec = tick_pos * tick_dt if tick_dt > 0 else 0.0
                offset_total = prev_records_duration + offset_in_rec
                comments.append(_comment_entry(offset_total, text, recording_start_dt))
            if tick_dt > 0 and n_ticks > 0:
                prev_records_duration += n_ticks * tick_dt
    except Exception as e:
        raise RuntimeError(f"Failed to extract comments from .adicht: {e}")

    comments.sort(key=lambda c: c['time_sec'])
    return comments


def parse_text_comments(filepath, recording_start_dt=None):
    """Parse comments from a plain text / CSV file.

    Accepted per-line formats (first match wins):
      * HH:MM:SS<sep>TEXT     (absolute clock time)
      * HH:MM<sep>TEXT        (absolute clock time)
      * <seconds><sep>TEXT    (offset from recording start)
    Separator is tab, comma, or semicolon. Header lines containing no time
    token are silently ignored.
    """
    import csv, re

    # Read all raw rows first (handle both tab-separated and CSV)
    rows = []
    try:
        with open(filepath, 'r', encoding='utf-8-sig', errors='replace') as fh:
            sample = fh.read(4096)
            fh.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters="\t,;")
            except Exception:
                dialect = csv.excel_tab
            reader = csv.reader(fh, dialect)
            for r in reader:
                if r and any(cell.strip() for cell in r):
                    rows.append([c.strip() for c in r])
    except Exception as e:
        raise RuntimeError(f"Failed to open comments file: {e}")

    time_re_hms = re.compile(r'^(\d{1,2}):(\d{2}):(\d{2})(?:[.,](\d+))?$')
    time_re_hm = re.compile(r'^(\d{1,2}):(\d{2})$')
    num_re = re.compile(r'^-?\d+(?:[.,]\d+)?$')

    # Recording-start seconds-since-midnight for HH:MM conversions.
    rec_start_sec = None
    if recording_start_dt is not None:
        try:
            rec_start_sec = (recording_start_dt.hour * 3600
                             + recording_start_dt.minute * 60
                             + recording_start_dt.second
                             + recording_start_dt.microsecond / 1e6)
        except Exception:
            rec_start_sec = None

    comments = []
    for row in rows:
        if not row:
            continue
        time_tok = row[0]
        text = (row[1] if len(row) > 1 else '').strip()
        if len(row) > 2:
            # Join any trailing columns (e.g. CSV with commas in text)
            text = ','.join(cell for cell in row[1:] if cell)

        offset_sec = None
        m = time_re_hms.match(time_tok)
        if m:
            h, mi, s = int(m.group(1)), int(m.group(2)), int(m.group(3))
            frac = float('0.' + m.group(4)) if m.group(4) else 0.0
            abs_sec = h * 3600 + mi * 60 + s + frac
            if rec_start_sec is not None:
                offset_sec = abs_sec - rec_start_sec
                # If the comment is earlier than recording_start, assume it is
                # actually on the following day (midnight-crossing recording).
                if offset_sec < -3600:
                    offset_sec += 24 * 3600
            else:
                offset_sec = abs_sec
        else:
            m = time_re_hm.match(time_tok)
            if m:
                h, mi = int(m.group(1)), int(m.group(2))
                abs_sec = h * 3600 + mi * 60
                if rec_start_sec is not None:
                    offset_sec = abs_sec - rec_start_sec
                    if offset_sec < -3600:
                        offset_sec += 24 * 3600
                else:
                    offset_sec = abs_sec
            elif num_re.match(time_tok):
                try:
                    offset_sec = float(time_tok.replace(',', '.'))
                except Exception:
                    offset_sec = None

        if offset_sec is None or not text:
            continue
        comments.append(_comment_entry(offset_sec, text, recording_start_dt))

    comments.sort(key=lambda c: c['time_sec'])
    return comments


class CommentPickerDialog(QDialog):
    """Modal popup that lets the user pick one of the imported comments.

    Returns the selected comment dict via `self.chosen_comment` (None if
    cancelled). Intentionally lightweight so it can be re-opened from any
    HH:MM field without feeling heavy.
    """

    def __init__(self, comments, parent=None, title="Select Comment", dark_mode=False):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.setMinimumWidth(460)
        self.setMinimumHeight(360)
        self.chosen_comment = None
        self._comments = list(comments or [])

        if dark_mode:
            self.setStyleSheet("""
                QDialog { background-color: #0D1117; }
                QLabel { color: #E6EDF3; background: transparent; }
                QLineEdit {
                    padding: 4px 6px; border: 1px solid #30363D;
                    background: #161B22; color: #E6EDF3; border-radius: 4px;
                }
                QLineEdit:focus { border: 1px solid #58A6FF; }
                QListWidget {
                    background: #161B22; color: #E6EDF3;
                    border: 1px solid #30363D; border-radius: 4px;
                }
                QListWidget::item:selected { background: #1F6FEB; color: white; }
                QPushButton {
                    background: #21262D; color: #E6EDF3; border: 1px solid #30363D;
                    border-radius: 4px; padding: 5px 12px;
                }
                QPushButton:hover { background: #30363D; }
            """)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(10, 10, 10, 10)
        lay.setSpacing(6)

        hint = QLabel("Double-click a comment to use its time. "
                      "Search filters by text or time.")
        hint.setStyleSheet("color: gray; font-size: 9pt;")
        lay.addWidget(hint)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Filter…")
        self.search_edit.textChanged.connect(self._refresh_list)
        lay.addWidget(self.search_edit)

        self.list_widget = QListWidget()
        self.list_widget.itemDoubleClicked.connect(self._on_accept)
        lay.addWidget(self.list_widget, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        self.btn_ok = QPushButton("Use Selected")
        self.btn_ok.setDefault(True)
        self.btn_ok.clicked.connect(self._on_accept)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_ok)
        btn_row.addWidget(btn_cancel)
        lay.addLayout(btn_row)

        self._refresh_list()

    def _refresh_list(self):
        needle = self.search_edit.text().strip().lower()
        self.list_widget.clear()
        for c in self._comments:
            label = f"[{c['hhmmss']}]  {c['text']}"
            if needle and needle not in label.lower():
                continue
            item = QListWidgetItem(label)
            item.setData(Qt.UserRole, c)
            self.list_widget.addItem(item)
        if self.list_widget.count() > 0:
            self.list_widget.setCurrentRow(0)

    def _on_accept(self):
        item = self.list_widget.currentItem()
        if item is None:
            self.reject()
            return
        self.chosen_comment = item.data(Qt.UserRole)
        self.accept()


class CommentPickerButton(QPushButton):
    """Small dropdown-style button next to an HH:MM QLineEdit.

    When clicked, opens CommentPickerDialog over the parent window and fills
    the associated QLineEdit with the selected comment's HH:MM (or HH:MM:SS).
    Disables itself automatically when no comments are available.

    Uses a simple text glyph (no emoji) so rendering is consistent across
    platforms / fonts and the button never shows clipped characters.
    """

    def __init__(self, target_edit, get_comments_callable, parent=None,
                 dark_mode=False, use_seconds=False, title="Select Comment"):
        super().__init__("▾", parent)
        self._edit = target_edit
        self._get_comments = get_comments_callable
        self._dark_mode = dark_mode
        self._use_seconds = use_seconds
        self._title = title
        self.setToolTip("Pick from imported recording comments")
        # Slightly wider than the old emoji button so the glyph is centred
        # with visible side padding even in dense grids.
        self.setFixedWidth(34)
        # Align with sibling QLineEdit height on most themes.
        self.setMinimumHeight(22)
        f = self.font()
        f.setBold(True)
        self.setFont(f)
        self.setCursor(Qt.PointingHandCursor)
        self.clicked.connect(self._on_click)
        self.refresh_state()

    def refresh_state(self):
        """Enable/disable based on whether any comments are available."""
        try:
            comments = self._get_comments() or []
        except Exception:
            comments = []
        self.setEnabled(bool(comments))
        if comments:
            self.setToolTip(f"Pick from {len(comments)} recording comments")
        else:
            self.setToolTip("No comments imported. Use 'Import Comments…' to load them.")

    def _on_click(self):
        try:
            comments = self._get_comments() or []
        except Exception:
            comments = []
        if not comments:
            return
        dlg = CommentPickerDialog(comments, parent=self.window(),
                                  title=self._title, dark_mode=self._dark_mode)
        if dlg.exec_() == QDialog.Accepted and dlg.chosen_comment:
            c = dlg.chosen_comment
            self._edit.setText(c['hhmmss'] if self._use_seconds else c['hhmm'])


class CommentTimeCombo(QComboBox):
    """Editable combobox that doubles as an HH:MM analysis-time picker.

    Looks and behaves like the existing 'Recording Type' combobox. Users
    can type a time directly or open the dropdown to pick from imported
    recording comments. The dropdown re-populates on every open so newly
    imported comments appear without having to close/re-open the dialog.

    Exposes `.text()`/`.setText()`/`.setPlaceholderText()` so it is a
    drop-in replacement for the QLineEdit fields used by existing code.
    """

    def __init__(self, initial_text="", get_comments=None, parent=None,
                 use_seconds=False, placeholder=""):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self._get_comments = get_comments or (lambda: [])
        self._use_seconds = use_seconds
        self.setMinimumHeight(22)
        self.setEditText(str(initial_text or ""))
        if placeholder:
            self.setPlaceholderText(placeholder)
        self.activated.connect(self._on_activated)

    def showPopup(self):
        current_text = self.currentText()
        self.blockSignals(True)
        self.clear()
        try:
            comments = self._get_comments() or []
        except Exception:
            comments = []
        for c in comments:
            time_str = c.get('hhmmss' if self._use_seconds else 'hhmm', '') or ''
            raw_text = (c.get('text') or '').strip()
            hhmmss = c.get('hhmmss', time_str)
            display = f"{hhmmss}  —  {raw_text}" if raw_text else hhmmss
            self.addItem(display, userData=time_str)
        self.setEditText(current_text)
        self.blockSignals(False)
        super().showPopup()

    def _on_activated(self, index):
        if index < 0:
            return
        data = self.itemData(index)
        if isinstance(data, str) and data:
            self.setEditText(data)

    # ----- QLineEdit-compatible API used by existing dialog code ----------
    def text(self):
        return self.currentText()

    def setText(self, s):
        self.setEditText(str(s or ""))

    def setPlaceholderText(self, s):
        le = self.lineEdit()
        if le is not None:
            le.setPlaceholderText(str(s or ""))


# -----------------------------------------------------------------------------
# Advanced Export Dialog - Per-animal export settings
# -----------------------------------------------------------------------------
class AdvancedExportDialog(QDialog):
    """Dialog for assigning per-animal analysis windows and custom basal periods."""

    def __init__(self, animals_data, global_settings, recording_type, parent=None,
                 dark_mode=False, existing_settings=None, get_comments=None):
        super().__init__(parent)
        self.animals_data = animals_data
        self.global_settings = global_settings
        self.recording_type = recording_type
        self.dark_mode = dark_mode
        self.existing_settings = existing_settings or {}
        self.result_settings = {}
        self._animal_widgets = {}
        # Callable returning the current global list of imported comments
        # (the same list used by the main Export dialog). Per-animal fields
        # gain a 📝 picker button when this callable returns a non-empty list.
        self._get_comments = get_comments or (lambda: [])
        self._picker_buttons = []

        self.setWindowTitle("Advanced Export Settings")
        self.setModal(True)
        self.setMinimumWidth(580)
        self.setMinimumHeight(350)
        self._setup_ui()

    def _setup_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(8)
        main_layout.setContentsMargins(12, 12, 12, 12)

        if self.dark_mode:
            self.setStyleSheet("""
                QDialog { background-color: #0D1117; }
                QLabel { color: #E6EDF3; background: transparent; }
                QLineEdit, QSpinBox {
                    padding: 3px 6px; min-height: 22px; border-radius: 4px;
                    border: 1px solid #30363D; background: #161B22; color: #E6EDF3;
                }
                QLineEdit:focus, QSpinBox:focus { border: 1px solid #58A6FF; }
                QGroupBox {
                    border: 1px solid #30363D; border-radius: 6px;
                    margin-top: 10px; padding-top: 14px; color: #E6EDF3; font-weight: 600;
                }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }
                QCheckBox { color: #E6EDF3; }
                QCheckBox::indicator { border: 1px solid #30363D; background: #161B22; }
                QCheckBox::indicator:checked { background: #238636; border-color: #238636; }
                QPushButton {
                    background-color: #21262D; color: #E6EDF3; border: 1px solid #30363D;
                    border-radius: 4px; padding: 6px 12px;
                }
                QPushButton:hover { background-color: #30363D; }
            """)
        else:
            self.setStyleSheet("""
                QGroupBox {
                    border: 1px solid palette(mid); border-radius: 6px;
                    margin-top: 10px; padding-top: 14px; font-weight: 600;
                }
                QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }
                QLineEdit, QSpinBox {
                    padding: 3px 6px; min-height: 22px; border-radius: 4px;
                    border: 1px solid palette(mid); background: palette(base);
                }
            """)

        info_lbl = QLabel("Configure per-animal analysis windows. Uncheck 'Use Global' to customise.")
        info_lbl.setWordWrap(True)
        main_layout.addWidget(info_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.NoFrame)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(6)

        is_injection = self.recording_type == "injection"

        for animal_name in self.animals_data:
            ex = self.existing_settings.get(animal_name, {})
            has_existing = ex and not ex.get('use_global', True)

            group = QGroupBox(animal_name)
            g_lay = QGridLayout(group)
            g_lay.setVerticalSpacing(4)
            g_lay.setHorizontalSpacing(8)

            cb_global = QCheckBox("Use global settings")
            cb_global.setChecked(not has_existing)
            g_lay.addWidget(cb_global, 0, 0, 1, 2)

            widgets = {}

            # HH:MM fields become editable combo-boxes so their dropdown
            # also acts as the recording-comment picker.
            get_cmts = self._get_comments

            if is_injection:
                lbl_inj = QLabel("Injection Time:")
                edt_inj = CommentTimeCombo(
                    ex.get('injection_time', self.global_settings.get('injection_time', '09:00')) if has_existing else self.global_settings.get('injection_time', '09:00'),
                    get_cmts,
                )
                lbl_end = QLabel("End Time:")
                edt_end = CommentTimeCombo(
                    ex.get('injection_end', self.global_settings.get('injection_end', '12:00')) if has_existing else self.global_settings.get('injection_end', '12:00'),
                    get_cmts,
                )
                lbl_int = QLabel("Interval (min):")
                edt_int = QLineEdit(str(ex.get('interval_minutes', self.global_settings.get('interval_minutes', 20)) if has_existing else self.global_settings.get('interval_minutes', 20)))
                edt_int.setPlaceholderText("e.g. 20")
                lbl_bs = QLabel("Custom Basal Start:")
                edt_bs = CommentTimeCombo(
                    ex.get('custom_basal_start', '') if has_existing else '',
                    get_cmts,
                    placeholder="e.g. 08:35 (leave empty for default)",
                )
                lbl_be = QLabel("Custom Basal End:")
                edt_be = CommentTimeCombo(
                    ex.get('custom_basal_end', '') if has_existing else '',
                    get_cmts,
                    placeholder="e.g. 08:55 (leave empty for default)",
                )

                rows_def = [
                    (lbl_inj, edt_inj),
                    (lbl_end, edt_end),
                    (lbl_int, edt_int),
                    (lbl_bs, edt_bs),
                    (lbl_be, edt_be),
                ]
                row = 1
                for lbl, w in rows_def:
                    g_lay.addWidget(lbl, row, 0)
                    g_lay.addWidget(w, row, 1)
                    row += 1

                widgets = {
                    'cb_global': cb_global,
                    'injection_time': edt_inj, 'injection_end': edt_end,
                    'interval_edit': edt_int,
                    'custom_basal_start': edt_bs, 'custom_basal_end': edt_be,
                    'all_fields': [lbl_inj, edt_inj,
                                   lbl_end, edt_end,
                                   lbl_int, edt_int,
                                   lbl_bs, edt_bs,
                                   lbl_be, edt_be],
                }
            else:
                lbl_start = QLabel("Start Time:")
                edt_start = CommentTimeCombo(
                    ex.get('basal_start', self.global_settings.get('basal_start', '09:00')) if has_existing else self.global_settings.get('basal_start', '09:00'),
                    get_cmts,
                )
                lbl_end = QLabel("End Time:")
                edt_end = CommentTimeCombo(
                    ex.get('basal_end', self.global_settings.get('basal_end', '12:00')) if has_existing else self.global_settings.get('basal_end', '12:00'),
                    get_cmts,
                )
                lbl_int = QLabel("Interval (min):")
                edt_int = QLineEdit(str(ex.get('interval_minutes', self.global_settings.get('interval_minutes', 20)) if has_existing else self.global_settings.get('interval_minutes', 20)))
                edt_int.setPlaceholderText("e.g. 20")

                rows_def = [
                    (lbl_start, edt_start),
                    (lbl_end, edt_end),
                    (lbl_int, edt_int),
                ]
                row = 1
                for lbl, w in rows_def:
                    g_lay.addWidget(lbl, row, 0)
                    g_lay.addWidget(w, row, 1)
                    row += 1

                widgets = {
                    'cb_global': cb_global,
                    'basal_start': edt_start, 'basal_end': edt_end,
                    'interval_edit': edt_int,
                    'all_fields': [lbl_start, edt_start,
                                   lbl_end, edt_end,
                                   lbl_int, edt_int],
                }

            def _toggle_fields(checked, w=widgets):
                for f in w['all_fields']:
                    f.setEnabled(not checked)

            cb_global.stateChanged.connect(_toggle_fields)
            _toggle_fields(cb_global.isChecked())

            self._animal_widgets[animal_name] = widgets
            scroll_layout.addWidget(group)

        scroll_layout.addStretch()
        scroll.setWidget(scroll_widget)
        main_layout.addWidget(scroll, 1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("OK")
        btn_ok.setDefault(True)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setCursor(Qt.PointingHandCursor)
        btn_row.addWidget(btn_ok)
        btn_row.addWidget(btn_cancel)
        main_layout.addLayout(btn_row)

        btn_ok.clicked.connect(self._on_ok)
        btn_cancel.clicked.connect(self.reject)

    def _on_ok(self):
        is_injection = self.recording_type == "injection"
        for animal_name, w in self._animal_widgets.items():
            if w['cb_global'].isChecked():
                self.result_settings[animal_name] = {'use_global': True}
            else:
                s = {'use_global': False}
                if is_injection:
                    s['injection_time'] = w['injection_time'].text().strip()
                    s['injection_end'] = w['injection_end'].text().strip()
                    s['interval_minutes'] = int(w['interval_edit'].text().strip() or 20)
                    s['custom_basal_start'] = w['custom_basal_start'].text().strip()
                    s['custom_basal_end'] = w['custom_basal_end'].text().strip()
                else:
                    s['basal_start'] = w['basal_start'].text().strip()
                    s['basal_end'] = w['basal_end'].text().strip()
                    s['interval_minutes'] = int(w['interval_edit'].text().strip() or 20)
                self.result_settings[animal_name] = s
        self.accept()

    def get_settings(self):
        return self.result_settings


class EnhancedEEGPlotter(QWidget):
    def __init__(self, model, eeg0, eeg1, sfreq, recording_start,
                 interval_length=0.6, overlap_length=0.3,
                 sequence_length=100, spectral_threshold=0,
                 channel_names=None, edf_filename=None, model_type="cwt_only",
                 unet_model=None, unet_channel=0):
        super().__init__()
        # Core data attributes
        self.model = model
        self.cwt_model_preloaded = None  # Store preloaded CWT model
        
        # Track whether EDF is loaded
        self.edf_loaded = eeg0 is not None
        self.edf_raw = None  # Store full EDF for channel switching
        self.edf_path = None
        
        # Multi-animal session storage
        self.animals_data = {}  # {animal_name: {'events': [], 'channels': (ch1_name, ch2_name), 'channel_indices': (idx1, idx2)}}
        self.current_animal_name = None
        self.animal_counter = 1  # For auto-generating animal names
        
        # Handle no data case (app starts without EDF)
        if eeg0 is None:
            self.eeg0 = np.zeros(1000)  # Placeholder
            self.eeg1 = np.zeros(1000)
            self.sfreq = 512.0  # Default
            self.recording_start_time = datetime.now()
            self.total_samples = 1000
            self.single_channel_view = False
        else:
            self.eeg0 = eeg0
            # Handle single channel view
            self.single_channel_view = False
            if eeg1 is None:
                self.single_channel_view = True
                self.eeg1 = eeg0  # Duplicate for model compatibility
            else:
                self.eeg1 = eeg1
            self.sfreq = sfreq
            self.recording_start_time = recording_start
            self.total_samples = len(self.eeg0)
            
        self.interval_length = interval_length
        self.overlap_length = overlap_length
        self.sequence_length = sequence_length
        self.spectral_threshold = spectral_threshold
        self.channel_names = channel_names if channel_names is not None else ["Channel 1", "Channel 2"]
        self.edf_filename = edf_filename or "EEG Recording"
        
        # Model pipeline configuration
        self.model_type = model_type  # "unet_only", "cwt_only", or "cwt_unet"
        # Track whether the user has explicitly chosen a model
        self.model_chosen = False
        # Require both overlapping tokens to agree before accepting an SWD.
        # Defaults to True for CWT-based pipelines (matches the previous hard-
        # coded behaviour) and False for UNET-only. The user can override this
        # from Settings -> Advanced Settings -> CWT Hybrid CNN Settings.
        self.require_overlap_agreement = self.model_type in ("cwt_only", "cwt_unet")
        self.unet_model_preloaded = unet_model  # Pre-loaded UNET model
        self.unet_preprocessing_channel = unet_channel  # Which channel for UNET

        # Plot/navigation state
        self.display_window_seconds = 20
        self.current_pos_sec = 0

        # Event storage
        self.manual_events = []
        self.merged_swd = []  # Store the raw merged events (before refinement)
        self.refined_swd = []  # Store the refined events (after border analysis)
        self.all_events = []  # Combined list for easier management

        # Selection/edit state
        self.selected_event_idx = None
        self.region_patches = []
        self.current_edit_event = None
        self.dragging_start = False
        self.dragging_end = False
        self.edit_mode_active = True
        self.active_event_type = None  # 'auto' or 'manual'
        
        # Display state
        self.dark_mode = False
        self.show_unrefined = False  # Whether to display unrefined events

        # PSD analysis dialog
        self.psd_dialog = None

        # UNET state
        self.unet_model = None
        self.unet_model_path = None
        self.unet_predictions = None  # 1D float, fs=100 Hz
        self.unet_intervals = []      # list of (start,end) at 100 Hz
        self.unet_fs = 100.0
        self.unet_channel = 0  # 0 for Channel 1, 1 for Channel 2
        # Storage for pre-refinement autos
        self.refined_swd_base = []
        
        # UNET settings (only the ones actually used)
        self.unet_prediction_threshold = 0.5
        self.unet_min_duration = 1.0
        self.unet_gap_threshold = 0.75
        
        # Plot-only bandpass filter settings
        self.filter_signal_enabled = False
        self.filter_low_cut_hz = 1.0
        self.filter_high_cut_hz = 40.0
        self._filtered_eeg0 = None
        self._filtered_eeg1 = None
        
        # Export settings with defaults
        self.export_basal_start = "09:00"
        self.export_basal_end = "12:00"
        self.export_injection_time = "00:20"
        self.export_recording_type = "basal"
        self.export_animal_id = "Animal_1"
        self.export_analyzed = True
        self.per_animal_export_settings = {}

        # Imported recording comments (from .adicht or text file). List of
        # dicts with keys 'time_sec', 'hhmm', 'hhmmss', 'text'. Populated on
        # demand via the Export dialog; persisted with animals_data so
        # switching animals keeps the appropriate comments.
        self.adicht_comments = []
        # Per-saved-animal cache of comments (loaded when animal is activated).
        self._animal_comments_cache = {}

        # Initialize UI components (show UI immediately)
        self.init_ui()

        # Only start preprocessing if EDF is loaded
        if self.edf_loaded:
            # Defer heavy preprocessing to allow UI to render first
            QTimer.singleShot(0, self._start_background_preprocessing)
            self.update_plot()
        else:
            # Show welcome screen
            self._show_welcome_screen()

    def _start_background_preprocessing(self):
        """Prepare for model selection WITHOUT running any preprocessing yet.

        Preprocessing used to run both CWT and UNET automatically on load. That
        wastes time/CPU when the user only wants one model. Instead we now let
        the user pick a model first; the chosen model's pipeline is loaded
        on-demand in on_model_selected(). Here we just reset per-recording state
        and enable the model buttons so a choice can be made.
        """
        self._cwt_done = False
        self._unet_done = False
        # Track whether each pipeline actually completed successfully.
        # This is distinct from finding events: a successful run with 0
        # detected SWDs must NOT be treated as a failure.
        self._cwt_succeeded = False
        self._unet_succeeded = False
        # The model the user is currently loading (None when idle).
        self._pending_model_type = None

        # Initialize storage
        self.cwt_token_predictions = None
        self.cwt_merged_swd = []
        self.cwt_refined_swd = []
        self.unet_predictions = None
        self.unet_intervals = []
        self.unet_refined_swd = []

        # Hide the loading bar (nothing is computing yet) and let the user
        # choose a model. The chosen pipeline loads on demand.
        try:
            self.progress_bar.hide()
        except Exception:
            pass
        self._set_model_buttons_enabled(True)
        try:
            if hasattr(self, 'status_widget'):
                self.status_widget.show()
            if hasattr(self, 'progress_label'):
                self.progress_label.setText("Select a model (CWT, Combined or UNET) to run detection")
        except Exception:
            pass

    def _set_model_buttons_enabled(self, enabled):
        """Enable/disable the three model-selection buttons together.

        A model that has no backing network available stays disabled regardless
        of `enabled` (e.g. UNET buttons when the UNET model failed to load)."""
        cwt_ok = self.model is not None
        unet_ok = getattr(self, 'unet_model_preloaded', None) is not None
        try:
            if hasattr(self, 'btn_model_cwt'):
                self.btn_model_cwt.setEnabled(bool(enabled) and cwt_ok)
            if hasattr(self, 'btn_model_cwt_unet'):
                self.btn_model_cwt_unet.setEnabled(bool(enabled) and cwt_ok and unet_ok)
            if hasattr(self, 'btn_model_unet'):
                self.btn_model_unet.setEnabled(bool(enabled) and unet_ok)
        except Exception:
            pass

    def _start_cwt_preprocessing(self):
        """Start CWT preprocessing in background"""
        if self.model is None:
            self._cwt_done = True
            self._check_all_complete()
            return
            
        self.worker_thread = QThread(self)
        # Use a default power percentile if the spinbox isn't ready or accessible
        try:
            pp = self.power_percentile_spin.value()
        except Exception:
            pp = 25

        overlap_flag = bool(getattr(self, 'require_overlap_agreement',
                                    self.model_type in ("cwt_only", "cwt_unet")))
        self.worker = PreprocessWorker(
            self.eeg0, self.eeg1, self.sfreq, 
            self.interval_length, self.overlap_length, self.sequence_length,
            self.model, pp,
            require_overlap_agreement=overlap_flag,
        )
        self.worker.moveToThread(self.worker_thread)
        self.worker.progress.connect(self._on_cwt_progress)
        self.worker.finished.connect(self._on_cwt_finished)
        self.worker.error.connect(self._on_cwt_error)
        self.worker_thread.started.connect(self.worker.run)
        self.worker_thread.start()

    def _on_cwt_progress(self, done, total, eta, phase):
        """Handle CWT preprocessing progress"""
        if total <= 0:
            return

        # The progress bar runs in indeterminate (busy) mode; use the label to
        # communicate which phase of the CWT pipeline is active.
        try:
            if hasattr(self, 'progress_label'):
                phase_txt = "Extracting CWT features" if phase == 'features' else "Running CWT model"
                self.progress_label.setText(f"{phase_txt}...")
        except Exception:
            pass

        QApplication.processEvents()

    def _on_cwt_error(self, msg):
        print(f"CWT Error: {msg}")
        self._cwt_done = True
        if hasattr(self, 'worker_thread'):
            self.worker_thread.quit()
            self.worker_thread.wait()
        self._check_all_complete()

    def _on_cwt_finished(self, result):
        try:
            (token_preds, merged_swd, refined_swd) = result
            self.cwt_token_predictions = token_preds
            self.cwt_merged_swd = merged_swd
            self.cwt_refined_swd = refined_swd
            # New CWT baseline invalidates any cached user edits on CWT models.
            if hasattr(self, '_model_state_cache'):
                self._model_state_cache.pop('cwt_only', None)
                self._model_state_cache.pop('cwt_unet', None)
            self._cwt_succeeded = True
            print(f"CWT Complete: {len(refined_swd)} events")
        except Exception as e:
            print(f"Error storing CWT results: {e}")
        finally:
            if hasattr(self, 'worker_thread'):
                self.worker_thread.quit()
                self.worker_thread.wait()
            self._cwt_done = True
            self._check_all_complete()

    def _start_unet_preprocessing(self):
        """Start UNET preprocessing in background"""
        if not hasattr(self, 'unet_model_preloaded') or self.unet_model_preloaded is None:
            self._unet_done = True
            self._check_all_complete()
            return
        
        if not hasattr(self, 'unet_model') or self.unet_model is None:
            self.unet_model = self.unet_model_preloaded
        
        if not hasattr(self, 'unet_channel_name'):
             self.unet_channel_name = "Channel 1"
        
        signal = self.eeg0 if self.unet_channel == 0 else self.eeg1
        
        self.unet_thread = QThread(self)
        self.unet_worker = UNETWorker(
            signal, self.sfreq, self.unet_model,
            unet_fs=100.0, segment_len=1024, min_duration=1.0, gap_threshold=0.75
        )
        self.unet_worker.moveToThread(self.unet_thread)
        self.unet_worker.finished.connect(self._on_unet_preprocessing_finished)
        self.unet_worker.error.connect(self._on_unet_preprocessing_error)
        self.unet_thread.started.connect(self.unet_worker.run)
        self.unet_thread.start()

    def _on_unet_preprocessing_finished(self, result):
        try:
            self.unet_predictions = result['predictions']
            self.unet_confidence_scores = result.get('confidence')
            self.unet_intervals = result['intervals']
            
            scale = self.sfreq / self.unet_fs
            self.unet_refined_swd = []
            for s_idx, e_idx in self.unet_intervals:
                s_sample = int(round(s_idx * scale))
                e_sample = int(round((e_idx + 1) * scale))
                duration = (e_sample - s_sample) / self.sfreq
                event = {
                    'start_sample': s_sample,
                    'end_sample': e_sample,
                    'duration_sec': duration,
                    'is_auto': True
                }
                self.unet_refined_swd.append(event)
            # New UNET baseline invalidates any cached user edits on models that
            # use UNET predictions.
            if hasattr(self, '_model_state_cache'):
                self._model_state_cache.pop('unet_only', None)
                self._model_state_cache.pop('cwt_unet', None)
            print(f"UNET Complete: {len(self.unet_intervals)} intervals")
        except Exception as e:
            print(f"Error storing UNET results: {e}")
        finally:
            if hasattr(self, 'unet_thread'):
                self.unet_thread.quit()
                self.unet_thread.wait()
            self._unet_done = True
            self._check_all_complete()
    
    def _on_unet_preprocessing_error(self, msg):
        print(f"UNET preprocessing error: {msg}")
        self._unet_done = True
        self._check_all_complete()

    def _check_all_complete(self):
        """Route pipeline completion into the on-demand model-load flow.

        Preprocessing now only ever runs because the user picked a model, so a
        completing pipeline just needs to advance that pending selection."""
        self._maybe_finalize_pending()

    def _maybe_finalize_pending(self):
        """Finish an on-demand model load once its required pipelines report done."""
        mt = getattr(self, '_pending_model_type', None)
        if mt is None:
            return
        needs_cwt = mt in ("cwt_only", "cwt_unet")
        needs_unet = mt in ("unet_only", "cwt_unet")

        # Wait until every pipeline this model needs has finished.
        if needs_cwt and not self._cwt_done:
            return
        if needs_unet and not self._unet_done:
            return

        # All required pipelines finished; verify they actually succeeded.
        failed = []
        if needs_cwt and not self._cwt_succeeded:
            failed.append("CWT")
        if needs_unet and not self._unet_succeeded:
            failed.append("UNET")

        self._pending_model_type = None
        try:
            self.progress_bar.hide()
        except Exception:
            pass

        if failed:
            self._set_model_buttons_enabled(True)
            try:
                if hasattr(self, 'progress_label'):
                    self.progress_label.setText("Model loading failed - try another model")
            except Exception:
                pass
            QMessageBox.warning(
                self, "Preprocessing Failed",
                f"{' & '.join(failed)} preprocessing failed. Please try again or pick another model.")
            return

        try:
            if hasattr(self, 'status_widget'):
                self.status_widget.hide()
        except Exception:
            pass
        self._finalize_model_selection(mt)

    def _current_model_event_colors(self):
        """Return (auto_event_fill_hex, accent_hex) based on current model type.
        Colors are muted in dark mode for better eye comfort."""
        try:
            if self.dark_mode:
                # Muted colors for dark mode
                if self.model_type == "unet_only":
                    return ("#4A9B6E", "#3A7A55")  # Muted emerald green
                elif self.model_type == "cwt_unet":
                    return ('#8B7A4A', "#6B5F3A")  # Muted gold
                elif self.model_type == "cwt_only":
                    return ("#CC5555", "#993333")  # Muted red
            else:
                # Original bright colors for light mode
                if self.model_type == "unet_only":
                    return ("#50C878", "#2EA043")  # Emerald green theme
                elif self.model_type == "cwt_unet":
                    return ('#AC994F', "#756436")  # Gold theme
                elif self.model_type == "cwt_only":
                    return ("#FF0000", "#CC0000")  # Red theme
        except Exception:
            pass
        return ("#2F2F2F", "#000000")  # Fallback dark gray
    
    def _apply_slider_color(self, accent_hex):
        """Apply a given color to the position slider handle without changing other theme styles.
        Uses muted colors in dark mode."""
        try:
            if not hasattr(self, 'slider') or self.slider is None:
                return
            
            # Use muted hover color in dark mode
            hover_bg = '#B8860B' if self.dark_mode else '#FFD700'
            hover_border = '#D0D0D0' if self.dark_mode else '#FFF'
            
            self.slider.setStyleSheet(f"""
                QSlider::handle:horizontal {{
                    background: {accent_hex};
                    border: 2px solid {accent_hex};
                    width: 18px;
                    height: 18px;
                    margin: -7px 0;
                    border-radius: 9px;
                }}
                QSlider::handle:horizontal:hover {{
                    background: {hover_bg};
                    border: 2px solid {hover_border};
                }}
            """)
        except Exception:
            pass
    
    def _start_unet_preprocessing(self):
        """Start UNET preprocessing in background"""
        if not hasattr(self, 'unet_model_preloaded') or self.unet_model_preloaded is None:
            # Skip UNET if model not available (marks as done/not-succeeded).
            self._unet_done = True
            self._check_all_complete()
            return
        
        # Set UNET model and run predictions
        if not hasattr(self, 'unet_model') or self.unet_model is None:
            self.unet_model = self.unet_model_preloaded
        
        # Use first channel as default for preprocessing
        if not hasattr(self, 'unet_channel') or self.unet_channel is None:
            self.unet_channel = 0
            self.unet_channel_name = "Channel 1"
        
        # Don't update UI for UNET start since CWT drives the progress bar
        # self.progress_label.setText("Step 2/2: UNET Preprocessing...")
        # self.progress_bar.setValue(50)
        
        # Select signal based on channel
        signal = self.eeg0 if self.unet_channel == 0 else self.eeg1
        
        # Start UNET worker
        self.unet_thread = QThread(self)
        self.unet_worker = UNETWorker(
            signal=signal,
            sfreq=self.sfreq,
            model=self.unet_model,
            unet_fs=self.unet_fs,
            segment_len=1024,
            min_duration=self.unet_min_duration,
            gap_threshold=self.unet_gap_threshold
        )
        self.unet_worker.moveToThread(self.unet_thread)
        self.unet_worker.finished.connect(self._on_unet_preprocessing_finished)
        self.unet_worker.error.connect(self._on_unet_preprocessing_error)
        self.unet_thread.started.connect(self.unet_worker.run)
        self.unet_thread.start()
    
    def _on_unet_preprocessing_finished(self, result):
        """Store UNET results"""
        try:
            self.unet_predictions = result['predictions']
            self.unet_confidence_scores = result.get('confidence')
            self.unet_intervals = result['intervals']
            
            # Convert UNET intervals to event format for UNET-only mode
            scale = self.sfreq / self.unet_fs
            self.unet_refined_swd = []
            for s_idx, e_idx in self.unet_intervals:
                s_sample = int(round(s_idx * scale))
                e_sample = int(round((e_idx + 1) * scale))
                duration = (e_sample - s_sample) / self.sfreq
                
                event = {
                    'start_sample': s_sample,
                    'end_sample': e_sample,
                    'duration_sec': duration,
                    'is_auto': True
                }
                self.unet_refined_swd.append(event)
            
            self._unet_succeeded = True
            print(f"UNET Complete: {len(self.unet_intervals)} intervals found")
            # self.progress_bar.setValue(100) # Let CWT drive the progress bar
        except Exception as e:
            print(f"Error storing UNET results: {e}")
        finally:
            self.unet_thread.quit()
            self.unet_thread.wait()
        
        # Mark UNET as done and check if CWT is also done
        self._unet_done = True
        self._check_all_complete()
        
        # self._on_all_preprocessing_complete()
    
    def _on_unet_preprocessing_error(self, msg):
        """Handle UNET preprocessing error"""
        print(f"UNET preprocessing error: {msg}")
        # self.progress_label.setText("UNET preprocessing failed, using CWT only")
        # self.progress_bar.setValue(100)
        self._unet_done = True
        self._check_all_complete()
    
    def _on_all_preprocessing_complete(self):
        """Called when both CWT and UNET preprocessing are complete"""
        self.progress_bar.hide()
        try:
            self.status_widget.hide()
        except Exception:
            pass
        
        # Now user can select model - results are already computed
        try:
            self.btn_model_unet.setEnabled(True)
            self.btn_model_cwt.setEnabled(True)
            self.btn_model_cwt_unet.setEnabled(True)
        except Exception:
            pass

    # ------------------------- UNET Refinement Logic -------------------------
    def on_unet_refine_toggled(self, _state):
        if self.cb_unet_refine.isChecked():
            # Enable UNET refinement
            self.apply_unet_refinement_if_enabled()
        else:
            # Disable UNET refinement - revert to base events
            if hasattr(self, 'refined_swd_base') and self.refined_swd_base is not None:
                self.refined_swd = list(self.refined_swd_base)
            # Update status label
            if hasattr(self, 'unet_status_lbl'):
                self.unet_status_lbl.setText("UNET refinement: Inactive")
                self.unet_status_lbl.setStyleSheet("color: gray; font-size: 9pt;")
        
        self.update_combined_events()
        self.update_filtered_events()
        self.update_plot()

    def apply_unet_refinement_if_enabled(self):
        """If checkbox is on and UNET predictions exist, refine borders by
        intersecting each auto SWD with UNET positive regions. Keeps manual
        events unchanged. Reverts to base autos when off.
        
        NEW BEHAVIOR: Instead of dividing events into multiple pieces,
        this function now trims the outer borders of each event to match
        the intersection with UNET predictions. Each event remains as a
        single unit with adjusted start/end boundaries.
        
        ENHANCED: Now supports confidence weighting and refinement modes."""
        # If we do not have a base reference yet, set it
        if not hasattr(self, 'refined_swd_base') or self.refined_swd_base is None:
            self.refined_swd_base = list(self.refined_swd)

        # Start from base autos
        self.refined_swd = list(self.refined_swd_base)

        if not self.cb_unet_refine.isChecked():
            return
        if self.unet_predictions is None or self.unet_intervals is None:
            return

        # Safety check: ensure we have base events to work with
        if not self.refined_swd_base:
            return

        # Convert UNET intervals (in 100 Hz) to sample indices at EEG sfreq
        scale = self.sfreq / float(self.unet_fs)
        unet_spans = []
        for s_idx, e_idx in (self.unet_intervals or []):
            s_sample = int(round(s_idx * scale))
            e_sample = int(round((e_idx + 1) * scale))
            unet_spans.append((s_sample, e_sample))
        # Merge UNET spans to simplify
        unet_spans = self._merge_spans(unet_spans)

        refined = []
        for ev in self.refined_swd_base:
            # Safety check: ensure event has valid start and end samples
            if 'start_sample' not in ev or 'end_sample' not in ev:
                continue
                
            s, e = ev['start_sample'], ev['end_sample']
            
            # Safety check: ensure valid sample indices
            if s < 0 or e < 0 or s >= e:
                continue
                
            # Find intersections with UNET spans for this event
            overlaps = self._intersections_with_spans(s, e, unet_spans)
            
            if not overlaps:
                # No UNET overlap - keep original event
                refined.append(ev.copy())
                continue
            
            # Find the overall span that covers all UNET intersections for this event
            # This ensures we don't divide the event, only trim its borders
            min_start = min(overlap[0] for overlap in overlaps)
            max_end = max(overlap[1] for overlap in overlaps)
            
            # Ensure the trimmed event is still valid
            if max_end > min_start:
                new_ev = ev.copy()
                new_ev['start_sample'] = min_start
                new_ev['end_sample'] = max_end
                new_ev['duration_sec'] = (max_end - min_start) / self.sfreq
                
                # Only keep events that are at least 1 second long
                if new_ev['duration_sec'] >= 1.0:
                    refined.append(new_ev)
                    # Debug info: show how much the event was trimmed
                    original_dur = (e - s) / self.sfreq
                    trimmed_dur = new_ev['duration_sec']
                    print(f"UNET refinement: Event trimmed from {original_dur:.2f}s to {trimmed_dur:.2f}s")
                else:
                    # Event too short after trimming, keep original
                    refined.append(ev.copy())
            else:
                # If trimming failed, keep original event
                refined.append(ev.copy())
        
        # Update the refined events list
        if refined:
            self.refined_swd = refined
            print(f"UNET refinement completed: {len(refined)} events refined")
            # Update status label
            if hasattr(self, 'unet_status_lbl'):
                self.unet_status_lbl.setText(f"UNET refinement: Active ({len(refined)} events refined)")
                self.unet_status_lbl.setStyleSheet("color: green; font-size: 9pt;")
        else:
            # No events were refined
            if hasattr(self, 'unet_status_lbl'):
                self.unet_status_lbl.setText("UNET refinement: No events refined")
                self.unet_status_lbl.setStyleSheet("color: orange; font-size: 9pt;")
        # Note: No need to merge overlapping events here since we're not creating new divisions

    def _merge_spans(self, spans):
        if not spans:
            return []
        spans = sorted(spans, key=lambda x: x[0])
        merged = [list(spans[0])]
        for s, e in spans[1:]:
            last = merged[-1]
            if s <= last[1]:
                last[1] = max(last[1], e)
            else:
                merged.append([s, e])
        return [(s, e) for s, e in merged]

    def _intersections_with_spans(self, s, e, spans):
        """
        Find all intersections between event (s, e) and UNET spans.
        Returns list of (start, end) tuples representing the covered regions.
        """
        res = []
        for us, ue in spans:
            # Check if UNET span overlaps with event
            if ue <= s or us >= e:
                continue
            
            # Calculate intersection
            inter_s = max(s, us)
            inter_e = min(e, ue)
            
            if inter_e > inter_s:
                res.append((inter_s, inter_e))
        
        # Sort intersections by start time for better processing
        res.sort(key=lambda x: x[0])
        return res

    def init_ui(self):
        """Initialize the user interface components with modern tabbed design."""
        self.setWindowTitle("🧠 AutoSWD Detector")
        self.setGeometry(50, 50, 1800, 1000)
        main = QVBoxLayout(self)
        # If starting in dark mode, attempt to apply dark title bar on Windows
        try:
            if getattr(self, 'dark_mode', False):
                self._apply_win_dark_titlebar(self.winId())
        except Exception:
            pass
        main.setSpacing(0)
        main.setContentsMargins(0, 0, 0, 0)
        # Ensure the main widget can receive key events when needed
        self.setFocusPolicy(Qt.StrongFocus)

        # ===== Quick Action Toolbar =====
        toolbar = QToolBar("Quick Actions")
        toolbar.setMovable(False)
        toolbar.setIconSize(QSize(20, 20))
        
        # File action - opens popup menu on click (no dropdown arrow)
        self.action_file = QAction(" 📁 File ", self)
        self.action_file.setToolTip("File operations")
        toolbar.addAction(self.action_file)
        
        # Create File menu (shown on click, not as dropdown)
        self.file_menu = QMenu(self)
        
        self.action_load_edf = self.file_menu.addAction("📂 Load EDF...")
        self.action_load_edf.setShortcut(QKeySequence("Ctrl+O"))
        self.action_load_edf.setToolTip("Load an EDF file")
        self.action_load_edf.triggered.connect(self.load_edf_file)
        
        self.action_select_channels = self.file_menu.addAction("🔀 Select Channels...")
        self.action_select_channels.setToolTip("Select channels for current animal")
        self.action_select_channels.triggered.connect(self.select_channels_dialog)
        self.action_select_channels.setEnabled(False)
        
        self.file_menu.addSeparator()
        
        self.action_save_animal = self.file_menu.addAction("💾 Save Animal")
        self.action_save_animal.setToolTip("Save current animal's events to session")
        self.action_save_animal.triggered.connect(self.save_current_animal)
        self.action_save_animal.setEnabled(False)
        
        self.action_new_animal = self.file_menu.addAction("➕ New Animal")
        self.action_new_animal.setToolTip("Save current animal and start analyzing a new one")
        self.action_new_animal.triggered.connect(self.start_new_animal)
        self.action_new_animal.setEnabled(False)
        
        self.file_menu.addSeparator()

        # --- Recording comments (imported from .adicht or text/CSV) ---------
        # Kept here (File section) so the main toolbar / plot area stay clean.
        self.action_import_comments = self.file_menu.addAction("📝 Import Comments...")
        self.action_import_comments.setToolTip(
            "Import recording comments from a LabChart .adicht file "
            "or a plain text/CSV dump. Imported comments are drawn on the "
            "plot and on the minimap, and can be used to fill analysis "
            "HH:MM fields in the Export dialog."
        )
        self.action_import_comments.triggered.connect(self._import_comments_dialog)

        self.action_comments_list = self.file_menu.addAction("📋 Comments List...")
        self.action_comments_list.setToolTip("Browse imported comments and jump to their time")
        self.action_comments_list.triggered.connect(self.show_comments_popup)
        self.action_comments_list.setEnabled(False)

        self.action_clear_comments = self.file_menu.addAction("🗑 Clear Comments")
        self.action_clear_comments.setToolTip("Remove all imported recording comments")
        self.action_clear_comments.triggered.connect(self._clear_comments)
        self.action_clear_comments.setEnabled(False)

        self.file_menu.addSeparator()

        # Show saved animals submenu
        self.saved_animals_menu = self.file_menu.addMenu("📋 Saved Animals")
        self.saved_animals_menu.setEnabled(False)
        
        self.action_manage_animals = self.file_menu.addAction("⚙️ Manage Saved Animals...")
        self.action_manage_animals.setToolTip("Rename or delete saved animals")
        self.action_manage_animals.triggered.connect(self.show_manage_animals_dialog)
        
        # Connect File action to show menu
        self.action_file.triggered.connect(self._show_file_menu)
        
        toolbar.addSeparator()
        
        # Export action - direct export to Excel
        self.action_export = QAction("💾 Export", self)
        self.action_export.setToolTip("Export to Excel (Ctrl+S)")
        self.action_export.setShortcut(QKeySequence("Ctrl+S"))
        toolbar.addAction(self.action_export)
        
        toolbar.addSeparator()
        
        # Events list action (moved to top)
        self.action_events = QAction("📝 Events List", self)
        self.action_events.setToolTip("Show events list")
        toolbar.addAction(self.action_events)
        
        # Statistics action (moved to top)
        self.action_statistics = QAction("📊 Statistics", self)
        self.action_statistics.setToolTip("Show statistics")
        toolbar.addAction(self.action_statistics)
        
        toolbar.addSeparator()
        
        # Settings action
        self.action_settings = QAction("⚙️ Settings", self)
        self.action_settings.setToolTip("Open Settings")
        toolbar.addAction(self.action_settings)
        
        # Dark mode toggle
        self.action_dark_mode = QAction("🌓 Dark Mode", self)
        self.action_dark_mode.setCheckable(True)
        self.action_dark_mode.setToolTip("Toggle Dark Mode")
        toolbar.addAction(self.action_dark_mode)
        
        toolbar.addSeparator()
        
        # Controls action
        self.action_help = QAction("🎮 Controls", self)
        self.action_help.setToolTip("Show controls")
        toolbar.addAction(self.action_help)
        
        main.addWidget(toolbar)
        
        # Initialize model-themed accents for minimap
        try:
            auto_hex, _ = self._current_model_event_colors()
            # If method exists later, set after minimap creation
        except Exception:
            auto_hex = "#FF7F50"

        # ===== Main Content Area - Stacked Widget for Welcome/Plot =====
        from PyQt5.QtWidgets import QStackedWidget
        self.content_stack = QStackedWidget()
        
        # ===== Welcome Screen (Index 0) =====
        self.welcome_widget = self._create_welcome_screen()
        self.content_stack.addWidget(self.welcome_widget)
        
        # ===== Plot Container (Index 1) =====
        plot_container = QWidget()
        self.plot_container = plot_container
        plot_layout = QVBoxLayout(plot_container)
        plot_layout.setContentsMargins(5, 5, 5, 5)
        plot_layout.setSpacing(3)
        
        # Timeline minimap (compact)
        self.timeline_minimap = TimelineMinimap()
        self.timeline_minimap.setFixedHeight(40)
        try:
            self.timeline_minimap.set_auto_color(auto_hex)
        except Exception:
            pass
        self.timeline_minimap.clicked.connect(self.on_minimap_clicked)
        plot_layout.addWidget(self.timeline_minimap)
        
        # Plot canvas with toolbar
        self.fig, _ = plt.subplots(figsize=(18, 10))
        self.canvas = FigureCanvas(self.fig)
        plot_layout.addWidget(self.canvas)
        
        self.content_stack.addWidget(plot_container)
        
        main.addWidget(self.content_stack)
        
        # ===== Spacious Control Bar =====
        control_bar = QWidget()
        control_bar.setMinimumHeight(140)
        control_bar.setMaximumHeight(160)
        control_layout = QVBoxLayout(control_bar)
        control_layout.setContentsMargins(15, 10, 15, 10)
        control_layout.setSpacing(12)
        
        # Row 1: Position slider (prominent and spacious)
        slider_layout = QHBoxLayout()
        slider_layout.setContentsMargins(0, 0, 0, 0)
        position_label = QLabel("Position:")
        position_label.setMargin(0)
        position_label.setContentsMargins(0, 0, 0, 0)
        position_label.setStyleSheet("background-color: transparent;")
        position_label.setMinimumWidth(70)
        slider_layout.addWidget(position_label)
        self.slider = QSlider(Qt.Horizontal)
        self.slider.setMinimumHeight(28)
        # Start with gray knob while preprocessing/no model is selected
        self._apply_slider_color("#808080")
        slider_layout.addWidget(self.slider)
        control_layout.addLayout(slider_layout)
        # Row 2: All controls with generous spacing
        main_control_row = QHBoxLayout()
        main_control_row.setSpacing(15)
        
        # Event navigation (left)
        event_nav_layout = QHBoxLayout()
        event_nav_layout.setSpacing(8)
        self.btn_prev_event = QPushButton("⏮ Prev Event")
        self.btn_prev_event.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_prev_event.setFixedHeight(40)
        self.btn_prev_event.setMinimumWidth(100)
        self.btn_prev_event.setToolTip("Previous Event (←)")
        event_nav_layout.addWidget(self.btn_prev_event)
        
        self.btn_next_event = QPushButton("Next Event ⏭")
        self.btn_next_event.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_next_event.setFixedHeight(40)
        self.btn_next_event.setMinimumWidth(110)
        self.btn_next_event.setToolTip("Next Event (→)")
        event_nav_layout.addWidget(self.btn_next_event)
        main_control_row.addLayout(event_nav_layout)
        
        # Operations group
        ops_layout = QHBoxLayout()
        ops_layout.setSpacing(8)
        # Merge SWDs button
        self.btn_merge_swds = QPushButton("🔗 Merge SWDs")
        self.btn_merge_swds.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_merge_swds.setFixedHeight(40)
        self.btn_merge_swds.setMinimumWidth(135)
        self.btn_merge_swds.setToolTip("Merge with nearest SWD within 20s (if any)")
        ops_layout.addWidget(self.btn_merge_swds)
        # Event Filter toggle button
        self.btn_event_filter = QPushButton("⏱ Event Filter")
        self.btn_event_filter.setCheckable(True)
        self.btn_event_filter.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_event_filter.setFixedHeight(40)
        self.btn_event_filter.setMinimumWidth(135)
        self.btn_event_filter.setToolTip("Filter events by duration")
        ops_layout.addWidget(self.btn_event_filter)
        
        
        # View controls - more spacious
        view_layout = QHBoxLayout()
        view_layout.setSpacing(8)
        view_layout.setContentsMargins(0, 0, 0, 0)
        view_label = QLabel("Window:")
        view_label.setMargin(0)
        view_label.setContentsMargins(0, 0, 0, 0)
        view_label.setStyleSheet("background-color: transparent;")
        view_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        view_label.setMinimumWidth(60)
        view_layout.addWidget(view_label)
        
        self.win_spin = QDoubleSpinBox()
        self.win_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.win_spin.setRange(1, 120)
        self.win_spin.setValue(self.display_window_seconds)
        self.win_spin.setSuffix("s")
        self.win_spin.setFixedWidth(90)
        self.win_spin.setFixedHeight(40)
        self.win_spin.setToolTip("Window size in seconds")
        view_layout.addWidget(self.win_spin)
        
        y_label = QLabel("Y-Axis:")
        y_label.setMargin(0)
        y_label.setContentsMargins(0, 0, 0, 0)
        y_label.setStyleSheet("background-color: transparent;")
        y_label.setMinimumWidth(50)
        y_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        view_layout.addWidget(y_label)
        
        self.ylim_input = QDoubleSpinBox()
        self.ylim_input.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        # Fewer decimals (no long "0,4000000" zero run) and a 10x larger arrow
        # step (0.1 mV instead of 0.01 mV) for quicker adjustment.
        self.ylim_input.setRange(0.01, 10.0)
        self.ylim_input.setSingleStep(0.1)
        self.ylim_input.setValue(0.4)
        self.ylim_input.setDecimals(2)
        self.ylim_input.setPrefix("±")
        self.ylim_input.setSuffix(" mV")
        self.ylim_input.setFixedWidth(110)
        self.ylim_input.setFixedHeight(40)
        self.ylim_input.setToolTip("Y-axis range in millivolts (0.01–10 mV, arrows step 0.1 mV)")
        view_layout.addWidget(self.ylim_input)
        
        self.btn_ylim_auto = QPushButton("Auto Scale")
        self.btn_ylim_auto.setCheckable(True)
        # Use size policy to let the button expand as needed for best fit
        self.btn_ylim_auto.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_ylim_auto.setFixedHeight(40)
        self.btn_ylim_auto.setToolTip("Auto-scale Y-axis")
        view_layout.addWidget(self.btn_ylim_auto)
        
        # Scroll sensitivity control
        scroll_label = QLabel("Scroll:")
        scroll_label.setMargin(0)
        scroll_label.setContentsMargins(0, 0, 0, 0)
        scroll_label.setStyleSheet("background-color: transparent;")
        scroll_label.setMinimumWidth(50)
        scroll_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        view_layout.addWidget(scroll_label)
        
        self.scroll_sensitivity = QDoubleSpinBox()
        self.scroll_sensitivity.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.scroll_sensitivity.setRange(0.1, 5.0)
        self.scroll_sensitivity.setValue(1.0)
        self.scroll_sensitivity.setSingleStep(0.1)
        self.scroll_sensitivity.setDecimals(1)
        self.scroll_sensitivity.setFixedWidth(70)
        self.scroll_sensitivity.setFixedHeight(40)
        self.scroll_sensitivity.setToolTip("Mouse wheel scroll sensitivity")
        view_layout.addWidget(self.scroll_sensitivity)
        
        main_control_row.addLayout(view_layout)
        
        
        # Model Selection - spacious
        model_layout = QHBoxLayout()
        model_layout.setSpacing(8)
        model_layout.setContentsMargins(0, 0, 0, 0)
        model_label = QLabel("Model:")
        model_label.setMargin(0)
        model_label.setContentsMargins(0, 0, 0, 0)
        model_label.setStyleSheet("background-color: transparent;")
        model_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        model_label.setMinimumWidth(55)
        model_layout.addWidget(model_label)
        
        self.btn_model_cwt = QPushButton("CWT")
        self.btn_model_cwt.setCheckable(True)
        self.btn_model_cwt.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_model_cwt.setFixedHeight(40)
        self.btn_model_cwt.setMinimumWidth(85)
        self.btn_model_cwt.setToolTip("CWT Only")
        self.btn_model_cwt.setEnabled(False)
        self.btn_model_cwt.clicked.connect(lambda: self.on_model_selected("cwt_only"))
        model_layout.addWidget(self.btn_model_cwt)
        
        self.btn_model_cwt_unet = QPushButton("Combined")
        self.btn_model_cwt_unet.setCheckable(True)
        self.btn_model_cwt_unet.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_model_cwt_unet.setFixedHeight(40)
        self.btn_model_cwt_unet.setMinimumWidth(95)
        self.btn_model_cwt_unet.setToolTip("CWT + UNET")
        self.btn_model_cwt_unet.setEnabled(False)
        self.btn_model_cwt_unet.clicked.connect(lambda: self.on_model_selected("cwt_unet"))
        model_layout.addWidget(self.btn_model_cwt_unet)
        
        self.btn_model_unet = QPushButton("UNET")
        self.btn_model_unet.setCheckable(True)
        self.btn_model_unet.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_model_unet.setFixedHeight(40)
        self.btn_model_unet.setMinimumWidth(90)
        self.btn_model_unet.setToolTip("UNET Only")
        self.btn_model_unet.setEnabled(False)
        self.btn_model_unet.clicked.connect(lambda: self.on_model_selected("unet_only"))
        model_layout.addWidget(self.btn_model_unet)
        main_control_row.addLayout(model_layout)
        
        
        # Edit actions - always enabled unless disabled in settings
        edit_layout = QHBoxLayout()
        edit_layout.setSpacing(8)
        
        self.btn_add = QPushButton("➕ Add Event")
        self.btn_add.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_add.setFixedHeight(40)
        self.btn_add.setMinimumWidth(100)
        self.btn_add.setEnabled(True)
        self.btn_add.setToolTip("Add Event (A)")
        edit_layout.addWidget(self.btn_add)
        
        self.btn_del = QPushButton("🗑️ Delete")
        self.btn_del.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_del.setFixedHeight(40)
        self.btn_del.setMinimumWidth(85)
        self.btn_del.setEnabled(False)
        self.btn_del.setToolTip("Delete Event (Del)")
        edit_layout.addWidget(self.btn_del)
        main_control_row.addLayout(edit_layout)
        
        
        # Analysis button - spacious
        self.btn_psd = QPushButton("📈 PSD Analysis")
        self.btn_psd.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_psd.setFixedHeight(40)
        self.btn_psd.setMinimumWidth(135)
        self.btn_psd.setEnabled(False)
        self.btn_psd.setToolTip("PSD Analysis (P)")
        main_control_row.addWidget(self.btn_psd)
        
        # Push following controls to the right
        main_control_row.addStretch()
        # Add operations (Event Filter, Merge) to the far right
        main_control_row.addLayout(ops_layout)
        
        control_layout.addLayout(main_control_row)
        
        main.addWidget(control_bar)

        # ===== Enhanced Status Bar =====
        self.status_widget = QWidget()
        # Blend the status bar into the window background (no white block).
        # Scope the rule to this widget only so the child progress bar keeps
        # its own themed style.
        self.status_widget.setObjectName("statusWidget")
        self.status_widget.setStyleSheet("QWidget#statusWidget { background: transparent; }")
        status_layout = QHBoxLayout(self.status_widget)
        status_layout.setContentsMargins(10, 5, 10, 5)
        
        # Status text (model-selection hint / on-demand pipeline loading state).
        # Keep the background transparent so it blends into the status bar
        # instead of showing a white block next to the progress bar.
        self.progress_label = QLabel("")
        self.progress_label.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.progress_label.setStyleSheet("background: transparent; border: none;")
        status_layout.addWidget(self.progress_label)

        # Indeterminate progress bar (replaces status label and determinate progress bar)
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 0)  # Indeterminate mode
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.progress_bar.setMinimumWidth(520)
        self.progress_bar.hide()  # Hidden by default, shown when processing
        status_layout.addWidget(self.progress_bar, 1)
        
        # Current position indicator
        self.position_indicator = QLabel("📍 0.0s")
        self.position_indicator.hide()
        
        # Model indicator
        self.model_indicator = QLabel("🔬 No Model")
        self.model_indicator.hide()
        
        # Event count badge
        self.event_count_badge = QLabel("📊 0 Events")
        self.event_count_badge.hide()
        
        main.addWidget(self.status_widget)

        # Hidden/internal controls (no longer visible in UI but needed for functionality)
        self.cb_show_auto = QCheckBox()
        self.cb_show_auto.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_show_auto.setChecked(True)
        self.cb_show_auto.hide()
        
        self.cb_show_unrefined = QCheckBox()
        self.cb_show_unrefined.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_show_unrefined.hide()
        
        self.cb_dark_mode = QCheckBox()
        self.cb_dark_mode.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_dark_mode.hide()
        
        self.cb_show_unet = QCheckBox()
        self.cb_show_unet.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        # Do not show UNET overlays until a model is chosen
        self.cb_show_unet.setChecked(False)
        self.cb_show_unet.hide()
        
        self.cb_unet_post = QCheckBox()
        self.cb_unet_post.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_unet_post.setChecked(True)
        self.cb_unet_post.hide()
        
        self.cb_unet_refine = QCheckBox()
        self.cb_unet_refine.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_unet_refine.setChecked(False)
        self.cb_unet_refine.setEnabled(False)
        self.cb_unet_refine.hide()

        self.btn_unet = QPushButton()
        self.btn_unet.hide()
        self.btn_unet_settings = QPushButton()
        self.btn_unet_settings.hide()
        
        # Hidden threshold controls
        self.gap_spin = QSpinBox()
        self.gap_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.gap_spin.setRange(1, 10)
        self.gap_spin.setValue(4)
        self.gap_spin.hide()
        
        self.th_spin = QDoubleSpinBox()
        self.th_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.th_spin.setRange(0.1, 0.9)
        self.th_spin.setSingleStep(0.05)
        self.th_spin.setValue(0.5)
        self.th_spin.hide()
        
        self.power_percentile_spin = QSpinBox()
        self.power_percentile_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.power_percentile_spin.setRange(1, 50)
        self.power_percentile_spin.setValue(25)
        self.power_percentile_spin.hide()
        
        # Y-axis zoom buttons (hidden, functionality kept)
        self.btn_ylim_in = QPushButton()
        self.btn_ylim_in.hide()
        self.btn_ylim_out = QPushButton()
        self.btn_ylim_out.hide()
        
        # Duration filter controls (hidden)
        self.cb_duration_filter = QCheckBox()
        self.cb_duration_filter.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_duration_filter.hide()
        self.duration_min_spin = QDoubleSpinBox()
        self.duration_min_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.duration_min_spin.setRange(0.1, 60.0)
        self.duration_min_spin.setValue(1.0)
        self.duration_min_spin.hide()
        self.duration_max_spin = QDoubleSpinBox()
        self.duration_max_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.duration_max_spin.setRange(0.1, 60.0)
        self.duration_max_spin.setValue(10.0)
        self.duration_max_spin.hide()
        
        # Additional hidden controls
        
        self.btn_merge_nearest = QPushButton()
        self.btn_merge_nearest.hide()
        self.btn_convert = QPushButton()
        self.btn_convert.hide()
        self.btn_exp = QPushButton()
        self.btn_exp.hide()
        self.btn_settings = QPushButton()
        self.btn_settings.hide()
        
        # Event/stats labels (now shown in popups)
        self.event_info_lbl = QLabel()
        self.event_info_lbl.hide()
        self.info_lbl = QLabel()
        self.info_lbl.hide()
        self.stats_lbl = QLabel()
        self.stats_lbl.hide()
        self.stats_mean_lbl = QLabel()
        self.stats_mean_lbl.hide()
        self.stats_total_lbl = QLabel()
        self.stats_total_lbl.hide()
        self.stats_sd_lbl = QLabel()
        self.stats_sd_lbl.hide()
        self.stats_min_lbl = QLabel()
        self.stats_min_lbl.hide()
        self.stats_max_lbl = QLabel()
        self.stats_max_lbl.hide()
        self.stats_count_lbl = QLabel()
        self.stats_count_lbl.hide()
        
        # UNET status label
        self.unet_status_lbl = QLabel()
        self.unet_status_lbl.hide()

        # Key assignments
        # Values are integer combinations of Qt.Key_* | modifier flags
        # (Qt.ControlModifier, Qt.ShiftModifier, Qt.AltModifier). A plain key
        # value without modifier bits still works as before.
        self.key_assignments = {
            'Add Event': int(Qt.Key_A),
            'Delete Event': int(Qt.Key_Delete),
            'Merge Adjacent SWDs': int(Qt.Key_M),
            'Previous': int(Qt.Key_PageUp),
            'Next': int(Qt.Key_PageDown),
            'Previous Event': int(Qt.Key_Left),
            'Next Event': int(Qt.Key_Right),
            'PSD Analysis': int(Qt.Key_P),
            'Undo': int(Qt.ControlModifier) | int(Qt.Key_Z),
            'Redo': int(Qt.ControlModifier) | int(Qt.ShiftModifier) | int(Qt.Key_Z),
        }
        self.key_assignments_file = 'key_assignments.json'
        self.load_key_assignments()
        self.key_labels = {}

        # Undo / redo history. Each snapshot captures the edit-relevant state
        # (per-type event lists + selection) so the user can roll back any
        # destructive editing action.
        self._undo_stack = []
        self._redo_stack = []
        self._max_undo = 200

        # Per-model cache of user-edited auto events so switching between
        # models (e.g. CWT -> UNET -> CWT) preserves manual tweaks to the
        # auto predictions for each model.
        self._model_state_cache = {}

        # Connect signals
        self.connect_signals()
        
        # Initialize filtered events list
        self.filtered_events = []

        # Apply initial styling
        self.apply_modern_stylesheet()
    
    def _create_separator(self):
        """Create a vertical separator line"""
        line = QFrame()
        line.setFrameShape(QFrame.VLine)
        line.setFrameShadow(QFrame.Sunken)
        line.setFixedWidth(2)
        return line
    
    # ------------------------------------------------------------------
    # Recording-comment workflow (import / clear / list / refresh UI)
    # ------------------------------------------------------------------
    def _import_comments_dialog(self):
        """Pick a file and import recording comments from it.

        Supports LabChart .adicht (needs optional `adi` package) and plain
        text/CSV dumps. Imported comments are shown on the plot + minimap
        and fed to the Export dialog's HH:MM pickers.
        """
        path, _ = QFileDialog.getOpenFileName(
            self, "Import Recording Comments",
            "",
            "Supported (*.adicht *.txt *.csv *.tsv);;"
            "LabChart (*.adicht);;"
            "Text / CSV (*.txt *.csv *.tsv);;"
            "All Files (*)"
        )
        if not path:
            return
        ext = os.path.splitext(path)[1].lower()
        rec_start = getattr(self, 'recording_start_time', None)
        try:
            if ext == '.adicht':
                comments = parse_adicht_comments(path, rec_start)
            else:
                comments = parse_text_comments(path, rec_start)
        except ImportError as e:
            QMessageBox.warning(self, "Missing Dependency", str(e))
            return
        except Exception as e:
            QMessageBox.critical(self, "Import Failed", str(e))
            return
        if not comments:
            QMessageBox.information(
                self, "No Comments Found",
                "The selected file did not contain any recognizable comments."
            )
            return

        self.adicht_comments = comments
        if not hasattr(self, '_animal_comments_cache'):
            self._animal_comments_cache = {}
        if getattr(self, 'current_animal_name', None):
            self._animal_comments_cache[self.current_animal_name] = list(comments)
        self._refresh_comments_ui()
        try:
            self.show_toast(f"Imported {len(comments)} comments", "success")
        except Exception:
            pass

    def _clear_comments(self):
        """Remove all imported recording comments."""
        if not getattr(self, 'adicht_comments', None):
            return
        self.adicht_comments = []
        if (hasattr(self, '_animal_comments_cache')
                and getattr(self, 'current_animal_name', None)):
            self._animal_comments_cache.pop(self.current_animal_name, None)
        self._refresh_comments_ui()
        try:
            self.show_toast("Recording comments cleared", "info")
        except Exception:
            pass

    def _refresh_comments_ui(self):
        """Reflect current self.adicht_comments everywhere it is visible:
        File-menu action enablement, the plot, and the minimap.
        """
        n = len(getattr(self, 'adicht_comments', []) or [])
        if hasattr(self, 'action_comments_list'):
            self.action_comments_list.setEnabled(n > 0)
            self.action_comments_list.setText(
                "📋 Comments List..." + (f" ({n})" if n > 0 else "")
            )
        if hasattr(self, 'action_clear_comments'):
            self.action_clear_comments.setEnabled(n > 0)
        # Re-draw plot and minimap so marker lines are added/removed.
        try:
            if getattr(self, 'edf_loaded', False):
                self.update_plot()
        except Exception:
            pass
        try:
            if hasattr(self, 'timeline_minimap') and self.timeline_minimap is not None:
                self.update_timeline_minimap()
        except Exception:
            pass

    def _jump_to_time_sec(self, time_sec):
        """Centre the view on a given time offset (seconds from recording start)."""
        try:
            total = self.total_samples / self.sfreq
        except Exception:
            total = 0
        if total <= 0:
            return
        half = self.display_window_seconds / 2.0
        target = max(0.0, min(time_sec - half, max(0.0, total - self.display_window_seconds)))
        self.current_pos_sec = target
        try:
            self.slider.blockSignals(True)
            self.slider.setValue(int(self.current_pos_sec))
            self.slider.blockSignals(False)
        except Exception:
            pass
        self.update_plot()
        self.update_timeline_minimap()

    def show_comments_popup(self):
        """List imported comments with a 'jump-to-time' double-click action."""
        comments = getattr(self, 'adicht_comments', []) or []
        if not comments:
            QMessageBox.information(
                self, "No Comments",
                "No recording comments loaded.\n"
                "Use File → Import Comments… to load some first."
            )
            return

        dialog = QDialog(self)
        dialog.setWindowTitle(f"📋 Comments ({len(comments)})")
        dialog.setGeometry(120, 120, 560, 460)
        try:
            if self.dark_mode:
                self._apply_win_dark_titlebar(dialog.winId())
        except Exception:
            pass
        if self.dark_mode:
            dialog.setStyleSheet("""
                QDialog { background-color: #0D1117; color: #E6EDF3; }
                QLabel { color: #E6EDF3; }
                QLineEdit {
                    padding: 4px 6px; border: 1px solid #30363D;
                    background: #161B22; color: #E6EDF3; border-radius: 4px;
                }
                QLineEdit:focus { border: 1px solid #58A6FF; }
                QTableWidget {
                    background-color: #0D1117; color: #E6EDF3;
                    border: 1px solid #30363D; gridline-color: #21262D;
                    alternate-background-color: #161B22;
                }
                QTableWidget::item:selected {
                    background-color: #1C3A1F; color: #7EE787;
                }
                QHeaderView::section {
                    background-color: #161B22; color: #8B949E;
                    border-bottom: 2px solid #30363D;
                }
                QPushButton {
                    background: #21262D; color: #E6EDF3;
                    border: 1px solid #30363D; border-radius: 4px;
                    padding: 5px 12px;
                }
                QPushButton:hover { background: #30363D; }
            """)

        lay = QVBoxLayout(dialog)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(8)

        hint = QLabel("Click a comment to jump to that time. "
                      "Search filters by time or text.")
        hint.setStyleSheet("color: gray; font-size: 9pt;")
        lay.addWidget(hint)

        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Filter…")
        lay.addWidget(search_edit)

        table = QTableWidget()
        table.setColumnCount(3)
        table.setHorizontalHeaderLabels(["#", "Time", "Comment"])
        table.setSelectionBehavior(QAbstractItemView.SelectRows)
        table.setSelectionMode(QAbstractItemView.SingleSelection)
        table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.verticalHeader().setVisible(False)
        hdr = table.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)
        lay.addWidget(table, 1)

        def _populate():
            needle = search_edit.text().strip().lower()
            table.setRowCount(0)
            for i, c in enumerate(comments):
                label = f"{c['hhmmss']} {c['text']}"
                if needle and needle not in label.lower():
                    continue
                r = table.rowCount()
                table.insertRow(r)
                item_num = QTableWidgetItem(str(i + 1))
                item_time = QTableWidgetItem(c['hhmmss'])
                item_text = QTableWidgetItem(c['text'])
                # Store the seconds offset so we can navigate on activation
                item_num.setData(Qt.UserRole, float(c.get('time_sec', 0.0)))
                item_time.setData(Qt.UserRole, float(c.get('time_sec', 0.0)))
                item_text.setData(Qt.UserRole, float(c.get('time_sec', 0.0)))
                table.setItem(r, 0, item_num)
                table.setItem(r, 1, item_time)
                table.setItem(r, 2, item_text)
            if table.rowCount() > 0:
                table.selectRow(0)

        def _jump(item):
            if item is None:
                return
            t = item.data(Qt.UserRole)
            if t is None:
                return
            self._jump_to_time_sec(float(t))

        # Single-click on a row jumps directly to that comment's time.
        # itemClicked fires once per mouse release so filter navigation
        # (arrow keys) still works without an unwanted extra jump.
        table.itemClicked.connect(_jump)
        search_edit.textChanged.connect(_populate)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)
        btn_close = QPushButton("Close")
        btn_close.setCursor(Qt.PointingHandCursor)
        btn_close.setDefault(True)
        btn_close.clicked.connect(dialog.close)
        btn_row.addWidget(btn_close)
        lay.addLayout(btn_row)

        _populate()
        dialog.exec_()

    def show_events_popup(self):
        """Show simple events list in a popup dialog, dark mode fixes for rows, minimal controls"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📝 Events List")
        dialog.setGeometry(100, 100, 900, 600)
        # Apply dark titlebar if needed
        try:
            if self.dark_mode:
                self._apply_win_dark_titlebar(dialog.winId())
        except Exception:
            pass

        # Strong full-table dark mode, including alternate rows
        if self.dark_mode:
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #0D1117;
                    color: #E6EDF3;
                }
                QLabel {
                    color: #E6EDF3;
                }
                QTableWidget {
                    background-color: #0D1117;
                    color: #E6EDF3;
                    border: 1px solid #30363D;
                    gridline-color: #21262D;
                    alternate-background-color: #161B22;
                }
                QTableWidget::item {
                    color: #E6EDF3;
                    background-color: transparent;
                }
                QTableWidget::item:selected {
                    background-color: #1C3A1F;
                    color: #7EE787;
                }
                QHeaderView::section {
                    background-color: #161B22;
                    color: #8B949E;
                    border-bottom: 2px solid #30363D;
                }
            """)

        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Event table - minimal columns
        event_table = QTableWidget()
        event_table.setColumnCount(4)
        event_table.setHorizontalHeaderLabels(["#", "Event Start","Event End", "Duration (s)"])
        event_table.horizontalHeader().setStretchLastSection(True)
        event_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        event_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        event_table.setSelectionBehavior(QTableWidget.SelectRows)
        event_table.setEditTriggers(QTableWidget.NoEditTriggers)
        event_table.setAlternatingRowColors(True)
        if self.dark_mode:
            event_table.setAlternatingRowColors(True)
            event_table.setStyleSheet(
                """
                QTableWidget { alternate-background-color: #161B22; background-color: #0D1117; color: #E6EDF3; }
                QTableCornerButton::section { background-color: #161B22; border: 1px solid #30363D; }
                QHeaderView::section { background-color: #161B22; color: #8B949E; border-bottom: 2px solid #30363D; cursor: pointer; }
                QHeaderView::section:hover { background-color: #21262D; }
                """
            )

        # Populate table
        all_events = []
        for idx, ev in enumerate(self.manual_events):
            all_events.append((idx, 'manual', ev))
        for idx, ev in enumerate(self.refined_swd):
            all_events.append((idx, 'auto', ev))
        all_events.sort(key=lambda x: x[2]['start_sample'])
        
        # Store events list for sorting
        events_data = list(all_events)
        current_sort_column = [1]  # Default sort by start time
        current_sort_order = [Qt.AscendingOrder]

        def populate_table(sorted_events):
            event_table.setRowCount(len(sorted_events))
            for row, (idx, ev_type, ev) in enumerate(sorted_events):
                item0 = QTableWidgetItem(str(row + 1))
                item0.setData(Qt.UserRole, (ev_type, idx))  # Store event identity
                event_table.setItem(row, 0, item0)
                
                start_time = (self.recording_start_time + timedelta(seconds=ev['start_sample'] / self.sfreq)).strftime('%H:%M:%S')
                event_table.setItem(row, 1, QTableWidgetItem(start_time))
                
                end_time = (self.recording_start_time + timedelta(seconds=ev['end_sample'] / self.sfreq)).strftime('%H:%M:%S')
                event_table.setItem(row, 2, QTableWidgetItem(end_time))
                
                event_table.setItem(row, 3, QTableWidgetItem(f"{ev['duration_sec']:.2f}"))
        
        def on_header_clicked(column):
            nonlocal events_data
            # Toggle sort order if same column clicked again
            if current_sort_column[0] == column:
                current_sort_order[0] = Qt.DescendingOrder if current_sort_order[0] == Qt.AscendingOrder else Qt.AscendingOrder
            else:
                current_sort_column[0] = column
                current_sort_order[0] = Qt.AscendingOrder
            
            reverse = current_sort_order[0] == Qt.DescendingOrder
            
            if column == 0:  # Sort by row number (original order)
                events_data.sort(key=lambda x: x[2]['start_sample'], reverse=reverse)
            elif column == 1:  # Sort by start time
                events_data.sort(key=lambda x: x[2]['start_sample'], reverse=reverse)
            elif column == 2:  # Sort by end time
                events_data.sort(key=lambda x: x[2]['end_sample'], reverse=reverse)
            elif column == 3:  # Sort by duration
                events_data.sort(key=lambda x: x[2]['duration_sec'], reverse=reverse)
            
            populate_table(events_data)
        
        event_table.horizontalHeader().sectionClicked.connect(on_header_clicked)
        
        # Initial population
        populate_table(events_data)
        
        layout.addWidget(event_table)

        # Bottom: only a close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_close = QPushButton("Close")
        btn_close.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)

        # Clicking (any column of) a row focuses to that event as before
        def on_table_click(row, col):
            # Metadata stored in the first column
            info_item = event_table.item(row, 0)
            if info_item:
                ev_type, idx = info_item.data(Qt.UserRole)
                self.selected_event_idx = idx
                self.active_event_type = ev_type
                events = self.manual_events if ev_type == 'manual' else self.refined_swd
                if idx < len(events):
                    ev = events[idx]
                    center_time = (ev['start_sample'] + ev['end_sample']) / (2 * self.sfreq)
                    self.current_pos_sec = max(0, center_time - self.display_window_seconds / 2)
                    self.update_plot()
                    dialog.accept()

        event_table.cellClicked.connect(on_table_click)

        dialog.exec_()
    
    def show_statistics_popup(self):
        """Show statistics in a popup dialog - Modern Dashboard Style"""
        dialog = QDialog(self)
        dialog.setWindowTitle("Statistics Dashboard")
        dialog.setMinimumSize(780, 780)
        dialog.resize(780, 820)
        try:
            if self.dark_mode:
                self._apply_win_dark_titlebar(dialog.winId())
        except Exception:
            pass
        
        # Colors for light/dark mode
        if self.dark_mode:
            bg_color = "#0D1117"
            card_bg = "#161B22"
            card_border = "#30363D"
            text_color = "#E6EDF3"
            text_secondary = "#8B949E"
            section_bg = "#161B22"
        else:
            bg_color = "#F5F5F5"
            card_bg = "#FFFFFF"
            card_border = "#E0E0E0"
            text_color = "#1F2937"
            text_secondary = "#6B7280"
            section_bg = "#FFFFFF"
        
        dialog.setStyleSheet(f"""
            QDialog {{
                background-color: {bg_color};
            }}
            QLabel {{
                color: {text_color};
            }}
        """)
        
        main_layout = QVBoxLayout(dialog)
        main_layout.setContentsMargins(25, 25, 25, 25)
        main_layout.setSpacing(20)
        
        # Calculate stats
        auto_cnt = len(self.refined_swd)
        manu_cnt = len(self.manual_events)
        total_cnt = auto_cnt + manu_cnt
        all_durations = [r['duration_sec'] for r in (self.refined_swd + self.manual_events)]
        total_dur = sum(all_durations) if all_durations else 0
        mean_dur = np.mean(all_durations) if all_durations else 0
        sd_dur = np.std(all_durations, ddof=1) if len(all_durations) > 1 else 0
        min_dur = np.min(all_durations) if all_durations else 0
        max_dur = np.max(all_durations) if all_durations else 0
        rec_len = self.total_samples / self.sfreq
        rec_len_min = rec_len / 60.0
        event_rate = (total_cnt / rec_len_min) if rec_len_min > 0 else 0
        
        # Determine model name for badge
        model_info = {
            "unet_only": ("UNET", "#27AE60"),
            "cwt_only": ("CWT", "#E74C3C"),
            "cwt_unet": ("CWT+UNET", "#F39C12"),
        }
        model_text, model_color = model_info.get(
            getattr(self, 'model_type', ''), 
            ("No Model", "#95A5A6")
        )
        
        # Header with title and mode badge
        header_layout = QHBoxLayout()
        title_label = QLabel("Statistics Dashboard")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {text_color};")
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        
        model_badge = QLabel(model_text)
        model_badge.setStyleSheet(f"""
            background-color: {model_color};
            color: white;
            padding: 6px 16px;
            border-radius: 12px;
            font-weight: bold;
            font-size: 13px;
        """)
        header_layout.addWidget(model_badge)
        main_layout.addLayout(header_layout)
        
        # === First row of stat cards ===
        row1_layout = QHBoxLayout()
        row1_layout.setSpacing(15)
        
        # Total Events card
        card1 = self._create_stat_card(
            str(total_cnt), "Total Events", f"Auto: {auto_cnt} | Manual: {manu_cnt}",
            "#3498DB", card_bg, card_border, text_color, text_secondary
        )
        row1_layout.addWidget(card1)
        
        # Total Duration card
        card2 = self._create_stat_card(
            f"{total_dur:.1f}s", "Total Duration", f"Recording: {rec_len_min:.1f} min",
            "#27AE60", card_bg, card_border, text_color, text_secondary
        )
        row1_layout.addWidget(card2)
        
        # Mean Duration card
        card3 = self._create_stat_card(
            f"{mean_dur:.2f}s", "Mean Duration", f"SD: {sd_dur:.2f}s",
            "#9B59B6", card_bg, card_border, text_color, text_secondary
        )
        row1_layout.addWidget(card3)
        
        main_layout.addLayout(row1_layout)
        
        # === Second row of stat cards ===
        row2_layout = QHBoxLayout()
        row2_layout.setSpacing(15)
        
        # Minimum card
        card4 = self._create_stat_card(
            f"{min_dur:.2f}s", "Minimum", "Shortest event",
            "#3498DB", card_bg, card_border, text_color, text_secondary
        )
        row2_layout.addWidget(card4)
        
        # Maximum card
        card5 = self._create_stat_card(
            f"{max_dur:.2f}s", "Maximum", "Longest event",
            "#9B59B6", card_bg, card_border, text_color, text_secondary
        )
        row2_layout.addWidget(card5)
        
        # Rate card
        card6 = self._create_stat_card(
            f"{event_rate:.2f}", "Rate", "events/min",
            "#E74C3C", card_bg, card_border, text_color, text_secondary
        )
        row2_layout.addWidget(card6)
        
        main_layout.addLayout(row2_layout)
        
        # === Duration Distribution Section ===
        dist_frame = QFrame()
        dist_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {section_bg};
                border: 1px solid {card_border};
                border-radius: 10px;
            }}
        """)
        dist_layout = QVBoxLayout(dist_frame)
        dist_layout.setContentsMargins(20, 15, 20, 15)
        dist_layout.setSpacing(12)
        
        dist_title = QLabel("Duration Distribution")
        dist_title.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {text_color}; border: none;")
        dist_layout.addWidget(dist_title)
        
        # Calculate duration distribution
        bins = [(0, 2, "0-2s", "#3498DB"), (2, 5, "2-5s", "#27AE60"), 
                (5, 10, "5-10s", "#F39C12"), (10, float('inf'), ">10s", "#E74C3C")]
        
        for low, high, label, color in bins:
            count = sum(1 for d in all_durations if low <= d < high) if all_durations else 0
            pct = (count / total_cnt * 100) if total_cnt > 0 else 0
            
            row_layout = QHBoxLayout()
            row_layout.setSpacing(10)
            
            label_widget = QLabel(label)
            label_widget.setFixedWidth(50)
            label_widget.setStyleSheet(f"color: {text_secondary}; font-size: 13px; border: none;")
            row_layout.addWidget(label_widget)
            
            # Progress bar
            progress = QProgressBar()
            progress.setRange(0, 100)
            progress.setValue(int(pct))
            progress.setTextVisible(False)
            progress.setFixedHeight(20)
            progress.setStyleSheet(f"""
                QProgressBar {{
                    background-color: {'#30363D' if self.dark_mode else '#E8E8E8'};
                    border: none;
                    border-radius: 5px;
                }}
                QProgressBar::chunk {{
                    background-color: {color};
                    border-radius: 5px;
                }}
            """)
            row_layout.addWidget(progress, 1)
            
            pct_label = QLabel(f"{pct:.0f}% ({count})")
            pct_label.setFixedWidth(80)
            pct_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            pct_label.setStyleSheet(f"color: {text_secondary}; font-size: 13px; border: none;")
            row_layout.addWidget(pct_label)
            
            dist_layout.addLayout(row_layout)
        
        main_layout.addWidget(dist_frame)
        
        # === Recording Information Section ===
        info_frame = QFrame()
        info_frame.setStyleSheet(f"""
            QFrame {{
                background-color: {section_bg};
                border: 1px solid {card_border};
                border-radius: 10px;
            }}
        """)
        info_layout = QVBoxLayout(info_frame)
        info_layout.setContentsMargins(20, 15, 20, 15)
        info_layout.setSpacing(8)
        
        info_title = QLabel("Recording Information")
        info_title.setStyleSheet(f"font-size: 16px; font-weight: bold; color: {text_color}; border: none;")
        info_layout.addWidget(info_title)
        
        # Info rows
        info_items = [
            ("File:", getattr(self, 'edf_filename', 'N/A')),
            ("Sampling Rate:", f"{self.sfreq} Hz"),
            ("Duration:", f"{rec_len_min:.2f} minutes"),
            ("Start Time:", self.recording_start_time.strftime('%Y-%m-%d %H:%M:%S')),
            ("Channels:", ", ".join(self.channel_names) if hasattr(self, 'channel_names') else "N/A"),
        ]
        
        for label_text, value_text in info_items:
            row_layout = QHBoxLayout()
            label_widget = QLabel(label_text)
            label_widget.setFixedWidth(120)
            label_widget.setStyleSheet(f"color: {text_secondary}; font-size: 13px; font-weight: 500; border: none;")
            row_layout.addWidget(label_widget)
            
            value_widget = QLabel(str(value_text))
            value_widget.setStyleSheet(f"color: {text_color}; font-size: 13px; border: none;")
            value_widget.setWordWrap(True)
            row_layout.addWidget(value_widget, 1)
            
            info_layout.addLayout(row_layout)
        
        main_layout.addWidget(info_frame)
        
        main_layout.addStretch()
        
        # Close button
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_close = QPushButton("Close")
        btn_close.setMinimumSize(120, 40)
        btn_close.setCursor(Qt.PointingHandCursor)
        btn_close.setStyleSheet(f"""
            QPushButton {{
                background-color: #E67E22;
                color: white;
                border: none;
                border-radius: 8px;
                font-size: 14px;
                font-weight: bold;
                padding: 10px 30px;
            }}
            QPushButton:hover {{
                background-color: #D35400;
            }}
        """)
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_close)
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)
        
        dialog.exec_()
    
    def _create_stat_card(self, value, title, subtitle, value_color, bg_color, border_color, text_color, text_secondary):
        """Helper to create a statistic card for the dashboard"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {bg_color};
                border: 1px solid {border_color};
                border-radius: 10px;
            }}
        """)
        card.setMinimumHeight(100)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(20, 15, 20, 15)
        layout.setSpacing(4)
        
        # Value (large, colored)
        value_label = QLabel(value)
        value_label.setStyleSheet(f"font-size: 32px; font-weight: bold; color: {value_color}; border: none;")
        layout.addWidget(value_label)
        
        # Title
        title_label = QLabel(title)
        title_label.setStyleSheet(f"font-size: 14px; font-weight: 600; color: {text_color}; border: none;")
        layout.addWidget(title_label)
        
        # Subtitle
        subtitle_label = QLabel(subtitle)
        subtitle_label.setStyleSheet(f"font-size: 12px; color: {text_secondary}; border: none;")
        layout.addWidget(subtitle_label)
        
        return card
    
    def apply_modern_stylesheet(self):
        """Apply modern cross-platform stylesheet"""
        if self.dark_mode:
            return  # Dark mode has its own stylesheet
        
        # Get model-themed accent colors for light mode
        _, accent_color = self._current_model_event_colors()
        accent_light = accent_color  # Use the accent directly
        accent_border = accent_color
        
        # Modern light mode stylesheet with model-themed accents
        style = f"""
            QWidget {{
                font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            }}
            QGroupBox {{
                border: 2px solid #E5E7EB;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 16px;
                background-color: #FFFFFF;
                font-weight: 600;
                font-size: 13px;
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                subcontrol-position: top left;
                padding: 4px 12px;
                color: #1F2937;
            }}
            QPushButton {{
                background-color: #F3F4F6;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                padding: 8px 16px;
                color: #1F2937;
                font-weight: 500;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background-color: #E5E7EB;
                border-color: #9CA3AF;
            }}
            QPushButton:pressed {{
                background-color: #D1D5DB;
            }}
            QPushButton:checked {{
                background-color: {accent_light};
                border-color: {accent_border};
                color: white;
            }}
            QPushButton:disabled {{
                background-color: #F9FAFB;
                color: #9CA3AF;
                border-color: #E5E7EB;
            }}
            QLineEdit, QSpinBox, QDoubleSpinBox {{
                background-color: white;
                border: 1px solid #D1D5DB;
                border-radius: 6px;
                padding: 6px 10px;
                font-size: 13px;
            }}
            QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {{
                border-color: {accent_light};
                outline: none;
            }}
            QSlider::groove:horizontal {{
                border: 1px solid #D1D5DB;
                height: 6px;
                background: #F3F4F6;
                border-radius: 3px;
            }}
            QSlider::handle:horizontal {{
                background: {accent_light};
                border: 2px solid {accent_border};
                width: 18px;
                height: 18px;
                margin: -7px 0;
                border-radius: 9px;
            }}
            QSlider::handle:horizontal:hover {{
                background: #FFD700;
                border: 2px solid #FFF;
            }}
            QCheckBox {{
                spacing: 6px;
                font-size: 13px;
            }}
            QCheckBox::indicator {{
                width: 18px;
                height: 18px;
                border: 2px solid #D1D5DB;
                border-radius: 4px;
                background: white;
            }}
            QCheckBox::indicator:checked {{
                background-color: {accent_light};
                border-color: {accent_light};
                image: none;
            }}
            QTableWidget {{
                border: 1px solid #E5E7EB;
                border-radius: 6px;
                background-color: white;
                gridline-color: #F3F4F6;
                font-size: 13px;
            }}
            QTableWidget::item {{
                padding: 6px;
            }}
            QTableWidget::item:selected {{
                background-color: #DBEAFE;
                color: #1E40AF;
            }}
            QHeaderView::section {{
                background-color: #F9FAFB;
                padding: 8px;
                border: none;
                border-bottom: 2px solid #E5E7EB;
                font-weight: 600;
                font-size: 12px;
            }}
            QProgressBar {{
                border: 1px solid #E5E7EB;
                border-radius: 6px;
                text-align: center;
                background-color: #F3F4F6;
                font-size: 12px;
            }}
            QProgressBar::chunk {{
                background-color: #238636;  /* constant green */
                border-radius: 5px;
            }}
            QToolBar {{
                background-color: #F9FAFB;
                border-bottom: 1px solid #E5E7EB;
                padding: 4px;
                spacing: 8px;
            }}
            QToolButton {{
                background-color: transparent;
                border: none;
                border-radius: 4px;
                padding: 8px 12px;
                font-size: 13px;
            }}
            QToolButton:hover {{
                background-color: #E5E7EB;
            }}
            QToolButton:checked {{
                background-color: {accent_light};
                color: white;
            }}
            QFrame {{
                background-color: white;
            }}
        """
        self.setStyleSheet(style)
    
    # =========================================================================
    # File Operations and Multi-Animal Workflow Methods
    # =========================================================================
    
    def _show_file_menu(self):
        """Show the File menu below the toolbar action"""
        # Get the toolbar widget for the action
        toolbar = self.findChild(QToolBar, "Quick Actions")
        if toolbar:
            # Find the position of the File action widget
            for action in toolbar.actions():
                if action == self.action_file:
                    widget = toolbar.widgetForAction(action)
                    if widget:
                        pos = widget.mapToGlobal(widget.rect().bottomLeft())
                        self.file_menu.exec_(pos)
                        return
        # Fallback: show at cursor position
        self.file_menu.exec_(QCursor.pos())
    
    def _create_welcome_screen(self):
        """Create an aesthetic welcome screen widget"""
        welcome = QWidget()
        welcome.setObjectName("welcomeScreen")
        
        # Main layout
        layout = QVBoxLayout(welcome)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Center content area
        center_widget = QWidget()
        center_layout = QVBoxLayout(center_widget)
        center_layout.setAlignment(Qt.AlignCenter)
        center_layout.setSpacing(20)
        
        # Logo/Icon area with brain wave visualization
        icon_label = QLabel("🧠")
        icon_label.setAlignment(Qt.AlignCenter)
        icon_label.setStyleSheet("font-size: 72px; background: transparent;")
        center_layout.addWidget(icon_label)
        
        # App title
        title = QLabel("AutoSWD")
        title.setAlignment(Qt.AlignCenter)
        title.setObjectName("welcomeTitle")
        center_layout.addWidget(title)
        
        # Subtitle
        subtitle = QLabel("Automated Spike-Wave Discharge Detection")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setObjectName("welcomeSubtitle")
        center_layout.addWidget(subtitle)
        
        center_layout.addSpacing(35)
        
        # Workflow cards container
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(25)
        cards_layout.setAlignment(Qt.AlignCenter)
        
        workflow_steps = [
            ("📂", "Load", "Import EDF files"),
            ("📊", "Analyze", "Detect SWD events"),
            ("💾", "Export", "Save results to Excel"),
        ]
        
        for icon, title_text, desc in workflow_steps:
            card = QFrame()
            card.setObjectName("featureCard")
            card.setFixedSize(140, 110)
            card_layout = QVBoxLayout(card)
            card_layout.setAlignment(Qt.AlignCenter)
            card_layout.setSpacing(6)
            
            card_icon = QLabel(icon)
            card_icon.setAlignment(Qt.AlignCenter)
            card_icon.setStyleSheet("font-size: 32px; background: transparent; border: none;")
            card_layout.addWidget(card_icon)
            
            card_title = QLabel(title_text)
            card_title.setAlignment(Qt.AlignCenter)
            card_title.setObjectName("cardTitle")
            card_layout.addWidget(card_title)
            
            card_desc = QLabel(desc)
            card_desc.setAlignment(Qt.AlignCenter)
            card_desc.setObjectName("cardDesc")
            card_layout.addWidget(card_desc)
            
            cards_layout.addWidget(card)
        
        center_layout.addLayout(cards_layout)
        
        center_layout.addSpacing(45)
        
        # Load button
        btn_container = QHBoxLayout()
        btn_container.setAlignment(Qt.AlignCenter)
        
        self.welcome_load_btn = QPushButton("   Load EDF File   ")
        self.welcome_load_btn.setObjectName("loadButton")
        self.welcome_load_btn.setCursor(Qt.PointingHandCursor)
        self.welcome_load_btn.setFixedSize(200, 50)
        self.welcome_load_btn.clicked.connect(self.load_edf_file)
        btn_container.addWidget(self.welcome_load_btn)
        
        center_layout.addLayout(btn_container)
        
        # Hint text
        hint = QLabel("or press Ctrl+O")
        hint.setAlignment(Qt.AlignCenter)
        hint.setObjectName("hintText")
        center_layout.addWidget(hint)
        
        layout.addStretch(1)
        layout.addWidget(center_widget)
        layout.addStretch(1)
        
        # Apply styling
        self._apply_welcome_style(welcome)
        
        return welcome
    
    def _apply_welcome_style(self, welcome_widget):
        """Apply styling to welcome screen based on current mode"""
        if self.dark_mode:
            welcome_widget.setStyleSheet("""
                #welcomeScreen {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #0D1117, stop:0.5 #161B22, stop:1 #0D1117);
                }
                #welcomeTitle {
                    font-size: 36px;
                    font-weight: bold;
                    color: #E6EDF3;
                    background: transparent;
                }
                #welcomeSubtitle {
                    font-size: 16px;
                    color: #8B949E;
                    background: transparent;
                }
                #featureCard {
                    background-color: rgba(33, 38, 45, 0.8);
                    border: 1px solid #30363D;
                    border-radius: 12px;
                }
                #featureCard:hover {
                    border-color: #58A6FF;
                    background-color: rgba(33, 38, 45, 1);
                }
                #cardTitle {
                    font-size: 13px;
                    font-weight: 600;
                    color: #E6EDF3;
                    background: transparent;
                    border: none;
                }
                #cardDesc {
                    font-size: 11px;
                    color: #8B949E;
                    background: transparent;
                    border: none;
                }
                #loadButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #238636, stop:1 #2EA043);
                    border: none;
                    border-radius: 8px;
                    color: white;
                    font-size: 16px;
                    font-weight: 600;
                }
                #loadButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #2EA043, stop:1 #3FB950);
                }
                #loadButton:pressed {
                    background: #238636;
                }
                #hintText {
                    font-size: 12px;
                    color: #6E7681;
                    background: transparent;
                }
            """)
        else:
            welcome_widget.setStyleSheet("""
                #welcomeScreen {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                        stop:0 #F8FAFC, stop:0.5 #EEF2FF, stop:1 #F8FAFC);
                }
                #welcomeTitle {
                    font-size: 36px;
                    font-weight: bold;
                    color: #1E293B;
                    background: transparent;
                }
                #welcomeSubtitle {
                    font-size: 16px;
                    color: #64748B;
                    background: transparent;
                }
                #featureCard {
                    background-color: rgba(255, 255, 255, 0.9);
                    border: 1px solid #E2E8F0;
                    border-radius: 12px;
                }
                #featureCard:hover {
                    border-color: #3B82F6;
                    background-color: white;
                    box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
                }
                #cardTitle {
                    font-size: 13px;
                    font-weight: 600;
                    color: #1E293B;
                    background: transparent;
                    border: none;
                }
                #cardDesc {
                    font-size: 11px;
                    color: #64748B;
                    background: transparent;
                    border: none;
                }
                #loadButton {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #3B82F6, stop:1 #6366F1);
                    border: none;
                    border-radius: 8px;
                    color: white;
                    font-size: 16px;
                    font-weight: 600;
                }
                #loadButton:hover {
                    background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                        stop:0 #2563EB, stop:1 #4F46E5);
                }
                #loadButton:pressed {
                    background: #1D4ED8;
                }
                #hintText {
                    font-size: 12px;
                    color: #94A3B8;
                    background: transparent;
                }
            """)
    
    def _show_welcome_screen(self):
        """Show the welcome screen"""
        if hasattr(self, 'content_stack'):
            self.content_stack.setCurrentIndex(0)
        # Disable controls that require EDF
        self._set_controls_enabled(False)
        
    def _show_plot_view(self):
        """Show the plot view (after EDF is loaded and channels selected)"""
        if hasattr(self, 'content_stack'):
            self.content_stack.setCurrentIndex(1)
        # Enable controls
        self._set_controls_enabled(True)
        
    def _set_controls_enabled(self, enabled):
        """Enable/disable controls based on whether EDF is loaded"""
        # File menu items
        self.action_select_channels.setEnabled(enabled and self.edf_raw is not None)
        self.action_save_animal.setEnabled(enabled)
        self.action_new_animal.setEnabled(enabled and self.edf_raw is not None)
        
        # Navigation controls
        if hasattr(self, 'slider'):
            self.slider.setEnabled(enabled)
        if hasattr(self, 'btn_prev_event'):
            self.btn_prev_event.setEnabled(enabled)
        if hasattr(self, 'btn_next_event'):
            self.btn_next_event.setEnabled(enabled)
        if hasattr(self, 'btn_add'):
            self.btn_add.setEnabled(enabled)
        if hasattr(self, 'btn_psd'):
            self.btn_psd.setEnabled(enabled)
            
    def load_edf_file(self):
        """Load an EDF file and show channel selection dialog"""
        if getattr(self, '_resampling', False):
            self.show_toast("Still downsampling the previous recording, please wait…", "warning")
            return

        eeg_path, _ = QFileDialog.getOpenFileName(
            self, "Select EDF File", "", "EDF Files (*.edf);;All Files (*)")
        
        if not eeg_path:
            return  # User cancelled
            
        try:
            # Load the EDF file
            raw = mne.io.read_raw_edf(eeg_path, preload=True, verbose=False)
            print(f"Loaded EEG with {len(raw.ch_names)} channels: {raw.ch_names}")

            # Downsample recordings sampled above 1000 Hz down to 1000 Hz.
            # The resample applies an anti-aliasing low-pass filter with a
            # cutoff at the new Nyquist frequency (target_fs / 2 = 500 Hz) so
            # that decimation cannot fold high-frequency energy back into the
            # retained band. This runs in a background thread so the UI stays
            # responsive on long recordings.
            target_fs = 1000.0
            orig_fs = float(raw.info['sfreq'])
            if orig_fs > target_fs:
                self._start_resample(raw, eeg_path, orig_fs, target_fs)
            else:
                self._finish_edf_load(raw, eeg_path)

        except Exception as e:
            QMessageBox.critical(self, "EDF Loading Error", f"Failed to load EDF file:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def _start_resample(self, raw, eeg_path, orig_fs, target_fs):
        """Downsample the raw recording in a background thread, keeping the UI
        responsive, then continue loading once finished."""
        # Prevent re-entry (e.g. loading another file) while resampling runs.
        self._resampling = True
        if hasattr(self, 'action_load_edf'):
            self.action_load_edf.setEnabled(False)

        # Show the indeterminate progress bar as a busy indicator with a label.
        try:
            self.status_widget.show()
        except Exception:
            pass
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFormat(
            f"Downsampling {orig_fs:.0f} Hz → {int(target_fs)} Hz (anti-aliased)…")
        self.progress_bar.show()
        QApplication.processEvents()

        self._pending_edf_path = eeg_path
        self._resample_orig_fs = orig_fs
        self._resample_target_fs = target_fs

        self.resample_thread = QThread(self)
        self.resample_worker = ResampleWorker(raw, target_fs)
        self.resample_worker.moveToThread(self.resample_thread)
        self.resample_worker.finished.connect(self._on_resample_finished)
        self.resample_worker.error.connect(self._on_resample_error)
        self.resample_thread.started.connect(self.resample_worker.run)
        self.resample_thread.start()

    def _on_resample_finished(self, raw):
        """Continue EDF loading after background downsampling completes."""
        if hasattr(self, 'resample_thread'):
            self.resample_thread.quit()
            self.resample_thread.wait()

        orig_fs = getattr(self, '_resample_orig_fs', 0.0)
        target_fs = getattr(self, '_resample_target_fs', 1000.0)
        eeg_path = getattr(self, '_pending_edf_path', None)

        self._resampling = False
        self.progress_bar.hide()
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFormat("")
        if hasattr(self, 'action_load_edf'):
            self.action_load_edf.setEnabled(True)

        print(f"Downsampled from {orig_fs:.1f} Hz to {target_fs:.1f} Hz (anti-aliased)")
        self.show_toast(
            f"Recording downsampled {orig_fs:.0f} → {int(target_fs)} Hz (anti-aliased)",
            "info")

        try:
            self._finish_edf_load(raw, eeg_path)
        except Exception as e:
            QMessageBox.critical(self, "EDF Loading Error", f"Failed to load EDF file:\n{str(e)}")
            import traceback
            traceback.print_exc()

    def _on_resample_error(self, msg):
        """Handle a failure during background downsampling."""
        if hasattr(self, 'resample_thread'):
            self.resample_thread.quit()
            self.resample_thread.wait()
        self._resampling = False
        self.progress_bar.hide()
        self.progress_bar.setTextVisible(False)
        self.progress_bar.setFormat("")
        if hasattr(self, 'action_load_edf'):
            self.action_load_edf.setEnabled(True)
        print(f"Resample error: {msg}")
        QMessageBox.critical(self, "Downsampling Error", f"Failed to downsample recording:\n{msg}")

    def _finish_edf_load(self, raw, eeg_path):
        """Store the loaded (and possibly downsampled) raw object and continue
        to channel selection."""
        # Store the raw object for channel switching
        self.edf_raw = raw
        self.edf_path = eeg_path
        self.edf_filename = os.path.basename(eeg_path)

        # Get start time
        start_time = raw.info.get('meas_date')
        if isinstance(start_time, tuple):
            start_time = start_time[0]
        if start_time is None:
            # Use current time as fallback
            start_time = datetime.now()
            self.show_toast("No measurement date in EDF, using current time", "warning")
        self.recording_start_time = start_time

        # Enable channel selection
        self.action_select_channels.setEnabled(True)

        # Show channel selection dialog
        self.select_channels_dialog()
            
    def select_channels_dialog(self):
        """Show dialog to select channels for the current animal"""
        if self.edf_raw is None:
            self.show_toast("Please load an EDF file first", "warning")
            return
            
        ch_names = list(self.edf_raw.ch_names)
        print(f"Showing channel selection dialog with {len(ch_names)} channels")
        
        # Generate default animal name from first word of first channel
        default_name = f"Animal_{self.animal_counter}"
        if self.current_animal_name:
            default_name = self.current_animal_name
        elif ch_names:
            first_word = ch_names[0].split()[0].strip() if ch_names[0].strip() else ""
            if first_word:
                default_name = first_word
            
        # Show channel selection dialog
        dialog = ChannelSelectionDialog(
            ch_names, 
            parent=self,
            default_animal_name=default_name,
            allow_single_channel=True,
            title="Select Channels for Animal"
        )
        
        # Ensure dialog is visible and on top
        dialog.raise_()
        dialog.activateWindow()
        
        result = dialog.exec_()
        print(f"Dialog result: {result} (Accepted={QDialog.Accepted})")
        
        if result == QDialog.Accepted:
            selection = dialog.get_selection()
            print(f"Channel selection: {selection}")
            self._apply_channel_selection(selection)
        else:
            print("Dialog was cancelled or closed")
            
    def _apply_channel_selection(self, selection):
        """Apply the channel selection and start analysis"""
        try:
            ch1_name, ch2_name = selection['channels']
            ch1_idx, ch2_idx = selection['channel_indices']
            animal_name = selection['animal_name']
            unet_channel = selection['unet_channel']
            
            # Reset model selection - treat this like a fresh start
            self.model_chosen = False
            self.model_type = None
            
            # Disable model buttons until preprocessing is complete (same as first load)
            if hasattr(self, 'btn_model_cwt'):
                self.btn_model_cwt.setEnabled(False)
                self.btn_model_cwt.setChecked(False)
            if hasattr(self, 'btn_model_unet'):
                self.btn_model_unet.setEnabled(False)
                self.btn_model_unet.setChecked(False)
            if hasattr(self, 'btn_model_cwt_unet'):
                self.btn_model_cwt_unet.setEnabled(False)
                self.btn_model_cwt_unet.setChecked(False)
            
            # Reset UNET display options
            if hasattr(self, 'cb_show_unet'):
                self.cb_show_unet.setChecked(False)
            if hasattr(self, 'cb_unet_refine'):
                self.cb_unet_refine.setChecked(False)
                self.cb_unet_refine.setEnabled(False)
            
            # Show loading bar
            self.progress_bar.show()
            QApplication.processEvents()
            
            # Track that this is a fresh animal with no unsaved changes
            self._animal_has_unsaved_changes = False
            self._last_saved_events_hash = None
            
            # Get data from raw
            data, _ = self.edf_raw[:]
            
            self.eeg0 = data[ch1_idx]
            if ch2_idx is None or ch2_name is None:
                # Single channel mode
                self.eeg1 = self.eeg0.copy()
                self.single_channel_view = True
                self.channel_names = [ch1_name]
            else:
                self.eeg1 = data[ch2_idx]
                self.single_channel_view = False
                self.channel_names = [ch1_name, ch2_name]
                
            self.sfreq = self.edf_raw.info['sfreq']
            self.total_samples = len(self.eeg0)
            self.current_animal_name = animal_name
            self.unet_channel = unet_channel
            self.export_animal_id = animal_name
            
            # Clear previous events
            self.manual_events = []
            self.merged_swd = []
            self.refined_swd = []
            self.all_events = []
            self.selected_event_idx = None
            
            # Clear preprocessing results
            self.cwt_token_predictions = None
            self.cwt_merged_swd = []
            self.cwt_refined_swd = []
            self.unet_predictions = None
            self.unet_intervals = []
            self.unet_refined_swd = []
            # Reset per-pipeline success flags for the new recording so a
            # prior animal's success state can't leak across loads.
            self._cwt_succeeded = False
            self._unet_succeeded = False
            # New animal means a fresh detection baseline; invalidate caches
            # and history so old edits can't leak across animals.
            if hasattr(self, '_model_state_cache'):
                self._model_state_cache.clear()
            if hasattr(self, '_undo_stack'):
                self._undo_stack.clear()
            if hasattr(self, '_redo_stack'):
                self._redo_stack.clear()
            # Fresh recording -> clear imported comments until the user imports
            # them again. This avoids leaking HH:MM picks across animals.
            self.adicht_comments = []
            # Reflect the cleared state in the File menu action states.
            try:
                self._refresh_comments_ui()
            except Exception:
                pass
            
            # Reset filtered data
            self._filtered_eeg0 = None
            self._filtered_eeg1 = None
            
            # Mark EDF as loaded
            self.edf_loaded = True
            
            # Update window title
            self.setWindowTitle(f"AutoSWD Detector - {self.edf_filename} - {animal_name}")
            
            # Enable file menu items
            self.action_save_animal.setEnabled(True)
            self.action_new_animal.setEnabled(True)
            
            # Show plot view
            self._show_plot_view()
            
            # Update slider range
            total_seconds = self.total_samples / self.sfreq
            self.slider.setMaximum(int(max(0, total_seconds - self.display_window_seconds)))
            self.slider.setValue(0)
            self.current_pos_sec = 0
            
            # Reset slider color to default (gray) since no model is selected
            self._apply_slider_color("#9CA3AF")
            
            # Reset minimap to default color and clear events
            if hasattr(self, 'timeline_minimap'):
                self.timeline_minimap.set_auto_color("#9CA3AF")
                self.update_timeline_minimap()
            
            # Start preprocessing
            self._start_background_preprocessing()
            
            # Update plot
            self.update_plot()
            
            self.show_toast(f"Loaded channels for {animal_name}", "success")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to apply channel selection:\n{str(e)}")
            import traceback
            traceback.print_exc()
            
    def _compute_events_hash(self):
        """Compute a simple hash of current events to track changes"""
        all_events = self._get_all_events_for_export()
        # Create a simple representation that changes when events change
        event_repr = [(ev.get('start_sample', 0), ev.get('end_sample', 0), ev.get('is_auto', False)) 
                      for ev in all_events]
        return hash(tuple(event_repr))
    
    def _has_unsaved_changes(self):
        """Check if there are unsaved changes to the current animal"""
        current_hash = self._compute_events_hash()
        return current_hash != getattr(self, '_last_saved_events_hash', None)
    
    def save_current_animal(self):
        """Save the current animal's events to the session"""
        if not self.edf_loaded:
            self.show_toast("No EDF loaded", "warning")
            return
            
        if not self.current_animal_name:
            # Ask for name if not set
            name, ok = QInputDialog.getText(self, "Animal Name", "Enter animal name:", 
                                           text=f"Animal_{self.animal_counter}")
            if not ok or not name.strip():
                return
            self.current_animal_name = name.strip()
            
        # Collect all events
        all_events = self._get_all_events_for_export()
        
        # Get channel info
        if self.single_channel_view:
            channels = (self.channel_names[0], None)
            channel_indices = (0, None)
        else:
            channels = tuple(self.channel_names[:2])
            channel_indices = (0, 1)
            
        # Store in animals_data
        self.animals_data[self.current_animal_name] = {
            'events': all_events.copy(),
            'channels': channels,
            'channel_indices': channel_indices,
            'sfreq': self.sfreq,
            'recording_start': self.recording_start_time,
            'model_type': getattr(self, 'model_type', None),
            'auto_event_count': len(self.refined_swd),
            'manual_event_count': len(self.manual_events),
            'adicht_comments': list(getattr(self, 'adicht_comments', []) or []),
        }
        # Mirror into the per-animal comment cache so switching animals later
        # restores the right set.
        if not hasattr(self, '_animal_comments_cache'):
            self._animal_comments_cache = {}
        self._animal_comments_cache[self.current_animal_name] = list(
            getattr(self, 'adicht_comments', []) or []
        )
        
        # Update the saved hash to track changes from this point
        self._last_saved_events_hash = self._compute_events_hash()
        
        # Update saved animals menu
        self._update_saved_animals_menu()
        
        self.show_toast(f"Saved {len(all_events)} events for {self.current_animal_name}", "success")
        
    def _get_all_events_for_export(self):
        """Get all events (auto + manual) for export"""
        all_events = []
        
        # Add refined auto events
        for ev in self.refined_swd:
            all_events.append(ev.copy())
            
        # Add manual events
        for ev in self.manual_events:
            all_events.append(ev.copy())
            
        # Sort by start time
        all_events.sort(key=lambda x: x.get('start_sample', 0))
        
        return all_events
        
    def start_new_animal(self):
        """Save current animal and start analyzing a new one"""
        if not self.edf_loaded:
            self.show_toast("No EDF loaded", "warning")
            return
            
        # Only ask to save if there are unsaved changes
        if self._has_unsaved_changes() and len(self._get_all_events_for_export()) > 0:
            reply = QMessageBox.question(
                self, "Save Current Animal?",
                f"Save events for {self.current_animal_name or 'current animal'} before switching?",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            
            if reply == QMessageBox.Cancel:
                return
            elif reply == QMessageBox.Yes:
                self.save_current_animal()
                
        # Increment counter for new animal name
        self.animal_counter += 1
        self.current_animal_name = None
        
        # Show channel selection dialog for new animal
        self.select_channels_dialog()
        
    def _update_saved_animals_menu(self):
        """Update the saved animals submenu"""
        self.saved_animals_menu.clear()
        
        if not self.animals_data:
            self.saved_animals_menu.setEnabled(False)
            return
            
        self.saved_animals_menu.setEnabled(True)
        
        for animal_name, data in self.animals_data.items():
            event_count = len(data.get('events', []))
            channels = data.get('channels', ('?', '?'))
            ch_str = f"{channels[0]}" if channels[1] is None else f"{channels[0]} + {channels[1]}"
            
            action = self.saved_animals_menu.addAction(
                f"{animal_name} ({event_count} events) - {ch_str}"
            )
            action.setData(animal_name)
            action.triggered.connect(lambda checked, name=animal_name: self._load_saved_animal(name))
    
    def show_manage_animals_dialog(self):
        """Show dialog to manage (rename/delete) saved animals"""
        if not self.animals_data:
            self.show_toast("No saved animals to manage", "warning")
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("Manage Saved Animals")
        dialog.setMinimumWidth(500)
        dialog.setMinimumHeight(350)
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Apply dark mode styling if active
        if self.dark_mode:
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #0D1117;
                    color: #E6EDF3;
                }
                QLabel {
                    color: #E6EDF3;
                    font-size: 13px;
                }
                QListWidget {
                    background-color: #161B22;
                    border: 1px solid #30363D;
                    border-radius: 8px;
                    color: #E6EDF3;
                    font-size: 13px;
                    padding: 5px;
                }
                QListWidget::item {
                    padding: 10px;
                    border-radius: 4px;
                    margin: 2px 0;
                }
                QListWidget::item:selected {
                    background-color: #238636;
                    color: white;
                }
                QListWidget::item:hover {
                    background-color: #21262D;
                }
                QPushButton {
                    background-color: #21262D;
                    border: 1px solid #30363D;
                    border-radius: 6px;
                    color: #E6EDF3;
                    padding: 8px 16px;
                    font-size: 13px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background-color: #30363D;
                    border-color: #8B949E;
                }
            """)
        else:
            dialog.setStyleSheet("""
                QDialog {
                    background-color: #FFFFFF;
                }
                QLabel {
                    color: #1F2937;
                    font-size: 13px;
                }
                QListWidget {
                    background-color: #F9FAFB;
                    border: 1px solid #E5E7EB;
                    border-radius: 8px;
                    color: #1F2937;
                    font-size: 13px;
                    padding: 5px;
                }
                QListWidget::item {
                    padding: 10px;
                    border-radius: 4px;
                    margin: 2px 0;
                }
                QListWidget::item:selected {
                    background-color: #3B82F6;
                    color: white;
                }
                QListWidget::item:hover {
                    background-color: #E5E7EB;
                }
                QPushButton {
                    background-color: #F3F4F6;
                    border: 1px solid #D1D5DB;
                    border-radius: 6px;
                    color: #1F2937;
                    padding: 8px 16px;
                    font-size: 13px;
                    font-weight: 500;
                }
                QPushButton:hover {
                    background-color: #E5E7EB;
                    border-color: #9CA3AF;
                }
            """)
        
        # Title
        title_label = QLabel("Manage Saved Animals")
        title_label.setStyleSheet(f"font-size: 18px; font-weight: bold; color: {'#E6EDF3' if self.dark_mode else '#1F2937'}; margin-bottom: 5px;")
        layout.addWidget(title_label)
        
        # Instructions
        info_label = QLabel("Select an animal to rename or delete:")
        layout.addWidget(info_label)
        
        # List of animals
        self.animals_list = QListWidget()
        self.animals_list.setSelectionMode(QListWidget.SingleSelection)
        
        for animal_name, data in self.animals_data.items():
            event_count = len(data.get('events', []))
            channels = data.get('channels', ('?', '?'))
            ch_str = f"{channels[0]}" if channels[1] is None else f"{channels[0]} + {channels[1]}"
            
            item = QListWidgetItem(f"  {animal_name}  •  {event_count} events  •  {ch_str}")
            item.setData(Qt.UserRole, animal_name)
            self.animals_list.addItem(item)
        
        layout.addWidget(self.animals_list)
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        rename_btn = QPushButton("✏️ Rename")
        rename_btn.setCursor(Qt.PointingHandCursor)
        rename_btn.clicked.connect(lambda: self._rename_selected_animal(dialog))
        btn_layout.addWidget(rename_btn)
        
        delete_btn = QPushButton("🗑️ Delete")
        delete_btn.setCursor(Qt.PointingHandCursor)
        if self.dark_mode:
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #8B2635;
                    border: 1px solid #A63446;
                    color: white;
                }
                QPushButton:hover {
                    background-color: #A63446;
                }
            """)
        else:
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #DC2626;
                    border: 1px solid #B91C1C;
                    color: white;
                }
                QPushButton:hover {
                    background-color: #B91C1C;
                }
            """)
        delete_btn.clicked.connect(lambda: self._delete_selected_animal(dialog))
        btn_layout.addWidget(delete_btn)
        
        btn_layout.addStretch()
        
        close_btn = QPushButton("Close")
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.clicked.connect(dialog.accept)
        btn_layout.addWidget(close_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec_()
    
    def _rename_selected_animal(self, dialog):
        """Rename the selected animal"""
        selected_items = self.animals_list.selectedItems()
        if not selected_items:
            self.show_toast("Please select an animal to rename", "warning")
            return
            
        old_name = selected_items[0].data(Qt.UserRole)
        
        new_name, ok = QInputDialog.getText(
            dialog, "Rename Animal", 
            f"Enter new name for '{old_name}':",
            text=old_name
        )
        
        if ok and new_name.strip() and new_name.strip() != old_name:
            new_name = new_name.strip()
            
            # Check if name already exists
            if new_name in self.animals_data:
                self.show_toast(f"Animal '{new_name}' already exists", "warning")
                return
                
            # Rename in data
            self.animals_data[new_name] = self.animals_data.pop(old_name)
            
            # Update if this is the current animal
            if self.current_animal_name == old_name:
                self.current_animal_name = new_name
                self.setWindowTitle(f"AutoSWD Detector - {self.edf_filename} - {new_name}")
            
            # Update menu and list
            self._update_saved_animals_menu()
            
            # Update list item
            selected_items[0].setText(f"{new_name} - {len(self.animals_data[new_name].get('events', []))} events")
            selected_items[0].setData(Qt.UserRole, new_name)
            
            self.show_toast(f"Renamed '{old_name}' to '{new_name}'", "success")
    
    def _delete_selected_animal(self, dialog):
        """Delete the selected animal"""
        selected_items = self.animals_list.selectedItems()
        if not selected_items:
            self.show_toast("Please select an animal to delete", "warning")
            return
            
        animal_name = selected_items[0].data(Qt.UserRole)
        
        reply = QMessageBox.question(
            dialog, "Confirm Delete",
            f"Are you sure you want to delete '{animal_name}'?\nThis cannot be undone.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            # Remove from data
            del self.animals_data[animal_name]
            
            # Update menu
            self._update_saved_animals_menu()
            
            # Remove from list
            row = self.animals_list.row(selected_items[0])
            self.animals_list.takeItem(row)
            
            self.show_toast(f"Deleted '{animal_name}'", "success")
            
            # Close dialog if no more animals
            if not self.animals_data:
                dialog.accept()
            
    def _load_saved_animal(self, animal_name):
        """Load a previously saved animal for viewing/editing"""
        if animal_name not in self.animals_data:
            self.show_toast(f"Animal {animal_name} not found", "error")
            return
            
        # Only ask to save if there are unsaved changes
        if self.current_animal_name and self.current_animal_name != animal_name and self._has_unsaved_changes():
            reply = QMessageBox.question(
                self, "Save Current Animal?",
                f"Save events for {self.current_animal_name} before switching?",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel
            )
            
            if reply == QMessageBox.Cancel:
                return
            elif reply == QMessageBox.Yes:
                self.save_current_animal()
                
        # Load the saved animal
        data = self.animals_data[animal_name]
        
        # Restore channel data from EDF
        channels = data.get('channels', (None, None))
        ch_names = list(self.edf_raw.ch_names)
        
        try:
            ch1_idx = ch_names.index(channels[0]) if channels[0] else 0
            ch2_idx = ch_names.index(channels[1]) if channels[1] else None
        except ValueError:
            self.show_toast("Channel names not found in EDF", "error")
            return
            
        raw_data, _ = self.edf_raw[:]
        self.eeg0 = raw_data[ch1_idx]
        
        if ch2_idx is not None:
            self.eeg1 = raw_data[ch2_idx]
            self.single_channel_view = False
            self.channel_names = [channels[0], channels[1]]
        else:
            self.eeg1 = self.eeg0.copy()
            self.single_channel_view = True
            self.channel_names = [channels[0]]
            
        self.sfreq = data.get('sfreq', self.edf_raw.info['sfreq'])
        self.total_samples = len(self.eeg0)
        self.current_animal_name = animal_name
        self.export_animal_id = animal_name
        self.recording_start_time = data.get('recording_start', self.recording_start_time)
        
        # Restore events
        events = data.get('events', [])
        self.refined_swd = [ev for ev in events if ev.get('is_auto', False)]
        self.manual_events = [ev for ev in events if not ev.get('is_auto', False)]
        self.all_events = events.copy()

        # Restore imported recording comments for this animal (if any)
        cached_comments = None
        if hasattr(self, '_animal_comments_cache'):
            cached_comments = self._animal_comments_cache.get(animal_name)
        if cached_comments is None:
            cached_comments = data.get('adicht_comments', [])
        self.adicht_comments = list(cached_comments or [])
        try:
            self._refresh_comments_ui()
        except Exception:
            pass
        
        # Clear filtered data
        self._filtered_eeg0 = None
        self._filtered_eeg1 = None
        
        # Update window title
        self.setWindowTitle(f"AutoSWD Detector - {self.edf_filename} - {animal_name}")
        
        # Update slider range
        total_seconds = self.total_samples / self.sfreq
        self.slider.setMaximum(int(max(0, total_seconds - self.display_window_seconds)))
        
        # Update plot
        self.update_plot()
        
        self.show_toast(f"Loaded {animal_name} with {len(events)} events", "success")
        
    def connect_signals(self):
        """Connect all signal handlers."""
        # Toolbar actions
        self.action_export.triggered.connect(self.on_export_clicked)
        self.action_events.triggered.connect(self.show_events_popup)
        self.action_statistics.triggered.connect(self.show_statistics_popup)
        self.action_settings.triggered.connect(self.on_settings_clicked)
        self.action_dark_mode.triggered.connect(self.on_dark_mode_toggle)
        self.action_help.triggered.connect(self.show_keyboard_shortcuts)
        
        # Navigation
        self.slider.valueChanged.connect(self.on_slider_changed)
        self.win_spin.valueChanged.connect(self.on_window_size_changed)
        # Event navigation
        self.btn_prev_event.clicked.connect(self.on_prev_event_clicked)
        self.btn_next_event.clicked.connect(self.on_next_event_clicked)
        # Operations
        self.btn_merge_swds.clicked.connect(self.on_merge_swds_clicked)
        self.btn_event_filter.toggled.connect(self.on_event_filter_toggled)
        
        # Y-axis controls
        self.btn_ylim_auto.clicked.connect(self.on_ylim_auto_toggle)
        self.ylim_input.valueChanged.connect(self.on_ylim_manual_change)
        
        # Edit mode removed (always on unless disabled in settings)
        
        # Event actions
        self.btn_add.clicked.connect(self.on_add_event_clicked)
        self.btn_del.clicked.connect(self.on_delete_event_clicked)
        self.btn_psd.clicked.connect(self.on_psd_analysis_clicked)
        
        # Display options (hidden checkboxes)
        self.cb_show_auto.stateChanged.connect(self.update_plot)
        self.cb_show_unrefined.stateChanged.connect(self.on_show_unrefined_changed)
        self.cb_dark_mode.stateChanged.connect(self.on_dark_mode_changed)
        self.cb_show_unet.stateChanged.connect(self.update_plot)
        self.cb_unet_post.stateChanged.connect(self.update_plot)
        self.cb_unet_refine.stateChanged.connect(self.on_unet_refine_toggled)

        # Filter controls removed from control panel; nothing to connect here
        
        # Canvas events
        self.canvas.mpl_connect('pick_event', self.on_pick_region)
        self.canvas.mpl_connect('button_press_event', self.on_mouse_click)
        self.canvas.mpl_connect('button_release_event', self.on_mouse_release)
        self.canvas.mpl_connect('motion_notify_event', self.on_mouse_move)
        
        # Duration filter controls (hidden)
        self.cb_duration_filter.stateChanged.connect(self.on_duration_filter_changed)
        self.duration_min_spin.valueChanged.connect(self.update_filtered_events)
        self.duration_max_spin.valueChanged.connect(self.update_filtered_events)
    
    def show_toast(self, message, notification_type="info", duration=3000):
        """Show a toast notification"""
        toast = ToastNotification(self, message, notification_type, duration)
        toast.show_toast()
    
    def show_keyboard_shortcuts(self):
        """Open keyboard shortcuts editor"""
        self.show_key_assignments_dialog()
    
    def on_dark_mode_toggle(self):
        """Toggle dark mode from toolbar action"""
        self.dark_mode = self.action_dark_mode.isChecked()
        self.cb_dark_mode.setChecked(self.dark_mode)
        self.on_dark_mode_changed(self.dark_mode)

    # --------------------- Filter helpers ---------------------
    def _apply_plot_bandpass_filter(self):
        """Apply mne.filter.filter_data to both channels for plotting only.
        Caches results to avoid repeated filtering when unchanged."""
        try:
            l_freq = float(self.filter_low_cut_hz) if self.filter_low_cut_hz and self.filter_low_cut_hz > 0 else None
            h_freq = float(self.filter_high_cut_hz) if self.filter_high_cut_hz and self.filter_high_cut_hz < (self.sfreq/2) else None
            if l_freq is not None and h_freq is not None and l_freq >= h_freq:
                h_freq = min(max(l_freq + 0.1, 0.2), self.sfreq/2 - 0.1)
            self._filtered_eeg0 = mne.filter.filter_data(self.eeg0.astype(np.float64), sfreq=self.sfreq, l_freq=l_freq, h_freq=h_freq, verbose=False)
            if hasattr(self, 'single_channel_view') and self.single_channel_view:
                self._filtered_eeg1 = self._filtered_eeg0
            else:
                self._filtered_eeg1 = mne.filter.filter_data(self.eeg1.astype(np.float64), sfreq=self.sfreq, l_freq=l_freq, h_freq=h_freq, verbose=False)
        except Exception as e:
            print(f"Filter error: {e}")
            self._filtered_eeg0 = None
            self._filtered_eeg1 = None

    def on_filter_toggle(self, state):
        self.filter_signal_enabled = bool(state)
        if self.filter_signal_enabled:
            self.filter_low_cut_hz = float(self.filter_low_spin.value())
            self.filter_high_cut_hz = float(self.filter_high_spin.value())
            self._apply_plot_bandpass_filter()
        self.update_plot()

    def on_filter_params_changed(self, _value):
        if not hasattr(self, 'cb_filter_signal'):
            return
        if not self.cb_filter_signal.isChecked():
            return
        self.filter_low_cut_hz = float(self.filter_low_spin.value())
        self.filter_high_cut_hz = float(self.filter_high_spin.value())
        self._apply_plot_bandpass_filter()
        self.update_plot()
    
    def on_minimap_clicked(self, time_sec):
        """Handle minimap clicks to jump to position"""
        self.current_pos_sec = max(0, min(time_sec, self.total_samples / self.sfreq - self.display_window_seconds))
        slider_max = max(1, int(self.total_samples / self.sfreq - self.display_window_seconds))
        self.slider.blockSignals(True)
        self.slider.setValue(int(self.current_pos_sec))
        self.slider.blockSignals(False)
        self.update_plot()
        self.update_timeline_minimap()
    
    def update_timeline_minimap(self):
        """Update the timeline minimap with current state"""
        total_duration = self.total_samples / self.sfreq
        
        # Collect events for minimap
        minimap_events = []
        
        # Check if duration filter is active
        filter_active = self.cb_duration_filter.isChecked()
        
        if filter_active:
            # Only show filtered events in minimap
            filtered_event_keys = {(ev['type'], ev['index']) for ev in self.filtered_events}
            
            for idx, ev in enumerate(self.refined_swd):
                if ('auto', idx) in filtered_event_keys:
                    start = ev['start_sample'] / self.sfreq
                    end = ev['end_sample'] / self.sfreq
                    minimap_events.append((start, end, 'auto'))
            
            for idx, ev in enumerate(self.manual_events):
                if ('manual', idx) in filtered_event_keys:
                    start = ev['start_sample'] / self.sfreq
                    end = ev['end_sample'] / self.sfreq
                    minimap_events.append((start, end, 'manual'))
        else:
            # Show all events
            for ev in self.refined_swd:
                start = ev['start_sample'] / self.sfreq
                end = ev['end_sample'] / self.sfreq
                minimap_events.append((start, end, 'auto'))
            
            for ev in self.manual_events:
                start = ev['start_sample'] / self.sfreq
                end = ev['end_sample'] / self.sfreq
                minimap_events.append((start, end, 'manual'))

        # Imported recording comments as zero-width 'comment' markers.
        # Rendered by the minimap as thin amber ticks on top of events.
        for c in getattr(self, 'adicht_comments', []) or []:
            t = float(c.get('time_sec', 0.0) or 0.0)
            if 0.0 <= t <= total_duration:
                minimap_events.append((t, t, 'comment'))

        self.timeline_minimap.set_data(
            total_duration,
            self.current_pos_sec,
            self.display_window_seconds,
            minimap_events
        )
    
    def update_status_indicators(self):
        """Update all status bar indicators"""
        # Position indicator
        current_time = (self.recording_start_time + timedelta(seconds=self.current_pos_sec)).strftime('%H:%M:%S')
        self.position_indicator.setText(f"📍 {current_time}")
        
        # Model indicator
        model_text = {
            "unet_only": "🟢 UNET",
            "cwt_only": "🔴 CWT",
            "cwt_unet": "🟡 CWT+UNET",
            "": "🔬 No Model"
        }.get(self.model_type, "🔬 No Model")
        self.model_indicator.setText(model_text)
        
        # Event count badge
        total = len(self.refined_swd) + len(self.manual_events)
        self.event_count_badge.setText(f"📊 {total} Events")

    # Advanced toggle removed; all controls are visible

    def load_and_preprocess(self):
        """Load and preprocess the EEG data."""
        self.progress_bar.show()
        QApplication.processEvents()
        
        # Process data and get predictions
        intervals0 = divide_into_intervals(self.eeg0, self.sfreq, self.interval_length, self.overlap_length)
        intervals1 = divide_into_intervals(self.eeg1, self.sfreq, self.interval_length, self.overlap_length)
        intervals_pairs = list(zip(intervals0, intervals1))
        
        # Process intervals
        def progress_callback(done, total, eta):
            # Progress bar is indeterminate, no need to update value
            pass
            QApplication.processEvents()
        
        features = process_intervals_parallel(intervals_pairs, self.sfreq, progress_callback)
        print("\nGrouping intervals...")
        sequences = group_intervals_with_overlap(features, self.sequence_length)
        
        # Get predictions
        print("Running model...")
        token_preds = self.model.predict(sequences, verbose=1)
        token_preds = token_preds.reshape(-1)
        
        # Store token predictions for later reanalysis
        self.token_predictions = token_preds
        
        # Merge predictions into regions
        thres = self.th_spin.value()
        gap = self.gap_spin.value()
        # Overlap agreement requires both overlapping tokens to predict SWD.
        # The flag is a user-visible setting (Advanced Settings); we keep the
        # model-type default so behaviour is unchanged until the user toggles.
        use_overlap_agreement = bool(getattr(self, 'require_overlap_agreement',
                                             self.model_type in ("cwt_only", "cwt_unet")))
        self.merged_swd = merge_adjacent_swds(token_preds, threshold=thres, max_gap_intervals=gap, require_overlap_agreement=use_overlap_agreement)
        
        # Convert raw merged regions to sample indices
        step = int((self.interval_length - self.overlap_length) * self.sfreq)
        for region in self.merged_swd:
            start_idx = region['start_idx']
            end_idx = region['end_idx']
            region['start_sample'] = start_idx * step
            region['end_sample'] = (end_idx + 1) * step + int(self.interval_length * self.sfreq) - step
            region['duration_sec'] = (region['end_sample'] - region['start_sample']) / self.sfreq
            region['is_auto'] = True
        
        # Refine borders
        power_percentile = self.power_percentile_spin.value()
        try:
            self.refined_swd = refine_swd_borders(
                self.eeg0, self.sfreq, self.merged_swd, power_percentile
            )
        except Exception as e:
            print(f"Warning: Border refinement failed: {e}")
            print("Using unrefined events instead")
            # Fall back to unrefined events
            self.refined_swd = []
            for region in self.merged_swd:
                refined_event = region.copy()
                refined_event['duration_sec'] = (region['end_sample'] - region['start_sample']) / self.sfreq
                if refined_event['duration_sec'] >= 1.0:
                    self.refined_swd.append(refined_event)
        
        # Filter by minimum duration
        self.refined_swd = [r for r in self.refined_swd if r['duration_sec'] >= 1.0]
        
        # Final merge step: merge events closer than 1 second apart
        try:
            self.refined_swd = merge_overlapping_events(self.refined_swd, self.sfreq, min_overlap_sec=-1.0)
        except Exception as e:
            print(f"Warning: Event merging failed: {e}")
            # Continue with unmerged events
        
        # Reset selection if it was an auto event
        if self.active_event_type == 'auto':
            self.selected_event_idx = None
            self.active_event_type = None
            self.current_edit_event = None
            self.event_info_lbl.setText("")
        
        # Update displays
        self.progress_bar.hide()
        # Keep a copy before optional UNET refinement
        self.refined_swd_base = list(self.refined_swd)
        # Apply UNET refinement if enabled
        try:
            self.apply_unet_refinement_if_enabled()
        except Exception as e:
            print(f"Warning: UNET refinement failed: {e}")
            # Continue without UNET refinement
        self.update_combined_events()
        self.update_filtered_events()
        self.update_plot()

    def update_combined_events(self):
        """Update the combined list of events for consistent selection/editing"""
        self.all_events = []
        for i, ev in enumerate(self.manual_events):
            ev_copy = ev.copy()
            ev_copy['idx'] = i
            ev_copy['type'] = 'manual'
            ev_copy['is_auto'] = False
            self.all_events.append(ev_copy)
            
        for i, ev in enumerate(self.refined_swd):
            ev_copy = ev.copy()
            ev_copy['idx'] = i
            ev_copy['type'] = 'auto'
            ev_copy['is_auto'] = True
            self.all_events.append(ev_copy)
        
        # Only reset selection if the selected event no longer exists
        if (self.selected_event_idx is not None and self.active_event_type is not None):
            # Check if the currently selected event still exists
            events = self.manual_events if self.active_event_type == 'manual' else self.refined_swd
            if self.selected_event_idx >= len(events):
                # Selected event no longer exists, reset selection
                self.selected_event_idx = None
                self.active_event_type = None
                self.event_info_lbl.setText("ℹ️ No event selected")
            
        self.update_info_display()
        self.update_filtered_events()
        self.update_status_indicators()
        self.update_timeline_minimap()

    def update_info_display(self):
        """Update the information display with event counts and durations."""
        auto_cnt = len(self.refined_swd)
        manu_cnt = len(self.manual_events)
        total_cnt = auto_cnt + manu_cnt
        
        # Update button states
        self.btn_psd.setEnabled(self.selected_event_idx is not None)

    def update_plot(self):
        """Update the plot with current data and settings."""
        # Clear the figure first
        self.fig.clear()
        
        # Determine colors based on dark mode
        if self.dark_mode:
            text_color = '#E6EDF3'
            grid_color = '#30363D'
            bg_color = '#0D1117'
            plot_bg = '#161B22'
        else:
            plt.style.use('default')
            text_color = 'black'
            grid_color = '#CCCCCC'
            bg_color = 'white'
            plot_bg = 'white'
        
        # Set figure background color
        self.fig.patch.set_facecolor(bg_color)
        
        # Calculate SWD statistics
        auto_count = len(self.refined_swd)
        manual_count = len(self.manual_events)
        total_count = auto_count + manual_count
        
        auto_duration = sum(ev['duration_sec'] for ev in self.refined_swd)
        manual_duration = sum(ev['duration_sec'] for ev in self.manual_events)
        total_duration = auto_duration + manual_duration
        
        # Calculate filtered statistics when filter is active
        if self.cb_duration_filter.isChecked():
            filtered_count = len(self.filtered_events)
            filtered_auto = sum(1 for ev in self.filtered_events if ev['type'] == 'auto')
            filtered_manual = sum(1 for ev in self.filtered_events if ev['type'] == 'manual')
            filtered_duration = sum(ev['event']['duration_sec'] for ev in self.filtered_events)
            stats_text = f"<b>SWD Statistics (Filtered):</b> {filtered_count}/{total_count} events (Auto: {filtered_auto}/{auto_count}, Manual: {filtered_manual}/{manual_count}) | " \
                        f"Duration: {filtered_duration:.1f}s/{total_duration:.1f}s | Avg: {filtered_duration/max(1,filtered_count):.1f}s"
        else:
            stats_text = f"<b>SWD Statistics:</b> {total_count} events (Auto: {auto_count}, Manual: {manual_count}) | " \
                        f"Total Duration: {total_duration:.1f}s | Avg Duration: {total_duration/max(1,total_count):.1f}s"
        
        # Create a title with SWD statistics
        title = f"SWD Events: {total_count} (Auto: {auto_count}, Manual: {manual_count}) | " \
                f"Duration: {total_duration:.1f}s (Auto: {auto_duration:.1f}s, Manual: {manual_duration:.1f}s)"
        
        # Update the stats label
        self.stats_lbl.setText(stats_text)
        
        # Check if single channel view
        single_channel = getattr(self, 'single_channel_view', False)
        
        # Create subplots
        if single_channel:
            ax1 = self.fig.add_subplot(1,1,1)
            ax2 = None
            axes = [ax1]
        else:
            ax1 = self.fig.add_subplot(2,1,1)
            ax2 = self.fig.add_subplot(2,1,2)
            axes = [ax1, ax2]
        
        # Set background colors for axes
        for ax in axes:
            ax.set_facecolor(plot_bg)
        
        # Apply dark mode styling
        if self.dark_mode:
            subtle = '#484F58'  # More visible border
            for ax in axes:
                for side in ('top','bottom','left','right'):
                    ax.spines[side].set_color(subtle)
                ax.tick_params(colors=text_color, labelcolor=text_color)
                ax.xaxis.label.set_color(text_color)
                ax.yaxis.label.set_color(text_color)

        # Channel labels as left-aligned subplot titles
        ch1_label = self.channel_names[0] if isinstance(self.channel_names, (list, tuple)) and len(self.channel_names) >= 1 else "Channel 1"
        ch2_label = self.channel_names[1] if isinstance(self.channel_names, (list, tuple)) and len(self.channel_names) >= 2 else "Channel 2"
        label_color = text_color if not self.dark_mode else '#D0D0D0'
        ax1.set_title(ch1_label, loc='left', color=label_color, fontsize=10)
        if ax2:
            ax2.set_title(ch2_label, loc='left', color=label_color, fontsize=10)
        
        # Get display window data
        start_s = int(self.current_pos_sec*self.sfreq)
        end_s = min(start_s+int(self.display_window_seconds*self.sfreq), self.total_samples)
        t = np.arange(start_s,end_s)/self.sfreq

        # Choose raw or filtered signal for plotting
        eeg0_src = self.eeg0
        eeg1_src = self.eeg1
        if self.filter_signal_enabled:
            # Ensure cached filtered signals up to date using current settings (from Settings)
            if (self._filtered_eeg0 is None or self._filtered_eeg1 is None):
                self._apply_plot_bandpass_filter()
            if self._filtered_eeg0 is not None and self._filtered_eeg1 is not None:
                eeg0_src = self._filtered_eeg0
                eeg1_src = self._filtered_eeg1
        
        # Calculate y-limits from navigation controls
        if self.btn_ylim_auto.isChecked():
            # Auto-scale mode - share the same y-limits across both channels
            # Very small floor so low-amplitude signals can zoom in much further
            min_range = 1e-10
            data1 = eeg0_src[start_s:end_s]
            p99_1 = np.percentile(np.abs(data1), 99) if len(data1) > 0 else 1e-5
            if ax2:
                data2 = eeg1_src[start_s:end_s]
                p99_2 = np.percentile(np.abs(data2), 99) if len(data2) > 0 else 1e-5
                p99_shared = max(p99_1, p99_2)
            else:
                p99_shared = p99_1
            y_max = max(p99_shared * 1.3, min_range)
            y1_lim = (-y_max, y_max)
            y2_lim = (-y_max, y_max)
        else:
            # Manual y-limits from spinbox (convert mV to V)
            ylim = self.ylim_input.value() / 1000.0
            y1_lim = (-ylim, ylim)
            y2_lim = (-ylim, ylim)
        
        # Apply y-limits to both axes
        try:
            ax1.set_ylim(y1_lim)
            if ax2:
                ax2.set_ylim(y2_lim)
        except Exception:
            pass
        # Set x-limits to ensure full coverage
        ax1.set_xlim(self.current_pos_sec, self.current_pos_sec + self.display_window_seconds)
        if ax2:
            ax2.set_xlim(self.current_pos_sec, self.current_pos_sec + self.display_window_seconds)
        
        # Plot EEG signals
        ch1_label = self.channel_names[0] if isinstance(self.channel_names, (list, tuple)) and len(self.channel_names) >= 1 else "Channel 1"
        ch2_label = self.channel_names[1] if isinstance(self.channel_names, (list, tuple)) and len(self.channel_names) >= 2 else "Channel 2"
        # Use softer, muted colors for better eye comfort
        signal_color = '#5A8FBA' if self.dark_mode else '#3366AA'  # Softer blue tones
        ax1.plot(t, eeg0_src[start_s:end_s], label=ch1_label, color=signal_color, linewidth=0.8)
        if ax2:
            ax2.plot(t, eeg1_src[start_s:end_s], label=ch2_label, color=signal_color, linewidth=0.8)
        
        # Clear region patches
        self.region_patches = []
        
        # Show unrefined events if enabled
        if self.show_unrefined and hasattr(self, 'merged_swd'):
            for idx, ev in enumerate(self.merged_swd):
                s = ev.get('start_sample', 0)/self.sfreq
                e = ev.get('end_sample', 0)/self.sfreq
                if e < self.current_pos_sec or s > self.current_pos_sec + self.display_window_seconds:
                    continue
                    
                vs = max(s, self.current_pos_sec)
                ve = min(e, self.current_pos_sec + self.display_window_seconds)
                
                # Make events pickable in edit mode
                # Use softer orange in dark mode
                unrefined_color = '#B8860B' if self.dark_mode else 'orange'  # Dark goldenrod in dark mode
                for ax in axes:
                    p = ax.axvspan(vs, ve, color=unrefined_color, alpha=0.2, linestyle='--', linewidth=0.5)
        
        # Display auto events if enabled
        show_auto = self.cb_show_auto.isChecked()
        
        # Build set of filtered event keys for quick lookup when filter is active
        filter_active = self.cb_duration_filter.isChecked()
        if filter_active:
            filtered_event_keys = {(ev['type'], ev['index']) for ev in self.filtered_events}
        else:
            filtered_event_keys = None
        
        # Plot auto events with improved colors
        if show_auto:
            # Color scheme based on model type - toned down in dark mode
            if self.model_type == "unet_only":
                event_color = '#4A9B6E' if self.dark_mode else '#50C878'  # Muted emerald in dark mode
                text_color = '#2D5F3F' if self.dark_mode else '#2EA043'   # Darker green for text
            elif self.model_type == "cwt_unet":
                event_color = '#8B7A4A' if self.dark_mode else '#AC994F'  # Muted gold in dark mode
                text_color = '#5C4E2F' if self.dark_mode else '#756436'   # Darker gold for text
            elif self.model_type == "cwt_only":
                event_color = '#CC5555' if self.dark_mode else '#FF0000'  # Muted red in dark mode
                text_color = '#993333' if self.dark_mode else '#CC0000'   # Darker red for text
            else:
                event_color = '#2F2F2F'  # Fallback dark gray
                text_color = '#000000'   # Black for text
            
            for idx, ev in enumerate(self.refined_swd):
                s = ev['start_sample']/self.sfreq
                e = ev['end_sample']/self.sfreq
                if e < self.current_pos_sec or s > self.current_pos_sec + self.display_window_seconds:
                    continue
                # Skip events not in filter when filter is active
                if filter_active and ('auto', idx) not in filtered_event_keys:
                    continue
                vs = max(s, self.current_pos_sec)
                ve = min(e, self.current_pos_sec + self.display_window_seconds)
                for ax in axes:
                    p = ax.axvspan(vs, ve, facecolor=event_color, alpha=0.35, edgecolor=text_color, linewidth=0.5)
                    mid = (vs + ve) / 2
                    # Show event start time in HH:MM:SS, number, and duration with icon
                    event_time = (self.recording_start_time + timedelta(seconds=ev['start_sample']/self.sfreq)).strftime('%H:%M:%S')
                    # Use current y-limits for positioning
                    y_top = ax.get_ylim()[1]
                    # Use softer text colors in dark mode
                    text_color_fg = '#E0E0E0' if self.dark_mode else 'white'
                    edge_color = '#B0B0B0' if self.dark_mode else 'white'
                    ax.text(mid, y_top * 0.9, f" {idx+1} | {event_time} | {ev['duration_sec']:.2f}s", 
                           color=text_color_fg, fontsize=8, ha='center', va='top', weight='bold',
                           bbox=dict(boxstyle='round,pad=0.3', facecolor=text_color, alpha=0.9, edgecolor=edge_color, linewidth=1))
                    if self.edit_mode_active:
                        p.set_picker(True)
                        self.region_patches.append((p, idx, 'auto'))
                        if (self.selected_event_idx == idx and self.active_event_type == 'auto'):
                            # Use softer gold in dark mode
                            selection_color = '#B8860B' if self.dark_mode else '#FFD700'  # Dark goldenrod in dark mode
                            p.set_edgecolor(selection_color)
                            p.set_linewidth(3)
        
        # Always show manual events with deep blue color (unless filtered out)
        for idx, ev in enumerate(self.manual_events):
            s = ev['start_sample']/self.sfreq
            e = ev['end_sample']/self.sfreq
            if e < self.current_pos_sec or s > self.current_pos_sec + self.display_window_seconds:
                continue
            # Skip events not in filter when filter is active
            if filter_active and ('manual', idx) not in filtered_event_keys:
                continue
            
            vs = max(s, self.current_pos_sec)
            ve = min(e, self.current_pos_sec + self.display_window_seconds)
            
            # Use softer blue in dark mode
            manual_color = '#5B7FC7' if self.dark_mode else '#4169E1'  # Muted royal blue in dark mode
            manual_dark = '#3A5A8F' if self.dark_mode else '#1E40AF'   # Darker blue for text
            
            for ax in axes:
                p = ax.axvspan(vs, ve, facecolor=manual_color, alpha=0.35, edgecolor=manual_dark, linewidth=0.5)
                mid = (vs + ve) / 2
                event_time = (self.recording_start_time + timedelta(seconds=ev['start_sample']/self.sfreq)).strftime('%H:%M:%S')
                # Use current y-limits for positioning
                y_top = ax.get_ylim()[1]
                # Use softer text colors in dark mode
                text_color_fg = '#E0E0E0' if self.dark_mode else 'white'
                edge_color = '#B0B0B0' if self.dark_mode else 'white'
                ax.text(mid, y_top * 0.9, f"{idx+1} | {event_time} | {ev['duration_sec']:.2f}s", 
                       color=text_color_fg, fontsize=8, ha='center', va='top', weight='bold',
                       bbox=dict(boxstyle='round,pad=0.3', facecolor=manual_dark, alpha=0.9, edgecolor=edge_color, linewidth=1))
                
                # In edit mode, make events pickable
                if self.edit_mode_active:
                    p.set_picker(True)
                    self.region_patches.append((p, idx, 'manual'))
                    
                    # Highlight selected event
                    if (self.selected_event_idx == idx and 
                        self.active_event_type == 'manual'):
                        # Use softer gold in dark mode
                        selection_color = '#B8860B' if self.dark_mode else '#FFD700'  # Dark goldenrod in dark mode
                        p.set_edgecolor(selection_color)
                        p.set_linewidth(3)

        # --- Imported recording comments ------------------------------------
        # Draw dashed vertical lines at each comment inside the current window
        # plus a small amber-boxed label near the top edge. Lines span both
        # channels by using each axes' own y-limits, so they never look
        # misplaced when the user zooms.
        comments_list = getattr(self, 'adicht_comments', []) or []
        if comments_list:
            win_start = self.current_pos_sec
            win_end = self.current_pos_sec + self.display_window_seconds
            cmt_line = '#F5B041' if self.dark_mode else '#D97706'
            cmt_text_bg = '#3B2E14' if self.dark_mode else '#FEF3C7'
            cmt_text_fg = '#FDE68A' if self.dark_mode else '#92400E'
            for c in comments_list:
                t = float(c.get('time_sec', 0.0) or 0.0)
                if t < win_start or t > win_end:
                    continue
                text = str(c.get('text', '') or '')
                # Truncate long texts so they don't overlap the waveforms.
                if len(text) > 40:
                    text = text[:37] + '…'
                for ax in axes:
                    ax.axvline(t, color=cmt_line, linestyle='--',
                               linewidth=1.0, alpha=0.85, zorder=5)
                # Anchor the label only on the top axis to avoid duplicates.
                y_top = ax1.get_ylim()[1]
                ax1.text(t, y_top * 0.98,
                         f"💬 {c.get('hhmmss', '')}  {text}".strip(),
                         rotation=90, ha='right', va='top',
                         fontsize=8, color=cmt_text_fg,
                         bbox=dict(boxstyle='round,pad=0.25',
                                   facecolor=cmt_text_bg, alpha=0.85,
                                   edgecolor=cmt_line, linewidth=0.6),
                         zorder=6, clip_on=True)

        # UNET overlays color based on model type - toned down in dark mode
        if self.model_type == "unet_only":
            # Use consistent emerald tones for UNET visuals - muted in dark mode
            unet_color = '#4A9B6E' if self.dark_mode else '#50C878'
            unet_edge_color = '#2D5F3F' if self.dark_mode else '#2EA043'
        elif self.model_type == "cwt_unet":
            unet_color = '#4A9B6E' if self.dark_mode else '#50C878'  # Muted emerald in dark mode
            unet_edge_color = '#2D5F3F' if self.dark_mode else '#2E8B57'
        else:  # Should not show UNET overlay for cwt_only, but keep default
            unet_color = '#4A9B6E' if self.dark_mode else '#50C878'  # Muted emerald in dark mode
            unet_edge_color = '#2D5F3F' if self.dark_mode else '#2E8B57'
        # Only draw UNET overlays in combined mode to avoid duplicate visuals in UNET-only
        if (self.model_type == "cwt_unet") and self.cb_show_unet.isChecked() and (self.unet_predictions is not None) and self.model_chosen:
            # Determine visible time window in seconds
            win_start_sec = self.current_pos_sec
            win_end_sec = self.current_pos_sec + self.display_window_seconds
            if self.cb_unet_post.isChecked() and self.unet_intervals:
                # Postprocessed intervals in 100 Hz samples
                for (s_idx, e_idx) in self.unet_intervals:
                    s_sec = s_idx / self.unet_fs
                    e_sec = (e_idx + 1) / self.unet_fs
                    if e_sec < win_start_sec or s_sec > win_end_sec:
                        continue
                    vs = max(s_sec, win_start_sec)
                    ve = min(e_sec, win_end_sec)
                    for ax in axes:
                        ax.axvspan(vs, ve, facecolor=unet_color, alpha=0.25, edgecolor=unet_edge_color, linewidth=0.5)
            else:
                # Raw thresholded intervals using configurable threshold
                threshold = getattr(self, 'unet_prediction_threshold', 0.5)
                raw_bin = (self.unet_predictions > threshold).astype(np.int32)
                raw_intervals = unet_find_continuous_intervals(raw_bin)
                for (s_idx, e_idx) in raw_intervals:
                    s_sec = s_idx / self.unet_fs
                    e_sec = (e_idx + 1) / self.unet_fs
                    if e_sec < win_start_sec or s_sec > win_end_sec:
                        continue
                    vs = max(s_sec, win_start_sec)
                    ve = min(e_sec, win_end_sec)
                    for ax in axes:
                        ax.axvspan(vs, ve, facecolor=unet_color, alpha=0.20,
                                   edgecolor=unet_edge_color, linewidth=0.5, linestyle='--')
        
        # Update plot labels and styling
        ct = self.recording_start_time + timedelta(seconds=self.current_pos_sec)
        et = ct + timedelta(seconds=self.display_window_seconds)
        position_time = f"Time: {ct.strftime('%H:%M:%S')} to {et.strftime('%H:%M:%S')}"
        
        # Add timestamps at regular intervals
        n_ticks = min(10, int(self.display_window_seconds))
        if n_ticks > 0:
            tick_positions = np.linspace(self.current_pos_sec, self.current_pos_sec + self.display_window_seconds, n_ticks+1)
            tick_labels = [(self.recording_start_time + timedelta(seconds=pos)).strftime('%H:%M:%S') for pos in tick_positions]
        else:
            tick_positions = [self.current_pos_sec, self.current_pos_sec + self.display_window_seconds]
            tick_labels = [(self.recording_start_time + timedelta(seconds=pos)).strftime('%H:%M:%S') for pos in tick_positions]
        
        # Format axes
        ax1.set_ylabel("Amplitude", color=text_color, fontsize=10)
        if ax2:
            ax2.set_ylabel("Amplitude", color=text_color, fontsize=10)
            ax2.set_xlabel("Time (HH:MM:SS)", color=text_color, fontsize=10)
        else:
            ax1.set_xlabel("Time (HH:MM:SS)", color=text_color, fontsize=10)
        
        # Set x-axis ticks and labels with diagonal rotation for space efficiency
        ax1.set_xticks(tick_positions)
        if ax2:
            ax2.set_xticks(tick_positions)
            ax1.set_xticklabels([])  # Hide x-axis labels on top plot
            ax2.set_xticklabels(tick_labels, ha='center', fontsize=9)
        else:
            ax1.set_xticklabels(tick_labels, ha='center', fontsize=9)
        
        # Ensure axes use the same color scheme (lighter in dark mode)
        if not self.dark_mode:
            ax1.tick_params(colors=text_color)
            if ax2:
                ax2.tick_params(colors=text_color)
            for ax in axes:
                for side in ('bottom','top','right','left'):
                    ax.spines[side].set_color(text_color)
        
        # Add position time as text in the upper right corner
        ax1.text(0.99, 0.98, position_time, transform=ax1.transAxes, ha='right', va='top', 
                 fontsize=9, color=text_color, bbox=dict(facecolor='none', alpha=0.5, edgecolor='none'))
        
        # Legend removed - cleaner UI
        
        # Set grid
        ax1.grid(True, linestyle='--', alpha=0.3, color=grid_color)
        if ax2:
            ax2.grid(True, linestyle='--', alpha=0.3, color=grid_color)
        
        # Remove plot legends; show channel names at top-left outside plots instead
        
        # Set title after creating subplots: show EDF filename (without extension)
        # Use softer gray in dark mode instead of bright white
        main_title_color = '#D0D0D0' if self.dark_mode else 'black'
        edf_title = self.edf_filename if isinstance(self.edf_filename, str) else 'EEG Recording'
        try:
            edf_title_no_ext = os.path.splitext(edf_title)[0]
        except Exception:
            edf_title_no_ext = edf_title
        self.fig.suptitle(edf_title_no_ext, fontsize=11, fontweight='bold', color=main_title_color, y=0.98)
        
        # Improve layout with better spacing
        self.fig.tight_layout(rect=[0, 0, 1, 0.95])  # Leave a bit more space for title
        
        # Force canvas update
        self.canvas.draw_idle()
        
        # Update status indicators and timeline
        self.update_status_indicators()
        self.update_timeline_minimap()
        
        # Update status indicators and timeline/minimap accent
        self.update_status_indicators()
        self.update_timeline_minimap()
        # Update slider position and range
        max_pos = max(0, int(self.total_samples/self.sfreq - self.display_window_seconds))
        self.slider.setRange(0, max_pos)
        self.slider.blockSignals(True)
        self.slider.setValue(int(self.current_pos_sec))
        self.slider.blockSignals(False)

        # Make slider smoother
        self.slider.setSingleStep(1)
        self.slider.setPageStep(max(1, int(self.display_window_seconds // 2)))
        self.slider.setTickInterval(max(1, int(self.display_window_seconds // 4)))

    def on_duration_filter_changed(self, state):
        """Handle changes to duration filter checkbox"""
        enabled = bool(state)
        self.duration_min_spin.setEnabled(enabled)
        self.duration_max_spin.setEnabled(enabled)
        self.update_filtered_events()
    
    def on_event_filter_toggled(self, checked):
        """Open/close the duration filter popup and toggle filter state"""
        if checked:
            # Show popup to set min/max durations
            dlg = QDialog(self)
            dlg.setWindowTitle("⏱ Event Duration Filter")
            lay = QVBoxLayout(dlg)
            try:
                if self.dark_mode:
                    self._apply_win_dark_titlebar(dlg.winId())
            except Exception:
                pass
            row = QHBoxLayout()
            row.addWidget(QLabel("Min Duration (s):"))
            min_spin = QDoubleSpinBox()
            min_spin.setRange(0.1, 60.0)
            min_spin.setValue(self.duration_min_spin.value())
            row.addWidget(min_spin)
            row2 = QHBoxLayout()
            row2.addWidget(QLabel("Max Duration (s):"))
            max_spin = QDoubleSpinBox()
            max_spin.setRange(0.1, 60.0)
            max_spin.setValue(self.duration_max_spin.value())
            row2.addWidget(max_spin)
            lay.addLayout(row)
            lay.addLayout(row2)
            btns = QHBoxLayout()
            apply_btn = QPushButton("Apply Filter")
            clear_btn = QPushButton("Clear Filter")
            btns.addWidget(apply_btn)
            btns.addWidget(clear_btn)
            lay.addLayout(btns)
            
            def apply_and_close():
                self.duration_min_spin.setValue(min_spin.value())
                self.duration_max_spin.setValue(max_spin.value())
                self.cb_duration_filter.setChecked(True)
                self.update_filtered_events()
                dlg.accept()
            def clear_and_close():
                self.cb_duration_filter.setChecked(False)
                self.update_filtered_events()
                self.btn_event_filter.setChecked(False)
                dlg.reject()
            apply_btn.clicked.connect(apply_and_close)
            clear_btn.clicked.connect(clear_and_close)
            dlg.exec_()
            # Keep button pressed while filter remains enabled
            if not self.cb_duration_filter.isChecked():
                self.btn_event_filter.setChecked(False)
        else:
            # Disable filter
            self.cb_duration_filter.setChecked(False)
            self.update_filtered_events()
        
    def update_filtered_events(self):
        """Update the filtered events list based on duration filter"""
        all_events = []
        
        # Combine all events with their type and index
        for i, ev in enumerate(self.manual_events):
            all_events.append({
                'event': ev,
                'type': 'manual',
                'index': i,
                'start_time': ev['start_sample'] / self.sfreq
            })
            
        for i, ev in enumerate(self.refined_swd):
            all_events.append({
                'event': ev,
                'type': 'auto', 
                'index': i,
                'start_time': ev['start_sample'] / self.sfreq
            })
        
        # Sort by start time
        all_events.sort(key=lambda x: x['start_time'])
        
        # Apply duration filter if enabled
        if self.cb_duration_filter.isChecked():
            min_dur = self.duration_min_spin.value()
            max_dur = self.duration_max_spin.value()
            self.filtered_events = [
                ev for ev in all_events 
                if min_dur <= ev['event']['duration_sec'] <= max_dur
            ]
        else:
            self.filtered_events = all_events
        
        # Refresh plot and minimap to show only filtered events when filter is active
        if hasattr(self, 'canvas') and self.canvas is not None:
            self.update_plot()
        if hasattr(self, 'timeline_minimap') and self.timeline_minimap is not None:
            self.update_timeline_minimap()

    def get_current_filtered_event_index(self):
        """Get the index of the currently selected event in the filtered list"""
        if self.selected_event_idx is None or self.active_event_type is None:
            return None
            
        for i, filtered_ev in enumerate(self.filtered_events):
            if (filtered_ev['type'] == self.active_event_type and 
                filtered_ev['index'] == self.selected_event_idx):
                return i
        return None

    def on_slider_changed(self, value):
        """Handle position slider changes with smooth update."""
        self.current_pos_sec = value
        self.update_plot()
        
    def on_window_size_changed(self, value):
        """Handle window size spinbox changes."""
        self.display_window_seconds = value
        self.update_plot()
        
    # Removed window prev/next buttons from UI (keyboard shortcuts still supported)
        
    def on_reanalyze_clicked(self):
        """Reanalyze SWD events with current parameters"""
        # Get current parameters
        thr = self.th_spin.value()
        gap = self.gap_spin.value()
        power_percentile = self.power_percentile_spin.value()
        
        # Update status
        self.progress_bar.show()
        QApplication.processEvents()
        
        # Make sure we have token predictions
        if not hasattr(self, 'token_predictions') or self.token_predictions is None:
            QMessageBox.warning(self, "Reanalysis Error", "No token predictions available for reanalysis")
            return
            
        # Reanalyze merged regions using the user-configurable overlap flag.
        use_overlap_agreement = bool(getattr(self, 'require_overlap_agreement',
                                             self.model_type in ("cwt_only", "cwt_unet")))
        self.merged_swd = merge_adjacent_swds(self.token_predictions, 
                                             threshold=thr, max_gap_intervals=gap, require_overlap_agreement=use_overlap_agreement)
        
        # Convert raw merged regions to sample indices
        step = int((self.interval_length - self.overlap_length) * self.sfreq)
        for region in self.merged_swd:
            start_idx = region['start_idx']
            end_idx = region['end_idx']
            region['start_sample'] = start_idx * step
            region['end_sample'] = (end_idx + 1) * step + int(self.interval_length * self.sfreq) - step
            region['duration_sec'] = (region['end_sample'] - region['start_sample']) / self.sfreq
            region['is_auto'] = True
        
        # Refine borders
        try:
            self.refined_swd = refine_swd_borders(
                self.eeg0, self.sfreq, self.merged_swd, power_percentile
            )
        except Exception as e:
            print(f"Warning: Border refinement failed during reanalysis: {e}")
            print("Using unrefined events instead")
            # Fall back to unrefined events
            self.refined_swd = []
            for region in self.merged_swd:
                refined_event = region.copy()
                refined_event['duration_sec'] = (region['end_sample'] - region['start_sample']) / self.sfreq
                if refined_event['duration_sec'] >= 1.0:
                    self.refined_swd.append(refined_event)
        
        # Filter by minimum duration
        self.refined_swd = [r for r in self.refined_swd if r['duration_sec'] >= 1.0]
        
        # Final merge step: merge events closer than 1 second apart
        try:
            self.refined_swd = merge_overlapping_events(self.refined_swd, self.sfreq, min_overlap_sec=-1.0)
        except Exception as e:
            print(f"Warning: Event merging failed during reanalysis: {e}")
            # Continue with unmerged events
        
        # Reset selection if it was an auto event
        if self.active_event_type == 'auto':
            self.selected_event_idx = None
            self.active_event_type = None
            self.current_edit_event = None
            self.event_info_lbl.setText("")
        
        # Update displays
        self.progress_bar.hide()
        # Keep a copy before optional UNET refinement
        self.refined_swd_base = list(self.refined_swd)
        try:
            self.apply_unet_refinement_if_enabled()
        except Exception as e:
            print(f"Warning: UNET refinement failed during reanalysis: {e}")
            # Continue without UNET refinement
        self.update_combined_events()
        self.update_filtered_events()
        self.update_plot()

    def on_edit_mode_changed(self, state):
        """Handle changes to edit mode state."""
        self.edit_mode_active = bool(state)
        
        # Update button states
        self.btn_add.setEnabled(self.edit_mode_active)
        self.btn_del.setEnabled(self.edit_mode_active and self.selected_event_idx is not None)
        self.btn_convert.setEnabled(self.edit_mode_active and len(self.refined_swd) > 0)
        self.btn_psd.setEnabled(self.edit_mode_active and self.selected_event_idx is not None)
        
        # In edit mode, always edit both auto and manual events
        if self.edit_mode_active:
            self.active_event_type = 'both'
        else:
            # Reset selection when exiting edit mode
            self.selected_event_idx = None
            self.active_event_type = None
            self.event_info_lbl.setText("")
        
        # Update combined events and plot
        self.update_combined_events()
        self.update_plot()

    def set_edit_mode_enabled(self, enabled):
        """Compatibility wrapper used by Settings: enable/disable edit mode."""
        try:
            self.cb_edit.setChecked(bool(enabled))
        except Exception:
            pass
        self.on_edit_mode_changed(bool(enabled))
        
    def on_mouse_release(self, event):
        """Handle mouse release to end dragging and finalize border editing"""
        if self.dragging_start or self.dragging_end:
            self.dragging_start = False
            self.dragging_end = False
            # Now merge overlapping events after editing
            if self.active_event_type == 'manual':
                self.manual_events = merge_overlapping_events(self.manual_events, self.sfreq)
            elif self.active_event_type == 'auto':
                self.refined_swd = merge_overlapping_events(self.refined_swd, self.sfreq)
            self.update_combined_events()
        self.update_event_info()
        self.update_plot()
        
    def on_mouse_move(self, event):
        """Handle mouse move for resizing events"""
        if not (self.dragging_start or self.dragging_end) or event.xdata is None:
            return
        # Get new time position
        new_pos = max(0, min(self.total_samples/self.sfreq, event.xdata))
        new_sample = int(new_pos * self.sfreq)
        updated = False
        if self.dragging_start:
            # Prevent start from going beyond end
            if new_sample < self.current_edit_event['end_sample'] - int(0.5 * self.sfreq):
                self.current_edit_event['start_sample'] = new_sample
                updated = True
        elif self.dragging_end:
            # Prevent end from going before start
            if new_sample > self.current_edit_event['start_sample'] + int(0.5 * self.sfreq):
                self.current_edit_event['end_sample'] = new_sample
                updated = True
        if updated:
            # Update duration
            dur = (self.current_edit_event['end_sample'] - self.current_edit_event['start_sample']) / self.sfreq
            self.current_edit_event['duration_sec'] = dur
            # Do not merge or finalize here, just update plot for smooth feedback
            self.update_event_info()
            self.update_plot()

    def update_event_info(self):
        """Update the event information display."""
        if self.selected_event_idx is None:
            self.event_info_lbl.setText("")
            self.btn_del.setEnabled(False)
            self.btn_psd.setEnabled(False)
            return
            
        self.btn_del.setEnabled(True)
        self.btn_psd.setEnabled(True)  # Enable PSD analysis for any selected event
        
        # Get the right event based on type
        events = self.manual_events if self.active_event_type == 'manual' else self.refined_swd
        if 0 <= self.selected_event_idx < len(events):
            ev = events[self.selected_event_idx]
            start_time = self.recording_start_time + timedelta(seconds=ev['start_sample']/self.sfreq)
            end_time = self.recording_start_time + timedelta(seconds=ev['end_sample']/self.sfreq)
            
            # Build info string with basic properties
            info = [
                f"Selected: {self.active_event_type.capitalize()} #{self.selected_event_idx+1}",
                f"Start: {start_time.strftime('%H:%M:%S')}",
                f"End: {end_time.strftime('%H:%M:%S')}",
                f"Duration: {ev['duration_sec']:.2f}s"
            ]
            
            # Add additional properties if available
            if 'amplitude_stats' in ev:
                stats = ev['amplitude_stats']
                info.extend([
                    f"Mean Amp: {stats['mean_amplitude']:.3f}",
                    f"Max Amp: {stats['max_amplitude']:.3f}",
                    f"P-P Amp: {stats['peak_to_peak']:.3f}"
                ])
                
            if 'dominant_freq' in ev:
                info.append(f"Dominant Freq: {ev['dominant_freq']:.1f} Hz")
                
            if 'power_ratio' in ev:
                info.append(f"Power Ratio: {ev['power_ratio']:.2f}")
                
            if 'confidence' in ev:
                info.append(f"Confidence: {ev['confidence']:.2f}")
                
            self.event_info_lbl.setText(" | ".join(info))

    def on_export_clicked(self):
        """Export all events to Excel"""
        has_current_events = bool(self.manual_events or self.refined_swd)
        has_saved_animals = bool(getattr(self, 'animals_data', {}))
        if not has_current_events and not has_saved_animals:
            self.show_toast("No events to export", "warning")
            return
        
        # Open Export Settings popup
        dlg = QDialog(self)
        dlg.setWindowTitle("📤 Export Settings")
        dlg.setModal(True)
        dlg.setMinimumWidth(400)
        try:
            if self.dark_mode:
                self._apply_win_dark_titlebar(dlg.winId())
        except Exception:
            pass
        lay = QVBoxLayout(dlg)
        lay.setSpacing(6)
        lay.setContentsMargins(12, 12, 12, 12)
        
        # Compact field styles - explicit colors for dark mode
        if self.dark_mode:
            dlg.setStyleSheet("""
                QDialog { background-color: #0D1117; }
                QLabel { margin: 0px; padding: 0px; background: transparent; color: #E6EDF3; font-weight: 500; }
                QLineEdit, QComboBox, QSpinBox {
                    padding: 3px 6px; margin: 0px; min-height: 22px; border-radius: 4px;
                    border: 1px solid #30363D; background: #161B22; color: #E6EDF3;
                }
                QLineEdit:focus, QComboBox:focus, QSpinBox:focus { border: 1px solid #58A6FF; }
                QGroupBox { 
                    border: 1px solid #30363D; border-radius: 4px; margin-top: 6px; padding-top: 6px; color: #E6EDF3;
                }
                QCheckBox { color: #E6EDF3; }
                QCheckBox::indicator { border: 1px solid #30363D; background: #161B22; }
                QCheckBox::indicator:checked { background: #238636; border-color: #238636; }
                QPushButton { 
                    background-color: #21262D; color: #E6EDF3; border: 1px solid #30363D; 
                    border-radius: 4px; padding: 6px 12px; 
                }
                QPushButton:hover { background-color: #30363D; }
                QPushButton:pressed { background-color: #484F58; }
                QFrame { color: #30363D; }
                QComboBox::drop-down { border: none; }
                QComboBox QAbstractItemView { background: #161B22; color: #E6EDF3; selection-background-color: #1C3A1F; }
            """)
        else:
            dlg.setStyleSheet("""
                QLabel { margin: 0px; padding: 0px; background: transparent; color: palette(text); font-weight: 500; }
                QLineEdit, QComboBox, QSpinBox {
                    padding: 3px 6px; margin: 0px; min-height: 22px; border-radius: 4px;
                    border: 1px solid palette(mid); background: palette(base);
                }
                QLineEdit:focus, QComboBox:focus, QSpinBox:focus { border: 1px solid palette(highlight); }
                QGroupBox { 
                    border: 1px solid palette(mid); border-radius: 4px; margin-top: 6px; padding-top: 6px; 
                }
            """)

        # --- Recording Comments status line ---------------------------------
        # Imports are done from the File menu ("Import Comments…") so this
        # dialog stays focused on export. The small label just reminds the
        # user whether comments are available inside the time dropdowns
        # below.
        lbl_comments_status = QLabel()
        lbl_comments_status.setStyleSheet("font-size: 9pt;")
        lay.addWidget(lbl_comments_status)

        def _update_comments_status():
            n = len(getattr(self, 'adicht_comments', []) or [])
            if n > 0:
                lbl_comments_status.setText(
                    f"💬  {n} recording comment(s) loaded — open any time dropdown to pick one."
                )
                lbl_comments_status.setStyleSheet("color: #3FB950; font-size: 9pt;")
            else:
                lbl_comments_status.setText(
                    "💬  No recording comments loaded. Use File → Import Comments… to load them."
                )
                lbl_comments_status.setStyleSheet("color: gray; font-size: 9pt;")

        # 1. Main Options
        cb_analyzed = QCheckBox("Export Analyzed Excel")
        cb_analyzed.setChecked(getattr(self, 'export_analyzed', True))
        lay.addWidget(cb_analyzed)

        # Export All Saved Animals option (only show if there are saved animals)
        cb_export_all = QCheckBox("Export All Saved Animals Combined")
        saved_count = len(getattr(self, 'animals_data', {}))
        # Auto-check when saved animals exist but current animal has no events
        cb_export_all.setChecked(saved_count > 0 and not has_current_events)
        if saved_count > 0:
            cb_export_all.setText(f"Export All Saved Animals Combined ({saved_count} saved)")
            cb_export_all.setEnabled(True)
        else:
            cb_export_all.setText("Export All Saved Animals (none saved)")
            cb_export_all.setEnabled(False)
        lay.addWidget(cb_export_all)
        
        # Grid for general settings
        grid_gen = QGridLayout()
        grid_gen.setVerticalSpacing(6)
        grid_gen.setHorizontalSpacing(10)
        
        grid_gen.addWidget(QLabel("Interval (min):"), 0, 0)
        interval_edit = QLineEdit(str(getattr(self, 'export_interval_minutes', 20)))
        interval_edit.setPlaceholderText("e.g. 20")
        grid_gen.addWidget(interval_edit, 0, 1)

        grid_gen.addWidget(QLabel("Recording Type:"), 1, 0)
        export_type_combo = QComboBox(); export_type_combo.addItems(["basal", "injection"]) 
        export_type_combo.setCurrentText(getattr(self, 'export_recording_type', "basal"))
        grid_gen.addWidget(export_type_combo, 1, 1)
        
        grid_gen.addWidget(QLabel("Animal ID:"), 2, 0)
        animal_edit = QLineEdit(getattr(self, 'export_animal_id', "Animal_1"))
        grid_gen.addWidget(animal_edit, 2, 1)
        
        lay.addLayout(grid_gen)
        
        # Advanced per-animal settings button (only when export all is active)
        btn_advanced = QPushButton("Advanced Settings...")
        btn_advanced.setCursor(Qt.PointingHandCursor)
        btn_advanced.setToolTip("Configure per-animal analysis windows and custom basal periods")
        btn_advanced.setVisible(False)
        lay.addWidget(btn_advanced)

        def _open_advanced_dialog():
            rec_type = export_type_combo.currentText()
            gs = {
                'basal_start': edt_b_start.text().strip() if hasattr(edt_b_start, 'text') else getattr(self, 'export_basal_start', '09:00'),
                'basal_end': edt_b_end.text().strip() if hasattr(edt_b_end, 'text') else getattr(self, 'export_basal_end', '12:00'),
                'injection_time': edt_inj_time.text().strip() if hasattr(edt_inj_time, 'text') else getattr(self, 'export_injection_time', '09:00'),
                'injection_end': edt_inj_end.text().strip() if hasattr(edt_inj_end, 'text') else getattr(self, 'export_injection_end', '12:00'),
                'interval_minutes': int(interval_edit.text().strip() or 20),
                'basal_duration': int(edt_pre_dur.text().strip() or 20),
            }
            adv_dlg = AdvancedExportDialog(
                self.animals_data, gs, rec_type, parent=dlg,
                dark_mode=self.dark_mode,
                existing_settings=self.per_animal_export_settings,
                get_comments=lambda: getattr(self, 'adicht_comments', []) or [],
            )
            try:
                if self.dark_mode:
                    self._apply_win_dark_titlebar(adv_dlg.winId())
            except Exception:
                pass
            if adv_dlg.exec_() == QDialog.Accepted:
                self.per_animal_export_settings = adv_dlg.get_settings()

        btn_advanced.clicked.connect(_open_advanced_dialog)

        # Update visibility when export all is toggled
        def update_export_all_visibility():
            export_all = cb_export_all.isChecked()
            animal_edit.setEnabled(not export_all)
            btn_advanced.setVisible(export_all and saved_count > 0)
            
        cb_export_all.stateChanged.connect(update_export_all_visibility)
        update_export_all_visibility()
        
        # Separator line
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setFrameShadow(QFrame.Sunken)
        lay.addWidget(line)

        # 2. Dynamic Settings Area
        params_widget = QWidget()
        params_lay = QGridLayout(params_widget)
        params_lay.setContentsMargins(0,0,0,0)
        params_lay.setVerticalSpacing(6)
        params_lay.setHorizontalSpacing(10)
        lay.addWidget(params_widget)

        # Fields – HH:MM fields are editable combo-boxes so their dropdown
        # also serves as the recording-comment picker (matches the style of
        # the 'Recording Type' combo above).
        get_cmts = lambda: getattr(self, 'adicht_comments', []) or []

        # Basal Mode Fields
        lbl_b_start = QLabel("Basal Start (HH:MM):")
        edt_b_start = CommentTimeCombo(getattr(self, 'export_basal_start', "09:00"), get_cmts)
        lbl_b_end = QLabel("Basal End (HH:MM):")
        edt_b_end = CommentTimeCombo(getattr(self, 'export_basal_end', "12:00"), get_cmts)

        # Injection Mode Fields
        lbl_inj_time = QLabel("Injection Time (HH:MM):")
        edt_inj_time = CommentTimeCombo(getattr(self, 'export_injection_time', "09:00"), get_cmts)
        lbl_pre_dur = QLabel("Basal Duration (min):")
        edt_pre_dur = QLineEdit(str(getattr(self, 'export_basal_duration', 20)))
        edt_pre_dur.setPlaceholderText("e.g. 20")
        lbl_inj_end = QLabel("End Time (HH:MM):")
        edt_inj_end = CommentTimeCombo(getattr(self, 'export_injection_end', "12:00"), get_cmts)

        # Layout: col 0 = label, col 1 = input (editable combo or line edit)
        # Basal rows
        params_lay.addWidget(lbl_b_start, 0, 0); params_lay.addWidget(edt_b_start, 0, 1)
        params_lay.addWidget(lbl_b_end, 1, 0); params_lay.addWidget(edt_b_end, 1, 1)

        # Injection rows
        params_lay.addWidget(lbl_inj_time, 2, 0); params_lay.addWidget(edt_inj_time, 2, 1)
        params_lay.addWidget(lbl_pre_dur, 3, 0); params_lay.addWidget(edt_pre_dur, 3, 1)
        params_lay.addWidget(lbl_inj_end, 4, 0); params_lay.addWidget(edt_inj_end, 4, 1)

        # Reflect any already-loaded comments in the status label.
        _update_comments_status()

        def update_visibility():
            is_analyzed = cb_analyzed.isChecked()
            is_basal = export_type_combo.currentText() == "basal"
            
            # General controls
            interval_edit.setEnabled(is_analyzed)
            export_type_combo.setEnabled(is_analyzed)
            animal_edit.setEnabled(is_analyzed)
            params_widget.setEnabled(is_analyzed)
            
            # Basal specific
            lbl_b_start.setVisible(is_basal); edt_b_start.setVisible(is_basal)
            lbl_b_end.setVisible(is_basal); edt_b_end.setVisible(is_basal)

            # Injection specific
            is_inj = not is_basal
            lbl_inj_time.setVisible(is_inj); edt_inj_time.setVisible(is_inj)
            lbl_pre_dur.setVisible(is_inj); edt_pre_dur.setVisible(is_inj)
            lbl_inj_end.setVisible(is_inj); edt_inj_end.setVisible(is_inj)
            
            # Adjust layout size with a slight delay to allow layout to reflow
            # But immediately adjust to minimum size hint
            dlg.adjustSize()
            
            # Force resize to compact if basal (remove excess height from hidden widgets)
            if is_basal:
                dlg.resize(dlg.minimumSizeHint())

        cb_analyzed.stateChanged.connect(update_visibility)
        export_type_combo.currentTextChanged.connect(update_visibility)
        
        # Initial state
        update_visibility()
        
        # Buttons
        buttons = QHBoxLayout()
        buttons.setContentsMargins(0, 6, 0, 0)
        btn_ok = QPushButton("Export")
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setDefault(True)
        btn_cancel = QPushButton("Cancel")
        btn_cancel.setCursor(Qt.PointingHandCursor)
        buttons.addStretch(); buttons.addWidget(btn_ok); buttons.addWidget(btn_cancel)
        lay.addLayout(buttons)
        
        def on_cancel():
            dlg.reject()
        btn_cancel.clicked.connect(on_cancel)
        
        def on_ok():
            self.export_recording_type = export_type_combo.currentText()
            self.export_basal_start = edt_b_start.text().strip()
            self.export_basal_end = edt_b_end.text().strip()
            self.export_injection_time = edt_inj_time.text().strip()
            self.export_basal_duration = int(edt_pre_dur.text().strip() or 20)
            self.export_injection_end = edt_inj_end.text().strip()
            self.export_animal_id = animal_edit.text().strip() or "Animal_1"
            self.export_analyzed = cb_analyzed.isChecked()
            self.export_interval_minutes = int(interval_edit.text().strip() or 20)
            self.export_all_animals = cb_export_all.isChecked()
            dlg.accept()
        btn_ok.clicked.connect(on_ok)
        
        if dlg.exec_() != QDialog.Accepted:
            return
        
        # Use export settings
        analyzed = getattr(self, 'export_analyzed', True)
        export_all = getattr(self, 'export_all_animals', False)
        
        # Smart default filename
        if export_all and len(self.animals_data) > 1:
            default_export_name = os.path.splitext(getattr(self, 'edf_filename', 'export'))[0] + ".xlsx"
        else:
            aname = (getattr(self, 'export_animal_id', None) or
                     getattr(self, 'current_animal_name', None) or 'export')
            default_export_name = aname + ".xlsx"
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Export Events", default_export_name, "Excel Files (*.xlsx);;All Files (*)")
        if not filename:
            return
        if not filename.lower().endswith('.xlsx'):
            filename += '.xlsx'
        
        try:
            if export_all and self.animals_data:
                # Export all saved animals combined
                self._export_all_animals_excel(filename, analyzed)
            else:
                # Export current animal only
                # Combine and sort events
                all_events = self.manual_events + self.refined_swd
                all_events.sort(key=lambda x: x['start_sample'])
                
                if not analyzed:
                    export_to_excel(all_events, self.sfreq, self.recording_start_time, filename)
                else:
                    # Validate settings; if missing, open Settings on Export tab
                    if not hasattr(self, 'export_recording_type'):
                        self.on_settings_clicked()
                        self.show_toast("Set Export settings in Settings → Analysis → Export Settings", "info")
                        return
                    self._export_analyzed_excel(filename, all_events)
                self.show_toast(f"✓ Exported {len(all_events)} events successfully!", "success")
        except Exception as e:
            self.show_toast(f"Export failed: {str(e)}", "error")
            import traceback
            traceback.print_exc()

    def _events_to_dataframe_for_analysis(self, events):
        """Convert events to list of [Start, End, Duration] using absolute times."""
        rows = []
        for ev in events:
            start_sec = ev['start_sample'] / self.sfreq
            end_sec = ev['end_sample'] / self.sfreq
            start_dt = self.recording_start_time + timedelta(seconds=start_sec)
            end_dt = self.recording_start_time + timedelta(seconds=end_sec)
            duration = max(0.0, end_sec - start_sec)
            rows.append([start_dt.time(), end_dt.time(), duration])
        df = pd.DataFrame(rows, columns=[0, 1, 2])
        return df

    def _export_analyzed_excel(self, save_path, events):
        """Generate analyzer-style Excel (single animal) using analyzer helpers."""
        # Build events DataFrame
        df = self._events_to_dataframe_for_analysis(events)
        # Ensure settings
        rec_type = getattr(self, 'export_recording_type', 'basal')
        animal_id = getattr(self, 'export_animal_id', os.path.splitext(getattr(self, 'edf_filename', 'Animal_1'))[0])
        interval_minutes = getattr(self, 'export_interval_minutes', 20)
        
        if rec_type == 'basal':
            result, seizures_divided, total_stats = self._analyze_seizures_basal(
                df, self.export_basal_start, self.export_basal_end, interval_minutes
            )
        else:
            basal_duration = getattr(self, 'export_basal_duration', 20)
            # Default end time: injection + 3h if not specified (or parse failed)
            inj_end = getattr(self, 'export_injection_end', '')
            
            result, seizures_divided, total_stats = self._analyze_seizures_injection(
                df, self.export_injection_time, interval_minutes, basal_duration, inj_end
            )
        # Create workbook
        workbook = Workbook()
        workbook.remove(workbook.active)
        # Combined sheet
        processed_results = {animal_id: result}
        title = os.path.splitext(os.path.basename(save_path))[0]
        self._create_combined_sheet(workbook, processed_results, title)
        # Animal sheet
        self._create_animal_sheet(workbook, animal_id, seizures_divided, rec_type, animal_id[:31])
        # Details sheet
        ch = tuple(self.channel_names[:2]) if not self.single_channel_view else (self.channel_names[0], None)
        single_info = {animal_id: {
            'channels': ch,
            'model_type': getattr(self, 'model_type', None),
            'events': events,
            'auto_event_count': len(self.refined_swd),
            'manual_event_count': len(self.manual_events),
        }}
        gs = {
            'interval_minutes': interval_minutes,
            'basal_start': getattr(self, 'export_basal_start', ''),
            'basal_end': getattr(self, 'export_basal_end', ''),
            'injection_time': getattr(self, 'export_injection_time', ''),
            'injection_end': getattr(self, 'export_injection_end', ''),
            'basal_duration': getattr(self, 'export_basal_duration', 20),
        }
        self._create_details_sheet(workbook, single_info, rec_type, gs)
        # Save
        workbook.save(save_path)

    def _export_all_animals_excel(self, save_path, analyzed=True):
        """Export all saved animals to a combined Excel file.
        
        If analyzed=True: Creates a combined Analysis sheet with all animals,
                         plus individual sheets for each animal's events.
        If analyzed=False: Creates separate sheets for each animal with raw events.
        """
        if not self.animals_data:
            self.show_toast("No saved animals to export", "warning")
            return
            
        rec_type = getattr(self, 'export_recording_type', 'basal')
        interval_minutes = getattr(self, 'export_interval_minutes', 20)
        
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        if analyzed:
            # Create combined Analysis sheet like EDF_to_Excel_SWD_Analyzer.py
            processed_results = {}
            all_seizures_divided = {}
            
            for animal_name, animal_data in self.animals_data.items():
                events = animal_data.get('events', [])
                sfreq = animal_data.get('sfreq', self.sfreq)
                recording_start = animal_data.get('recording_start', self.recording_start_time)
                
                # Convert events to DataFrame
                rows = []
                for ev in events:
                    start_sec = ev['start_sample'] / sfreq
                    end_sec = ev['end_sample'] / sfreq
                    start_dt = recording_start + timedelta(seconds=start_sec)
                    end_dt = recording_start + timedelta(seconds=end_sec)
                    duration = max(0.0, end_sec - start_sec)
                    rows.append([start_dt.time(), end_dt.time(), duration])
                df = pd.DataFrame(rows, columns=[0, 1, 2])
                
                # Per-animal settings override
                pa = self.per_animal_export_settings.get(animal_name, {})
                use_global = pa.get('use_global', True)

                if rec_type == 'basal':
                    b_start = self.export_basal_start if use_global else pa.get('basal_start', self.export_basal_start)
                    b_end = self.export_basal_end if use_global else pa.get('basal_end', self.export_basal_end)
                    iv_min = interval_minutes if use_global else pa.get('interval_minutes', interval_minutes)
                    result, seizures_divided, _ = self._analyze_seizures_basal(df, b_start, b_end, iv_min)
                else:
                    if use_global:
                        basal_dur = getattr(self, 'export_basal_duration', 20)
                        inj_end = getattr(self, 'export_injection_end', '')
                        result, seizures_divided, _ = self._analyze_seizures_injection(
                            df, self.export_injection_time, interval_minutes, basal_dur, inj_end
                        )
                    else:
                        cbs = pa.get('custom_basal_start', '')
                        cbe = pa.get('custom_basal_end', '')
                        fallback_dur = getattr(self, 'export_basal_duration', 20) if not (cbs and cbe) else 0
                        result, seizures_divided, _ = self._analyze_seizures_injection(
                            df,
                            pa.get('injection_time', self.export_injection_time),
                            pa.get('interval_minutes', interval_minutes),
                            fallback_dur,
                            pa.get('injection_end', getattr(self, 'export_injection_end', '')),
                            custom_basal_start=cbs,
                            custom_basal_end=cbe
                        )

                processed_results[animal_name] = result
                all_seizures_divided[animal_name] = seizures_divided
            
            # Create combined Analysis sheet
            title = os.path.splitext(os.path.basename(save_path))[0]
            self._create_combined_sheet(workbook, processed_results, title)
            
            # Create individual animal sheets
            for animal_name, seizures_divided in all_seizures_divided.items():
                sheet_name = animal_name[:31]
                self._create_animal_sheet(workbook, animal_name, seizures_divided, rec_type, sheet_name)

            # Analysis Details sheet
            gs = {
                'interval_minutes': interval_minutes,
                'basal_start': getattr(self, 'export_basal_start', ''),
                'basal_end': getattr(self, 'export_basal_end', ''),
                'injection_time': getattr(self, 'export_injection_time', ''),
                'injection_end': getattr(self, 'export_injection_end', ''),
                'basal_duration': getattr(self, 'export_basal_duration', 20),
            }
            self._create_details_sheet(workbook, self.animals_data, rec_type, gs)

            self.show_toast(f"✓ Exported {len(self.animals_data)} animals (analyzed) successfully!", "success")
        else:
            # Raw export: each animal gets its own sheet with raw events
            for animal_name, animal_data in self.animals_data.items():
                events = animal_data.get('events', [])
                sfreq = animal_data.get('sfreq', self.sfreq)
                recording_start = animal_data.get('recording_start', self.recording_start_time)
                
                # Create sheet for this animal
                sheet_name = animal_name[:31]
                sheet = workbook.create_sheet(title=sheet_name)
                
                # Header
                headers = ['Start Time', 'End Time', 'Duration']
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col)
                    cell.value = header
                    cell.font = Font(bold=True)
                    
                # Data rows
                for row_idx, ev in enumerate(events, 2):
                    st = recording_start + timedelta(seconds=ev['start_sample'] / sfreq)
                    et = recording_start + timedelta(seconds=ev['end_sample'] / sfreq)
                    dur = (ev['end_sample'] - ev['start_sample']) / sfreq
                    
                    # Format with comma decimal separator
                    st_ms = f"{int(st.microsecond // 1000):03d}"
                    et_ms = f"{int(et.microsecond // 1000):03d}"
                    
                    sheet.cell(row=row_idx, column=1).value = f"{st:%H:%M:%S},{st_ms}"
                    sheet.cell(row=row_idx, column=2).value = f"{et:%H:%M:%S},{et_ms}"
                    sheet.cell(row=row_idx, column=3).value = f"{dur:.2f}".replace('.', ',')
                    
                    # Set as text to prevent Excel from reformatting
                    for col in range(1, 4):
                        sheet.cell(row=row_idx, column=col).number_format = '@'
                        
                # Set column widths
                sheet.column_dimensions['A'].width = 18
                sheet.column_dimensions['B'].width = 18
                sheet.column_dimensions['C'].width = 12
                
            self.show_toast(f"✓ Exported {len(self.animals_data)} animals (raw) successfully!", "success")
        
        # Save workbook
        workbook.save(save_path)

    # ==== Analysis Details sheet ====
    @staticmethod
    def _hhmm_subtract(time_str, minutes):
        """Subtract minutes from an HH:MM string and return HH:MM."""
        try:
            parts = time_str.split(':')
            h, m = int(parts[0]), int(parts[1])
            total = h * 60 + m - minutes
            if total < 0:
                total += 24 * 60
            return f"{total // 60:02d}:{total % 60:02d}"
        except Exception:
            return time_str

    def _create_details_sheet(self, workbook, animals_info, rec_type, global_settings):
        """Add an 'Analysis Details' sheet summarizing all export parameters per animal."""
        sheet = workbook.create_sheet(title="Analysis Details")

        # Fills matching the combined Analysis sheet
        title_fill = PatternFill(start_color="dfcd7f", end_color="dfcd7f", fill_type="solid")
        section_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF")
        animal_id_fill = PatternFill(start_color="FFE800", end_color="FFE800", fill_type="solid")
        interval_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        basal_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
        label_font = Font(bold=True)

        model_names = {"cwt_only": "CWT", "unet_only": "UNET", "cwt_unet": "Combined (CWT+UNET)"}
        per_animal_settings = getattr(self, 'per_animal_export_settings', {})

        # -- File Info (compact key-value pairs) --
        row = 1
        info_items = [
            ("EDF Filename", getattr(self, 'edf_filename', 'N/A')),
            ("Recording Start", str(getattr(self, 'recording_start_time', 'N/A'))),
            ("Sampling Rate (Hz)", str(getattr(self, 'sfreq', 'N/A'))),
            ("Recording Type", rec_type),
            ("Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        for label, value in info_items:
            sheet.cell(row=row, column=1, value=label).font = label_font
            sheet.cell(row=row, column=2, value=str(value))
            row += 1

        # -- Per-Animal Details table --
        row += 1
        if rec_type == 'basal':
            headers = ["Animal", "Model", "Channels", "Start Time", "End Time",
                       "Interval (min)", "Total Events", "Auto Events", "Manual Events"]
        else:
            headers = ["Animal", "Model", "Channels", "Injection Time", "End Time",
                       "Interval (min)", "Basal Start", "Basal End",
                       "Total Events", "Auto Events", "Manual Events"]

        # Section title row
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        cell = sheet.cell(row=row, column=1, value="Per-Animal Details")
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal='center')
        row += 1

        # Header row
        header_row = row
        for c, h in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            if h in ("Animal",):
                cell.fill = animal_id_fill
            elif h in ("Basal Start", "Basal End"):
                cell.fill = basal_fill
            else:
                cell.fill = interval_fill
        row += 1

        # Track max content width per column (start with header widths)
        col_widths = [len(h) for h in headers]

        # Data rows
        for aname, ainfo in animals_info.items():
            pa = per_animal_settings.get(aname, {})
            use_global = pa.get('use_global', True)

            ch = ainfo.get('channels', ('?', '?'))
            ch_str = ch[0] if ch[1] is None else f"{ch[0]} + {ch[1]}"
            mtype = ainfo.get('model_type', None)
            model_str = model_names.get(mtype, mtype or 'N/A')
            total = len(ainfo.get('events', []))
            auto_cnt = ainfo.get('auto_event_count', sum(1 for e in ainfo.get('events', []) if e.get('is_auto', False)))
            manual_cnt = ainfo.get('manual_event_count', total - auto_cnt)

            if rec_type == 'basal':
                bs = global_settings.get('basal_start', '') if use_global else pa.get('basal_start', '')
                be = global_settings.get('basal_end', '') if use_global else pa.get('basal_end', '')
                iv = global_settings.get('interval_minutes', 20) if use_global else pa.get('interval_minutes', 20)
                vals = [aname, model_str, ch_str, bs, be, iv, total, auto_cnt, manual_cnt]
            else:
                inj_t = global_settings.get('injection_time', '') if use_global else pa.get('injection_time', '')
                inj_end = global_settings.get('injection_end', '') if use_global else pa.get('injection_end', '')
                iv = global_settings.get('interval_minutes', 20) if use_global else pa.get('interval_minutes', 20)
                cbs = '' if use_global else pa.get('custom_basal_start', '')
                cbe = '' if use_global else pa.get('custom_basal_end', '')
                if not cbs:
                    dur = global_settings.get('basal_duration', 20)
                    cbs = self._hhmm_subtract(inj_t, dur)
                    cbe = inj_t
                vals = [aname, model_str, ch_str, inj_t, inj_end, iv, cbs, cbe,
                        total, auto_cnt, manual_cnt]

            for c, v in enumerate(vals, 1):
                sheet.cell(row=row, column=c, value=v).alignment = Alignment(horizontal='center')
                col_widths[c - 1] = max(col_widths[c - 1], len(str(v)))
            row += 1

        # Auto-fit column widths (padding of 3 chars)
        for c in range(len(headers)):
            sheet.column_dimensions[get_column_letter(c + 1)].width = col_widths[c] + 3

    # ==== Analyzer helpers (ported for identical formatting) ====
    def _time_to_timedelta_local(self, t):
        if isinstance(t, datetime):
            return pd.Timedelta(hours=t.hour, minutes=t.minute, seconds=t.second, microseconds=t.microsecond)
        return pd.Timedelta(hours=t.hour, minutes=t.minute, seconds=t.second) if hasattr(t, 'hour') else pd.NaT

    def _clean_duration_column_local(self, column):
        cleaned = []
        for item in column:
            if isinstance(item, str):
                item = item.replace(',', '.')
                try:
                    cleaned.append(float(item))
                except ValueError:
                    cleaned.append(np.nan)
            elif isinstance(item, (int, float)):
                cleaned.append(item)
            else:
                cleaned.append(np.nan)
        return pd.Series(cleaned)

    def _analyze_seizures_basal(self, df, recording_start, recording_end, interval_minutes=20):
        df = df.copy()
        df[0] = df[0].apply(self._time_to_timedelta_local)
        df[1] = df[1].apply(self._time_to_timedelta_local)
        df[2] = self._clean_duration_column_local(df[2])
        df = df[df[2] >= 1.0]
        interval_duration = pd.Timedelta(minutes=interval_minutes)
        recording_start = pd.to_timedelta(recording_start+":00") if len(recording_start) == 5 else pd.to_timedelta(recording_start)
        recording_end = pd.to_timedelta(recording_end+":00") if len(recording_end) == 5 else pd.to_timedelta(recording_end)

        # Handle midnight crossing
        if recording_end < recording_start:
            recording_end += pd.Timedelta(days=1)
            # Adjust event times that appear early (next day) by adding 24h
            # We assume events < recording_start belong to next day
            df.loc[df[0] < recording_start, 0] += pd.Timedelta(days=1)
            df.loc[df[1] < recording_start, 1] += pd.Timedelta(days=1)

        intervals = pd.timedelta_range(start=recording_start, end=recording_end, freq=interval_duration)
        interval_labels = [f'{i*interval_minutes}-{(i+1)*interval_minutes}' for i in range(len(intervals)-1)]
        interval_means, interval_sds, interval_ses, interval_cum, interval_counts = [], [], [], [], []
        seizures_divided = {}
        for i, start in enumerate(intervals[:-1]):
            end = start + interval_duration
            # Include any event that OVERLAPS the interval (not just those whose start
            # falls inside it), then clip its start/end/duration to the interval
            # boundaries. This prevents events that cross an interval boundary from
            # being reported with a duration longer than the interval itself.
            mask = (df[0] < end) & (df[1] > start)
            iv = df[mask].copy()
            if not iv.empty:
                iv[0] = iv[0].clip(lower=start)
                iv[1] = iv[1].clip(upper=end)
                iv[2] = (iv[1] - iv[0]).dt.total_seconds()
                mean_duration = iv[2].mean(); sd_duration = iv[2].std(); se_duration = sd_duration/np.sqrt(iv.shape[0])
                cumulative_duration = iv[2].sum(); count = iv.shape[0]
            else:
                mean_duration = sd_duration = se_duration = cumulative_duration = count = 0
            interval_means.append(mean_duration); interval_sds.append(sd_duration); interval_ses.append(se_duration)
            interval_cum.append(cumulative_duration); interval_counts.append(count); seizures_divided[interval_labels[i]] = iv
        result = pd.DataFrame({
            'Interval Label': interval_labels,
            'Mean Duration': np.round(interval_means, 2),
            'SD': np.round(interval_sds, 2),
            'SE': np.round(interval_ses, 2),
            'Cumulative Duration': np.round(interval_cum, 2),
            'Seizure Count': interval_counts
        })
        total_stats = {
            'Total Mean Duration': round(df[2].mean() if not df[2].empty else 0, 2),
            'Total SD': round(df[2].std() if not df[2].empty else 0, 2),
            'Total SE': round(df[2].std()/np.sqrt(df.shape[0]) if df.shape[0] > 0 else 0, 2),
            'Total Cumulative Duration': round(df[2].sum() if not df[2].empty else 0, 2),
            'Total Count': df.shape[0]
        }
        return result, seizures_divided, total_stats

    def _analyze_seizures_injection(self, df, injection_time, interval_minutes=20, basal_duration_min=20, end_time_str='',
                                     custom_basal_start='', custom_basal_end=''):
        df = df.copy()
        df[0] = df[0].apply(self._time_to_timedelta_local)
        df[1] = df[1].apply(self._time_to_timedelta_local)
        df[2] = self._clean_duration_column_local(df[2])
        df = df[df[2] >= 1.0]
        interval_duration = pd.Timedelta(minutes=interval_minutes)
        
        # Parse times
        injection_td = pd.to_timedelta(injection_time+":00") if len(injection_time) == 5 else pd.to_timedelta(injection_time)
        
        # Determine End Time (defaults to Injection + 3h if empty)
        if end_time_str and len(end_time_str) >= 5:
            end_td = pd.to_timedelta(end_time_str+":00") if len(end_time_str) == 5 else pd.to_timedelta(end_time_str)
            if end_td < injection_td:
                end_td += pd.Timedelta(days=1)
        else:
            end_td = injection_td + pd.Timedelta(hours=3)

        # Handle midnight crossing for events
        if end_td >= pd.Timedelta(hours=24):
             df.loc[df[0] < (injection_td - pd.Timedelta(hours=12)), 0] += pd.Timedelta(days=1)
             df.loc[df[1] < (injection_td - pd.Timedelta(hours=12)), 1] += pd.Timedelta(days=1)
        elif injection_td < pd.Timedelta(minutes=basal_duration_min):
             df.loc[df[0] > (injection_td + pd.Timedelta(hours=12)), 0] -= pd.Timedelta(days=1)
             df.loc[df[1] > (injection_td + pd.Timedelta(hours=12)), 1] -= pd.Timedelta(days=1)

        # Determine basal window: custom explicit range or computed from basal_duration_min
        if custom_basal_start and custom_basal_end:
            basal_start_td = pd.to_timedelta(custom_basal_start+":00") if len(custom_basal_start) == 5 else pd.to_timedelta(custom_basal_start)
            basal_end_td = pd.to_timedelta(custom_basal_end+":00") if len(custom_basal_end) == 5 else pd.to_timedelta(custom_basal_end)
            if basal_end_td < basal_start_td:
                basal_end_td += pd.Timedelta(days=1)
            effective_basal_min = max(1, int((basal_end_td - basal_start_td).total_seconds() / 60))
            basal_start = basal_start_td
            basal_end_for_mask = basal_end_td
        else:
            basal_start = injection_td - pd.Timedelta(minutes=basal_duration_min)
            basal_end_for_mask = injection_td
            effective_basal_min = basal_duration_min
        
        interval_means, interval_sds, interval_ses, interval_cum, interval_counts = [], [], [], [], []
        interval_labels = []
        seizures_divided = {}

        # -- Process Basal - Single interval --
        basal_lbl = "Basal"
        # Overlap-based inclusion + clipping so events straddling the basal window
        # boundaries are reported with the correct in-window duration.
        mask = (df[0] < basal_end_for_mask) & (df[1] > basal_start)
        basal_seizures = df[mask].copy()

        if not basal_seizures.empty:
            basal_seizures[0] = basal_seizures[0].clip(lower=basal_start)
            basal_seizures[1] = basal_seizures[1].clip(upper=basal_end_for_mask)
            basal_seizures[2] = (basal_seizures[1] - basal_seizures[0]).dt.total_seconds()
            mean_duration = basal_seizures[2].mean(); sd_duration = basal_seizures[2].std(); se_duration = sd_duration/np.sqrt(basal_seizures.shape[0])
            cumulative_duration = basal_seizures[2].sum(); count = basal_seizures.shape[0]
        else:
            mean_duration = sd_duration = se_duration = cumulative_duration = count = 0
        
        interval_means.append(mean_duration); interval_sds.append(sd_duration); interval_ses.append(se_duration)
        interval_cum.append(cumulative_duration); interval_counts.append(count)
        interval_labels.append(basal_lbl)
        seizures_divided[basal_lbl] = basal_seizures

        # Part 2: Post-injection
        curr = injection_td
        while curr < end_td:
            next_step = curr + interval_duration
            
            start_min = int((curr - injection_td).total_seconds() / 60)
            end_min = int((next_step - injection_td).total_seconds() / 60)
            lbl = f"{start_min}-{end_min}"
            
            # Overlap-based inclusion + clipping to the [curr, next_step) interval.
            # Prevents a seizure that spans interval boundaries from being reported
            # with its full (pre-binning) duration inside a single interval.
            mask = (df[0] < next_step) & (df[1] > curr)
            iv = df[mask].copy()

            if not iv.empty:
                iv[0] = iv[0].clip(lower=curr)
                iv[1] = iv[1].clip(upper=next_step)
                iv[2] = (iv[1] - iv[0]).dt.total_seconds()
                mean_duration = iv[2].mean(); sd_duration = iv[2].std(); se_duration = sd_duration/np.sqrt(iv.shape[0])
                cumulative_duration = iv[2].sum(); count = iv.shape[0]
            else:
                mean_duration = sd_duration = se_duration = cumulative_duration = count = 0
                
            interval_means.append(mean_duration); interval_sds.append(sd_duration); interval_ses.append(se_duration)
            interval_cum.append(cumulative_duration); interval_counts.append(count)
            interval_labels.append(lbl)
            seizures_divided[lbl] = iv
            
            curr = next_step

        result = pd.DataFrame({
            'Interval Label': interval_labels,
            'Mean Duration': np.round(interval_means, 2),
            'SD': np.round(interval_sds, 2),
            'SE': np.round(interval_ses, 2),
            'Cumulative Duration': np.round(interval_cum, 2),
            'Seizure Count': interval_counts
        })
        total_stats = {
            'Total Mean Duration': round(df[2].mean() if not df[2].empty else 0, 2),
            'Total SD': round(df[2].std() if not df[2].empty else 0, 2),
            'Total SE': round(df[2].std()/np.sqrt(df.shape[0]) if df.shape[0] > 0 else 0, 2),
            'Total Cumulative Duration': round(df[2].sum() if not df[2].empty else 0, 2),
            'Total Count': df.shape[0]
        }
        return result, seizures_divided, total_stats

    def _create_combined_sheet(self, workbook, processed_results, title):
        cumulative_duration_data = []
        mean_duration_data = []
        number_of_swds_data = []
        all_interval_labels = set()
        for result in processed_results.values():
            all_interval_labels.update(result['Interval Label'].tolist())
        
        # Sort intervals numerically
        # Sort order: "Basal X", then numeric ranges "0-20", "20-40", etc.
        def sort_key(label):
            if str(label).startswith('Basal'):
                return -1
            try:
                # Extract first number from range string "X-Y"
                return int(label.split('-')[0])
            except (ValueError, IndexError):
                return float('inf')  # Put other non-numeric labels at end

        all_interval_labels = sorted(list(all_interval_labels), key=sort_key)
        
        for animal_name, result in processed_results.items():
            animal_cumulative_duration = {'Animal ID': animal_name}
            animal_mean_duration = {'Animal ID': animal_name}
            animal_number_of_swds = {'Animal ID': animal_name}
            for col in ['DOB', 'Strain', 'Gender', 'Weight (gr)', 'GEN.', 'Stx Date', 'EEG Rec Date']:
                animal_cumulative_duration[col] = ''
                animal_mean_duration[col] = ''
                animal_number_of_swds[col] = ''
            for interval_label in all_interval_labels:
                if interval_label in result['Interval Label'].values:
                    row = result[result['Interval Label'] == interval_label].iloc[0]
                    animal_cumulative_duration[interval_label] = row['Cumulative Duration']
                    animal_mean_duration[interval_label] = row['Mean Duration']
                    animal_number_of_swds[interval_label] = row['Seizure Count']
                else:
                    animal_cumulative_duration[interval_label] = 0
                    animal_mean_duration[interval_label] = 0
                    animal_number_of_swds[interval_label] = 0
            animal_cumulative_duration['Cumulative'] = None
            animal_cumulative_duration['Mean'] = None
            animal_mean_duration['Cumulative'] = None
            animal_mean_duration['Mean'] = None
            animal_number_of_swds['Cumulative'] = None
            animal_number_of_swds['Mean'] = None
            cumulative_duration_data.append(animal_cumulative_duration)
            mean_duration_data.append(animal_mean_duration)
            number_of_swds_data.append(animal_number_of_swds)
        cumulative_duration_df = pd.DataFrame(cumulative_duration_data)
        mean_duration_df = pd.DataFrame(mean_duration_data)
        number_of_swds_df = pd.DataFrame(number_of_swds_data)
        required_cols = ['DOB', 'Strain', 'Gender', 'Weight (gr)', 'GEN.', 'Stx Date', 'EEG Rec Date']
        additional_columns = required_cols
        for df_ in [cumulative_duration_df, mean_duration_df, number_of_swds_df]:
            for col in required_cols:
                if col not in df_.columns:
                    df_[col] = ""
        cols_order = ['Animal ID'] + additional_columns + [col for col in cumulative_duration_df.columns if col not in ['Animal ID'] + additional_columns]
        cumulative_duration_df = cumulative_duration_df[cols_order]
        mean_duration_df = mean_duration_df[cols_order]
        number_of_swds_df = number_of_swds_df[cols_order]
        worksheet = workbook.create_sheet('Analysis', 0)
        num_columns = len(cumulative_duration_df.columns)
        last_column_letter = get_column_letter(num_columns)
        worksheet.merge_cells(f'A1:{last_column_letter}1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="dfcd7f", end_color="dfcd7f", fill_type="solid")
        section_title_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        basal_fill = PatternFill(start_color="FA8072", end_color="FA8072", fill_type="solid")
        interval_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        animal_id_fill = PatternFill(start_color="FFE800", end_color="FFE800", fill_type="solid")
        kumulatif_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        mean_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        red_font = Font(color='FF0000')
        section_title_font = Font(bold=True, color='FFFFFF')
        def write_section(df, section_title, start_row):
            worksheet.cell(row=start_row, column=1).value = section_title
            worksheet.cell(row=start_row, column=1).fill = section_title_fill
            worksheet.cell(row=start_row, column=1).font = section_title_font
            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=num_columns)
            worksheet.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
            start_row += 1
            df = df.fillna(0)
            for col_idx, column in enumerate(df.columns, 1):
                worksheet.cell(row=start_row, column=col_idx).value = column
            for r_idx, row_data in enumerate(df.itertuples(index=False), start=start_row+1):
                for c_idx, value in enumerate(row_data, start=1):
                    worksheet.cell(row=r_idx, column=c_idx).value = value
            data_start_row = start_row + 1
            data_end_row = start_row + df.shape[0]
            kumulatif_col_idx = None
            mean_col_idx = None
            interval_col_indices = []
            for col_idx, column in enumerate(df.columns, 1):
                column_width = max(len(str(column)), 12)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = column_width
                header_cell = worksheet.cell(row=start_row, column=col_idx)
                header_cell.font = Font(bold=True)
                header_cell.alignment = Alignment(horizontal='center')
                if column == 'Animal ID' or column in additional_columns:
                    header_cell.fill = animal_id_fill
                elif str(column).startswith('Basal'):
                    header_cell.fill = basal_fill
                    interval_col_indices.append(col_idx)
                elif column == 'Cumulative':
                    header_cell.fill = kumulatif_fill
                    kumulatif_col_idx = col_idx
                elif column == 'Mean':
                    header_cell.fill = mean_fill
                    mean_col_idx = col_idx
                elif '-' in str(column) or str(column).isdigit():
                    header_cell.fill = interval_fill
                    interval_col_indices.append(col_idx)
                for row_idx in range(data_start_row, data_end_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
            for row_idx in range(data_start_row, data_end_row + 1):
                interval_cells = [f"{get_column_letter(idx)}{row_idx}" for idx in interval_col_indices]
                if kumulatif_col_idx is not None:
                    kumulatif_cell = worksheet.cell(row=row_idx, column=kumulatif_col_idx)
                    kumulatif_cell.value = f"=SUM({','.join(interval_cells)})"
                    kumulatif_cell.number_format = '0.00'
                if mean_col_idx is not None:
                    mean_cell = worksheet.cell(row=row_idx, column=mean_col_idx)
                    mean_cell.value = f"=AVERAGE({','.join(interval_cells)})"
                    mean_cell.number_format = '0.00'
            
            # Only add statistics rows (n, Mean, SD, SE) if there's more than 1 animal
            # These stats are meaningless for a single animal
            if df.shape[0] > 1:
                stats_start_row = data_end_row + 3
                stats_titles = ['n', 'Mean', 'SD', 'SE']
                for i, stat_title in enumerate(stats_titles):
                    cell = worksheet.cell(row=stats_start_row + i, column=1)
                    cell.value = stat_title
                    cell.font = red_font
                for col_idx, col_name in enumerate(df.columns[len(additional_columns)+1:], start=len(additional_columns)+2):
                    col_letter = get_column_letter(col_idx)
                    col_cells = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                    n_cell = worksheet.cell(row=stats_start_row, column=col_idx)
                    n_cell.value = f"=COUNT({col_cells})"
                    n_cell.font = red_font
                    mean_cell_stat = worksheet.cell(row=stats_start_row + 1, column=col_idx)
                    mean_cell_stat.value = f"=AVERAGE({col_cells})"
                    mean_cell_stat.font = red_font
                    mean_cell_stat.number_format = '0.00'
                    sd_cell = worksheet.cell(row=stats_start_row + 2, column=col_idx)
                    sd_cell.value = f"=STDEV({col_cells})"
                    sd_cell.font = red_font
                    sd_cell.number_format = '0.00'
                    se_cell = worksheet.cell(row=stats_start_row + 3, column=col_idx)
                    se_cell.value = f"=IF({col_letter}{stats_start_row}>0, {col_letter}{stats_start_row + 2}/SQRT({col_letter}{stats_start_row}), 0)"
                    se_cell.font = red_font
                    se_cell.number_format = '0.00'
                return stats_start_row + 4
            else:
                # Single animal - skip stats rows
                return data_end_row + 2
        row = 2
        row = write_section(cumulative_duration_df, 'SWD Analysis/Cumulative Duration (sec)', row)
        row = write_section(mean_duration_df, 'SWD Analysis/Mean Duration of SWDs (sec)', row)
        row = write_section(number_of_swds_df, 'SWD Analysis/Number of SWDs', row)

    def _create_animal_sheet(self, workbook, animal_id, seizures_divided, recording_type, sheet_name):
        sheet = workbook.create_sheet(title=sheet_name[:31])
        basal_fill = PatternFill(start_color="FA8070", end_color="FA8070", fill_type="solid")
        interval_fill_1 = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
        interval_fill_2 = PatternFill(start_color="dfcd7f", end_color="dfcd7f", fill_type="solid")
        red_font = Font(color='FF0000', bold=True)
        header_font_white = Font(bold=True, color='FFFFFF')
        header_font_black = Font(bold=True, color='000000')
        intervals = list(seizures_divided.keys())
        max_seizures = max([seizures_divided[iv].shape[0] for iv in intervals]) if intervals else 0
        stats_row_start = 3 + max_seizures + 4
        col_idx = 1
        fill_colors = [interval_fill_1, interval_fill_2]
        for idx, interval in enumerate(intervals):
            seizures = seizures_divided[interval]
            num_seizures = seizures.shape[0]
            header_cell = sheet.cell(row=1, column=col_idx)
            sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+3)
            header_cell.value = interval
            header_cell.alignment = Alignment(horizontal='center')
            header_cell.font = Font(bold=True)
            header_cell.fill = basal_fill if str(interval).startswith('Basal') else fill_colors[idx % 2]
            header_cell.font = header_font_white if header_cell.fill == interval_fill_1 else header_font_black
            sub_headers = ['Sel Start', 'Sel End', 'Sel Duration', '']
            for i, sub_header in enumerate(sub_headers):
                cell = sheet.cell(row=2, column=col_idx + i)
                cell.value = sub_header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            for row_idx in range(num_seizures):
                sel_start = seizures.iloc[row_idx, 0]
                sel_end = seizures.iloc[row_idx, 1]
                sel_duration = seizures.iloc[row_idx, 2]
                sel_start = (datetime.min + sel_start).time().strftime('%H:%M:%S') if pd.notnull(sel_start) else ''
                sel_end = (datetime.min + sel_end).time().strftime('%H:%M:%S') if pd.notnull(sel_end) else ''
                data_row = 3 + row_idx
                sheet.cell(row=data_row, column=col_idx).value = sel_start
                sheet.cell(row=data_row, column=col_idx).alignment = Alignment(horizontal='left')
                sheet.cell(row=data_row, column=col_idx+1).value = sel_end
                sheet.cell(row=data_row, column=col_idx+1).alignment = Alignment(horizontal='left')
                sheet.cell(row=data_row, column=col_idx+2).value = sel_duration if not pd.isna(sel_duration) else 0
                sheet.cell(row=data_row, column=col_idx+2).number_format = '0.00'
                sheet.cell(row=data_row, column=col_idx+2).alignment = Alignment(horizontal='left')
            stats_row = stats_row_start
            stats = ['Total', 'Mean', 'SD', 'SE', 'Count']
            for i, stat in enumerate(stats):
                stat_label_cell = sheet.cell(row=stats_row + i, column=col_idx)
                stat_label_cell.value = stat
                stat_label_cell.font = red_font
                stat_label_cell.alignment = Alignment(horizontal='left')
                duration_col_letter = get_column_letter(col_idx + 2)
                data_start_row = 3
                data_end_row = 2 + num_seizures
                duration_cells = f"{duration_col_letter}{data_start_row}:{duration_col_letter}{data_end_row}"
                value_cell = sheet.cell(row=stats_row + i, column=col_idx + 2)
                value_cell.font = red_font
                value_cell.number_format = '0.00'
                if num_seizures > 0:
                    if stat == 'Total':
                        value_cell.value = f"=SUM({duration_cells})"
                    elif stat == 'Mean':
                        value_cell.value = f"=AVERAGE({duration_cells})"
                    elif stat == 'SD':
                        value_cell.value = f"=STDEV({duration_cells})"
                    elif stat == 'Count':
                        value_cell.value = f"=COUNT({duration_cells})"
                    elif stat == 'SE':
                        sd_cell_ref = f"{duration_col_letter}{stats_row + i - 2}"
                        count_cell_ref = f"{duration_col_letter}{stats_row + i + 1}"
                        value_cell.value = f"=IF({count_cell_ref}>0, {sd_cell_ref}/SQRT({count_cell_ref}), 0)"
                else:
                    value_cell.value = 0
            for i in range(4):
                sheet.column_dimensions[get_column_letter(col_idx + i)].width = 15
            col_idx += 4
        for row in sheet.iter_rows(min_row=1, max_row=stats_row_start + len(['Total','Mean','SD','SE','Count']) + 1):
            sheet.row_dimensions[row[0].row].height = 20

    def on_convert_to_manual(self):
        """Convert all automatic detections to manual events"""
        if not self.edit_mode_active:
            return

        self._snapshot_state()

        # Copy auto events to manual list, removing the is_auto flag
        for ev in self.refined_swd:
            manual_ev = ev.copy()
            manual_ev['is_auto'] = False
            self.manual_events.append(manual_ev)
            
        # Clear the auto list
        self.refined_swd = []
        
        # Merge any overlapping converted events
        self.manual_events = merge_overlapping_events(self.manual_events, self.sfreq)
        
        # Reset selection
        self.selected_event_idx = None
        self.active_event_type = None
        self.current_edit_event = None
        self.event_info_lbl.setText("")
        
        # Update displays
        self.update_combined_events()
        self.update_plot()

    def on_clear_events_clicked(self):
        """Clear all events of the appropriate type"""
        if not self.edit_mode_active:
            return
            
        # Show confirmation dialog
        reply = QMessageBox.question(
            self, "Confirm Clear Events", 
            "Are you sure you want to clear all events? This cannot be undone.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self._snapshot_state()
            show_auto = self.cb_show_auto.isChecked()

            # Clear manual events
            self.manual_events = []

            # Clear auto events if showing
            if show_auto:
                self.refined_swd = []
            
            # Reset selection
            self.selected_event_idx = None
            self.active_event_type = None
            self.current_edit_event = None
            self.event_info_lbl.setText("")
            
            # Update displays
            self.update_combined_events()
            self.update_plot()

    def on_delete_event_clicked(self):
        """Delete the currently selected event"""
        if self.selected_event_idx is None or not self.edit_mode_active:
            return

        self._snapshot_state()

        event_deleted = False
        if self.active_event_type == 'manual':
            if 0 <= self.selected_event_idx < len(self.manual_events):
                del self.manual_events[self.selected_event_idx]
                event_deleted = True
        elif self.active_event_type == 'auto':
            if 0 <= self.selected_event_idx < len(self.refined_swd):
                del self.refined_swd[self.selected_event_idx]
                event_deleted = True
        
        if event_deleted:
            self.show_toast("Event deleted", "success")
                
        # Reset selection
        self.selected_event_idx = None
        self.active_event_type = None
        self.current_edit_event = None
        self.event_info_lbl.setText("ℹ️ No event selected")
        
        # Update displays
        self.update_combined_events()
        self.update_plot()

    def on_add_event_clicked(self):
        """Add a new manual event at the current view position"""
        if not self.edit_mode_active:
            return

        self._snapshot_state()

        # Create a new event starting at the current position with a default 3-second duration
        start_sample = int(self.current_pos_sec * self.sfreq)
        end_sample = int((self.current_pos_sec + 3.0) * self.sfreq)
        
        # Make sure it doesn't exceed file bounds
        end_sample = min(end_sample, self.total_samples)
        
        new_event = {
            'start_sample': start_sample,
            'end_sample': end_sample,
            'duration_sec': (end_sample - start_sample) / self.sfreq,
            'is_auto': False  # Always create as manual event
        }
        
        # Add the event
        self.manual_events.append(new_event)
        
        # Check for overlaps and merge if needed
        self.manual_events = merge_overlapping_events(self.manual_events, self.sfreq)
        
        # Find index of the new (possibly merged) event
        for i, ev in enumerate(self.manual_events):
            if ev['start_sample'] <= start_sample and ev['end_sample'] >= end_sample:
                self.selected_event_idx = i
                self.active_event_type = 'manual'
                self.current_edit_event = ev
                break
        
        self.show_toast("New event added", "success")
        
        # Update displays
        self.update_combined_events()
        self.update_event_info()
        self.update_plot()

    def on_event_type_changed(self):
        """Handle changes in the event type dropdown selection."""
        # Reset selection when changing event types
        self.selected_event_idx = None
        self.active_event_type = None
        self.event_info_lbl.setText("")
        self.update_plot()

    def on_pick_region(self, event):
        """Handle mouse picking on region patches"""
        if not self.edit_mode_active:
            return
        
        # Find which region was picked
        for patch, idx, ev_type in self.region_patches:
            if event.artist == patch:
                # Always allow selection of both auto and manual events in edit mode
                self.selected_event_idx = idx
                self.active_event_type = ev_type
                
                # Get the events list based on type
                events = self.manual_events if ev_type == 'manual' else self.refined_swd
                
                if 0 <= idx < len(events):
                    # Set current edit event for dragging operations
                    self.current_edit_event = events[idx]
                    
                    # Update info display and plot
                    self.update_event_info()
                    self.update_plot()
                break

    def on_mouse_click(self, event):
        """Handle mouse click events for event editing (single-click selects/moves)."""
        if not self.edit_mode_active or event.inaxes is None:
            return
        
        # Get click position in time
        x_pos = event.xdata
        click_sample = int(x_pos * self.sfreq)
        
        # Check if we clicked on an existing event first
        for patch, idx, ev_type in self.region_patches:
            if patch.contains_point((event.x, event.y)):
                self.selected_event_idx = idx
                self.active_event_type = ev_type
                
                # Get the relevant event list
                events = self.manual_events if ev_type == 'manual' else self.refined_swd
                if 0 <= idx < len(events):
                    self.current_edit_event = events[idx]
                    
                    # Determine if click is near start or end boundary (within 0.5 seconds)
                    start_sec = self.current_edit_event['start_sample'] / self.sfreq
                    end_sec = self.current_edit_event['end_sample'] / self.sfreq
                    
                    boundary_margin = 0.5
                    
                    if abs(x_pos - start_sec) < boundary_margin:
                        # Near start boundary - prepare to adjust start time
                        self._snapshot_state()
                        self.dragging_start = True
                        self.dragging_end = False
                    elif abs(x_pos - end_sec) < boundary_margin:
                        # Near end boundary - prepare to adjust end time
                        self._snapshot_state()
                        self.dragging_start = False
                        self.dragging_end = True
                    
                    self.update_event_info()
                    self.update_plot()
                    return
        
        # If we get here, we didn't click on an existing event
        # Create new manual event only on double-click
        if self.btn_add.isEnabled() and getattr(event, 'dblclick', False):
            self._snapshot_state()

            default_duration = 2.0  # seconds
            half_duration_samples = int(default_duration * self.sfreq / 2)

            new_event = {
                'start_sample': max(0, click_sample - half_duration_samples),
                'end_sample': min(self.total_samples, click_sample + half_duration_samples),
                'duration_sec': default_duration,
                'is_auto': False  # Always create as manual event
            }
            
            self.manual_events.append(new_event)
            
            # Check for overlaps and merge if needed
            self.manual_events = merge_overlapping_events(self.manual_events, self.sfreq)
            
            # Find the newly created event in the list (it might have merged)
            found = False
            for i, ev in enumerate(self.manual_events):
                if (ev['start_sample'] <= new_event['start_sample'] and 
                    ev['end_sample'] >= new_event['end_sample']):
                    self.selected_event_idx = i
                    self.active_event_type = 'manual'
                    self.current_edit_event = ev
                    found = True
                    break
            
            if found:
                # Setup for dragging the end boundary
                self.dragging_start = False
                self.dragging_end = True
            
            # Update display
            self.update_combined_events()
            self.update_event_info()
            self.update_plot()

    def on_show_unrefined_changed(self, state):
        """Handle changes to the 'Show Unrefined Events' checkbox."""
        self.show_unrefined = bool(state)
        self.update_plot()
        
    def on_dark_mode_changed(self, state):
        """Handle changes to the 'Dark Mode' checkbox."""
        self.dark_mode = bool(state)
        
        # Apply dark mode to the application
        if self.dark_mode:
            plt.style.use('dark_background')
            # Get model-themed accent colors
            _, accent_color = self._current_model_event_colors()
            # Comprehensive GitHub-style dark theme with model-themed accents
            self.setStyleSheet(f"""
                QWidget {{ 
                    background-color: #0D1117; 
                    color: #E6EDF3; 
                    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
                }}
                QGroupBox {{ 
                    border: 2px solid #30363D; 
                    border-radius: 8px; 
                    margin-top: 12px; 
                    background-color: #161B22;
                    padding-top: 16px;
                    font-weight: 600;
                    font-size: 13px;
                    color: #E6EDF3;
                }}
                QGroupBox::title {{ 
                    subcontrol-origin: margin; 
                    subcontrol-position: top left; 
                    padding: 4px 12px; 
                    color: #E6EDF3; 
                    background-color: #161B22;
                }}
                QPushButton {{ 
                    background-color: #21262D; 
                    border: 1px solid #30363D; 
                    border-radius: 6px; 
                    padding: 8px 16px; 
                    color: #E6EDF3;
                    font-weight: 500;
                    font-size: 13px;
                }}
                QPushButton:hover {{ background-color: #30363D; }}
                QPushButton:pressed {{ background-color: #161B22; }}
                QPushButton:checked {{ 
                    background-color: {accent_color}; 
                    border-color: {accent_color};
                    color: #E0E0E0;
                }}
                QPushButton:disabled {{
                    background-color: #0D1117;
                    color: #6E7681;
                    border-color: #21262D;
                }}
                QLabel {{ color: #E6EDF3; }}
                QCheckBox {{ 
                    spacing: 6px; 
                    color: #E6EDF3;
                    font-size: 13px;
                }}
                QCheckBox::indicator {{ 
                    width: 18px; 
                    height: 18px; 
                    border: 2px solid #30363D; 
                    border-radius: 4px; 
                    background: #0D1117; 
                }}
                QCheckBox::indicator:checked {{ 
                    background-color: {accent_color}; 
                    border-color: {accent_color}; 
                }}
                QSlider::groove:horizontal {{ 
                    border: 1px solid #30363D; 
                    height: 6px; 
                    background: #21262D; 
                    border-radius: 3px; 
                }}
                QSlider::handle:horizontal {{ 
                    background: {accent_color}; 
                    border: 2px solid {accent_color}; 
                    width: 18px; 
                    height: 18px; 
                    margin: -7px 0; 
                    border-radius: 9px; 
                }}
                QSlider::handle:horizontal:hover {{ 
                    background: #B8860B;
                    border: 2px solid #D0D0D0; 
                }}
                QLineEdit, QSpinBox, QDoubleSpinBox {{ 
                    background-color: #0D1117; 
                    border: 1px solid #30363D; 
                    border-radius: 6px; 
                    padding: 6px 10px; 
                    color: #E6EDF3;
                    font-size: 13px;
                }}
                QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus {{
                    border-color: {accent_color};
                    outline: none;
                }}
                QScrollBar:horizontal {{ 
                    background: #0D1117; 
                    height: 12px; 
                    margin: 0px 16px 0 16px; 
                }}
                QScrollBar::handle:horizontal {{ 
                    background: #30363D; 
                    min-width: 20px; 
                    border-radius: 4px; 
                }}
                QScrollBar:vertical {{ 
                    background: #0D1117; 
                    width: 12px; 
                    margin: 16px 0 16px 0; 
                }}
                QScrollBar::handle:vertical {{ 
                    background: #30363D; 
                    min-height: 20px; 
                    border-radius: 4px; 
                }}
                QTabWidget::pane {{
                    border: 1px solid #30363D;
                    border-radius: 8px;
                    background-color: #161B22;
                    padding: 5px;
                }}
                QTabBar::tab {{
                    background-color: #0D1117;
                    border: 1px solid #30363D;
                    padding: 10px 20px;
                    margin-right: 4px;
                    border-top-left-radius: 6px;
                    border-top-right-radius: 6px;
                    color: #8B949E;
                    font-weight: 500;
                    font-size: 13px;
                }}
                QTabBar::tab:selected {{
                    background-color: #161B22;
                    border-bottom-color: #161B22;
                    color: {accent_color};
                }}
                QTabBar::tab:hover:!selected {{
                    background-color: #21262D;
                    color: #E6EDF3;
                }}
                QTableWidget {{
                    border: 1px solid #30363D;
                    border-radius: 6px;
                    background-color: #0D1117;
                    gridline-color: #21262D;
                    color: #E6EDF3;
                    font-size: 13px;
                }}
                QTableWidget::item {{
                    padding: 6px;
                    color: #E6EDF3;
                }}
                QTableWidget::item:selected {{
                    background-color: #1C3A1F;
                    color: #7EE787;
                }}
                QHeaderView::section {{
                    background-color: #161B22;
                    padding: 8px;
                    border: none;
                    border-bottom: 2px solid #30363D;
                    color: #8B949E;
                    font-weight: 600;
                    font-size: 12px;
                }}
                QProgressBar {{
                    border: 1px solid #30363D;
                    border-radius: 6px;
                    text-align: center;
                    background-color: #0D1117;
                    color: #E6EDF3;
                    font-size: 12px;
                }}
                QProgressBar::chunk {{
                    background-color: #238636;  /* constant green */
                    border-radius: 5px;
                }}
                QToolBar {{
                    background-color: #161B22;
                    border-bottom: 1px solid #30363D;
                    padding: 4px;
                    spacing: 8px;
                }}
                QToolButton {{
                    background-color: transparent;
                    border: none;
                    border-radius: 4px;
                    padding: 8px 12px;
                    color: #E6EDF3;
                    font-size: 13px;
                }}
                QToolButton:hover {{
                    background-color: #21262D;
                }}
                QToolButton:checked {{
                    background-color: {accent_color};
                    color: #E0E0E0;
                }}
                QFrame {{
                    background-color: #161B22;
                    border: 1px solid #30363D;
                    border-radius: 8px;
                }}
                StatCard {{
                    background-color: #161B22;
                    border: 1px solid #30363D;
                    border-radius: 8px;
                }}
            """)
            # Force dark title bar on Windows
            try:
                self._apply_win_dark_titlebar(self.winId())
            except Exception:
                pass
        else:
            plt.style.use('default')
            self.apply_modern_stylesheet()
            # Revert Windows title bar to light
            try:
                self._disable_win_dark_titlebar(self.winId())
            except Exception:
                pass
        
        # Update the figure with new style
        self.fig.set_facecolor(plt.rcParams['figure.facecolor'])
        
        # Update timeline minimap dark mode
        if hasattr(self, 'timeline_minimap'):
            self.timeline_minimap.set_dark_mode(self.dark_mode)
            # Update minimap color to match current mode's model theme
            auto_hex, _ = self._current_model_event_colors()
            self.timeline_minimap.set_auto_color(auto_hex)
        
        # Update welcome screen styling for dark mode
        if hasattr(self, 'welcome_widget'):
            self._apply_welcome_style(self.welcome_widget)

        # The PSD dialog's matplotlib canvas does not follow Qt stylesheets, so
        # explicitly re-theme it when it is open.
        psd_dialog = getattr(self, 'psd_dialog', None)
        if psd_dialog is not None:
            try:
                if psd_dialog.isVisible():
                    psd_dialog.refresh_theme()
            except Exception:
                pass

        self.update_plot()
        self.update_timeline_minimap()

    def on_prev_event_clicked(self):
        """Move to the previous SWD event in filtered list (step by one)."""
        # Refresh filtered events to ensure we have current data
        self.update_filtered_events()
        if not self.filtered_events:
            return

        # Prefer index-based navigation to avoid skipping
        current_idx = self.get_current_filtered_event_index()
        if current_idx is not None:
            prev_idx = current_idx - 1
            if prev_idx >= 0:
                self._navigate_to_filtered_event(prev_idx)
                return

        # Fallback when nothing is selected: time-based
        reference_time = self.current_pos_sec + self.display_window_seconds/2
        prev_event = None
        for fev in reversed(self.filtered_events):
            if fev['start_time'] < reference_time:
                prev_event = fev
                break
        if prev_event:
            target_idx = self.filtered_events.index(prev_event)
            self._navigate_to_filtered_event(target_idx)
        else:
            pass

    def on_next_event_clicked(self):
        """Move to the next SWD event in filtered list (step by one)."""
        # Refresh filtered events to ensure we have current data
        self.update_filtered_events()
        if not self.filtered_events:
            return

        # Prefer index-based navigation to avoid skipping
        current_idx = self.get_current_filtered_event_index()
        if current_idx is not None:
            next_idx = current_idx + 1
            if next_idx < len(self.filtered_events):
                self._navigate_to_filtered_event(next_idx)
                return

        # Fallback when nothing is selected: time-based
        reference_time = self.current_pos_sec + self.display_window_seconds/2
        next_event = None
        for fev in self.filtered_events:
            if fev['start_time'] > reference_time:
                next_event = fev
                break
        if next_event:
            target_idx = self.filtered_events.index(next_event)
            self._navigate_to_filtered_event(target_idx)
        else:
            pass
            
    def _navigate_to_filtered_event(self, filtered_idx):
        """Navigate to a specific event in the filtered list"""
        if not (0 <= filtered_idx < len(self.filtered_events)):
            return
            
        target_event = self.filtered_events[filtered_idx]
        event_data = target_event['event']
        event_type = target_event['type']
        event_index = target_event['index']
        
        # Center the event in the view with 2 second margin
        new_pos = max(0, event_data['start_sample'] / self.sfreq - 2)
        self.current_pos_sec = new_pos
        
        # Select the event
        self.selected_event_idx = event_index
        self.active_event_type = event_type
        self.current_edit_event = event_data
        
        # Update displays
        self.update_event_info()
        self.update_plot()

    def _apply_win_dark_titlebar(self, win_id):
        """Enable dark title bar on Windows 10+ using DWM attribute if available."""
        try:
            if sys.platform != 'win32':
                return
            hwnd = int(win_id)
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            DWMWA_USE_IMMERSIVE_DARK_MODE_BEFORE_20H1 = 19
            value = ctypes.c_int(1)
            dwmapi = ctypes.windll.dwmapi
            # Try both attribute IDs for compatibility across versions
            dwmapi.DwmSetWindowAttribute(wintypes.HWND(hwnd), ctypes.c_int(DWMWA_USE_IMMERSIVE_DARK_MODE), ctypes.byref(value), ctypes.sizeof(value))
            dwmapi.DwmSetWindowAttribute(wintypes.HWND(hwnd), ctypes.c_int(DWMWA_USE_IMMERSIVE_DARK_MODE_BEFORE_20H1), ctypes.byref(value), ctypes.sizeof(value))
        except Exception:
            pass

    def _disable_win_dark_titlebar(self, win_id):
        """Disable dark title bar on Windows (revert to light)."""
        try:
            if sys.platform != 'win32':
                return
            hwnd = int(win_id)
            DWMWA_USE_IMMERSIVE_DARK_MODE = 20
            DWMWA_USE_IMMERSIVE_DARK_MODE_BEFORE_20H1 = 19
            value = ctypes.c_int(0)
            dwmapi = ctypes.windll.dwmapi
            dwmapi.DwmSetWindowAttribute(wintypes.HWND(hwnd), ctypes.c_int(DWMWA_USE_IMMERSIVE_DARK_MODE), ctypes.byref(value), ctypes.sizeof(value))
            dwmapi.DwmSetWindowAttribute(wintypes.HWND(hwnd), ctypes.c_int(DWMWA_USE_IMMERSIVE_DARK_MODE_BEFORE_20H1), ctypes.byref(value), ctypes.sizeof(value))
        except Exception:
            pass

    def on_psd_analysis_clicked(self):
        """Open the PSD Analysis dialog for the selected event"""
        if self.selected_event_idx is None:
            QMessageBox.warning(self, "No Event Selected", "Please select an event for PSD analysis.")
            return

        # Get the selected event with bounds checking
        try:
            if self.active_event_type == 'auto':
                if not self.refined_swd or self.selected_event_idx < 0 or self.selected_event_idx >= len(self.refined_swd):
                    QMessageBox.warning(self, "Invalid Event", "Selected event index is out of range.")
                    return
                selected_event = self.refined_swd[self.selected_event_idx]
            elif self.active_event_type == 'manual':
                if not self.manual_events or self.selected_event_idx < 0 or self.selected_event_idx >= len(self.manual_events):
                    QMessageBox.warning(self, "Invalid Event", "Selected event index is out of range.")
                    return
                selected_event = self.manual_events[self.selected_event_idx]
            else:
                QMessageBox.warning(self, "No Active Events", "No active event list available.")
                return
        except (IndexError, KeyError) as e:
            QMessageBox.warning(self, "Error", f"Failed to access event: {str(e)}")
            return

        # Reuse existing dialog if it exists and is visible, otherwise create a new one
        if self.psd_dialog is not None and self.psd_dialog.isVisible():
            # Dialog is already open, just update the event data
            self.psd_dialog.set_event_data(self.eeg0, self.eeg1, self.sfreq, selected_event)
        else:
            # Create a new dialog instance (or reuse if it exists but was closed)
            if self.psd_dialog is None:
                self.psd_dialog = PSDAnalysisDialog(self)
            # Set event data for the dialog
            self.psd_dialog.set_event_data(self.eeg0, self.eeg1, self.sfreq, selected_event)
            # Show the dialog
            self.psd_dialog.show()
            self.psd_dialog.raise_()
            self.psd_dialog.activateWindow()

    def on_unet_predict_clicked(self):
        """Run UNET pipeline: model selection, preprocessing, inference, postprocess"""
        if torch is None:
            QMessageBox.critical(self, "PyTorch Missing", "PyTorch is not installed. Install torch to use UNET predictions.")
            return
        try:
            # Use preloaded UNET if available; else prompt user
            model_path = None
            if self.unet_model is not None:
                model_path = getattr(self, 'unet_model_path', None)
            else:
                model_path, _ = QFileDialog.getOpenFileName(
                    self, "Select UNET Model File", "", "PyTorch Models (*.pt *.pth);;All Files (*)")
                if not model_path:
                    return
                self.unet_model_path = model_path

            # Select channel
            ch_item, ok = QInputDialog.getItem(
                self, "Select Channel for UNET", "Channel:", ["Channel 1", "Channel 2"], 0, False)
            if not ok:
                return
            self.unet_channel = 0 if ch_item == "Channel 1" else 1
            
            # Store the channel name for display
            self.unet_channel_name = ch_item

            # Load model if not already preloaded: TorchScript preferred, else plain state_dict
            loaded = self.unet_model if self.unet_model is not None else None
            load_error = None
            if loaded is None:
                try:
                    loaded = torch.jit.load(model_path, map_location='cpu')
                    loaded.eval()
                except Exception as e:
                    load_error = e
                    loaded = None
                if loaded is None:
                    try:
                        # Try common locations for model.py
                        import os, sys, importlib
                        model_module = None
                        # 1) Same directory
                        local_model = os.path.join(os.path.dirname(__file__), 'model.py')
                        if os.path.exists(local_model):
                            import importlib.util
                            spec = importlib.util.spec_from_file_location('model', local_model)
                            model_module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(model_module)
                        # 2) Model/ folder
                        if model_module is None:
                            model_dir = os.path.join(os.path.dirname(__file__), 'Model')
                            if os.path.isdir(model_dir):
                                sys.path.insert(0, model_dir)
                                try:
                                    model_module = importlib.import_module('model')
                                except Exception:
                                    model_module = None
                        # 3) Relevant Scripts/ folder
                        if model_module is None:
                            rel_dir = os.path.join(os.path.dirname(__file__), 'Relevant Scripts')
                            if os.path.isdir(rel_dir):
                                sys.path.insert(0, rel_dir)
                                try:
                                    model_module = importlib.import_module('model')
                                except Exception:
                                    model_module = None
                        # 4) Ask user to pick a Python file
                        if model_module is None:
                            py_path, _ = QFileDialog.getOpenFileName(
                                self, "Select UNET architecture file (model.py)", "", "Python Files (*.py);;All Files (*)")
                            if not py_path:
                                raise ImportError('No model architecture selected')
                            import importlib.util
                            spec = importlib.util.spec_from_file_location('model', py_path)
                            model_module = importlib.util.module_from_spec(spec)
                            spec.loader.exec_module(model_module)

                        NetClass = getattr(model_module, 'AttentionResUNet', None)
                        if NetClass is None:
                            raise ImportError('AttentionResUNet not found in selected model module')

                        net = NetClass(n_classes=1, in_channels=1, out_channels=1)
                        if model_path and model_path.lower().endswith(('.pt', '.pth')):
                            state_dict = torch.load(model_path, map_location='cpu')
                            net.load_state_dict(state_dict)
                        net.eval()
                        loaded = net
                    except Exception as e2:
                        tb = traceback.format_exc()
                        QMessageBox.critical(self, 'Model Load Error',
                                             f"Failed to load UNET as TorchScript and as state_dict.\nFirst error: {load_error}\nSecond error: {e2}\n\n{tb}")
                        return
            self.unet_model = loaded
            # Kick off UNET compute in background thread
            sig = self.eeg0 if self.unet_channel == 0 else self.eeg1
            self.progress_bar.show()
            self.btn_unet.setEnabled(False)
            # Prepare worker
            self.unet_thread = QThread(self)
            self.unet_worker = UNETWorker(
                signal=sig,
                sfreq=self.sfreq,
                model=self.unet_model,
                unet_fs=self.unet_fs,
                segment_len=1024,
                min_duration=self.unet_min_duration,
                gap_threshold=self.unet_gap_threshold
            )
            self.unet_worker.moveToThread(self.unet_thread)
            self.unet_thread.started.connect(self.unet_worker.run)
            self.unet_worker.finished.connect(self._on_unet_finished)
            self.unet_worker.error.connect(self._on_unet_error)
            self.unet_worker.finished.connect(self._cleanup_unet_thread)
            self.unet_worker.error.connect(self._cleanup_unet_thread)
            self.unet_thread.start()

        except Exception as e:
            tb = traceback.format_exc()
            QMessageBox.critical(self, "UNET Error", f"UNET prediction failed: {e}\n\n{tb}")

    def start_key_assignment(self, task):
        self.assigning_key_task = task
        # Ensure widget focus to receive key events
        self.setFocus()

    def load_key_assignments(self):
        try:
            with open(self.key_assignments_file, 'r') as f:
                data = json.load(f)
            for task, key in data.items():
                if task in self.key_assignments:
                    self.key_assignments[task] = key
        except Exception:
            pass

    def load_analyzer_settings(self):
        try:
            settings_path = os.path.join(os.path.dirname(__file__), "analyzer_settings.json")
            if os.path.exists(settings_path):
                with open(settings_path, 'r') as f:
                    settings = json.load(f)
                
                # Update UNET model path if relative
                unet_path = settings.get('unet_model_path')
                if unet_path and not os.path.isabs(unet_path):
                     settings['unet_model_path'] = os.path.abspath(os.path.join(os.path.dirname(__file__), unet_path))
                
                # Apply settings to internal state if needed
                # For now just ensuring path resolution is correct
                
                return settings
        except Exception as e:
            print(f"Error loading analyzer settings: {e}")
        return {}

    def save_key_assignments(self):
        try:
            with open(self.key_assignments_file, 'w') as f:
                json.dump(self.key_assignments, f)
        except Exception:
            pass

    def _friendly_key_name(self, key):
        try:
            return QKeySequence(key).toString()
        except Exception:
            return str(key)

    # === Undo / Redo ========================================================
    def _capture_state(self):
        """Build a snapshot dict of the edit-relevant state."""
        try:
            return {
                'manual_events': [dict(ev) for ev in self.manual_events],
                'refined_swd': [dict(ev) for ev in self.refined_swd],
                'refined_swd_base': [dict(ev) for ev in getattr(self, 'refined_swd_base', []) or []],
                'selected_event_idx': self.selected_event_idx,
                'active_event_type': self.active_event_type,
                'model_type': getattr(self, 'model_type', None),
            }
        except Exception:
            return None

    def _snapshot_state(self):
        """Push the current state onto the undo stack (clears redo)."""
        snap = self._capture_state()
        if snap is None:
            return
        self._undo_stack.append(snap)
        if len(self._undo_stack) > self._max_undo:
            # Drop oldest
            del self._undo_stack[0:len(self._undo_stack) - self._max_undo]
        self._redo_stack.clear()

    def _restore_state(self, snap):
        if not snap:
            return
        try:
            self.manual_events = [dict(ev) for ev in snap.get('manual_events', [])]
            self.refined_swd = [dict(ev) for ev in snap.get('refined_swd', [])]
            self.refined_swd_base = [dict(ev) for ev in snap.get('refined_swd_base', [])]
            self.selected_event_idx = snap.get('selected_event_idx')
            self.active_event_type = snap.get('active_event_type')
            self.current_edit_event = None
            # Keep the per-model cache in sync so a later model switch doesn't
            # resurrect the pre-undo state.
            mt = snap.get('model_type')
            if mt and hasattr(self, '_model_state_cache'):
                self._model_state_cache[mt] = {
                    'refined_swd': list(self.refined_swd),
                    'refined_swd_base': list(self.refined_swd_base),
                }
        except Exception as e:
            print(f"Restore state failed: {e}")

    def undo(self):
        if not getattr(self, '_undo_stack', None):
            try:
                self.show_toast("Nothing to undo", "info")
            except Exception:
                pass
            return
        current = self._capture_state()
        if current is not None:
            self._redo_stack.append(current)
        snap = self._undo_stack.pop()
        self._restore_state(snap)
        try:
            self.update_combined_events()
            self.update_filtered_events()
            self.update_event_info()
            self.update_plot()
            self.show_toast("Undo", "info")
        except Exception:
            pass

    def redo(self):
        if not getattr(self, '_redo_stack', None):
            try:
                self.show_toast("Nothing to redo", "info")
            except Exception:
                pass
            return
        current = self._capture_state()
        if current is not None:
            self._undo_stack.append(current)
        snap = self._redo_stack.pop()
        self._restore_state(snap)
        try:
            self.update_combined_events()
            self.update_filtered_events()
            self.update_event_info()
            self.update_plot()
            self.show_toast("Redo", "info")
        except Exception:
            pass

    def show_key_assignments_dialog(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Key Assignments")
        # Compact style to remove white margins in value labels and buttons
        dlg.setStyleSheet("""
            QLabel { margin: 0px; padding: 0px; background: transparent; }
            QPushButton { padding: 4px 10px; }
            QLineEdit, QComboBox, QDoubleSpinBox, QAbstractSpinBox { padding: 1px 4px; margin: 0px; }
        """)
        dlg.setModal(True)
        dlg.setFocusPolicy(Qt.StrongFocus)
        try:
            if self.dark_mode:
                self._apply_win_dark_titlebar(dlg.winId())
        except Exception:
            pass
        v = QVBoxLayout(dlg)
        info = QLabel("Click 'Assign' then press a key. Assigned keys are shown by name (e.g., A, F1).")
        info.setStyleSheet("color: gray; font-size: 9pt;")
        v.addWidget(info)
        # Track labels for live updates
        labels_map = {}
        for task in self.key_assignments:
            row = QHBoxLayout()
            lbl = QLabel(f"{task}:")
            key_lbl = QLabel(self._friendly_key_name(self.key_assignments[task]))
            labels_map[task] = key_lbl
            btn = QPushButton("Assign")
            btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            def handler(checked=False, t=task, label_ref=key_lbl):
                self.start_key_assignment(t)
                label_ref.setText("Press any key...")
                dlg.setFocus()
                dlg.activateWindow()
            btn.clicked.connect(handler)
            row.addWidget(lbl)
            row.addWidget(key_lbl)
            row.addWidget(btn)
            v.addLayout(row)
        close_row = QHBoxLayout()
        close_btn = QPushButton("Close")
        close_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        close_btn.clicked.connect(dlg.accept)
        close_row.addStretch(1)
        close_row.addWidget(close_btn)
        v.addLayout(close_row)
        # Install event filter to capture keys while dialog is focused
        self._key_assign_dialog = dlg
        self._key_labels = labels_map
        dlg.installEventFilter(self)
        # Also capture key events from all children of the dialog
        for child in dlg.findChildren(QObject):
            child.installEventFilter(self)
        # Ensure dialog can receive key events
        dlg.setFocus()
        try:
            dlg.exec_()
        finally:
            dlg.removeEventFilter(self)
            for child in dlg.findChildren(QObject):
                try:
                    child.removeEventFilter(self)
                except Exception:
                    pass
            self._key_assign_dialog = None
            self._key_labels = None
            self.assigning_key_task = None
    
    # --- key-assignment helpers (modifier-aware) -----------------------------
    _MOD_MASK = int(Qt.ControlModifier) | int(Qt.ShiftModifier) | int(Qt.AltModifier) | int(Qt.MetaModifier)
    _PURE_MODIFIER_KEYS = {int(Qt.Key_Control), int(Qt.Key_Shift), int(Qt.Key_Alt), int(Qt.Key_Meta)}

    def _event_key_combo(self, event):
        """Return an int combining event.key() with the supported modifier bits."""
        return int(event.key()) | (int(event.modifiers()) & self._MOD_MASK)

    def keyPressEvent(self, event):
        if hasattr(self, 'assigning_key_task') and self.assigning_key_task:
            # Ignore solo modifier presses so the user can press e.g. Ctrl+Z and
            # we only commit on the non-modifier key.
            if int(event.key()) in self._PURE_MODIFIER_KEYS:
                return
            self.key_assignments[self.assigning_key_task] = self._event_key_combo(event)
            self.save_key_assignments()
            self.assigning_key_task = None
            return

        pressed = self._event_key_combo(event)

        # Resolve each assignment with a fallback so older files without an
        # entry (e.g. missing 'Undo') still behave sensibly.
        get_k = self.key_assignments.get
        if pressed == get_k('Undo', int(Qt.ControlModifier) | int(Qt.Key_Z)):
            self.undo()
        elif pressed == get_k('Redo', int(Qt.ControlModifier) | int(Qt.ShiftModifier) | int(Qt.Key_Z)):
            self.redo()
        elif pressed == get_k('Previous', int(Qt.Key_PageUp)):
            # Window prev removed from UI; keep PageUp as scroll backward shortcut
            self.current_pos_sec = max(0, self.current_pos_sec - self.display_window_seconds)
            self.update_plot()
        elif pressed == get_k('Next', int(Qt.Key_PageDown)):
            # Window next removed from UI; keep PageDown as scroll forward shortcut
            self.current_pos_sec = min(self.total_samples/self.sfreq - self.display_window_seconds,
                                       self.current_pos_sec + self.display_window_seconds)
            self.update_plot()
        elif pressed == get_k('Previous Event', int(Qt.Key_Left)):
            self.on_prev_event_clicked()
        elif pressed == get_k('Next Event', int(Qt.Key_Right)):
            self.on_next_event_clicked()
        elif pressed == get_k('Add Event', int(Qt.Key_A)):
            if self.edit_mode_active:
                self.on_add_event_clicked()
        elif pressed == get_k('Delete Event', int(Qt.Key_Delete)):
            if self.edit_mode_active and self.selected_event_idx is not None:
                self.on_delete_event_clicked()
        elif pressed == get_k('Merge Adjacent SWDs', int(Qt.Key_M)):
            self.merge_adjacent_selected_swd()
        elif pressed == get_k('PSD Analysis', int(Qt.Key_P)):
            self.on_psd_analysis_clicked()
        else:
            super().keyPressEvent(event)

    def eventFilter(self, obj, event):
        # Capture key presses inside the key assignment dialog
        if getattr(self, '_key_assign_dialog', None) is not None:
            # Check if the event is from the dialog or any of its children
            if (obj is self._key_assign_dialog or
                (hasattr(obj, 'parent') and obj.parent() is self._key_assign_dialog)):
                if event.type() == QEvent.KeyPress and getattr(self, 'assigning_key_task', None):
                    # Skip solo modifier presses so the user can record combos.
                    if int(event.key()) in self._PURE_MODIFIER_KEYS:
                        return True
                    task = self.assigning_key_task
                    combo = self._event_key_combo(event)
                    self.key_assignments[task] = combo
                    try:
                        # Update dialog label
                        if getattr(self, '_key_labels', None) and task in self._key_labels:
                            self._key_labels[task].setText(self._friendly_key_name(combo))
                    except Exception:
                        pass
                    self.save_key_assignments()
                    self.assigning_key_task = None
                    return True
        return super().eventFilter(obj, event)

    def merge_adjacent_selected_swd(self):
        """Merge the selected SWD with the nearest forward event within 20 seconds."""
        # Use the same logic as on_merge_swds_clicked for consistency
        self.on_merge_swds_clicked()

    def on_merge_swds_clicked(self):
        """Merge the selected event with the nearest FORWARD event within a 20 s gap.
        Only events that come AFTER the selected event (or overlap it) are considered;
        backward merges are disabled per user request."""
        try:
            if self.selected_event_idx is None:
                self.show_toast("Select an event first", "warning")
                return
            # Work on the list based on selection type
            if self.active_event_type == 'manual':
                events = self.manual_events
            elif self.active_event_type == 'auto':
                events = self.refined_swd
            else:
                self.show_toast("No active event list", "warning")
                return
            if not events or self.selected_event_idx < 0 or self.selected_event_idx >= len(events):
                return
            
            # Store the selected event before sorting
            selected_event = events[self.selected_event_idx]
            selected_start = selected_event['start_sample']
            
            # Sort events by start_sample to ensure correct gap calculation
            events.sort(key=lambda x: x['start_sample'])
            
            # Find the new index of the selected event after sorting
            # Match by start_sample to handle cases where events might be equal
            found_idx = None
            for idx, ev in enumerate(events):
                if ev['start_sample'] == selected_start and ev['end_sample'] == selected_event['end_sample']:
                    found_idx = idx
                    break
            
            if found_idx is None:
                # Fallback: use the original index if still valid
                if self.selected_event_idx < len(events):
                    found_idx = self.selected_event_idx
                else:
                    self.show_toast("Could not locate selected event after sorting", "error")
                    return
            
            self.selected_event_idx = found_idx
            target = events[self.selected_event_idx]
            max_gap_samples = int(20 * self.sfreq)  # 20 seconds maximum gap

            # Find nearest FORWARD event with gap ≤ 20 seconds.
            nearest_j = None
            nearest_gap = None

            target_start = target['start_sample']
            target_end = target['end_sample']

            for j, ev in enumerate(events):
                if j == self.selected_event_idx:
                    continue

                ev_start = ev['start_sample']
                ev_end = ev['end_sample']

                # Only consider events that are forward in time relative to the
                # selected event. An event is "forward" if it starts at or after
                # the selected event's start (and its end is at or beyond the
                # selected event's end, so it extends the right boundary).
                # Overlapping events that extend forward are still allowed.
                gap = None
                if ev_start >= target_end:
                    # Clean forward neighbour (gap after target end)
                    gap = ev_start - target_end
                elif ev_start >= target_start and ev_end > target_end:
                    # Overlapping but extends forward
                    gap = 0
                # else: event is fully before / ends before target_end -> backward, skip

                if gap is not None and gap <= max_gap_samples:
                    if nearest_gap is None or gap < nearest_gap:
                        nearest_gap = gap
                        nearest_j = j

            if nearest_j is None:
                self.show_toast("No forward SWD within 20s gap to merge", "info")
                return

            # Snapshot state for undo before mutating the events list.
            self._snapshot_state()

            other = events[nearest_j]
            merged = {
                'start_sample': min(target['start_sample'], other['start_sample']),
                'end_sample': max(target['end_sample'], other['end_sample']),
                'duration_sec': (max(target['end_sample'], other['end_sample']) - min(target['start_sample'], other['start_sample'])) / self.sfreq,
                'is_auto': target.get('is_auto', False) and other.get('is_auto', False)
            }
            
            # Preserve other event properties if they exist
            for key in ['confidence', 'model_type']:
                if key in target:
                    merged[key] = target[key]
                elif key in other:
                    merged[key] = other[key]
            
            lo = min(self.selected_event_idx, nearest_j)
            hi = max(self.selected_event_idx, nearest_j)
            events[lo] = merged
            del events[hi]
            
            # Adjust selected index after deletion
            self.selected_event_idx = lo
            
            self.update_combined_events()
            self.update_filtered_events()
            self.update_plot()
            
            gap_sec = nearest_gap / self.sfreq
            self.show_toast(f"✓ Merged forward (gap: {gap_sec:.2f}s ≤ 20s)", "success")
        except Exception as e:
            print(f"Merge SWDs failed: {e}")
            import traceback
            traceback.print_exc()
            self.show_toast("Merge failed", "error")

    def merge_nearest_event(self):
        """Merge the selected event with the nearest FORWARD event within 20 seconds gap.
        Backward merges are intentionally disabled."""
        if self.selected_event_idx is None:
            return
        # Determine working list based on current selection type
        if self.active_event_type == 'manual':
            events = self.manual_events
        elif self.active_event_type == 'auto':
            events = self.refined_swd
        else:
            return
        idx = self.selected_event_idx
        if idx < 0 or idx >= len(events):
            return

        target = events[idx]
        max_gap_samples = int(20 * self.sfreq)  # 20 seconds maximum gap

        # Find nearest FORWARD event within 20 s
        nearest_j = None
        nearest_gap = None

        target_start = target['start_sample']
        target_end = target['end_sample']

        for j, ev in enumerate(events):
            if j == idx:
                continue

            ev_start = ev['start_sample']
            ev_end = ev['end_sample']

            gap = None
            if ev_start >= target_end:
                gap = ev_start - target_end
            elif ev_start >= target_start and ev_end > target_end:
                gap = 0
            # else: backward / contained -> skip

            if gap is not None and gap <= max_gap_samples:
                if nearest_gap is None or gap < nearest_gap:
                    nearest_gap = gap
                    nearest_j = j
        
        if nearest_j is None:
            return

        self._snapshot_state()

        ev2 = events[nearest_j]
        merged = {
            'start_sample': min(target['start_sample'], ev2['start_sample']),
            'end_sample': max(target['end_sample'], ev2['end_sample']),
            'duration_sec': (max(target['end_sample'], ev2['end_sample']) - min(target['start_sample'], ev2['start_sample'])) / self.sfreq,
            'is_auto': target.get('is_auto', False) and ev2.get('is_auto', False)
        }
        
        # Preserve other event properties if they exist
        for key in ['confidence', 'model_type']:
            if key in target:
                merged[key] = target[key]
            elif key in ev2:
                merged[key] = ev2[key]
        
        lo = min(idx, nearest_j)
        hi = max(idx, nearest_j)
        events[lo] = merged
        del events[hi]
        self.selected_event_idx = lo
        self.update_combined_events()
        self.update_event_info()
        self.update_plot()

    def setup_smooth_scrolling(self):
        # Replace QSlider with QScrollBar for smoother, continuous scrolling
        from PyQt5.QtWidgets import QScrollBar
        self.scrollbar = QScrollBar(Qt.Horizontal)
        self.scrollbar.setMinimum(0)
        self.scrollbar.setMaximum(int(self.total_samples/self.sfreq - self.display_window_seconds))
        self.scrollbar.setSingleStep(1)
        self.scrollbar.setPageStep(max(1, int(self.display_window_seconds // 2)))
        self.scrollbar.valueChanged.connect(self.on_scrollbar_changed)
        # Replace slider in layout
        parent = self.slider.parentWidget()
        layout = parent.layout() if parent else None
        if layout:
            layout.replaceWidget(self.slider, self.scrollbar)
            self.slider.hide()
        self.scrollbar.show()

    def on_scrollbar_changed(self, value):
        self.current_pos_sec = value
        self.update_plot()

    def wheelEvent(self, event):
        # Scroll up: previous window, Scroll down: next window with sensitivity control
        if event.angleDelta().y() > 0:
            # Scroll backward with sensitivity
            scroll_amount = self.display_window_seconds * self.scroll_sensitivity.value() * 0.5
            self.current_pos_sec = max(0, self.current_pos_sec - scroll_amount)
            self.update_plot()
        elif event.angleDelta().y() < 0:
            # Scroll forward with sensitivity
            scroll_amount = self.display_window_seconds * self.scroll_sensitivity.value() * 0.5
            max_pos = self.total_samples/self.sfreq - self.display_window_seconds
            self.current_pos_sec = min(max_pos, self.current_pos_sec + scroll_amount)
            self.update_plot()

    def on_unet_settings_clicked(self):
        """Open the UNET Settings dialog"""
        if not hasattr(self, 'unet_settings_dialog'):
            self.unet_settings_dialog = UNETSettingsDialog(self)
        
        # Show the dialog
        self.unet_settings_dialog.show()
        self.unet_settings_dialog.raise_()
        self.unet_settings_dialog.activateWindow()
    
    def on_settings_clicked(self):
        """Open the comprehensive Settings dialog"""
        dialog = ComprehensiveSettingsDialog(self, self.model_type)
        # Keep a reference so other dialogs (e.g. the PSD dialog) can keep the
        # shared bandpass-filter widgets in sync while this dialog is open.
        self.settings_dialog = dialog
        try:
            if self.dark_mode:
                self._apply_win_dark_titlebar(dialog.winId())
        except Exception:
            pass
        try:
            dialog.exec_()
        finally:
            self.settings_dialog = None
    
    def on_ylim_zoom_out(self):
        """Zoom out Y-axis"""
        current = self.ylim_input.value()
        self.ylim_input.setValue(min(current + 0.1, 10.0))
    
    def on_ylim_zoom_in(self):
        """Zoom in Y-axis"""
        current = self.ylim_input.value()
        self.ylim_input.setValue(max(current - 0.1, 0.01))
    
    def on_ylim_auto_toggle(self):
        """Toggle auto Y-axis scaling"""
        if self.btn_ylim_auto.isChecked():
            self.ylim_input.setEnabled(False)
        else:
            self.ylim_input.setEnabled(True)
        self.update_plot()
    
    def on_ylim_manual_change(self):
        """Handle manual Y-axis limit change"""
        self.update_plot()
    
    def _apply_model_predictions(self):
        """Apply the selected model's predictions to the display.

        Manual events are always preserved across model switches. In addition,
        user edits to the auto predictions (delete / move / merge etc.) are
        preserved *per model*: switching Model A -> B -> A restores the edits
        previously made while Model A was active instead of wiping back to the
        fresh baseline."""
        saved_manual = list(self.manual_events)

        if not hasattr(self, '_model_state_cache'):
            self._model_state_cache = {}
        cached = self._model_state_cache.get(self.model_type)

        if self.model_type == "unet_only":
            self.token_predictions = np.array([])
            self.merged_swd = []
            if cached:
                self.refined_swd = [dict(ev) for ev in cached.get('refined_swd', [])]
                self.refined_swd_base = [dict(ev) for ev in cached.get('refined_swd_base', [])]
            else:
                self.refined_swd = list(self.unet_refined_swd)
                self.refined_swd_base = list(self.unet_refined_swd)

            self.cb_show_unet.setChecked(True)
            self.cb_unet_refine.setEnabled(False)

        elif self.model_type == "cwt_only":
            self.token_predictions = self.cwt_token_predictions
            self.merged_swd = list(self.cwt_merged_swd)
            if cached:
                self.refined_swd = [dict(ev) for ev in cached.get('refined_swd', [])]
                self.refined_swd_base = [dict(ev) for ev in cached.get('refined_swd_base', [])]
            else:
                self.refined_swd = list(self.cwt_refined_swd)
                self.refined_swd_base = list(self.cwt_refined_swd)

            # Disable UNET options
            self.cb_show_unet.setChecked(False)
            self.cb_unet_refine.setEnabled(False)
            self.cb_unet_refine.setChecked(False)

        elif self.model_type == "cwt_unet":
            # Use CWT predictions with UNET refinement
            self.token_predictions = self.cwt_token_predictions
            self.merged_swd = list(self.cwt_merged_swd)

            # Don't show raw UNET predictions, only refined ones
            self.cb_show_unet.setChecked(False)
            self.cb_unet_refine.setEnabled(True)

            if cached:
                # Restore the exact post-edit state; DO NOT rerun refinement,
                # which would clobber the user's edits with the raw refinement
                # of the baseline.
                self.refined_swd = [dict(ev) for ev in cached.get('refined_swd', [])]
                self.refined_swd_base = [dict(ev) for ev in cached.get('refined_swd_base', [])]
                # Preserve the previously-chosen refinement checkbox state
                self.cb_unet_refine.setChecked(bool(cached.get('unet_refine_checked', True)))
            else:
                self.refined_swd = list(self.cwt_refined_swd)
                self.refined_swd_base = list(self.cwt_refined_swd)
                self.cb_unet_refine.setChecked(True)  # Auto-check for refinement
                # Apply UNET refinement from the fresh baseline
                self.apply_unet_refinement_if_enabled()

        # Restore manual events (preserved across model switches)
        self.manual_events = saved_manual

        # Update display and theme accents
        self.update_combined_events()
        self.update_filtered_events()
        # Update minimap accent to current auto color
        try:
            if hasattr(self, 'timeline_minimap'):
                auto_hex, _ = self._current_model_event_colors()
                self.timeline_minimap.set_auto_color(auto_hex)
        except Exception:
            pass
        self.update_plot()
    
    
    def on_model_selected(self, model_type):
        """Handle model selection.

        Preprocessing is now on-demand: only the pipeline(s) the chosen model
        needs are computed, and only the first time that model is picked.
          - cwt_only  -> CWT pipeline
          - unet_only -> UNET pipeline
          - cwt_unet  -> both (combined)
        Results are cached per pipeline, so switching back to a previously
        loaded model is instant (no recompute)."""
        if not getattr(self, 'edf_loaded', False):
            return

        # Ignore clicks while a load is already in flight.
        if getattr(self, '_pending_model_type', None) is not None:
            self._sync_model_button_checks()
            return

        needs_cwt = model_type in ("cwt_only", "cwt_unet")
        needs_unet = model_type in ("unet_only", "cwt_unet")

        # Verify the backing networks are actually available.
        if needs_cwt and self.model is None:
            QMessageBox.warning(self, "Model Unavailable", "The CWT model could not be loaded.")
            self._sync_model_button_checks()
            return
        if needs_unet and getattr(self, 'unet_model_preloaded', None) is None:
            QMessageBox.warning(self, "Model Unavailable", "The UNET model could not be loaded.")
            self._sync_model_button_checks()
            return

        # Before switching, persist the currently-active model's edited auto
        # predictions so we can restore them if the user comes back. Only cache
        # if a model was actually selected (i.e. this is a *switch*, not the
        # initial selection).
        prev_model_type = getattr(self, 'model_type', None)
        if getattr(self, 'model_chosen', False) and prev_model_type and prev_model_type != model_type:
            if not hasattr(self, '_model_state_cache'):
                self._model_state_cache = {}
            self._model_state_cache[prev_model_type] = {
                'refined_swd': [dict(ev) for ev in getattr(self, 'refined_swd', [])],
                'refined_swd_base': [dict(ev) for ev in getattr(self, 'refined_swd_base', []) or []],
                'unet_refine_checked': bool(self.cb_unet_refine.isChecked()) if hasattr(self, 'cb_unet_refine') else False,
            }

        # If every pipeline this model needs is already computed, switch instantly.
        cwt_ready = (not needs_cwt) or getattr(self, '_cwt_succeeded', False)
        unet_ready = (not needs_unet) or getattr(self, '_unet_succeeded', False)
        if cwt_ready and unet_ready:
            self._finalize_model_selection(model_type)
            return

        # Otherwise, load the missing pipeline(s) on demand.
        self._pending_model_type = model_type
        self._set_model_buttons_enabled(False)
        try:
            self.status_widget.show()
            self.progress_bar.show()
            self.progress_bar.setValue(0)
            if hasattr(self, 'progress_label'):
                label = {"cwt_only": "CWT", "unet_only": "UNET",
                         "cwt_unet": "CWT + UNET"}.get(model_type, "model")
                self.progress_label.setText(f"Loading {label} model pipeline...")
        except Exception:
            pass
        QApplication.processEvents()

        # Start CWT if required and not already computed.
        if needs_cwt and not getattr(self, '_cwt_succeeded', False):
            self._cwt_done = False
            QTimer.singleShot(50, self._start_cwt_preprocessing)
        else:
            self._cwt_done = True

        # Start UNET if required and not already computed.
        if needs_unet and not getattr(self, '_unet_succeeded', False):
            self._unet_done = False
            QTimer.singleShot(50, self._start_unet_preprocessing)
        else:
            self._unet_done = True

        # In case both were already done (shouldn't happen given the check above),
        # still advance the pending flow.
        QTimer.singleShot(60, self._maybe_finalize_pending)

    def _sync_model_button_checks(self):
        """Reset the checked state of model buttons to match the active model."""
        mt = getattr(self, 'model_type', None)
        try:
            self.btn_model_unet.setChecked(mt == "unet_only")
            self.btn_model_cwt_unet.setChecked(mt == "cwt_unet")
            self.btn_model_cwt.setChecked(mt == "cwt_only")
        except Exception:
            pass

    def _finalize_model_selection(self, model_type):
        """Switch the display to a model whose pipeline data is ready."""
        # Update model type
        self.model_type = model_type

        # Tidy the loading UI.
        try:
            self.progress_bar.hide()
            if hasattr(self, 'progress_label'):
                self.progress_label.setText("")
            if hasattr(self, 'status_widget'):
                self.status_widget.hide()
        except Exception:
            pass

        # Update button states (checkable toggle) and re-enable them.
        self._set_model_buttons_enabled(True)
        self._sync_model_button_checks()

        # Enable Settings button
        self.btn_settings.setEnabled(True)

        # Apply the selected model's predictions
        self._apply_model_predictions()
        self.model_chosen = True

        # Update theme accents based on model selection
        try:
            auto_hex, accent = self._current_model_event_colors()
            if hasattr(self, 'timeline_minimap'):
                self.timeline_minimap.set_auto_color(auto_hex)
            # Update slider handle color to model accent
            self._apply_slider_color(accent)
            # Update button checked colors by re-applying stylesheet
            if self.dark_mode:
                # Re-trigger dark mode to refresh colors
                self.on_dark_mode_changed(True)
            else:
                # Re-apply light mode stylesheet with new accent
                self.apply_modern_stylesheet()
        except Exception:
            pass

    def _cleanup_unet_thread(self, *_args):
        try:
            if hasattr(self, 'unet_thread') and self.unet_thread is not None:
                self.unet_thread.quit()
                self.unet_thread.wait()
        except Exception:
            pass
        finally:
            self.unet_thread = None
            self.unet_worker = None
            try:
                self.btn_unet.setEnabled(True)
            except Exception:
                pass

    def _on_unet_error(self, msg):
        QMessageBox.critical(self, "UNET Error", f"UNET prediction failed: {msg}")
        self.progress_bar.hide()

    def _on_unet_finished(self, result):
        # Update predictions/state
        self.unet_predictions = result['predictions']
        self.unet_confidence_scores = result.get('confidence')
        self.unet_intervals = result['intervals']

        self.progress_bar.hide()
        # Do not auto-enable overlays unless a UNET-capable model is active
        self.cb_show_unet.setChecked(self.model_chosen and (self.model_type in ["unet_only", "cwt_unet"]))
        # Enable UNET refinement toggle and apply if selected
        self.cb_unet_refine.setEnabled(True)

        if hasattr(self, 'unet_status_lbl'):
            self.unet_status_lbl.setText(f"UNET refinement: Ready ({len(self.unet_intervals)} intervals) - Channel: {getattr(self, 'unet_channel_name', 'Channel 1')}")
            self.unet_status_lbl.setStyleSheet("color: blue; font-size: 9pt;")

        # For UNET-only mode, convert UNET intervals to standard events
        if self.model_type == "unet_only":
            scale = self.sfreq / self.unet_fs
            self.refined_swd = []
            for s_idx, e_idx in self.unet_intervals:
                s_sample = int(round(s_idx * scale))
                e_sample = int(round((e_idx + 1) * scale))
                duration = (e_sample - s_sample) / self.sfreq
                
                event = {
                    'start_sample': s_sample,
                    'end_sample': e_sample,
                    'duration_sec': duration,
                    'is_auto': True
                }
                self.refined_swd.append(event)
            
            self.refined_swd_base = list(self.refined_swd)
            self.update_combined_events()
            self.update_filtered_events()
            self.progress_bar.hide()

        try:
            self.apply_unet_refinement_if_enabled()
        except Exception:
            pass
        self.update_plot()

    def _recalculate_unet_intervals(self):
        """Recalculate UNET intervals with current settings"""
        if self.unet_predictions is None:
            return
            
        try:
            # Use current settings for postprocessing
            self.unet_intervals = unet_postprocess_predictions(
                self.unet_predictions, 
                fs_hz=self.unet_fs,
                min_duration=self.unet_min_duration,
                gap_threshold=self.unet_gap_threshold
            )
            print(f"UNET intervals recalculated with new settings: {len(self.unet_intervals)} intervals")
        except Exception as e:
            print(f"Warning: Failed to recalculate UNET intervals: {e}")
            # Fall back to default postprocessing
            self.unet_intervals = unet_postprocess_predictions(
                self.unet_predictions, 
                fs_hz=self.unet_fs,
                min_duration=self.unet_min_duration,
                gap_threshold=self.unet_gap_threshold
            )







class UNETSettingsDialog(QDialog):
    """Dialog for UNET prediction settings and thresholds"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("UNET Settings")
        self.setGeometry(100, 100, 500, 400)
        self.parent = parent
        try:
            if getattr(parent, 'dark_mode', False):
                parent._apply_win_dark_titlebar(self.winId())
        except Exception:
            pass
        
        self.init_ui()
        self.load_settings()

    def load_settings(self):
        """Load initial values from parent settings"""
        if self.parent and hasattr(self.parent, 'settings'):
            s = self.parent.settings
            if 'unet_min_duration' in s:
                self.min_duration_spin.setValue(float(s['unet_min_duration']))
            if 'unet_gap_threshold' in s:
                self.gap_threshold_spin.setValue(float(s['unet_gap_threshold']))
            # Add other settings as needed

    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # UNET Prediction Settings
        pred_group = QGroupBox("UNET Prediction Settings")
        pred_layout = QVBoxLayout()
        
        # Prediction threshold for plotting
        thresh_layout = QHBoxLayout()
        thresh_layout.addWidget(QLabel("Prediction Threshold:"))
        self.pred_threshold_spin = QDoubleSpinBox()
        self.pred_threshold_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.pred_threshold_spin.setRange(0.1, 0.9)
        self.pred_threshold_spin.setSingleStep(0.05)
        self.pred_threshold_spin.setValue(0.5)
        self.pred_threshold_spin.setToolTip("Threshold for UNET predictions in plotting (0.5 = default)")
        thresh_layout.addWidget(self.pred_threshold_spin)
        pred_layout.addLayout(thresh_layout)
        
        # Postprocessing settings
        post_layout = QHBoxLayout()
        post_layout.addWidget(QLabel("Min Duration (s):"))
        self.min_duration_spin = QDoubleSpinBox()
        self.min_duration_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.min_duration_spin.setRange(0.1, 10.0)
        self.min_duration_spin.setSingleStep(0.1)
        self.min_duration_spin.setValue(1.0)
        self.min_duration_spin.setToolTip("Minimum duration for UNET intervals after postprocessing")
        post_layout.addWidget(self.min_duration_spin)
        
        post_layout.addWidget(QLabel("Gap Threshold (s):"))
        self.gap_threshold_spin = QDoubleSpinBox()
        self.gap_threshold_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.gap_threshold_spin.setRange(0.1, 5.0)
        self.gap_threshold_spin.setSingleStep(0.1)
        self.gap_threshold_spin.setValue(0.75)
        self.gap_threshold_spin.setToolTip("Gap threshold for merging UNET intervals")
        post_layout.addWidget(self.gap_threshold_spin)
        pred_layout.addLayout(post_layout)
        
        pred_group.setLayout(pred_layout)
        layout.addWidget(pred_group)
        
        # UNET Border Refinement Settings
        refine_group = QGroupBox("UNET Border Refinement Settings")
        refine_layout = QVBoxLayout()
        
        # Refinement threshold
        ref_thresh_layout = QHBoxLayout()
        ref_thresh_layout.addWidget(QLabel("Refinement Threshold:"))
        self.refine_threshold_spin = QDoubleSpinBox()
        self.refine_threshold_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.refine_threshold_spin.setRange(0.1, 0.9)
        self.refine_threshold_spin.setSingleStep(0.05)
        self.refine_threshold_spin.setValue(0.5)
        self.refine_threshold_spin.setToolTip("Threshold for UNET predictions during border refinement (visual only)")
        ref_thresh_layout.addWidget(self.refine_threshold_spin)
        refine_layout.addLayout(ref_thresh_layout)
        
        # Minimum overlap requirement
        overlap_layout = QHBoxLayout()
        overlap_layout.addWidget(QLabel("Min Overlap (%):"))
        self.min_overlap_spin = QSpinBox()
        self.min_overlap_spin.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.min_overlap_spin.setRange(10, 90)
        self.min_overlap_spin.setValue(50)
        self.min_overlap_spin.setSuffix("%")
        self.min_overlap_spin.setToolTip("Minimum overlap percentage (currently not used in code)")
        overlap_layout.addWidget(self.min_overlap_spin)
        refine_layout.addLayout(overlap_layout)
        

        
        refine_group.setLayout(refine_layout)
        layout.addWidget(refine_group)
        
        # Advanced Settings
        adv_group = QGroupBox("Advanced Settings")
        adv_layout = QVBoxLayout()
        
        # Channel selection for refinement
        ch_layout = QHBoxLayout()
        ch_layout.addWidget(QLabel("Refinement Channel:"))
        self.refine_channel_combo = QComboBox()
        self.refine_channel_combo.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.refine_channel_combo.addItems(["Channel 1", "Channel 2"])
        self.refine_channel_combo.setCurrentText("Channel 1")
        self.refine_channel_combo.setToolTip("Which channel to use for UNET border refinement (must match the channel used for UNET prediction)")
        ch_layout.addWidget(self.refine_channel_combo)
        adv_layout.addLayout(ch_layout)
        
        # UNET preprocessing channel display
        preprocess_layout = QHBoxLayout()
        preprocess_layout.addWidget(QLabel("UNET Preprocessing Channel:"))
        self.preprocess_channel_lbl = QLabel("Not set")
        self.preprocess_channel_lbl.setStyleSheet("color: gray; font-style: italic;")
        self.preprocess_channel_lbl.setToolTip("Shows which channel was used for UNET preprocessing (set when you run UNET predictions)")
        preprocess_layout.addWidget(self.preprocess_channel_lbl)
        adv_layout.addLayout(preprocess_layout)
        

        
        adv_group.setLayout(adv_layout)
        layout.addWidget(adv_group)
        
        # Buttons
        btn_layout = QHBoxLayout()
        self.apply_btn = QPushButton("Apply Settings")
        self.apply_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.reset_btn = QPushButton("Reset to Defaults")
        self.reset_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.close_btn = QPushButton("Close")
        self.close_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        
        btn_layout.addWidget(self.apply_btn)
        btn_layout.addWidget(self.reset_btn)
        btn_layout.addWidget(self.close_btn)
        layout.addLayout(btn_layout)
        
        # Connect signals
        self.apply_btn.clicked.connect(self.apply_settings)
        self.reset_btn.clicked.connect(self.reset_to_defaults)
        self.close_btn.clicked.connect(self.close)
        
        # Load current settings from parent if available
        self.load_current_settings()
        
        # Update preprocessing channel display if available
        if hasattr(self.parent, 'unet_channel_name'):
            self.preprocess_channel_lbl.setText(self.parent.unet_channel_name)
            self.preprocess_channel_lbl.setStyleSheet("color: blue; font-weight: bold;")
        else:
            self.preprocess_channel_lbl.setText("Not set")
            self.preprocess_channel_lbl.setStyleSheet("color: gray; font-style: italic;")
        
    def load_current_settings(self):
        """Load current UNET settings from parent widget"""
        if hasattr(self.parent, 'unet_prediction_threshold'):
            self.pred_threshold_spin.setValue(self.parent.unet_prediction_threshold)
        # kept visual-only; no backing attribute needed
        if hasattr(self.parent, 'unet_min_duration'):
            self.min_duration_spin.setValue(self.parent.unet_min_duration)
        if hasattr(self.parent, 'unet_gap_threshold'):
            self.gap_threshold_spin.setValue(self.parent.unet_gap_threshold)
        if hasattr(self.parent, 'unet_refinement_channel'):
            ch_map = {0: 0, 1: 1, 2: 2}
            if self.parent.unet_refinement_channel in ch_map:
                self.refine_channel_combo.setCurrentIndex(ch_map[self.parent.unet_refinement_channel])
            
    def apply_settings(self):
        """Apply the current settings to the parent widget"""
        if not self.parent:
            return
            
        # Store settings in parent
        self.parent.unet_prediction_threshold = self.pred_threshold_spin.value()
        self.parent.unet_min_duration = self.min_duration_spin.value()
        self.parent.unet_gap_threshold = self.gap_threshold_spin.value()
        self.parent.unet_refinement_channel = self.refine_channel_combo.currentIndex()
        
        # Apply settings immediately if UNET predictions exist
        if hasattr(self.parent, 'unet_predictions') and self.parent.unet_predictions is not None:
            # Recalculate UNET intervals with new settings
            self.parent._recalculate_unet_intervals()
            
            # Apply refinement if enabled
            if hasattr(self.parent, 'cb_unet_refine') and self.parent.cb_unet_refine.isChecked():
                self.parent.apply_unet_refinement_if_enabled()
                self.parent.update_plot()
        
        # Update preprocessing channel display
        if hasattr(self.parent, 'unet_channel_name'):
            self.preprocess_channel_lbl.setText(self.parent.unet_channel_name)
            self.preprocess_channel_lbl.setStyleSheet("color: blue; font-weight: bold;")
        else:
            self.preprocess_channel_lbl.setText("Not set")
            self.preprocess_channel_lbl.setStyleSheet("color: gray; font-style: italic;")
        
        QMessageBox.information(self, "Settings Applied", 
                              "UNET settings have been applied successfully!")
        
    def reset_to_defaults(self):
        """Reset all settings to default values"""
        self.pred_threshold_spin.setValue(0.5)
        self.refine_threshold_spin.setValue(0.5)
        self.min_duration_spin.setValue(1.0)
        self.gap_threshold_spin.setValue(0.75)
        self.min_overlap_spin.setValue(50)
        self.refine_channel_combo.setCurrentText("Channel 1")
        
        # Reset preprocessing channel display
        self.preprocess_channel_lbl.setText("Not set")
        self.preprocess_channel_lbl.setStyleSheet("color: gray; font-style: italic;")
        
        QMessageBox.information(self, "Defaults Reset", "All settings have been reset to default values.")

class ComprehensiveSettingsDialog(QDialog):
    """Comprehensive settings dialog with tabs for better organization"""
    
    def __init__(self, parent=None, model_type="cwt_only"):
        super().__init__(parent)
        self.setWindowTitle("⚙️ Settings")
        self.setGeometry(100, 100, 800, 650)
        self.parent = parent
        self.model_type = model_type
        try:
            if getattr(parent, 'dark_mode', False):
                parent._apply_win_dark_titlebar(self.winId())
        except Exception:
            pass
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # Create tabbed interface for settings
        self.tab_widget = QTabWidget()
        
        # Create tabs
        self._create_general_tab()
        self._create_analysis_tab()
        self._create_display_tab()
        
        layout.addWidget(self.tab_widget)
        
        # Buttons at the bottom
        btn_layout = QHBoxLayout()
        self.apply_btn = QPushButton("✓ Apply")
        self.apply_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.apply_btn.setMinimumHeight(35)
        self.reset_btn = QPushButton("Reset to Defaults")
        self.reset_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.reset_btn.setMinimumHeight(35)
        self.close_btn = QPushButton("Close")
        self.close_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.close_btn.setMinimumHeight(35)
        
        btn_layout.addWidget(self.apply_btn)
        btn_layout.addWidget(self.reset_btn)
        btn_layout.addStretch()
        btn_layout.addWidget(self.close_btn)
        layout.addLayout(btn_layout)
        
        self.apply_btn.clicked.connect(self.apply_all_settings)
        self.reset_btn.clicked.connect(self.reset_all_to_defaults)
        self.close_btn.clicked.connect(self.close)
    
    def _create_general_tab(self):
        """Create the General settings tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        # (Dark mode moved to toolbar; remove from settings)
        
        # Edit Mode toggle
        edit_group = QGroupBox("✏️ Edit Mode")
        edit_layout = QVBoxLayout()
        edit_info = QLabel("Editing is enabled by default. You can disable editing here:")
        edit_info.setStyleSheet("color: #6B7280; margin-bottom: 10px;")
        edit_layout.addWidget(edit_info)
        self.cb_disable_edit = QCheckBox("Disable editing (lock events)")
        self.cb_disable_edit.setChecked(not self.parent.edit_mode_active if self.parent else False)
        edit_layout.addWidget(self.cb_disable_edit)
        edit_group.setLayout(edit_layout)
        layout.addWidget(edit_group)

        # Reanalysis
        reanalysis_group = QGroupBox("🔄 Reanalysis")
        reanalysis_layout = QVBoxLayout()
        reanalysis_info = QLabel("Reanalyze the data with current settings:")
        reanalysis_info.setStyleSheet("color: #6B7280; margin-bottom: 10px;")
        reanalysis_layout.addWidget(reanalysis_info)
        self.btn_reanalyze = QPushButton("🔄 Reanalyze Data")
        self.btn_reanalyze.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_reanalyze.setMinimumHeight(40)
        self.btn_reanalyze.clicked.connect(self.on_reanalyze_clicked)
        reanalysis_layout.addWidget(self.btn_reanalyze)
        reanalysis_group.setLayout(reanalysis_layout)
        layout.addWidget(reanalysis_group)
        
        # UNET Channel Selection
        if self.parent and hasattr(self.parent, 'channel_names'):
            channel_group = QGroupBox("🔬 UNET Channel Selection")
            channel_layout = QVBoxLayout()
            channel_info = QLabel("Select which channel to use for UNET processing:")
            channel_info.setStyleSheet("color: #6B7280; margin-bottom: 10px;")
            channel_layout.addWidget(channel_info)
            
            ch_select_layout = QHBoxLayout()
            ch_select_layout.addWidget(QLabel("UNET Channel:"))
            self.unet_channel_combo = QComboBox()
            self.unet_channel_combo.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
            self.unet_channel_combo.addItems(self.parent.channel_names[:2])
            if hasattr(self.parent, 'unet_channel'):
                self.unet_channel_combo.setCurrentIndex(self.parent.unet_channel)
            ch_select_layout.addWidget(self.unet_channel_combo)
            ch_select_layout.addStretch()
            channel_layout.addLayout(ch_select_layout)
            
            channel_group.setLayout(channel_layout)
            layout.addWidget(channel_group)
        
        layout.addStretch()
        self.tab_widget.addTab(tab, "🏠 General")
    
    def _create_analysis_tab(self):
        """Create the Analysis parameters tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        # Add advanced settings
        self.init_advanced_settings(layout)
        
        # Plot Filter Settings
        filter_group = QGroupBox("EEG Plot Bandpass Filter (plot-only)")
        fl = QVBoxLayout()
        row = QHBoxLayout()
        self.cb_filter_plot_tab = QCheckBox("Enable bandpass filter on plotted signal")
        self.cb_filter_plot_tab.setChecked(False)
        row.addWidget(self.cb_filter_plot_tab)
        row.addStretch()
        fl.addLayout(row)
        row2 = QHBoxLayout()
        low_label = QLabel("Low cut (Hz):")
        row2.addWidget(low_label)
        self.filter_low_tab = QDoubleSpinBox()
        self.filter_low_tab.setRange(0.0, 2000.0)
        self.filter_low_tab.setDecimals(2)
        self.filter_low_tab.setSingleStep(0.5)
        self.filter_low_tab.setValue(self.parent.filter_low_cut_hz if self.parent else 1.0)
        row2.addWidget(self.filter_low_tab)
        high_label = QLabel("High cut (Hz):")
        row2.addWidget(high_label)
        self.filter_high_tab = QDoubleSpinBox()
        self.filter_high_tab.setRange(0.1, 2000.0)
        self.filter_high_tab.setDecimals(2)
        self.filter_high_tab.setSingleStep(0.5)
        self.filter_high_tab.setValue(self.parent.filter_high_cut_hz if self.parent else 40.0)
        row2.addWidget(self.filter_high_tab)
        row2.addStretch()
        fl.addLayout(row2)
        filter_group.setLayout(fl)
        layout.addWidget(filter_group)
        
        # Enable/disable frequency inputs based on checkbox
        # Also dim labels to match
        def _toggle_filter_inputs(state):
            enabled = bool(state)
            low_label.setEnabled(enabled)
            high_label.setEnabled(enabled)
            self.filter_low_tab.setEnabled(enabled)
            self.filter_high_tab.setEnabled(enabled)
        self.cb_filter_plot_tab.stateChanged.connect(_toggle_filter_inputs)
        _toggle_filter_inputs(self.cb_filter_plot_tab.isChecked())
        
        layout.addStretch()
        self.tab_widget.addTab(tab, "🔬 Analysis")
    
    def _create_display_tab(self):
        """Create the Display options tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        display_group = QGroupBox("📊 Display Options")
        display_layout = QVBoxLayout()
        
        self.cb_show_unrefined = QCheckBox("Show Unrefined Events")
        self.cb_show_unrefined.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_show_unrefined.setChecked(self.parent.show_unrefined if self.parent else False)
        display_layout.addWidget(self.cb_show_unrefined)
        
        self.cb_show_unet = QCheckBox("Show UNET Predictions Overlay")
        self.cb_show_unet.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_show_unet.setChecked(self.parent.cb_show_unet.isChecked() if self.parent else True)
        display_layout.addWidget(self.cb_show_unet)
        
        self.cb_unet_post = QCheckBox("Apply UNET Postprocessing")
        self.cb_unet_post.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.cb_unet_post.setChecked(self.parent.cb_unet_post.isChecked() if self.parent else True)
        display_layout.addWidget(self.cb_unet_post)
        
        display_group.setLayout(display_layout)
        layout.addWidget(display_group)
        
        layout.addStretch()
        self.tab_widget.addTab(tab, "👁 Display")
    
    def _create_keyboard_tab(self):
        """Create the Keyboard shortcuts tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(15)
        
        shortcuts_group = QGroupBox("⌨ Keyboard Shortcuts")
        shortcuts_layout = QVBoxLayout()
        
        shortcuts_info = QLabel("Configure keyboard shortcuts for common actions:")
        shortcuts_info.setStyleSheet("color: #6B7280; margin-bottom: 10px;")
        shortcuts_layout.addWidget(shortcuts_info)
        
        self.btn_edit_shortcuts = QPushButton("✏️ Edit Keyboard Shortcuts")
        self.btn_edit_shortcuts.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_edit_shortcuts.setMinimumHeight(40)
        self.btn_edit_shortcuts.clicked.connect(self.on_edit_shortcuts_clicked)
        shortcuts_layout.addWidget(self.btn_edit_shortcuts)
        
        shortcuts_group.setLayout(shortcuts_layout)
        layout.addWidget(shortcuts_group)
        
        layout.addStretch()
        self.tab_widget.addTab(tab, "⌨ Shortcuts")
    
    def on_reanalyze_clicked(self):
        """Handle reanalyze button click"""
        if self.parent and hasattr(self.parent, 'on_reanalyze_clicked'):
            reply = QMessageBox.question(
                self, "Confirm Reanalysis",
                "This will reanalyze the data with current settings. Continue?",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                try:
                    self.parent.on_reanalyze_clicked()
                    if hasattr(self.parent, 'show_toast'):
                        self.parent.show_toast("Data reanalyzed successfully", "success")
                except Exception:
                    pass
    
    def on_edit_shortcuts_clicked(self):
        """Open keyboard shortcuts editor"""
        if self.parent:
            self.parent.show_keyboard_shortcuts()
    
    def reset_all_to_defaults(self):
        """Reset all settings to default values"""
        reply = QMessageBox.question(
            self, "Confirm Reset",
            "Reset all settings to defaults?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            # Reset checkboxes
            self.cb_dark_mode.setChecked(False)
            if hasattr(self, 'cb_show_unrefined'):
                self.cb_show_unrefined.setChecked(False)
            if hasattr(self, 'cb_show_unet'):
                self.cb_show_unet.setChecked(True)
            if hasattr(self, 'cb_unet_post'):
                self.cb_unet_post.setChecked(True)
            
            # Reset spinboxes
            if hasattr(self, 'cwt_threshold'):
                self.cwt_threshold.setValue(0.5)
            if hasattr(self, 'cwt_gap'):
                self.cwt_gap.setValue(4)
            if hasattr(self, 'cwt_power'):
                self.cwt_power.setValue(25)
            if hasattr(self, 'unet_min_dur'):
                self.unet_min_dur.setValue(1.0)
            if hasattr(self, 'unet_gap'):
                self.unet_gap.setValue(0.75)
            
            QMessageBox.information(self, "Reset Complete", "All settings have been reset to defaults.")
    
    def init_advanced_settings(self, layout):
        """Initialize advanced settings for all models"""
        # Always show all model settings regardless of current selection
        
        # CWT Settings
        cwt_group = QGroupBox("CWT Hybrid CNN Settings")
        cwt_layout = QVBoxLayout()
        self.init_cwt_advanced(cwt_layout)
        cwt_group.setLayout(cwt_layout)
        layout.addWidget(cwt_group)
        
        # UNET Settings
        unet_group = QGroupBox("UNET Settings")
        unet_layout = QVBoxLayout()
        self.init_unet_advanced(unet_layout)
        unet_group.setLayout(unet_layout)
        layout.addWidget(unet_group)
    
    def init_cwt_advanced(self, layout):
        """CWT-specific settings"""
        
        thresh_l = QHBoxLayout()
        thresh_l.addWidget(QLabel("Threshold:"))
        self.cwt_threshold = QDoubleSpinBox()
        self.cwt_threshold.setRange(0.1, 0.9)
        self.cwt_threshold.setSingleStep(0.05)
        self.cwt_threshold.setValue(self.parent.th_spin.value() if self.parent else 0.5)
        thresh_l.addWidget(self.cwt_threshold)
        layout.addLayout(thresh_l)
        
        gap_l = QHBoxLayout()
        gap_l.addWidget(QLabel("Max Gap:"))
        self.cwt_gap = QSpinBox()
        self.cwt_gap.setRange(1, 10)
        self.cwt_gap.setValue(self.parent.gap_spin.value() if self.parent else 4)
        gap_l.addWidget(self.cwt_gap)
        layout.addLayout(gap_l)
        
        power_l = QHBoxLayout()
        power_l.addWidget(QLabel("Power %:"))
        self.cwt_power = QSpinBox()
        self.cwt_power.setRange(1, 50)
        self.cwt_power.setValue(self.parent.power_percentile_spin.value() if self.parent else 25)
        power_l.addWidget(self.cwt_power)
        layout.addLayout(power_l)
    
    def init_unet_advanced(self, layout):
        """UNET-specific settings"""
        
        dur_l = QHBoxLayout()
        dur_l.addWidget(QLabel("Min Duration (s):"))
        self.unet_min_dur = QDoubleSpinBox()
        self.unet_min_dur.setRange(0.1, 10.0)
        self.unet_min_dur.setValue(self.parent.unet_min_duration if self.parent else 1.0)
        dur_l.addWidget(self.unet_min_dur)
        layout.addLayout(dur_l)
        
        gap_l = QHBoxLayout()
        gap_l.addWidget(QLabel("Gap Threshold (s):"))
        self.unet_gap = QDoubleSpinBox()
        self.unet_gap.setRange(0.1, 5.0)
        self.unet_gap.setValue(self.parent.unet_gap_threshold if self.parent else 0.75)
        gap_l.addWidget(self.unet_gap)
        layout.addLayout(gap_l)
    
    def init_general_settings(self, layout):
        """Initialize general settings"""
        # Dark Mode
        dark_group = QGroupBox("Appearance")
        dark_layout = QVBoxLayout()
        self.cb_dark_mode = QCheckBox("Dark Mode")
        self.cb_dark_mode.setChecked(self.parent.dark_mode if self.parent else False)
        dark_layout.addWidget(self.cb_dark_mode)
        dark_group.setLayout(dark_layout)
        layout.addWidget(dark_group)
        
        # Advanced Settings button
        advanced_group = QGroupBox("Model Settings")
        advanced_layout = QVBoxLayout()
        self.btn_advanced_settings = QPushButton("Advanced Settings")
        self.btn_advanced_settings.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_advanced_settings.clicked.connect(self.on_advanced_settings)
        advanced_layout.addWidget(self.btn_advanced_settings)
        advanced_group.setLayout(advanced_layout)
        layout.addWidget(advanced_group)
        
        # UNET Channel
        if self.model_type in ["unet_only", "cwt_unet"]:
            unet_ch_group = QGroupBox("UNET Channel Selection")
            unet_ch_layout = QVBoxLayout()
            ch_l = QHBoxLayout()
            ch_l.addWidget(QLabel("Channel:"))
            self.unet_channel_combo = QComboBox()
            ch_names = self.parent.channel_names if self.parent and hasattr(self.parent, 'channel_names') else ["Channel 1", "Channel 2"]
            if len(ch_names) < 2:
                ch_names = ["Channel 1", "Channel 2"]
            self.unet_channel_combo.addItems([ch_names[0], ch_names[1]])
            if self.parent and hasattr(self.parent, 'unet_channel'):
                self.unet_channel_combo.setCurrentIndex(self.parent.unet_channel)
            ch_l.addWidget(self.unet_channel_combo)
            unet_ch_layout.addLayout(ch_l)
            unet_ch_group.setLayout(unet_ch_layout)
            layout.addWidget(unet_ch_group)
        
        # Reanalyze button
        reanalyze_group = QGroupBox("Analysis")
        reanalyze_layout = QVBoxLayout()
        self.btn_reanalyze = QPushButton("Reanalyze")
        self.btn_reanalyze.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_reanalyze.clicked.connect(self.on_reanalyze)
        reanalyze_layout.addWidget(self.btn_reanalyze)
        reanalyze_group.setLayout(reanalyze_layout)
        layout.addWidget(reanalyze_group)
        
        # Key Assignments button
        key_group = QGroupBox("Keyboard Shortcuts")
        key_layout = QVBoxLayout()
        self.btn_key_assignments = QPushButton("Configure Key Assignments")
        self.btn_key_assignments.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.btn_key_assignments.clicked.connect(self.on_key_assignments)
        key_layout.addWidget(self.btn_key_assignments)
        key_group.setLayout(key_layout)
        layout.addWidget(key_group)
        
        layout.addStretch()
    
    def on_advanced_settings(self):
        """Open advanced settings dialog"""
        from PyQt5.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout, QPushButton
        
        # Create advanced settings dialog
        adv_dialog = QDialog(self)
        adv_dialog.setWindowTitle("Advanced Settings")
        adv_dialog.setGeometry(150, 150, 600, 500)
        
        layout = QVBoxLayout(adv_dialog)
        
        # Add all advanced settings
        self.init_advanced_settings(layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        apply_btn = QPushButton("Apply")
        apply_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        close_btn = QPushButton("Close")
        close_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        btn_layout.addWidget(apply_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)
        
        # Connect buttons
        apply_btn.clicked.connect(lambda: self.apply_advanced_settings(adv_dialog))
        close_btn.clicked.connect(adv_dialog.close)
        
        adv_dialog.exec_()
    
    def apply_advanced_settings(self, dialog):
        """Apply advanced settings from the dialog"""
        if not self.parent:
            return
        
        # Apply CWT settings
        if hasattr(self, 'cwt_threshold'):
            self.parent.th_spin.setValue(self.cwt_threshold.value())
        if hasattr(self, 'cwt_gap'):
            self.parent.gap_spin.setValue(self.cwt_gap.value())
        if hasattr(self, 'cwt_power'):
            self.parent.power_percentile_spin.setValue(self.cwt_power.value())
        
        # Apply UNET settings
        if hasattr(self, 'unet_min_dur'):
            self.parent.unet_min_duration = self.unet_min_dur.value()
        if hasattr(self, 'unet_gap'):
            self.parent.unet_gap_threshold = self.unet_gap.value()
        
        dialog.close()
        QMessageBox.information(self, "Settings Applied", "Advanced settings have been applied!")
    
    def on_reanalyze(self):
        """Trigger reanalysis"""
        if self.parent:
            self.parent.on_reanalyze_clicked()
            QMessageBox.information(self, "Reanalysis", "Reanalysis triggered!")
    
    def on_key_assignments(self):
        """Open key assignments dialog"""
        if self.parent:
            self.parent.show_key_assignments_dialog()
    
    def apply_all_settings(self):
        """Apply all settings"""
        if not self.parent:
            return
        
        # Apply Edit Mode toggle
        if hasattr(self, 'cb_disable_edit'):
            self.parent.set_edit_mode_enabled(not self.cb_disable_edit.isChecked())

        # Apply CWT settings
        if self.model_type in ["cwt_only", "cwt_unet"] and hasattr(self, 'cwt_threshold'):
            self.parent.th_spin.setValue(self.cwt_threshold.value())
            self.parent.gap_spin.setValue(self.cwt_gap.value())
            self.parent.power_percentile_spin.setValue(self.cwt_power.value())
        
        # Apply UNET settings
        if self.model_type in ["unet_only", "cwt_unet"] and hasattr(self, 'unet_min_dur'):
            self.parent.unet_min_duration = self.unet_min_dur.value()
            self.parent.unet_gap_threshold = self.unet_gap.value()
            
            # Update UNET channel if changed
            if hasattr(self, 'unet_channel_combo'):
                new_channel = self.unet_channel_combo.currentIndex()
                if new_channel != self.parent.unet_channel:
                    self.parent.unet_channel = new_channel
                    self.parent.unet_channel_name = self.unet_channel_combo.currentText()
                    QMessageBox.information(self, "UNET Channel Changed", 
                                          f"UNET channel changed to {self.parent.unet_channel_name}.\n\n"
                                          "Note: This will take effect on next model selection or reanalysis.")
        
        # (Export settings moved to Export popup)
        
        # Apply plot filter settings
        if hasattr(self, 'cb_filter_plot_tab'):
            self.parent.filter_signal_enabled = self.cb_filter_plot_tab.isChecked()
            self.parent.filter_low_cut_hz = float(self.filter_low_tab.value())
            self.parent.filter_high_cut_hz = float(self.filter_high_tab.value())
            if self.parent.filter_signal_enabled:
                self.parent._apply_plot_bandpass_filter()
            self.parent.update_plot()
            # Keep an open PSD dialog's bandpass controls consistent with the
            # newly applied plot filter (single shared bandpass configuration).
            psd_dialog = getattr(self.parent, 'psd_dialog', None)
            if psd_dialog is not None and hasattr(psd_dialog, '_sync_bandpass_from_parent'):
                psd_dialog._sync_bandpass_from_parent()
        
        QMessageBox.information(self, "Settings Applied", "All settings have been applied!")
        self.close()

class AdvancedSettingsDialog(QDialog):
    """Context-aware advanced settings dialog based on selected model type"""
    
    def __init__(self, parent=None, model_type="cwt_only"):
        super().__init__(parent)
        self.setWindowTitle("Advanced Settings")
        self.setGeometry(100, 100, 600, 500)
        self.parent = parent
        self.model_type = model_type
        try:
            if getattr(parent, 'dark_mode', False):
                parent._apply_win_dark_titlebar(self.winId())
        except Exception:
            pass
        self.init_ui()
        
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # Show different settings based on model type
        if self.model_type == "unet_only":
            self.init_unet_settings(layout)
        elif self.model_type == "cwt_only":
            self.init_cwt_settings(layout)
        elif self.model_type == "cwt_unet":
            self.init_cwt_settings(layout)
            self.init_unet_settings(layout)
        
        # Buttons
        btn_layout = QHBoxLayout()
        self.apply_btn = QPushButton("Apply Settings")
        self.apply_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.reset_btn = QPushButton("Reset to Defaults")
        self.reset_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.close_btn = QPushButton("Close")
        self.close_btn.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        
        btn_layout.addWidget(self.apply_btn)
        btn_layout.addWidget(self.reset_btn)
        btn_layout.addWidget(self.close_btn)
        layout.addLayout(btn_layout)
        
        # Connect signals
        self.apply_btn.clicked.connect(self.apply_settings)
        self.reset_btn.clicked.connect(self.reset_to_defaults)
        self.close_btn.clicked.connect(self.close)
        
        # Load current settings
        self.load_current_settings()
    
    def init_cwt_settings(self, layout):
        """Initialize CWT-specific settings"""
        cwt_group = QGroupBox("CWT Hybrid CNN Settings")
        cwt_layout = QVBoxLayout()
        
        # Threshold controls
        thresh_layout = QHBoxLayout()
        thresh_layout.addWidget(QLabel("Prediction Threshold:"))
        self.cwt_threshold_spin = QDoubleSpinBox()
        self.cwt_threshold_spin.setRange(0.1, 0.9)
        self.cwt_threshold_spin.setSingleStep(0.05)
        self.cwt_threshold_spin.setValue(0.5)
        self.cwt_threshold_spin.setToolTip("Threshold for CWT predictions")
        thresh_layout.addWidget(self.cwt_threshold_spin)
        cwt_layout.addLayout(thresh_layout)
        
        # Max gap
        gap_layout = QHBoxLayout()
        gap_layout.addWidget(QLabel("Max Gap (intervals):"))
        self.cwt_gap_spin = QSpinBox()
        self.cwt_gap_spin.setRange(1, 10)
        self.cwt_gap_spin.setValue(4)
        self.cwt_gap_spin.setToolTip("Maximum gap between intervals to merge")
        gap_layout.addWidget(self.cwt_gap_spin)
        cwt_layout.addLayout(gap_layout)
        
        # Power percentile
        power_layout = QHBoxLayout()
        power_layout.addWidget(QLabel("Power Percentile:"))
        self.cwt_power_spin = QSpinBox()
        self.cwt_power_spin.setRange(1, 50)
        self.cwt_power_spin.setValue(25)
        self.cwt_power_spin.setToolTip("Power threshold percentile for border refinement")
        power_layout.addWidget(self.cwt_power_spin)
        cwt_layout.addLayout(power_layout)

        # Require Overlap Agreement - both overlapping tokens must predict SWD
        # before a region is kept. Reduces false positives on the CWT pipeline;
        # previously hard-coded to True for CWT models, now user-configurable.
        self.cwt_overlap_agreement_cb = QCheckBox("Require Overlap Agreement")
        self.cwt_overlap_agreement_cb.setChecked(False)
        self.cwt_overlap_agreement_cb.setToolTip(
            "If enabled, both overlapping prediction tokens must agree (>= threshold) "
            "before an interval is accepted as SWD. Lowers false positives but may "
            "also miss short events. Recommended ON for CWT models."
        )
        cwt_layout.addWidget(self.cwt_overlap_agreement_cb)
        
        # Interval parameters
        interval_group = QGroupBox("Interval Parameters")
        interval_layout = QVBoxLayout()
        
        int_len_layout = QHBoxLayout()
        int_len_layout.addWidget(QLabel("Interval Length (s):"))
        self.interval_length_spin = QDoubleSpinBox()
        self.interval_length_spin.setRange(0.1, 2.0)
        self.interval_length_spin.setSingleStep(0.1)
        self.interval_length_spin.setValue(0.6)
        int_len_layout.addWidget(self.interval_length_spin)
        interval_layout.addLayout(int_len_layout)
        
        overlap_layout = QHBoxLayout()
        overlap_layout.addWidget(QLabel("Overlap Length (s):"))
        self.overlap_length_spin = QDoubleSpinBox()
        self.overlap_length_spin.setRange(0.0, 1.0)
        self.overlap_length_spin.setSingleStep(0.1)
        self.overlap_length_spin.setValue(0.3)
        overlap_layout.addWidget(self.overlap_length_spin)
        interval_layout.addLayout(overlap_layout)
        
        seq_layout = QHBoxLayout()
        seq_layout.addWidget(QLabel("Sequence Length:"))
        self.sequence_length_spin = QSpinBox()
        self.sequence_length_spin.setRange(10, 200)
        self.sequence_length_spin.setValue(100)
        seq_layout.addWidget(self.sequence_length_spin)
        interval_layout.addLayout(seq_layout)
        
        interval_group.setLayout(interval_layout)
        cwt_layout.addWidget(interval_group)
        
        cwt_group.setLayout(cwt_layout)
        layout.addWidget(cwt_group)
    
    def init_unet_settings(self, layout):
        """Initialize UNET-specific settings"""
        unet_group = QGroupBox("UNET Settings")
        unet_layout = QVBoxLayout()
        
        # Prediction threshold
        pred_layout = QHBoxLayout()
        pred_layout.addWidget(QLabel("Prediction Threshold:"))
        self.unet_threshold_spin = QDoubleSpinBox()
        self.unet_threshold_spin.setRange(0.1, 0.9)
        self.unet_threshold_spin.setSingleStep(0.05)
        self.unet_threshold_spin.setValue(0.5)
        self.unet_threshold_spin.setToolTip("Threshold for UNET predictions")
        pred_layout.addWidget(self.unet_threshold_spin)
        unet_layout.addLayout(pred_layout)
        
        # Min duration
        dur_layout = QHBoxLayout()
        dur_layout.addWidget(QLabel("Min Duration (s):"))
        self.unet_min_duration_spin = QDoubleSpinBox()
        self.unet_min_duration_spin.setRange(0.1, 10.0)
        self.unet_min_duration_spin.setSingleStep(0.1)
        self.unet_min_duration_spin.setValue(1.0)
        dur_layout.addWidget(self.unet_min_duration_spin)
        unet_layout.addLayout(dur_layout)
        
        # Gap threshold
        gap_layout = QHBoxLayout()
        gap_layout.addWidget(QLabel("Gap Threshold (s):"))
        self.unet_gap_spin = QDoubleSpinBox()
        self.unet_gap_spin.setRange(0.1, 5.0)
        self.unet_gap_spin.setSingleStep(0.1)
        self.unet_gap_spin.setValue(0.75)
        gap_layout.addWidget(self.unet_gap_spin)
        unet_layout.addLayout(gap_layout)
        
        unet_group.setLayout(unet_layout)
        layout.addWidget(unet_group)
    
    def load_current_settings(self):
        """Load current settings from parent widget"""
        if not self.parent:
            return
        
        # Load CWT settings
        if self.model_type in ["cwt_only", "cwt_unet"] and hasattr(self, 'cwt_threshold_spin'):
            if hasattr(self.parent, 'th_spin'):
                self.cwt_threshold_spin.setValue(self.parent.th_spin.value())
            if hasattr(self.parent, 'gap_spin'):
                self.cwt_gap_spin.setValue(self.parent.gap_spin.value())
            if hasattr(self.parent, 'power_percentile_spin'):
                self.cwt_power_spin.setValue(self.parent.power_percentile_spin.value())
            if hasattr(self, 'cwt_overlap_agreement_cb') and hasattr(self.parent, 'require_overlap_agreement'):
                self.cwt_overlap_agreement_cb.setChecked(bool(self.parent.require_overlap_agreement))
            if hasattr(self.parent, 'interval_length'):
                self.interval_length_spin.setValue(self.parent.interval_length)
            if hasattr(self.parent, 'overlap_length'):
                self.overlap_length_spin.setValue(self.parent.overlap_length)
            if hasattr(self.parent, 'sequence_length'):
                self.sequence_length_spin.setValue(self.parent.sequence_length)
        
        # Load UNET settings
        if self.model_type in ["unet_only", "cwt_unet"] and hasattr(self, 'unet_threshold_spin'):
            if hasattr(self.parent, 'unet_prediction_threshold'):
                self.unet_threshold_spin.setValue(self.parent.unet_prediction_threshold)
            if hasattr(self.parent, 'unet_min_duration'):
                self.unet_min_duration_spin.setValue(self.parent.unet_min_duration)
            if hasattr(self.parent, 'unet_gap_threshold'):
                self.unet_gap_spin.setValue(self.parent.unet_gap_threshold)
    
    def apply_settings(self):
        """Apply settings to parent widget"""
        if not self.parent:
            return
        
        # Apply CWT settings
        if self.model_type in ["cwt_only", "cwt_unet"] and hasattr(self, 'cwt_threshold_spin'):
            if hasattr(self.parent, 'th_spin'):
                self.parent.th_spin.setValue(self.cwt_threshold_spin.value())
            if hasattr(self.parent, 'gap_spin'):
                self.parent.gap_spin.setValue(self.cwt_gap_spin.value())
            if hasattr(self.parent, 'power_percentile_spin'):
                self.parent.power_percentile_spin.setValue(self.cwt_power_spin.value())
            if hasattr(self, 'cwt_overlap_agreement_cb'):
                self.parent.require_overlap_agreement = bool(self.cwt_overlap_agreement_cb.isChecked())
            # Note: interval parameters require reanalysis, so we store them but don't auto-apply
            self.parent.interval_length = self.interval_length_spin.value()
            self.parent.overlap_length = self.overlap_length_spin.value()
            self.parent.sequence_length = self.sequence_length_spin.value()
        
        # Apply UNET settings
        if self.model_type in ["unet_only", "cwt_unet"] and hasattr(self, 'unet_threshold_spin'):
            self.parent.unet_prediction_threshold = self.unet_threshold_spin.value()
            self.parent.unet_min_duration = self.unet_min_duration_spin.value()
            self.parent.unet_gap_threshold = self.unet_gap_spin.value()
            
            # Recalculate UNET intervals if they exist
            if hasattr(self.parent, 'unet_predictions') and self.parent.unet_predictions is not None:
                self.parent._recalculate_unet_intervals()
                if hasattr(self.parent, 'cb_unet_refine') and self.parent.cb_unet_refine.isChecked():
                    self.parent.apply_unet_refinement_if_enabled()
                self.parent.update_plot()
        
        QMessageBox.information(self, "Settings Applied", 
                              "Advanced settings have been applied successfully!")
    
    def reset_to_defaults(self):
        """Reset all settings to defaults"""
        if hasattr(self, 'cwt_threshold_spin'):
            self.cwt_threshold_spin.setValue(0.5)
            self.cwt_gap_spin.setValue(4)
            self.cwt_power_spin.setValue(25)
            if hasattr(self, 'cwt_overlap_agreement_cb'):
                self.cwt_overlap_agreement_cb.setChecked(True)
            self.interval_length_spin.setValue(0.6)
            self.overlap_length_spin.setValue(0.3)
            self.sequence_length_spin.setValue(100)
        
        if hasattr(self, 'unet_threshold_spin'):
            self.unet_threshold_spin.setValue(0.5)
            self.unet_min_duration_spin.setValue(1.0)
            self.unet_gap_spin.setValue(0.75)
        
        QMessageBox.information(self, "Defaults Reset", 
                              "All settings have been reset to default values.")

def main():
    """Main application entry point"""
    app = QApplication(sys.argv)
    
    # Mac OS UI Optimization
    if sys.platform == 'darwin':
        font = QFont("Helvetica", 13)  # Use standard sans-serif, slightly larger for Retina
        app.setFont(font)
        # Fix button text being cut off
        app.setStyleSheet("""
            QPushButton { padding: 6px 12px; }
            QComboBox { padding: 4px 8px; }
            QLabel { font-size: 13pt; }
        """)
    else:
        # Windows default enhancement
        font = QFont("Segoe UI", 9)
        app.setFont(font)
    
    # Pre-load models that might be needed
    cwt_model = None
    unet_model = None
    
    # Try to load CWT model (for cwt_only and cwt_unet modes)
    try:
        if getattr(sys, 'frozen', False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.dirname(os.path.abspath(__file__))

        saved_dir = os.path.join(base_path, "CWT Hybrid CNN_savedmodel")
        
        # Fallback to absolute path if not found (dev environment)
        if not os.path.exists(saved_dir) and not getattr(sys, 'frozen', False):
             # Try local relative first
             local_saved = os.path.join(base_path, "SWD_Detector_2_Portable", "CWT Hybrid CNN_savedmodel")
             if os.path.exists(local_saved):
                 saved_dir = local_saved
             else:
                 saved_dir = r"C:\Users\TEKNOSA\Desktop\Auto SWD\Model\CWT Hybrid CNN_savedmodel"

        if os.path.exists(saved_dir):
            cwt_model = SavedModelPredictor(saved_dir)
            print(f"Loaded CWT SavedModel from: {saved_dir}")
        else:
            print(f"Warning: CWT model path not found at: {saved_dir}")

    except Exception as e:
        print(f"Warning: Could not load CWT model: {e}")
    
    # Try to load UNET model (for unet_only and cwt_unet modes)
    if torch is not None:
        try:
            if getattr(sys, 'frozen', False):
                script_dir = sys._MEIPASS
            else:
                script_dir = os.path.dirname(os.path.abspath(__file__))
            
            # Look for UNET model files
            candidates = [
                os.path.join(script_dir, 'attention_resunet_BCEloss.pth'),
                os.path.join(script_dir, 'unet.ts'),
                os.path.join(script_dir, 'unet.pt'),
                os.path.join(script_dir, 'unet.pth')
            ]
            unet_path = None
            for mp in candidates:
                if os.path.isfile(mp):
                    unet_path = mp
                    break
            
            if unet_path is not None:
                # Try loading as TorchScript first
                try:
                    unet_model = torch.jit.load(unet_path, map_location='cpu')
                    unet_model.eval()
                    print(f"Loaded UNET (TorchScript) from: {unet_path}")
                except Exception:
                    # Try state_dict loading
                    model_module = None
                    local_model = os.path.join(script_dir, 'model.py')
                    if os.path.exists(local_model):
                        import importlib.util
                        spec = importlib.util.spec_from_file_location('model', local_model)
                        model_module = importlib.util.module_from_spec(spec)
                        spec.loader.exec_module(model_module)
                    else:
                        model_dir = os.path.join(script_dir, 'Model')
                        if os.path.isdir(model_dir):
                            sys.path.insert(0, model_dir)
                            try:
                                model_module = importlib.import_module('model')
                            except Exception:
                                model_module = None
                    
                    if model_module is not None:
                        NetClass = getattr(model_module, 'AttentionResUNet', None)
                        if NetClass is not None:
                            net = NetClass(n_classes=1, in_channels=1, out_channels=1)
                            state_dict = torch.load(unet_path, map_location='cpu')
                            net.load_state_dict(state_dict)
                            net.eval()
                            unet_model = net
                            print(f"Loaded UNET (state_dict) from: {unet_path}")
        except Exception as e:
            print(f"Warning: Could not load UNET model: {e}")
    
    # Create plotter without EDF data - user will load via File menu
    plotter = EnhancedEEGPlotter(
        model=cwt_model,
        eeg0=None,  # No data initially
        eeg1=None,
        sfreq=None,
        recording_start=None,
        interval_length=0.6,
        overlap_length=0.3,
        sequence_length=100,
        spectral_threshold=0,
        channel_names=None,
        edf_filename=None,
        model_type="",  # Empty string - model selection happens in UI
        unet_model=unet_model,
        unet_channel=0
    )
    
    # Preload models
    if cwt_model is not None:
        plotter.cwt_model_preloaded = cwt_model
        plotter.model = cwt_model
    if unet_model is not None:
        plotter.unet_model_preloaded = unet_model
        plotter.unet_model = unet_model
    
    plotter.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main() 
